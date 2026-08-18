from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = SKILL_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import repair_fulfillment_outputs as repair


def jsonl_rows(path: Path) -> list[dict]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def product(
    asin: str,
    *,
    method: str = "",
    raw: str = "",
    has_rank: bool = True,
    page_number: int = 1,
    rank: int = 1,
) -> dict:
    return {
        "source_type": "storefront",
        "store_url": "https://www.amazon.com/s?me=TEST",
        "store_name": "test-store",
        "store_sort_order": "Newest Arrivals",
        "page_url": f"https://www.amazon.com/s?me=TEST&page={page_number}",
        "page_number": page_number,
        "rank": str(rank),
        "asin": asin,
        "title": f"Product {asin}",
        "product_url": f"https://www.amazon.com/dp/{asin}",
        "fulfillment_method": method,
        "fulfillment_method_raw": raw,
        "subcategory_bsr_ranks": (
            [{"rank": 130, "category_name": "Fruit Bowls"}] if has_rank else []
        ),
        "load_status": "ok",
    }


class HistoricalFulfillmentRepairTests(unittest.TestCase):
    def write_source(self, root: Path, records: list[dict]) -> tuple[Path, Path]:
        source = root / "old-job"
        source.mkdir()
        records_path = source / "records.jsonl"
        records_path.write_text(
            "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in records),
            encoding="utf-8",
        )
        return source, records_path

    def test_sidecar_repair_generates_five_artifacts_and_keeps_unknown_raw(self) -> None:
        records = [
            product("B000000001", raw="FBM卖家", rank=1),
            product("B000000002", raw="FBA卖家", rank=2),
            product("B000000003", raw="", rank=3),
            product("B000000004", raw="SFP", rank=4),
            product("B000000005", raw="FBMPlus", rank=5),
            product("B000000006", raw="FBM", rank=6),
            product("B000000007", method="AMZ", raw="AMZ", rank=7),
            product("B000000008", raw="FBM卖家", has_rank=False, rank=8),
            product("B000000001", raw="FBM卖家", page_number=2, rank=1),
            product("B000000009", raw="FBA Fee", rank=9),
            product("B000000010", raw="non-FBA", rank=10),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source, records_path = self.write_source(root, records)
            source_hash = file_sha256(records_path)
            source_text = records_path.read_text(encoding="utf-8")
            output = root / "old-job-repaired"

            report = repair.repair_job_outputs(
                source,
                output,
                expected_record_count=11,
                expected_unique_asin_count=10,
                expected_filtered_asin_count=7,
            )

            self.assertEqual(file_sha256(records_path), source_hash)
            self.assertEqual(records_path.read_text(encoding="utf-8"), source_text)
            self.assertEqual(
                sorted(path.name for path in output.iterdir()),
                sorted(repair.ARTIFACT_NAMES),
            )

            repaired_rows = jsonl_rows(output / repair.REPAIRED_RECORDS_NAME)
            by_asin = {row["asin"]: row for row in repaired_rows}
            self.assertEqual(by_asin["B000000001"]["fulfillment_method"], "FBM")
            self.assertEqual(by_asin["B000000002"]["fulfillment_method"], "FBA")
            self.assertEqual(by_asin["B000000003"]["fulfillment_method"], "")
            self.assertEqual(by_asin["B000000004"]["fulfillment_method"], "")
            self.assertEqual(by_asin["B000000004"]["fulfillment_method_raw"], "SFP")
            self.assertEqual(by_asin["B000000005"]["fulfillment_method"], "FBM")
            self.assertEqual(by_asin["B000000005"]["fulfillment_method_raw"], "FBMPlus")
            self.assertEqual(by_asin["B000000006"]["fulfillment_method"], "FBM")
            self.assertEqual(by_asin["B000000007"]["fulfillment_method"], "AMZ")
            self.assertEqual(by_asin["B000000009"]["fulfillment_method"], "FBA")
            self.assertEqual(by_asin["B000000009"]["fulfillment_method_raw"], "FBA Fee")
            self.assertEqual(by_asin["B000000010"]["fulfillment_method"], "")
            self.assertEqual(by_asin["B000000010"]["fulfillment_method_raw"], "non-FBA")

            filtered_rows = jsonl_rows(output / repair.FILTERED_RECORDS_NAME)
            self.assertEqual(
                {row["asin"] for row in filtered_rows},
                {
                    "B000000001",
                    "B000000003",
                    "B000000004",
                    "B000000005",
                    "B000000006",
                    "B000000007",
                    "B000000010",
                },
            )
            unknowns = {
                item["value"]: item["count"]
                for item in report["unknown_nonempty_raw_values"]
            }
            self.assertEqual(unknowns, {"SFP": 1, "non-FBA": 1})
            self.assertEqual(
                report["fulfillment_distribution"][
                    "raw_nonempty_canonical_empty_records"
                ],
                2,
            )
            self.assertTrue(report["source"]["source_integrity_verified"])

            saved_report = json.loads(
                (output / repair.REPORT_NAME).read_text(encoding="utf-8")
            )
            self.assertEqual(saved_report["counts"]["source_records"], 11)
            self.assertEqual(saved_report["counts"]["repaired_unique_asins"], 10)
            self.assertEqual(saved_report["counts"]["filtered_unique_asins"], 7)
            self.assertEqual(
                saved_report["filter"]["exclude_fulfillment_methods"], ["FBA"]
            )
            self.assertEqual(
                saved_report["fulfillment_repair_rule"]["recognized_prefixes"],
                ["FBA", "FBM", "AMZ"],
            )

            workbook = load_workbook(output / repair.FULL_WORKBOOK_NAME, data_only=False)
            sheet = workbook["ASIN去重总表"]
            headers = [cell.value for cell in sheet[1]]
            asin_column = headers.index("ASIN") + 1
            fulfillment_column = headers.index("配送方式") + 1
            excel_methods = {
                sheet.cell(row=row_index, column=asin_column).value: sheet.cell(
                    row=row_index, column=fulfillment_column
                ).value
                for row_index in range(2, sheet.max_row + 1)
            }
            self.assertEqual(excel_methods["B000000001"], "FBM")
            self.assertEqual(excel_methods["B000000002"], "FBA")
            self.assertEqual(excel_methods["B000000005"], "FBM")
            self.assertEqual(excel_methods["B000000007"], "AMZ")
            self.assertEqual(excel_methods["B000000009"], "FBA")

    def test_unknown_canonical_is_preserved_as_raw_and_allowed_as_non_fba(self) -> None:
        repaired, audit = repair.repair_records(
            [product("B000000001", method="SFP", raw="")]
        )
        self.assertEqual(repaired[0]["fulfillment_method"], "")
        self.assertEqual(repaired[0]["fulfillment_method_raw"], "SFP")
        self.assertEqual(audit["unknown_original_canonical_values"], [{"value": "SFP", "count": 1}])
        filtered, _rejections = repair.filter_non_fba_with_subcategory_rank(repaired)
        self.assertEqual([row["asin"] for row in filtered], ["B000000001"])

    def test_refuses_source_or_existing_output_and_expectation_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source, _records_path = self.write_source(
                root, [product("B000000001", raw="FBM卖家")]
            )
            with self.assertRaises(repair.RepairError):
                repair.repair_job_outputs(source, source)
            with self.assertRaises(repair.RepairError):
                repair.repair_job_outputs(source, source / "nested-output")

            existing_output = root / "existing"
            existing_output.mkdir()
            with self.assertRaises(repair.RepairError):
                repair.repair_job_outputs(source, existing_output)

            mismatched_output = root / "mismatch"
            with self.assertRaises(repair.RepairError):
                repair.repair_job_outputs(
                    source, mismatched_output, expected_record_count=2
                )
            self.assertFalse(mismatched_output.exists())


if __name__ == "__main__":
    unittest.main()
