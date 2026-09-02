#!/usr/bin/env python3
"""
Amazon image-search competitor crawler.

The crawler reads a product list, resolves each row to a main image, searches
Amazon Lens in a visible Chrome session, filters the visual-search candidates
with a vision model, then writes SellerSprite-enriched competitor rows.

It does not solve CAPTCHA or bypass verification. If Amazon or SellerSprite
verification appears, it saves state and waits for manual handling.
"""

from __future__ import annotations

import argparse
import base64
from copy import copy as copy_style
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import html as html_lib
import json
import math
import mimetypes
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from urllib.parse import parse_qs, quote_plus, urlparse

import requests
from selenium.common.exceptions import JavascriptException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

from amazon_category_rank_crawler import (
    BatchPauseScheduler,
    DeliveryLocationUnconfirmedError,
    JobRunLock,
    REQUESTED_DATA_FIELDS,
    SELLERSPRITE_EVIDENCE_FIELDS,
    UserFacingError,
    VerificationUnconfirmedError,
    append_jsonl,
    build_delivery_location_config,
    clean_url,
    classify_sellersprite_snapshot,
    config_bool,
    config_float,
    config_int,
    config_text,
    country_from_flag_code_or_text,
    detect_block,
    dump_json,
    ensure_amazon_delivery_location,
    ensure_dir,
    extract_by_selectors,
    extract_table_rows,
    handle_amazon_verification,
    load_json,
    normalize_header,
    normalize_space,
    now_iso,
    now_ts,
    open_amazon_page,
    parse_field_from_text,
    parse_table_row_fields,
    plugin_node_count,
    preload_page_data_with_scroll,
    read_jsonl,
    resolve_path,
    save_debug_snapshot,
    safe_sellersprite_readiness,
    sellersprite_block_reason,
    set_sellersprite_readiness,
    sleep_between_pages,
    slugify,
    start_driver,
    try_activate_plugin,
    sellersprite_login_required,
    wait_for_manual_clear,
    wait_for_amazon_products,
    wait_for_sellersprite_data,
    verification_unconfirmed_message,
    write_jsonl_atomic,
)
from amazon_front_crawler import pick_column, read_input_rows


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CONFIG = ROOT_DIR / "config" / "amazon_image_competitors.json"
ASIN_RE = re.compile(r"\b([A-Z0-9]{10})\b", re.I)
URL_ASIN_RE = re.compile(r"/(?:dp|gp/product)/([A-Z0-9]{10})(?:[/?#]|$)", re.I)

MARKETPLACE_DOMAINS = {
    "us": "amazon.com",
    "usa": "amazon.com",
    "美国": "amazon.com",
    "美国站": "amazon.com",
    "amazon.com": "amazon.com",
    "ca": "amazon.ca",
    "加拿大": "amazon.ca",
    "加拿大站": "amazon.ca",
    "amazon.ca": "amazon.ca",
    "mx": "amazon.com.mx",
    "墨西哥": "amazon.com.mx",
    "墨西哥站": "amazon.com.mx",
    "amazon.com.mx": "amazon.com.mx",
    "uk": "amazon.co.uk",
    "gb": "amazon.co.uk",
    "英国": "amazon.co.uk",
    "英国站": "amazon.co.uk",
    "amazon.co.uk": "amazon.co.uk",
    "de": "amazon.de",
    "德国": "amazon.de",
    "德国站": "amazon.de",
    "amazon.de": "amazon.de",
    "fr": "amazon.fr",
    "法国": "amazon.fr",
    "法国站": "amazon.fr",
    "amazon.fr": "amazon.fr",
    "it": "amazon.it",
    "意大利": "amazon.it",
    "意大利站": "amazon.it",
    "amazon.it": "amazon.it",
    "es": "amazon.es",
    "西班牙": "amazon.es",
    "西班牙站": "amazon.es",
    "amazon.es": "amazon.es",
    "jp": "amazon.co.jp",
    "日本": "amazon.co.jp",
    "日本站": "amazon.co.jp",
    "amazon.co.jp": "amazon.co.jp",
    "au": "amazon.com.au",
    "澳洲": "amazon.com.au",
    "澳大利亚": "amazon.com.au",
    "澳洲站": "amazon.com.au",
    "amazon.com.au": "amazon.com.au",
    "in": "amazon.in",
    "印度": "amazon.in",
    "印度站": "amazon.in",
    "amazon.in": "amazon.in",
    "nl": "amazon.nl",
    "荷兰": "amazon.nl",
    "荷兰站": "amazon.nl",
    "amazon.nl": "amazon.nl",
    "se": "amazon.se",
    "瑞典": "amazon.se",
    "瑞典站": "amazon.se",
    "amazon.se": "amazon.se",
    "pl": "amazon.pl",
    "波兰": "amazon.pl",
    "波兰站": "amazon.pl",
    "amazon.pl": "amazon.pl",
    "ae": "amazon.ae",
    "阿联酋": "amazon.ae",
    "阿联酋站": "amazon.ae",
    "amazon.ae": "amazon.ae",
    "sa": "amazon.sa",
    "沙特": "amazon.sa",
    "沙特站": "amazon.sa",
    "amazon.sa": "amazon.sa",
    "sg": "amazon.sg",
    "新加坡": "amazon.sg",
    "新加坡站": "amazon.sg",
    "amazon.sg": "amazon.sg",
    "br": "amazon.com.br",
    "巴西": "amazon.com.br",
    "巴西站": "amazon.com.br",
    "amazon.com.br": "amazon.com.br",
    "za": "amazon.co.za",
    "南非": "amazon.co.za",
    "南非站": "amazon.co.za",
    "amazon.co.za": "amazon.co.za",
}

INPUT_ALIASES = {
    "source_asin": ["asin", "ASIN", "源ASIN", "产品ASIN", "被抓取ASIN", "来源ASIN"],
    "product_url": ["product_url", "商品URL", "产品URL", "链接", "商品链接", "亚马逊链接"],
    "image_url": ["image_url", "主图URL", "图片URL", "主图链接", "图片链接"],
    "image_path": ["image_path", "主图路径", "图片路径", "本地图片", "本地图片路径"],
    "note": ["note", "备注"],
}

IMAGE_COMPETITOR_HEADERS = [
    "来源ASIN",
    "竞品ASIN",
    "评论数量",
    "评分值",
    "卖家名称",
    "品牌名称",
    "卖家所处国家",
    "近30天销量（子体）",
    "近30天销量（父体）",
    "FBA费用",
    "毛利率",
    "配送方式",
    "配送时长",
    "上架时间",
    "自然搜索词数量",
    "广告搜索词数量",
    "商品标题",
    "商品URL",
    "来源商品URL",
    "来源主图",
    "同款置信度",
    "同款判断原因",
    "抓取时间",
    "加载状态",
    "备注",
]

IMAGE_COMPETITOR_FIELD_TO_HEADER = {
    "source_asin": "来源ASIN",
    "asin": "竞品ASIN",
    "review_count": "评论数量",
    "rating_value": "评分值",
    "seller_name": "卖家名称",
    "brand_name": "品牌名称",
    "seller_country": "卖家所处国家",
    "sales_30_days_child": "近30天销量（子体）",
    "sales_30_days_parent": "近30天销量（父体）",
    "fba_fee": "FBA费用",
    "gross_margin": "毛利率",
    "fulfillment_method": "配送方式",
    "delivery_duration": "配送时长",
    "launch_date": "上架时间",
    "organic_keywords_count": "自然搜索词数量",
    "ad_keywords_count": "广告搜索词数量",
    "title": "商品标题",
    "product_url": "商品URL",
    "source_product_url": "来源商品URL",
    "source_image": "来源主图",
    "match_confidence": "同款置信度",
    "match_reason": "同款判断原因",
    "scraped_at": "抓取时间",
    "load_status": "加载状态",
    "note": "备注",
}

CANDIDATE_HEADERS = [
    "source_id",
    "source_asin",
    "asin",
    "title",
    "product_url",
    "candidate_image_url",
    "rank",
    "is_competitor",
    "confidence",
    "reason",
    "prescreen_similarity",
    "prescreen_status",
    "mini_is_same_product",
    "mini_confidence",
    "mini_reason",
    "load_status",
]

COUNT_COLUMN_HEADER = "相似竞品数量"
REVIEW_PRODUCT_URL_HEADER = "商品URL"
MINI_CONFIRMED_COUNT_FIELD = "mini_confirmed_same_product_count"
MINI_CONFIRMED_COUNT_HEADER = "mini复核确认同款数量"
EMBEDDING_GREATER_THAN_TEN_LABEL = "Embedding判断同款数量大于10"
COUNT_RESULT_FIELD_HEADERS = {
    "prescreen_visual_match_count": "视觉粗筛命中数",
    "processing_status": "处理状态",
    "same_product_confidence": "同款判断置信度",
    "match_reason": "同款判断说明",
}
COUNT_ONLY_REMOVED_HEADERS = frozenset(
    {
        "最佳页码",
        "最佳排名",
        "加载状态",
        "备注",
        MINI_CONFIRMED_COUNT_HEADER,
    }
)

EMBEDDING_CACHE: Dict[str, List[float]] = {}
DOUBAO_EMBEDDING_MODEL = "doubao-embedding-vision-251215"
DOUBAO_EMBEDDING_BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
DOUBAO_EMBEDDING_API_PATH = "embeddings/multimodal"
DOUBAO_MINI_MODEL = "doubao-seed-2-0-mini-260428"
DOUBAO_MINI_BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
DOUBAO_MINI_API_PATH = "chat/completions"
CASCADE_MATCH_SEMANTICS = "primary-product-cascade-v3-lens-context"
MINI_RESPONSE_SEMANTICS = "same-product-strict-json-v2"
SOURCE_RESULT_SHARD_SEMANTICS = "image-competitor-source-result-v1"
RETRYABLE_EMBEDDING_STATUS_CODES = {408, 429}


class EmbeddingProviderError(UserFacingError):
    """A recoverable source-level embedding failure.

    The current source must remain in state so a later run can retry it instead
    of persisting a misleading zero-competitor result.
    """


class FatalEmbeddingProviderError(EmbeddingProviderError):
    """A provider credential, model, or endpoint error that must stop the task."""


class MiniProviderError(EmbeddingProviderError):
    """A recoverable source-level Doubao Mini verification failure."""


class FatalMiniProviderError(FatalEmbeddingProviderError):
    """A Mini credential, model, or endpoint error that must stop the task."""


@dataclass
class ImageCompetitorRuntimeConfig:
    job_id: str
    outputs_root: Path
    products_file: Path
    marketplace: str
    marketplace_domain: str
    lens_url_template: str
    search_strategy: str
    find_similar_timeout: int
    lens_results_timeout: int
    max_candidates_per_source: int
    max_competitors_per_source: int
    result_mode: str
    min_match_confidence: float
    include_source_as_competitor: bool
    match_mode: str
    doubao_embedding_config_file: Optional[Path]
    embedding_provider: str
    embedding_api_key: str
    embedding_model: str
    embedding_base_url: str
    embedding_api_path: str
    embedding_encoding_format: str
    embedding_retry_attempts: int
    embedding_retry_backoff_seconds: float
    prescreen_min_similarity: float
    prescreen_max_matches: int
    doubao_mini_config_file: Optional[Path]
    mini_provider: str
    mini_api_key: str
    mini_model: str
    mini_base_url: str
    mini_api_path: str
    mini_batch_size: int
    mini_retry_attempts: int
    mini_retry_backoff_seconds: float
    vision_model: str
    openai_api_key_env: str
    openai_base_url: str
    openai_api_path: str
    vision_batch_size: int
    vision_timeout: int
    resume: bool
    browser_backend: str
    browser_mode: str
    chrome_binary: str
    chrome_user_data_dir: Path
    chrome_profile_directory: str
    debugger_address: str
    extension_path: Path
    activate_plugin: bool
    page_timeout: int
    plugin_timeout: int
    plugin_retry_attempts: int
    plugin_retry_wait_seconds: float
    plugin_retry_wait_seconds_max: float
    plugin_relaunch_retry_attempts: int
    plugin_relaunch_wait_seconds: float
    plugin_second_relaunch_retry_attempts: int
    plugin_second_relaunch_wait_seconds: float
    manual_pause_timeout: int
    delivery_location_enabled: bool
    delivery_locations_file: Path
    delivery_location_timeout: int
    delivery_locations: Dict[str, Dict[str, str]]
    delivery_location_fingerprint: str
    delay_seconds_min: float
    delay_seconds_max: float
    batch_pause_pages_min: int
    batch_pause_pages_max: int
    batch_pause_seconds_min: float
    batch_pause_seconds_max: float
    page_scroll_before_extract: bool
    page_scroll_max_rounds: int
    page_scroll_step_ratio: float
    page_scroll_wait_seconds: float
    page_scroll_stable_rounds: int
    sellersprite_required: bool
    sellersprite_min_enriched_records: int
    sellersprite_min_fields_per_record: int
    sellersprite_stable_checks: int
    save_debug_snapshots: bool
    sellersprite_on_lens: bool
    enrich_accepted_results: bool
    enrichment_page_timeout: int
    enrichment_plugin_timeout: int
    field_selectors: Dict[str, List[str]] = field(default_factory=dict)
    provider_metrics: Dict[str, int] = field(default_factory=dict)

    @property
    def lens_url(self) -> str:
        return self.lens_url_template.format(domain=self.marketplace_domain)

    @property
    def is_count_only(self) -> bool:
        return self.result_mode == "count_only"


@dataclass
class MatchEvaluation:
    accepted_records: List[Dict[str, Any]]
    decisions: Dict[str, Dict[str, Any]]
    prescreen_visual_match_count: int | str
    processing_status: str
    same_product_count: Optional[int]
    same_product_confidence: float | str
    match_reason: str
    provider_metrics: Dict[str, int] = field(default_factory=dict)


class ImageCompetitorStateStore:
    def __init__(self, path: Path, runtime: ImageCompetitorRuntimeConfig, initial_queue: Sequence[Dict[str, Any]]) -> None:
        self.path = path
        self.runtime = runtime
        self.initial_queue = list(initial_queue)
        self.data: Dict[str, Any] = {}

    def has_existing_result_rows(self) -> bool:
        aggregate_rows_exist = any(
            path.exists() and path.stat().st_size > 0
            for path in (
                self.path.with_name("records.jsonl"),
                self.path.with_name("candidates.jsonl"),
                self.path.with_name("counts.jsonl"),
            )
        )
        shard_dir = self.path.parent / "source_results"
        return aggregate_rows_exist or (
            shard_dir.exists() and any(shard_dir.glob("*.json"))
        )

    def load_or_create(self) -> None:
        if self.runtime.resume and self.path.exists():
            self.data = load_json(self.path)
            expected = vision_provider_fingerprint(self.runtime)
            existing = self.data.get("vision_provider_fingerprint")
            expected_plan = image_crawl_plan_fingerprint(self.runtime, self.initial_queue)
            existing_plan = self.data.get("crawl_plan_fingerprint")
            expected_delivery = self.runtime.delivery_location_fingerprint
            existing_delivery = str(self.data.get("delivery_location_fingerprint") or "")
            has_progress = bool(
                self.data.get("records_count")
                or self.data.get("completed_sources")
                or self.has_existing_result_rows()
            )
            if existing_delivery != expected_delivery and has_progress:
                raise UserFacingError(
                    "配送地址配置与已有断点不一致；"
                    "请保留旧输出并改用新的 job_id，避免混合结果。"
                )
            if existing != expected and has_progress:
                raise UserFacingError(
                    "视觉向量模型、端点、相似度阈值或配送地址配置与已有断点不一致；"
                    "请改用新的 job_id，避免混合旧结果。"
                )
            if existing_plan != expected_plan and has_progress:
                raise UserFacingError(
                    "输入商品、输入顺序或以图搜图计划与已有断点不一致；"
                    "请保留旧输出并改用新的 job_id，避免把旧数量写到其他 ASIN。"
                )
            if (
                existing != expected
                or existing_delivery != expected_delivery
                or existing_plan != expected_plan
            ):
                self.data = self._new_data(expected, expected_plan)
                self.flush()
            return
        if self.runtime.resume and self.has_existing_result_rows():
            raise UserFacingError(
                "已有图片竞品结果但缺少可验证的地址/模型断点指纹；"
                "请保留旧输出并改用新的 job_id，避免混合结果。"
            )
        self.data = self._new_data(
            vision_provider_fingerprint(self.runtime),
            image_crawl_plan_fingerprint(self.runtime, self.initial_queue),
        )
        self.flush()

    def _new_data(
        self,
        provider_fingerprint: Dict[str, Any],
        crawl_plan_fingerprint: Dict[str, Any],
    ) -> Dict[str, Any]:
        return {
            "job_id": self.runtime.job_id,
            "mode": "image_competitor",
            "created_at": now_iso(),
            "marketplace": self.runtime.marketplace_domain,
            "queue": self.initial_queue,
            "current": None,
            "completed_sources": [],
            "records_count": 0,
            "failures_count": 0,
            "vision_provider_fingerprint": provider_fingerprint,
            "crawl_plan_fingerprint": crawl_plan_fingerprint,
            "delivery_location_fingerprint": self.runtime.delivery_location_fingerprint,
        }

    def flush(self) -> None:
        self.data["updated_at"] = now_iso()
        dump_json(self.path, self.data)

    def next_work(self) -> Optional[Dict[str, Any]]:
        current = self.data.get("current")
        if current:
            return current
        queue = self.data.get("queue") or []
        if not queue:
            return None
        current = queue.pop(0)
        self.data["queue"] = queue
        self.data["current"] = current
        self.flush()
        return current

    def set_current(self, current: Optional[Dict[str, Any]]) -> None:
        self.data["current"] = current
        self.flush()

    def finish_current_source(
        self,
        reason: str = "",
        count: int = 0,
        result: Optional[Dict[str, Any]] = None,
    ) -> None:
        current = self.data.get("current")
        if current:
            source_id = str(current.get("source_id") or "")
            done = set(self.data.setdefault("completed_sources", []))
            done.add(source_id)
            self.data["completed_sources"] = sorted(done)
            self.data.setdefault("completed_source_reasons", {})[source_id] = reason
            if result is not None:
                self.data.setdefault("completed_source_results", {})[source_id] = dict(result)
        self.data["records_count"] = int(self.data.get("records_count") or 0) + count
        self.data["current"] = None
        self.flush()

    def log_failure(self) -> None:
        self.data["failures_count"] = int(self.data.get("failures_count") or 0) + 1
        self.flush()

    def mark_sellersprite_readiness(self, report: Dict[str, Any]) -> None:
        self.data["sellersprite_readiness"] = safe_sellersprite_readiness(report)
        self.flush()

    def mark_manual_pause(self, reason: str, page_url: str) -> None:
        self.data["manual_pause"] = {
            "paused_at": now_iso(),
            "reason": reason,
            "page_url": page_url,
            "current": self.data.get("current"),
        }
        self.flush()

    def clear_manual_pause(self) -> None:
        if self.data.pop("manual_pause", None) is not None:
            self.flush()


def prompt_marketplace() -> str:
    print("请输入 Amazon 站点，例如：美国站、amazon.com、德国站、amazon.de。")
    try:
        return input("站点：").strip()
    except EOFError:
        return ""


def normalize_marketplace(value: str) -> str:
    text = normalize_space(value).lower()
    if text in MARKETPLACE_DOMAINS:
        return MARKETPLACE_DOMAINS[text]
    text = text.removeprefix("https://").removeprefix("http://").strip("/")
    text = text.removeprefix("www.")
    if text in MARKETPLACE_DOMAINS:
        return MARKETPLACE_DOMAINS[text]
    if text.startswith("amazon.") or ".amazon." in text:
        return text
    raise UserFacingError(f"暂不识别的 Amazon 站点：{value}")


def parse_asin(value: str) -> str:
    text = normalize_space(value).upper()
    match = URL_ASIN_RE.search(text)
    if match:
        return match.group(1).upper()
    match = ASIN_RE.search(text)
    return match.group(1).upper() if match else ""


def product_url_for_asin(domain: str, asin: str) -> str:
    return f"https://www.{domain}/dp/{asin}"


def redact_sensitive_text(value: Any, secrets: Sequence[str] = ()) -> str:
    text = str(value)
    for secret in secrets:
        if secret:
            text = text.replace(secret, "[REDACTED]")
    text = re.sub(r"(?i)(authorization\s*[:=]\s*bearer\s+)[^\s,;]+", r"\1[REDACTED]", text)
    text = re.sub(r'(?i)(["\']?api[_-]?key["\']?\s*[:=]\s*["\']?)[^"\'\s,;}]+', r"\1[REDACTED]", text)
    return text


def add_provider_metric(runtime: ImageCompetitorRuntimeConfig, key: str, value: int = 1) -> None:
    runtime.provider_metrics[key] = int(runtime.provider_metrics.get(key) or 0) + int(value)


def provider_metric_snapshot(runtime: ImageCompetitorRuntimeConfig) -> Dict[str, int]:
    return {key: int(value) for key, value in runtime.provider_metrics.items()}


def provider_metric_delta(
    before: Dict[str, int],
    after: Dict[str, int],
) -> Dict[str, int]:
    return {
        key: int(after.get(key) or 0) - int(before.get(key) or 0)
        for key in sorted(set(before) | set(after))
        if int(after.get(key) or 0) - int(before.get(key) or 0)
    }


def _provider_url(base_url: str, api_path: str) -> str:
    parsed = urlparse(base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise UserFacingError("视觉向量配置中的 base_url 必须是有效的 http/https 地址。")
    if not api_path:
        raise UserFacingError("视觉向量配置中的 api_path 不能为空。")
    return f"{base_url.rstrip('/')}/{api_path.strip('/')}"


def _require_https_provider_url(base_url: str, api_path: str, label: str) -> str:
    endpoint = _provider_url(base_url, api_path)
    parsed = urlparse(endpoint)
    if parsed.scheme != "https":
        raise UserFacingError(f"{label}配置必须使用 HTTPS 端点，避免 API Key 明文传输。")
    if parsed.username or parsed.password or parsed.query or parsed.fragment:
        raise UserFacingError(
            f"{label}端点不能包含用户名、密码、查询参数或 URL 片段。"
        )
    return endpoint


def _read_private_provider_config(path: Path, label: str) -> Dict[str, Any]:
    if not path.exists() or not path.is_file():
        raise UserFacingError(
            f"没有找到{label}配置文件：{path}。请在本地专用配置文件中填写 API Key；"
            "不要在聊天中发送 Key。"
        )
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise UserFacingError(f"{label}配置文件不是有效 JSON：{path}。") from exc
    if not isinstance(raw, dict):
        raise UserFacingError(f"{label}配置文件必须是 JSON 对象：{path}。")
    if not normalize_space(str(raw.get("api_key") or "")):
        raise UserFacingError(
            f"{label} API Key 为空。请把您自己的 Key 填入 {path} 的 api_key 字段；"
            "不要在聊天中发送 Key。"
        )
    return raw


def _prepare_embedding_provider(runtime: ImageCompetitorRuntimeConfig) -> None:
    config_path = runtime.doubao_embedding_config_file
    if config_path is not None:
        raw = _read_private_provider_config(config_path, "豆包视觉向量")
        runtime.embedding_provider = "doubao"
        runtime.embedding_api_key = normalize_space(str(raw.get("api_key") or ""))
        runtime.embedding_model = normalize_space(str(raw.get("model") or DOUBAO_EMBEDDING_MODEL))
        runtime.embedding_base_url = normalize_space(
            str(raw.get("base_url") or DOUBAO_EMBEDDING_BASE_URL)
        ).rstrip("/")
        runtime.embedding_api_path = normalize_space(
            str(raw.get("api_path") or DOUBAO_EMBEDDING_API_PATH)
        ).strip("/")
        runtime.embedding_encoding_format = normalize_space(
            str(raw.get("encoding_format") or "float")
        ).lower()
        if runtime.embedding_encoding_format != "float":
            raise UserFacingError("豆包视觉向量配置 encoding_format 当前只支持 float。")
        if not runtime.embedding_model:
            raise UserFacingError("豆包视觉向量配置中的 model 不能为空。")
        _require_https_provider_url(
            runtime.embedding_base_url,
            runtime.embedding_api_path,
            "豆包视觉向量",
        )
        return

    api_key = os.environ.get(runtime.openai_api_key_env, "").strip()
    if not api_key:
        raise UserFacingError(
            f"缺少旧版视觉向量 API Key 环境变量：{runtime.openai_api_key_env}。"
            "建议改用 doubao_embedding_config_file，并在专用 JSON 文件中绑定您自己的 Key。"
        )
    runtime.embedding_provider = "legacy_openai"
    runtime.embedding_api_key = api_key
    runtime.embedding_model = runtime.vision_model
    runtime.embedding_base_url = runtime.openai_base_url
    runtime.embedding_api_path = runtime.openai_api_path
    runtime.embedding_encoding_format = "float"
    _provider_url(runtime.embedding_base_url, runtime.embedding_api_path)
    print(
        "弃用提示：embedding 模式正在使用旧 openai_* 配置；"
        "请迁移到 doubao_embedding_config_file 专用配置。",
        file=sys.stderr,
    )


def _prepare_mini_provider(runtime: ImageCompetitorRuntimeConfig) -> None:
    config_path = runtime.doubao_mini_config_file
    if config_path is None:
        raise UserFacingError(
            "cascade 模式必须配置 doubao_mini_config_file，且 API Key 只能保存在该本地文件中。"
        )
    raw = _read_private_provider_config(config_path, "豆包同款复核")
    runtime.mini_provider = "doubao"
    runtime.mini_api_key = normalize_space(str(raw.get("api_key") or ""))
    runtime.mini_model = normalize_space(str(raw.get("model") or DOUBAO_MINI_MODEL))
    runtime.mini_base_url = normalize_space(
        str(raw.get("base_url") or DOUBAO_MINI_BASE_URL)
    ).rstrip("/")
    runtime.mini_api_path = normalize_space(
        str(raw.get("api_path") or DOUBAO_MINI_API_PATH)
    ).strip("/")
    if not runtime.mini_model:
        raise UserFacingError("豆包同款复核配置中的 model 不能为空。")
    _require_https_provider_url(
        runtime.mini_base_url,
        runtime.mini_api_path,
        "豆包同款复核",
    )


def prepare_vision_provider(runtime: ImageCompetitorRuntimeConfig) -> None:
    """Resolve every required credential before dry-run or browser startup."""
    if runtime.match_mode == "chat":
        if not os.environ.get(runtime.openai_api_key_env):
            raise UserFacingError(
                f"缺少视觉模型 API Key，请先设置环境变量 {runtime.openai_api_key_env}。"
            )
        runtime.embedding_provider = "openai_chat"
        _provider_url(runtime.openai_base_url, runtime.openai_api_path)
        print(
            "弃用提示：chat 模式继续使用旧 openai_* 配置；"
            "新任务建议使用 cascade 模式。",
            file=sys.stderr,
        )
        return

    _prepare_embedding_provider(runtime)
    if runtime.match_mode == "cascade":
        _prepare_mini_provider(runtime)


def vision_provider_fingerprint(runtime: ImageCompetitorRuntimeConfig) -> Dict[str, Any]:
    if runtime.match_mode == "chat":
        public_config: Dict[str, Any] = {
            "match_mode": "chat",
            "provider": "openai_chat",
            "model": runtime.vision_model,
            "endpoint": _provider_url(runtime.openai_base_url, runtime.openai_api_path),
            "min_match_confidence": runtime.min_match_confidence,
            "delivery_location_fingerprint": runtime.delivery_location_fingerprint,
        }
    elif runtime.match_mode == "embedding":
        public_config = {
            "match_mode": "embedding",
            "provider": runtime.embedding_provider,
            "model": runtime.embedding_model,
            "endpoint": _provider_url(runtime.embedding_base_url, runtime.embedding_api_path),
            "encoding_format": runtime.embedding_encoding_format,
            "min_match_confidence": runtime.min_match_confidence,
            "delivery_location_fingerprint": runtime.delivery_location_fingerprint,
        }
    else:
        public_config = {
            "match_mode": "cascade",
            "pipeline_semantics": CASCADE_MATCH_SEMANTICS,
            "mini_response_semantics": MINI_RESPONSE_SEMANTICS,
            "embedding_provider": runtime.embedding_provider,
            "embedding_model": runtime.embedding_model,
            "embedding_endpoint": _provider_url(
                runtime.embedding_base_url,
                runtime.embedding_api_path,
            ),
            "embedding_encoding_format": runtime.embedding_encoding_format,
            "prescreen_min_similarity": runtime.prescreen_min_similarity,
            "prescreen_max_matches": runtime.prescreen_max_matches,
            "mini_provider": runtime.mini_provider,
            "mini_model": runtime.mini_model,
            "mini_endpoint": _provider_url(runtime.mini_base_url, runtime.mini_api_path),
            "mini_batch_size": runtime.mini_batch_size,
            "include_source_as_competitor": runtime.include_source_as_competitor,
            "max_candidates_per_source": runtime.max_candidates_per_source,
            "delivery_location_fingerprint": runtime.delivery_location_fingerprint,
        }
    serialized = json.dumps(public_config, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return {
        **public_config,
        "sha256": hashlib.sha256(serialized.encode("utf-8")).hexdigest(),
    }


def image_crawl_plan_fingerprint(
    runtime: ImageCompetitorRuntimeConfig,
    initial_queue: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    """Bind resumable output to the normalized inputs and result semantics."""
    input_file_hasher = hashlib.sha256()
    with runtime.products_file.open("rb") as input_file:
        for chunk in iter(lambda: input_file.read(1024 * 1024), b""):
            input_file_hasher.update(chunk)
    normalized_sources = [
        {
            "source_id": normalize_space(str(row.get("source_id") or "")),
            "source_asin": normalize_space(str(row.get("source_asin") or "")).upper(),
            "source_product_url": clean_url(str(row.get("source_product_url") or "")),
            "input_image_url": clean_url(str(row.get("input_image_url") or "")),
            "input_image_path": str(row.get("input_image_path") or "").strip(),
            "input_row": int(row.get("input_row") or 0),
        }
        for row in initial_queue
    ]
    sources_json = json.dumps(
        normalized_sources,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    public_plan: Dict[str, Any] = {
        "semantics": "image-competitor-crawl-plan-v3",
        "source_result_semantics": SOURCE_RESULT_SHARD_SEMANTICS,
        "marketplace_domain": runtime.marketplace_domain,
        "result_mode": runtime.result_mode,
        "match_mode": runtime.match_mode,
        "search_strategy": runtime.search_strategy,
        "lens_url_template": runtime.lens_url_template,
        "max_candidates_per_source": runtime.max_candidates_per_source,
        "include_source_as_competitor": runtime.include_source_as_competitor,
        "source_count": len(normalized_sources),
        "input_file_sha256": input_file_hasher.hexdigest(),
        "normalized_sources_sha256": hashlib.sha256(sources_json.encode("utf-8")).hexdigest(),
    }
    serialized = json.dumps(
        public_plan,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return {
        **public_plan,
        "sha256": hashlib.sha256(serialized.encode("utf-8")).hexdigest(),
    }


def build_image_runtime_config(config: Dict[str, Any], no_resume: bool) -> ImageCompetitorRuntimeConfig:
    marketplace = config_text(config, "marketplace")
    if not marketplace:
        marketplace = prompt_marketplace()
    if not marketplace:
        raise UserFacingError("执行以图搜图抓取前必须输入 Amazon 站点。")
    marketplace_domain = normalize_marketplace(marketplace)

    products_file = resolve_path(config_text(config, "products_file", "inputs/image_competitors.csv"))
    if not products_file.exists():
        raise UserFacingError(f"没有找到产品清单：{products_file}")

    extension_path_text = config_text(config, "extension_path")
    extension_path = resolve_path(extension_path_text) if extension_path_text else Path("")
    browser_backend = config_text(config, "browser_backend", "cdp").lower()
    if browser_backend not in {"cdp", "selenium"}:
        raise UserFacingError("配置项 `browser_backend` 只支持 cdp 或 selenium。")
    browser_mode = config_text(config, "browser_mode", "launch").lower()
    if browser_mode not in {"launch", "attach", "reuse"}:
        raise UserFacingError("以图搜图 browser_mode 只支持 launch、attach 或 reuse。")
    if browser_mode == "launch" and extension_path_text and not extension_path.exists():
        raise UserFacingError(f"没有找到卖家精灵扩展目录：{extension_path}")

    min_delay = config_float(config, "delay_seconds_min", 6)
    max_delay = config_float(config, "delay_seconds_max", 12)
    if max_delay < min_delay:
        max_delay = min_delay
    batch_pages_min = config_int(config, "batch_pause_pages_min", 20) or 0
    batch_pages_max = config_int(config, "batch_pause_pages_max", 36) or 0
    if batch_pages_max and batch_pages_max < batch_pages_min:
        batch_pages_max = batch_pages_min
    batch_seconds_min = config_float(config, "batch_pause_seconds_min", 60)
    batch_seconds_max = config_float(config, "batch_pause_seconds_max", 150)
    if batch_seconds_max < batch_seconds_min:
        batch_seconds_max = batch_seconds_min
    plugin_retry_wait_min = config_float(
        config,
        "plugin_retry_wait_seconds_min",
        config_float(config, "plugin_retry_wait_seconds", 10),
    )
    plugin_retry_wait_max = config_float(config, "plugin_retry_wait_seconds_max", 20)
    if plugin_retry_wait_min < 0 or plugin_retry_wait_max < 0:
        raise UserFacingError("配置项 plugin_retry_wait_seconds_min / plugin_retry_wait_seconds_max 不能小于 0。")
    if plugin_retry_wait_max < plugin_retry_wait_min:
        plugin_retry_wait_max = plugin_retry_wait_min
    page_scroll_max_rounds = max(config_int(config, "page_scroll_max_rounds", 18) or 0, 0)
    page_scroll_step_ratio = config_float(config, "page_scroll_step_ratio", 0.85) or 0.85
    if page_scroll_step_ratio <= 0:
        raise UserFacingError("配置项 page_scroll_step_ratio 必须大于 0。")
    page_scroll_wait_seconds = max(config_float(config, "page_scroll_wait_seconds", 1.0) or 0, 0)
    page_scroll_stable_rounds = max(config_int(config, "page_scroll_stable_rounds", 2) or 1, 1)
    delivery_config = build_delivery_location_config(config)

    raw_selectors = config.get("field_selectors") or {}
    field_selectors: Dict[str, List[str]] = {}
    if isinstance(raw_selectors, dict):
        for key, value in raw_selectors.items():
            if isinstance(value, list):
                field_selectors[key] = [str(item).strip() for item in value if str(item).strip()]

    min_confidence = config_float(config, "min_match_confidence", 0.70)
    if min_confidence < 0 or min_confidence > 1:
        raise UserFacingError("配置项 min_match_confidence 必须在 0-1 之间。")
    result_mode = (config_text(config, "result_mode", "detail").lower() or "detail").replace("-", "_")
    if result_mode not in {"detail", "count_only"}:
        raise UserFacingError("配置项 result_mode 只支持 detail 或 count_only。")
    default_match_mode = "embedding" if "doubao_embedding_config_file" in config else "chat"
    match_mode = config_text(config, "match_mode", default_match_mode).lower() or default_match_mode
    if match_mode not in {"embedding", "chat", "cascade"}:
        raise UserFacingError("配置项 match_mode 只支持 embedding、chat 或 cascade。")
    if match_mode == "cascade" and result_mode != "count_only":
        raise UserFacingError(
            "cascade 当前只支持 result_mode=count_only；"
            "早停来源必须在数量表中保留明确状态，不能写成明细模式的静默空结果。"
        )
    doubao_config_text = config_text(config, "doubao_embedding_config_file")
    if (
        match_mode in {"embedding", "cascade"}
        and "doubao_embedding_config_file" in config
        and not doubao_config_text
    ):
        raise UserFacingError(
            "配置项 doubao_embedding_config_file 已声明但为空；"
            "请填写专用 JSON 配置文件路径，并在文件中绑定您自己的 API Key。"
        )
    doubao_config_path = resolve_path(doubao_config_text) if doubao_config_text else None
    if match_mode == "cascade" and doubao_config_path is None:
        raise UserFacingError(
            "cascade 模式必须配置 doubao_embedding_config_file，不能回退旧环境变量。"
        )
    doubao_mini_config_text = config_text(config, "doubao_mini_config_file")
    if match_mode == "cascade" and not doubao_mini_config_text:
        raise UserFacingError(
            "cascade 模式必须配置 doubao_mini_config_file；"
            "请把 API Key 放在独立本地 JSON 文件中。"
        )
    doubao_mini_config_path = (
        resolve_path(doubao_mini_config_text) if doubao_mini_config_text else None
    )
    prescreen_min_similarity = config_float(config, "prescreen_min_similarity", 0.70)
    if not 0 <= prescreen_min_similarity <= 1:
        raise UserFacingError("配置项 prescreen_min_similarity 必须在 0-1 之间。")
    prescreen_max_matches = config_int(config, "prescreen_max_matches", 10) or 0
    if prescreen_max_matches < 1:
        raise UserFacingError("配置项 prescreen_max_matches 必须大于等于 1。")
    mini_batch_size = config_int(config, "mini_batch_size", 6) or 0
    if mini_batch_size < 1 or mini_batch_size > 6:
        raise UserFacingError("配置项 mini_batch_size 必须在 1-6 之间。")
    mini_retry_attempts = config_int(config, "mini_retry_attempts", 3) or 0
    if mini_retry_attempts < 1:
        raise UserFacingError("配置项 mini_retry_attempts 必须大于等于 1。")
    mini_retry_backoff_seconds = config_float(config, "mini_retry_backoff_seconds", 1.0)
    if mini_retry_backoff_seconds < 0:
        raise UserFacingError("配置项 mini_retry_backoff_seconds 不能小于 0。")
    max_candidates_per_source = config_int(config, "max_candidates_per_source", 24) or 24
    if match_mode == "cascade" and max_candidates_per_source <= prescreen_max_matches:
        raise UserFacingError(
            "cascade 模式要求 max_candidates_per_source 至少比 prescreen_max_matches 大 1，"
            "否则无法观察到第 11 个粗筛命中。"
        )
    vision_model = config_text(config, "vision_model", "gpt-5.4-mini")
    openai_base_url = config_text(config, "openai_base_url", "https://api.openai.com/v1").rstrip("/")
    openai_api_path = config_text(config, "openai_api_path", "responses").strip("/")
    embedding_retry_attempts = max(config_int(config, "embedding_retry_attempts", 3) or 0, 1)
    embedding_retry_backoff_seconds = config_float(config, "embedding_retry_backoff_seconds", 1.0)
    if embedding_retry_backoff_seconds < 0:
        raise UserFacingError("配置项 embedding_retry_backoff_seconds 不能小于 0。")
    sellersprite_on_lens = config_bool(config, "sellersprite_on_lens", False)
    enrich_accepted_results = config_bool(config, "enrich_accepted_results", True)
    sellersprite_required = config_bool(config, "sellersprite_required", result_mode != "count_only")
    if result_mode == "count_only":
        sellersprite_on_lens = False
        enrich_accepted_results = False
        sellersprite_required = False
    if sellersprite_required and not (sellersprite_on_lens or enrich_accepted_results):
        raise UserFacingError(
            "sellersprite_required=true 时，必须启用 sellersprite_on_lens "
            "或 enrich_accepted_results。"
        )

    return ImageCompetitorRuntimeConfig(
        job_id=slugify(config_text(config, "job_id") or f"amazon-image-competitors-{now_ts()}"),
        outputs_root=resolve_path(config_text(config, "outputs_root", "outputs")),
        products_file=products_file,
        marketplace=marketplace,
        marketplace_domain=marketplace_domain,
        lens_url_template=config_text(config, "lens_url_template", "https://www.{domain}/Lens/b?ie=UTF8&node=206517768011"),
        search_strategy=config_text(config, "search_strategy", "sellersprite_find_similar_first").lower()
        or "sellersprite_find_similar_first",
        find_similar_timeout=config_int(config, "find_similar_timeout", 12) or 12,
        lens_results_timeout=config_int(config, "lens_results_timeout", 60) or 60,
        max_candidates_per_source=max_candidates_per_source,
        max_competitors_per_source=config_int(config, "max_competitors_per_source", 12) or 12,
        result_mode=result_mode,
        min_match_confidence=min_confidence,
        include_source_as_competitor=config_bool(config, "include_source_as_competitor", False),
        match_mode=match_mode,
        doubao_embedding_config_file=doubao_config_path,
        embedding_provider="doubao" if doubao_config_path else "legacy_openai",
        embedding_api_key="",
        embedding_model=DOUBAO_EMBEDDING_MODEL if doubao_config_path else vision_model,
        embedding_base_url=DOUBAO_EMBEDDING_BASE_URL if doubao_config_path else openai_base_url,
        embedding_api_path=DOUBAO_EMBEDDING_API_PATH if doubao_config_path else openai_api_path,
        embedding_encoding_format="float",
        embedding_retry_attempts=embedding_retry_attempts,
        embedding_retry_backoff_seconds=embedding_retry_backoff_seconds,
        prescreen_min_similarity=prescreen_min_similarity,
        prescreen_max_matches=prescreen_max_matches,
        doubao_mini_config_file=doubao_mini_config_path,
        mini_provider="doubao",
        mini_api_key="",
        mini_model=DOUBAO_MINI_MODEL,
        mini_base_url=DOUBAO_MINI_BASE_URL,
        mini_api_path=DOUBAO_MINI_API_PATH,
        mini_batch_size=mini_batch_size,
        mini_retry_attempts=mini_retry_attempts,
        mini_retry_backoff_seconds=mini_retry_backoff_seconds,
        vision_model=vision_model,
        openai_api_key_env=config_text(config, "openai_api_key_env", "OPENAI_API_KEY") or "OPENAI_API_KEY",
        openai_base_url=openai_base_url,
        openai_api_path=openai_api_path,
        vision_batch_size=config_int(config, "vision_batch_size", 6) or 6,
        vision_timeout=config_int(config, "vision_timeout", 120) or 120,
        resume=False if no_resume else config_bool(config, "resume", True),
        browser_backend=browser_backend,
        browser_mode=browser_mode,
        chrome_binary=config_text(config, "chrome_binary"),
        chrome_user_data_dir=resolve_path(config_text(config, "chrome_user_data_dir", "chrome_profiles/category-rank-sellersprite")),
        chrome_profile_directory=config_text(config, "chrome_profile_directory", "Default") or "Default",
        debugger_address=config_text(config, "debugger_address", "127.0.0.1:9222"),
        extension_path=extension_path,
        activate_plugin=config_bool(config, "activate_plugin", True),
        page_timeout=config_int(config, "page_timeout", 90) or 90,
        plugin_timeout=config_int(config, "plugin_timeout", 120) or 120,
        plugin_retry_attempts=max(config_int(config, "plugin_retry_attempts", 5) or 0, 0),
        plugin_retry_wait_seconds=plugin_retry_wait_min,
        plugin_retry_wait_seconds_max=plugin_retry_wait_max,
        plugin_relaunch_retry_attempts=max(config_int(config, "plugin_relaunch_retry_attempts", 0) or 0, 0),
        plugin_relaunch_wait_seconds=max(config_float(config, "plugin_relaunch_wait_seconds", 300) or 0, 0),
        plugin_second_relaunch_retry_attempts=max(config_int(config, "plugin_second_relaunch_retry_attempts", 0) or 0, 0),
        plugin_second_relaunch_wait_seconds=max(config_float(config, "plugin_second_relaunch_wait_seconds", 600) or 0, 0),
        manual_pause_timeout=config_int(config, "manual_pause_timeout", 900) or 900,
        **delivery_config,
        delay_seconds_min=min_delay,
        delay_seconds_max=max_delay,
        batch_pause_pages_min=batch_pages_min,
        batch_pause_pages_max=batch_pages_max,
        batch_pause_seconds_min=batch_seconds_min,
        batch_pause_seconds_max=batch_seconds_max,
        page_scroll_before_extract=config_bool(config, "page_scroll_before_extract", True),
        page_scroll_max_rounds=page_scroll_max_rounds,
        page_scroll_step_ratio=page_scroll_step_ratio,
        page_scroll_wait_seconds=page_scroll_wait_seconds,
        page_scroll_stable_rounds=page_scroll_stable_rounds,
        sellersprite_required=sellersprite_required,
        sellersprite_min_enriched_records=max(
            config_int(config, "sellersprite_min_enriched_records", 1) or 1,
            1,
        ),
        sellersprite_min_fields_per_record=max(
            config_int(config, "sellersprite_min_fields_per_record", 2) or 2,
            1,
        ),
        sellersprite_stable_checks=max(
            config_int(config, "sellersprite_stable_checks", 3) or 3,
            1,
        ),
        save_debug_snapshots=config_bool(config, "save_debug_snapshots", True),
        sellersprite_on_lens=sellersprite_on_lens,
        enrich_accepted_results=enrich_accepted_results,
        enrichment_page_timeout=config_int(config, "enrichment_page_timeout", 25) or 25,
        enrichment_plugin_timeout=config_int(config, "enrichment_plugin_timeout", 20) or 20,
        field_selectors={} if result_mode == "count_only" else field_selectors,
    )


def load_products(path: Path, marketplace_domain: str, dedupe: bool = True) -> List[Dict[str, Any]]:
    products: List[Dict[str, Any]] = []
    seen = set()
    rows = read_input_rows(path)
    for index, row in enumerate(rows, start=2):
        source_asin = parse_asin(pick_column(row, INPUT_ALIASES["source_asin"]))
        product_url = pick_column(row, INPUT_ALIASES["product_url"])
        image_url = pick_column(row, INPUT_ALIASES["image_url"])
        image_path = pick_column(row, INPUT_ALIASES["image_path"])
        note = pick_column(row, INPUT_ALIASES["note"])
        if not source_asin and product_url:
            source_asin = parse_asin(product_url)
        if source_asin and not product_url:
            product_url = product_url_for_asin(marketplace_domain, source_asin)
        base_source_id = source_asin or product_url or image_url or image_path or f"row-{index}"
        source_id = base_source_id if dedupe else f"{base_source_id}#row-{index}"
        if dedupe and source_id in seen:
            continue
        if not any([source_asin, product_url, image_url, image_path]):
            continue
        products.append(
            {
                "source_id": source_id,
                "source_asin": source_asin,
                "source_product_url": product_url,
                "input_image_url": image_url,
                "input_image_path": image_path,
                "input_row": index,
                "input_note": note,
            }
        )
        if dedupe:
            seen.add(source_id)
    if not products:
        raise UserFacingError("产品清单必须至少包含 ASIN、商品URL、主图URL 或本地图片路径中的一列。")
    return products


def log_failure(
    failures_path: Path,
    state: ImageCompetitorStateStore,
    current: Dict[str, Any],
    reason: str,
    message: str,
    page_url: str = "",
) -> None:
    append_jsonl(
        failures_path,
        {
            "time": now_iso(),
            "source_id": current.get("source_id", ""),
            "source_asin": current.get("source_asin", ""),
            "source_product_url": current.get("source_product_url", ""),
            "input_row": current.get("input_row", ""),
            "page_url": page_url,
            "reason": reason,
            "message": message,
        },
    )
    state.log_failure()


def request_headers() -> Dict[str, str]:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
        ),
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    }


def guess_image_suffix(url: str, content_type: str = "") -> str:
    guessed = mimetypes.guess_extension((content_type or "").split(";")[0].strip())
    if guessed and guessed.lower() in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return ".jpg" if guessed.lower() == ".jpe" else guessed.lower()
    path_suffix = Path(urlparse(url).path).suffix.lower()
    if path_suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return path_suffix
    return ".jpg"


def download_image(url: str, target_dir: Path, stem: str, timeout: int = 45) -> Path:
    ensure_dir(target_dir)
    response = requests.get(url, headers=request_headers(), timeout=timeout)
    response.raise_for_status()
    suffix = guess_image_suffix(url, response.headers.get("content-type", ""))
    path = target_dir / f"{slugify(stem)}{suffix}"
    path.write_bytes(response.content)
    return path


def image_url_to_data_url(url: str, timeout: int = 45) -> str:
    response = requests.get(url, headers=request_headers(), timeout=timeout)
    response.raise_for_status()
    mime = (response.headers.get("content-type") or "image/jpeg").split(";")[0].strip() or "image/jpeg"
    if not mime.lower().startswith("image/"):
        raise UserFacingError(
            f"候选图片下载结果不是图片（Content-Type: {mime or 'unknown'}）。"
        )
    encoded = base64.b64encode(response.content).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def resolve_local_image_path(raw_path: str) -> Path:
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = ROOT_DIR / path
    if not path.exists():
        raise UserFacingError(f"没有找到本地图片：{path}")
    return path


def extract_main_image_url(driver: WebDriver) -> str:
    script = r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (url) => {
  try { return new URL(url, location.href).href; } catch (err) { return url || ''; }
};
const usableUrl = (url) => !!url && !/^data:image\//i.test(url) && !/grey-pixel|transparent-pixel/i.test(url);
const pickFromDynamic = (el) => {
  const raw = el.getAttribute('data-a-dynamic-image') || '';
  if (!raw) return '';
  try {
    const data = JSON.parse(raw);
    let best = '';
    let bestArea = 0;
    for (const [url, size] of Object.entries(data)) {
      if (!usableUrl(url)) continue;
      const width = Array.isArray(size) ? Number(size[0] || 0) : 0;
      const height = Array.isArray(size) ? Number(size[1] || 0) : 0;
      const area = width * height;
      if (url && area >= bestArea) {
        best = url;
        bestArea = area;
      }
    }
    return best;
  } catch (err) {
    return '';
  }
};
const selectors = [
  '#landingImage',
  '#imgTagWrapperId img',
  '#main-image',
  '#ebooksImgBlkFront',
  '#imageBlock img',
  'img[data-a-dynamic-image]',
  'img.a-dynamic-image'
];
for (const selector of selectors) {
  const el = document.querySelector(selector);
  if (!el) continue;
  const dynamic = pickFromDynamic(el);
  if (dynamic) return absUrl(dynamic);
  const sources = [
    el.getAttribute('data-old-hires') || '',
    el.currentSrc || '',
    el.src || '',
    el.getAttribute('data-src') || '',
    el.getAttribute('src') || ''
  ];
  for (const src of sources) {
    if (usableUrl(src)) return absUrl(src);
  }
}
const imgs = [...document.querySelectorAll('img')].map(img => ({
  src: img.currentSrc || img.src || img.getAttribute('data-src') || '',
  alt: norm(img.getAttribute('alt') || ''),
  area: (img.naturalWidth || img.width || 0) * (img.naturalHeight || img.height || 0)
})).filter(item => usableUrl(item.src) && item.area > 10000);
imgs.sort((a, b) => b.area - a.area);
return imgs.length ? absUrl(imgs[0].src) : '';
"""
    try:
        return normalize_space(str(driver.execute_script(script) or ""))
    except (JavascriptException, WebDriverException):
        return ""


def is_transient_navigation_error(exc: BaseException) -> bool:
    message = str(exc).lower()
    return (
        "execution context was destroyed" in message
        or "cannot find context with specified id" in message
        or "most likely because of a navigation" in message
    )


def safe_driver_current_url(driver: WebDriver) -> str:
    try:
        return str(getattr(driver, "current_url", "") or "")
    except WebDriverException:
        return ""


def safe_window_handles(driver: WebDriver) -> List[str]:
    try:
        return list(driver.window_handles)
    except WebDriverException:
        return []


def source_window_restore_order(driver: WebDriver) -> List[str]:
    """Snapshot existing tabs with the active tab preferred for restoration."""
    handles = safe_window_handles(driver)
    try:
        current_handle = str(driver.current_window_handle or "")
    except WebDriverException:
        current_handle = ""
    if current_handle in handles:
        handles = [handle for handle in handles if handle != current_handle]
        handles.append(current_handle)
    return handles


def claimed_crawler_window_handles(driver: WebDriver) -> set[str]:
    raw = getattr(driver, "_image_crawler_owned_window_handles", set())
    return {str(handle) for handle in raw if str(handle)}


def claim_crawler_window_handle(driver: WebDriver, handle: str) -> None:
    """Register a tab observed after this crawler's explicit click/upload."""
    normalized = str(handle or "")
    if not normalized:
        return
    claimed = claimed_crawler_window_handles(driver)
    claimed.add(normalized)
    setattr(driver, "_image_crawler_owned_window_handles", claimed)
    register = getattr(driver, "register_owned_window_handle", None)
    if callable(register):
        try:
            register(normalized)
        except WebDriverException:
            # A short-lived popup may disappear between enumeration and
            # registration. Keeping the logical claim is harmless and lets
            # cleanup discard it once the handle is gone.
            pass


def claim_new_crawler_window_handles(
    driver: WebDriver,
    before_handles: set[str],
) -> List[str]:
    new_handles = [
        handle
        for handle in safe_window_handles(driver)
        if handle not in before_handles
    ]
    for handle in new_handles:
        claim_crawler_window_handle(driver, handle)
    return new_handles


def close_claimed_crawler_windows(
    driver: WebDriver,
    claimed_before: set[str],
    restore_handles: Sequence[str],
) -> int:
    """Close only tabs claimed during the current product and restore its base tab."""
    claimed = claimed_crawler_window_handles(driver)
    targets = claimed.difference(claimed_before)
    closed = 0
    for handle in reversed(safe_window_handles(driver)):
        if handle not in targets:
            continue
        try:
            driver.switch_to.window(handle)
            driver.close()
            claimed.discard(handle)
            closed += 1
        except WebDriverException:
            continue
    setattr(driver, "_image_crawler_owned_window_handles", claimed)
    restore_available_window(driver, restore_handles)
    return closed


def restore_available_window(driver: WebDriver, preferred_handles: Sequence[str]) -> bool:
    try:
        available = set(driver.window_handles)
    except WebDriverException:
        return False
    for handle in reversed(list(preferred_handles)):
        if handle not in available:
            continue
        try:
            driver.switch_to.window(handle)
            return True
        except WebDriverException:
            continue
    return False


def open_image_amazon_page(
    driver: WebDriver,
    url: str,
    runtime: ImageCompetitorRuntimeConfig,
    state: Optional[ImageCompetitorStateStore] = None,
) -> None:
    deadline = time.monotonic() + min(max(float(runtime.page_timeout), 1), 15)
    while True:
        try:
            open_amazon_page(
                driver,
                url,
                runtime,
                on_manual_pause=(
                    (lambda reason, page_url: state.mark_manual_pause(reason, page_url))
                    if state is not None
                    else None
                ),
                on_manual_resume=state.clear_manual_pause if state is not None else None,
            )
            return
        except WebDriverException as exc:
            if not is_transient_navigation_error(exc) or time.monotonic() >= deadline:
                raise
            time.sleep(0.25)


def handle_image_verification(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    state: Optional[ImageCompetitorStateStore],
    reason: str,
) -> None:
    if state is not None:
        state.mark_manual_pause(reason, str(getattr(driver, "current_url", "") or ""))
    if wait_for_manual_clear(driver, reason, runtime.manual_pause_timeout):
        if state is not None:
            state.clear_manual_pause()
        return
    raise VerificationUnconfirmedError(verification_unconfirmed_message(reason))


def handle_image_sellersprite_block(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    state: Optional[ImageCompetitorStateStore],
) -> None:
    handle_image_verification(
        driver,
        runtime,
        state,
        sellersprite_block_reason(driver),
    )


def resolve_source_image(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    image_dir: Path,
    state: Optional[ImageCompetitorStateStore] = None,
) -> Path:
    raw_path = normalize_space(str(current.get("input_image_path") or ""))
    if raw_path:
        path = resolve_local_image_path(raw_path)
        current["source_image"] = str(path)
        return path

    image_url = normalize_space(str(current.get("input_image_url") or ""))
    if not image_url:
        product_url = normalize_space(str(current.get("source_product_url") or ""))
        if not product_url:
            source_asin = normalize_space(str(current.get("source_asin") or ""))
            if source_asin:
                product_url = product_url_for_asin(runtime.marketplace_domain, source_asin)
                current["source_product_url"] = product_url
        if not product_url:
            raise UserFacingError("该行没有可用的图片、ASIN 或商品链接。")
        open_image_amazon_page(driver, product_url, runtime, state)
        WebDriverWait(driver, runtime.page_timeout).until(lambda d: bool(extract_main_image_url(d)))
        image_url = extract_main_image_url(driver)
        if not image_url:
            raise UserFacingError("商品页没有提取到主图。")
        current["input_image_url"] = image_url

    image_path = download_image(image_url, image_dir, str(current.get("source_id") or "source"))
    current["source_image"] = image_url
    return image_path


def ensure_lens_supported(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    state: ImageCompetitorStateStore,
    failures_path: Path,
    debug_dir: Path,
) -> None:
    open_image_amazon_page(driver, runtime.lens_url, runtime, state)
    block_reason = detect_block(driver)
    if block_reason:
        try:
            handle_image_verification(driver, runtime, state, block_reason)
        except VerificationUnconfirmedError:
            if runtime.save_debug_snapshots:
                save_debug_snapshot(driver, debug_dir, f"lens_{block_reason}")
            raise

    def has_file_input(d: WebDriver) -> bool:
        try:
            return bool(d.execute_script("return !!document.querySelector('input[type=file]');"))
        except (JavascriptException, WebDriverException):
            return False

    try:
        WebDriverWait(driver, min(runtime.page_timeout, 45)).until(has_file_input)
    except TimeoutException as exc:
        if runtime.save_debug_snapshots:
            save_debug_snapshot(driver, debug_dir, f"lens_unsupported_{runtime.marketplace_domain}")
        append_jsonl(
            failures_path,
            {
                "time": now_iso(),
                "source_id": "",
                "source_asin": "",
                "page_url": driver.current_url,
                "reason": "lens_unsupported",
                "message": f"{runtime.marketplace_domain} 当前页面未检测到 Amazon 以图搜图上传入口。",
            },
        )
        raise UserFacingError(
            f"{runtime.marketplace_domain} 当前未检测到 Amazon 以图搜图上传入口，无法执行后续任务。"
        ) from exc


def upload_image_to_lens(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    image_path: Path,
    state: Optional[ImageCompetitorStateStore] = None,
) -> None:
    open_image_amazon_page(driver, runtime.lens_url, runtime, state)
    script = r"""
const input = document.querySelector('input[type=file]');
if (!input) return false;
input.removeAttribute('hidden');
input.style.display = 'block';
input.style.visibility = 'visible';
input.style.opacity = '1';
input.style.position = 'fixed';
input.style.left = '8px';
input.style.top = '8px';
input.style.zIndex = '2147483647';
return true;
"""
    def reveal_file_input(d: WebDriver) -> bool:
        try:
            return bool(d.execute_script(script))
        except (JavascriptException, WebDriverException) as exc:
            if is_transient_navigation_error(exc):
                return False
            raise

    ok = bool(WebDriverWait(driver, runtime.page_timeout).until(reveal_file_input))
    if not ok:
        raise UserFacingError("Amazon Lens 页面未找到图片上传控件。")
    before_handles = set(safe_window_handles(driver))
    before_url = safe_driver_current_url(driver)
    setattr(driver, "_image_upload_window_baseline", (before_handles, before_url))
    input_el = driver.find_element(By.CSS_SELECTOR, "input[type=file]")
    try:
        input_el.send_keys(str(image_path.resolve()))
    except WebDriverException as exc:
        claim_new_crawler_window_handles(driver, before_handles)
        # Setting the file starts a client-side navigation immediately. A
        # destroyed old context is evidence that the upload was dispatched;
        # the following Lens-result wait still verifies the destination.
        if not is_transient_navigation_error(exc):
            raise


def trigger_sellersprite_find_similar(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    state: Optional[ImageCompetitorStateStore] = None,
) -> bool:
    before_handle_order: List[str] = []
    claimed_before = claimed_crawler_window_handles(driver)
    product_url = normalize_space(str(current.get("source_product_url") or ""))
    source_asin = normalize_space(str(current.get("source_asin") or ""))
    if not product_url and source_asin:
        product_url = product_url_for_asin(runtime.marketplace_domain, source_asin)
        current["source_product_url"] = product_url
    if not product_url:
        return False
    try:
        current_url = normalize_space(str(getattr(driver, "current_url", "") or ""))
        if not (source_asin and source_asin.upper() in current_url.upper()):
            open_image_amazon_page(driver, product_url, runtime, state)
        WebDriverWait(driver, min(runtime.page_timeout, 45)).until(lambda d: bool(extract_main_image_url(d)))
        script = r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const fireMouse = (el, type) => el && el.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
const image = document.querySelector('#landingImage,#imgTagWrapperId img,#main-image,img[data-a-dynamic-image],img.a-dynamic-image');
if (image) {
  image.scrollIntoView({block: 'center', inline: 'center'});
  fireMouse(image, 'mouseover');
  fireMouse(image, 'mouseenter');
  fireMouse(image, 'mousemove');
}
return true;
"""
        driver.execute_script(script)
        time.sleep(1.5)
        before_handle_order = list(driver.window_handles)
        before_handles = set(before_handle_order)
        before_url = normalize_space(str(getattr(driver, "current_url", "") or ""))
        clicked = bool(
            driver.execute_script(
                r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const visible = (el) => {
  const r = el.getBoundingClientRect();
  const style = getComputedStyle(el);
  return r.width > 0 && r.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
};
const sellerspriteRoot = document.querySelector('#seller-sprite-extension-find-similar-listing');
if (!sellerspriteRoot) return false;
const candidates = [...sellerspriteRoot.querySelectorAll('.find-similar,button,a,p,span')]
  .filter(el => visible(el))
  .map(el => ({el, text: norm(el.innerText || el.textContent || el.getAttribute('aria-label') || el.getAttribute('title') || '')}))
  .filter(item => /^(找相似|Find Similar)$/i.test(item.text))
  .sort((a, b) => {
    const ar = a.el.getBoundingClientRect();
    const br = b.el.getBoundingClientRect();
    return (ar.width * ar.height) - (br.width * br.height);
  });
if (!candidates.length) return false;
candidates[0].el.click();
return true;
"""
            )
        )
        if not clicked:
            return False
        WebDriverWait(driver, min(runtime.page_timeout, runtime.find_similar_timeout)).until(
            lambda d: switch_to_find_similar_result(d, before_handles, before_url)
        )
        current_result_url = str(getattr(driver, "current_url", "") or "")
        on_manual_pause = (
            (lambda reason, page_url: state.mark_manual_pause(reason, page_url))
            if state is not None
            else None
        )
        on_manual_resume = state.clear_manual_pause if state is not None else None
        handle_amazon_verification(
            driver,
            runtime,
            on_manual_pause,
            on_manual_resume,
        )
        ensure_amazon_delivery_location(
            driver,
            runtime,
            original_url=current_result_url,
            on_manual_pause=on_manual_pause,
            on_manual_resume=on_manual_resume,
        )
        return True
    except (JavascriptException, TimeoutException, WebDriverException):
        claim_new_crawler_window_handles(driver, set(before_handle_order))
        close_claimed_crawler_windows(driver, claimed_before, before_handle_order)
        return False


def is_lens_result_url(url: str) -> bool:
    normalized = normalize_space(url)
    if not normalized:
        return False
    try:
        parsed = urlparse(normalized)
    except ValueError:
        return False
    path = (parsed.path or "").lower()
    if re.search(r"/(?:dp|gp/product)/", path, re.I):
        return False
    if re.search(r"/(?:lens|stylesnap)(?:/|$)", path, re.I):
        return True
    if not path.rstrip("/").endswith("/products"):
        return False
    query = {
        str(key).lower(): [str(value).lower() for value in values]
        for key, values in parse_qs(parsed.query, keep_blank_values=True).items()
    }
    return (
        "flow" in query.get("searchtype", [])
        or any("stylesnap" in value for value in query.get("modes", []))
    )


def switch_to_find_similar_result(
    driver: WebDriver,
    before_handles: set[str],
    before_url: str = "",
) -> bool:
    try:
        new_handles = claim_new_crawler_window_handles(driver, before_handles)
        if new_handles:
            driver.switch_to.window(new_handles[-1])
        url = normalize_space(str(getattr(driver, "current_url", "") or ""))
        if before_url and url == before_url and not new_handles:
            return False
        return is_lens_result_url(url)
    except (JavascriptException, WebDriverException):
        return False


def run_image_search(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    source_image_path: Path,
    state: Optional[ImageCompetitorStateStore] = None,
) -> str:
    if runtime.search_strategy in {"sellersprite_find_similar_first", "find_similar_first"}:
        if trigger_sellersprite_find_similar(driver, runtime, current, state):
            return "sellersprite_find_similar"
    upload_image_to_lens(driver, runtime, source_image_path, state)
    return "amazon_upload"


def lens_no_results_visible(driver: WebDriver) -> bool:
    try:
        return bool(
            driver.execute_script(
                r"""
const visible = (el) => {
  if (!el) return false;
  const style = getComputedStyle(el);
  const rect = el.getBoundingClientRect();
  return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
};
const explicit = document.querySelector('.no-styles-found,[data-testid="noStylesFound"],[data-testid="no-results"]');
if (visible(explicit)) return true;
const text = String(document.body?.innerText || '').replace(/\s+/g, ' ').trim();
return /Oops!\s*No styles found for this look\.?/i.test(text);
"""
            )
        )
    except (JavascriptException, WebDriverException):
        return False


def wait_for_lens_results(driver: WebDriver, runtime: ImageCompetitorRuntimeConfig) -> str:
    def result_status(d: WebDriver) -> str:
        try:
            baseline = getattr(d, "_image_upload_window_baseline", None)
            if isinstance(baseline, tuple) and len(baseline) == 2:
                before_handles, before_url = baseline
                switch_to_find_similar_result(
                    d,
                    set(before_handles),
                    str(before_url or ""),
                )
            if not is_lens_result_url(str(getattr(d, "current_url", "") or "")):
                return ""
            cards = extract_lens_candidate_cards(d, include_text=False)
            if cards:
                return "results"
            if lens_no_results_visible(d):
                return "no_results"
            return ""
        except WebDriverException:
            return ""

    try:
        return str(
            WebDriverWait(driver, min(runtime.page_timeout, runtime.lens_results_timeout)).until(
                result_status
            )
        )
    finally:
        baseline = getattr(driver, "_image_upload_window_baseline", None)
        if isinstance(baseline, tuple) and len(baseline) == 2:
            claim_new_crawler_window_handles(driver, set(baseline[0]))
        setattr(driver, "_image_upload_window_baseline", None)


def detect_block_after_navigation(driver: WebDriver, timeout_seconds: float = 15) -> str:
    """Retry block detection while an upload/click is replacing the page context."""
    deadline = time.monotonic() + max(0.5, float(timeout_seconds))
    while True:
        try:
            return detect_block(driver)
        except WebDriverException as exc:
            if not is_transient_navigation_error(exc) or time.monotonic() >= deadline:
                raise
            time.sleep(0.25)


def upload_and_wait_for_lens_results(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    state: Optional[ImageCompetitorStateStore] = None,
) -> str:
    upload_image_to_lens(driver, runtime, source_image_path, state)
    block_reason = detect_block_after_navigation(
        driver,
        min(runtime.page_timeout, 15),
    )
    if block_reason:
        handle_image_verification(driver, runtime, state, block_reason)
    return wait_for_lens_results(driver, runtime)


def wait_for_lens_sellersprite_data(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    timeout_seconds: Optional[float] = None,
) -> str:
    if not runtime.sellersprite_required:
        report = {
            "status": "not_required",
            "checked_at": now_iso(),
            "page_url": str(getattr(driver, "current_url", "") or ""),
        }
        set_sellersprite_readiness(driver, report)
        return "not_required"
    if runtime.activate_plugin:
        try_activate_plugin(driver)
    preload_page_data_with_scroll(driver, runtime)
    timeout = runtime.plugin_timeout if timeout_seconds is None else max(float(timeout_seconds), 1)
    deadline = time.time() + timeout
    stable_seen = 0
    last_signature = ""
    last_status = "data_loading"
    while time.time() < deadline:
        blocked_reason = detect_block(driver)
        if blocked_reason:
            report = {
                "status": "blocked",
                "checked_at": now_iso(),
                "page_url": str(getattr(driver, "current_url", "") or ""),
                "blocked": True,
                "blocked_reason": blocked_reason,
            }
            set_sellersprite_readiness(driver, report)
            return "blocked"
        cards = extract_lens_candidate_cards(driver, include_text=True)
        table_rows = extract_table_rows(driver)
        field_counts_by_asin: Dict[str, int] = {}
        for row in table_rows:
            asin = normalize_space(str(row.get("asin") or "")).upper()
            if not asin:
                continue
            parsed = parse_table_row_fields(row)
            count = sum(1 for field_name in SELLERSPRITE_EVIDENCE_FIELDS if parsed.get(field_name))
            field_counts_by_asin[asin] = max(field_counts_by_asin.get(asin, 0), count)
        for card in cards:
            asin = normalize_space(str(card.get("asin") or "")).upper()
            if not asin:
                continue
            text = str(card.get("text") or "")
            count = sum(
                1
                for field_name in SELLERSPRITE_EVIDENCE_FIELDS
                if parse_field_from_text(field_name, text)
            )
            field_counts_by_asin[asin] = max(field_counts_by_asin.get(asin, 0), count)
        node_count = plugin_node_count(driver)
        min_fields = max(runtime.sellersprite_min_fields_per_record, 1)
        report = {
            "status": "data_loading",
            "checked_at": now_iso(),
            "page_url": str(getattr(driver, "current_url", "") or ""),
            "plugin_nodes": node_count,
            "login_required": sellersprite_login_required(driver),
            "product_count": len({str(card.get("asin") or "") for card in cards if card.get("asin")}),
            "enriched_records": sum(1 for count in field_counts_by_asin.values() if count >= min_fields),
            "max_fields_per_record": max(field_counts_by_asin.values(), default=0),
            "stable_checks": 0,
            "blocked": False,
            "blocked_reason": "",
        }
        report["status"] = classify_sellersprite_snapshot(
            report,
            runtime.sellersprite_min_enriched_records,
            min_fields,
        )
        signature = (
            f"{node_count}:{len(cards)}:{len(table_rows)}:"
            + ",".join(f"{asin}:{count}" for asin, count in sorted(field_counts_by_asin.items()))
        )
        if report["status"] == "ready_candidate":
            stable_seen = stable_seen + 1 if signature == last_signature else 1
            report["stable_checks"] = stable_seen
            if stable_seen >= runtime.sellersprite_stable_checks:
                report["status"] = "ready"
                set_sellersprite_readiness(driver, report)
                return "ok"
            report["status"] = "data_loading"
        else:
            stable_seen = 0
        last_status = str(report["status"])
        set_sellersprite_readiness(driver, report)
        last_signature = signature
        time.sleep(1)
    return last_status


def strip_html_text(value: str) -> str:
    text = re.sub(r"(?is)<(script|style|svg)\b.*?</\1>", " ", value or "")
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return normalize_space(html_lib.unescape(text))


def first_html_attr(block: str, attr_name: str) -> str:
    pattern = rf"""{re.escape(attr_name)}\s*=\s*["']([^"']+)["']"""
    match = re.search(pattern, block, re.I)
    return html_lib.unescape(match.group(1)) if match else ""


def normalize_candidate_image_url(value: Any, page_url: str = "") -> str:
    """Keep only explicit image data URLs or plausible HTTP image resources."""
    raw = normalize_space(str(value or ""))
    if re.match(r"^data:image/", raw, re.I):
        return raw
    if raw.lower().startswith("data:"):
        return ""
    try:
        parsed = urlparse(raw)
    except (TypeError, ValueError):
        return ""
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""
    try:
        current = urlparse(normalize_space(page_url))
    except (TypeError, ValueError):
        current = None
    if current and current.scheme and current.netloc:
        same_page = (
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path.rstrip("/"),
            parsed.query,
        ) == (
            current.scheme.lower(),
            current.netloc.lower(),
            current.path.rstrip("/"),
            current.query,
        )
        if same_page:
            return ""
    if re.search(r"/(?:dp|gp/product)/", parsed.path, re.I):
        return ""
    if not (
        re.search(r"/images/", parsed.path, re.I)
        or re.search(r"\.(?:jpe?g|png|webp|gif|avif)$", parsed.path, re.I)
    ):
        return ""
    return raw


def normalize_candidate_cards_image_urls(
    cards: Sequence[Dict[str, Any]],
    page_url: str = "",
) -> List[Dict[str, Any]]:
    normalized_cards: List[Dict[str, Any]] = []
    for card in cards:
        if not isinstance(card, dict):
            continue
        copied = dict(card)
        copied["candidate_image_url"] = normalize_candidate_image_url(
            copied.get("candidate_image_url"),
            page_url,
        )
        normalized_cards.append(copied)
    return normalized_cards


def extract_lens_candidate_cards_from_html(page_html: str, include_text: bool = True) -> List[Dict[str, Any]]:
    article_pattern = re.compile(
        r"""<article\b(?=[^>]*\bdata-csa-c-item-type\s*=\s*["']asin["'])(?P<attrs>[^>]*)>(?P<body>.*?)</article>""",
        re.I | re.S,
    )
    cards: List[Dict[str, Any]] = []
    seen_asins = set()
    for match in article_pattern.finditer(page_html or ""):
        attrs = match.group("attrs") or ""
        body = match.group("body") or ""
        raw_item_id = first_html_attr(attrs, "data-csa-c-item-id")
        asin_match = (
            re.search(r"\basin\.([A-Z0-9]{10})\b", raw_item_id, re.I)
            or re.search(r"\b([A-Z0-9]{10})\b", raw_item_id, re.I)
            or re.search(r"\bASIN\s*:?\s*([A-Z0-9]{10})\b", strip_html_text(body), re.I)
        )
        if not asin_match:
            continue
        asin = asin_match.group(1).upper()
        if asin in seen_asins:
            continue
        product_match = re.search(
            rf"""href\s*=\s*["']([^"']*/(?:dp|gp/product)/{re.escape(asin)}[^"']*)["']""",
            body,
            re.I,
        )
        product_url = html_lib.unescape(product_match.group(1)) if product_match else ""
        image_match = re.search(r"""<img\b[^>]*\bsrc\s*=\s*["']([^"']*m\.media-amazon\.com/images/[^"']+)["']""", body, re.I)
        image_url = html_lib.unescape(image_match.group(1)) if image_match else ""
        title_match = re.search(r"""<h5\b[^>]*>(.*?)</h5>""", body, re.I | re.S)
        title = strip_html_text(title_match.group(1)) if title_match else ""
        rank = first_html_attr(attrs, "data-csa-c-posx") or str(len(cards) + 1)
        seen_asins.add(asin)
        cards.append(
            {
                "asin": asin,
                "title": title,
                "product_url": product_url,
                "candidate_image_url": image_url,
                "rank": str(rank),
                "seller_country_flag_code": "",
                "text": strip_html_text(body) if include_text else "",
            }
        )
    return cards


def extract_lens_candidate_cards(driver: WebDriver, include_text: bool = True) -> List[Dict[str, Any]]:
    script = r"""
const includeText = arguments[0];
const asinRe = /\b([A-Z0-9]{10})\b/;
const asinUrlRe = /\/(?:dp|gp\/product)\/([A-Z0-9]{10})(?:[/?#]|$)/i;
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const imageFromSrcset = (value) => {
  if (!value) return '';
  const parts = value.split(',').map(part => part.trim()).filter(Boolean);
  let best = '';
  let bestWidth = 0;
  for (const part of parts) {
    const items = part.split(/\s+/);
    const url = items[0] || '';
    const width = Number((items[1] || '').replace(/[^\d.]/g, '')) || 0;
    if (url && width >= bestWidth) {
      best = url;
      bestWidth = width;
    }
  }
  return best;
};
const usableImageUrl = (value) => {
  if (!value) return false;
  if (/^data:image\//i.test(value)) return true;
  try {
    const parsed = new URL(value, location.href);
    if (!/^https?:$/i.test(parsed.protocol)) return false;
    if (parsed.href.split('#')[0] === location.href.split('#')[0]) return false;
    if (/\/(?:dp|gp\/product)\//i.test(parsed.pathname)) return false;
    return /\/images\//i.test(parsed.pathname)
      || /\.(?:jpe?g|png|webp|gif|avif)$/i.test(parsed.pathname);
  } catch (err) {
    return false;
  }
};
const getAsin = (el) => {
  for (const attr of ['data-asin', 'asin', 'data-csa-c-asin', 'data-csa-c-item-id']) {
    const value = el.getAttribute(attr);
    if (!value) continue;
    const cleaned = value.trim();
    if (/^[A-Z0-9]{10}$/i.test(cleaned)) return cleaned.toUpperCase();
    const attrMatch = cleaned.match(/(?:^|[.:_-])([A-Z0-9]{10})(?:$|[.:_-])/i) || cleaned.match(/\basin\.([A-Z0-9]{10})\b/i);
    if (attrMatch) return attrMatch[1].toUpperCase();
  }
  for (const a of el.querySelectorAll('a[href]')) {
    const match = a.href.match(asinUrlRe);
    if (match) return match[1].toUpperCase();
  }
  const textMatch = norm(el.innerText || el.textContent || '').match(asinRe);
  return textMatch ? textMatch[1].toUpperCase() : '';
};
const getUrl = (el, asin) => {
  const links = [...el.querySelectorAll('a[href]')];
  for (const a of links) {
    if (asin && a.href.includes(asin)) return absUrl(a.href);
  }
  for (const a of links) {
    if (/\/(?:dp|gp\/product)\//i.test(a.href)) return absUrl(a.href);
  }
  return '';
};
const getTitle = (el) => {
  const selectors = ['h2 a span', 'h2 span', 'a.a-link-normal span', 'img[alt]', '.p13n-sc-truncate'];
  for (const selector of selectors) {
    const item = el.querySelector(selector);
    if (!item) continue;
    const text = norm(item.getAttribute('alt') || item.innerText || item.textContent || '');
    if (text && text.length > 5) return text;
  }
  return '';
};
const getImageUrl = (el) => {
  const imgs = [...el.querySelectorAll('img')];
  let best = '';
  let bestArea = 0;
  for (const img of imgs) {
    const src = [
      img.getAttribute('data-old-hires'),
      img.currentSrc,
      img.getAttribute('src'),
      img.getAttribute('data-src'),
      imageFromSrcset(img.getAttribute('srcset') || '')
    ].find(usableImageUrl) || '';
    if (!src) continue;
    const area = (img.naturalWidth || img.width || 0) * (img.naturalHeight || img.height || 0);
    if (!best || area >= bestArea) {
      best = src;
      bestArea = area;
    }
  }
  return best ? absUrl(best) : '';
};
const getSellerCountryFlagCode = (el) => {
  const flagElements = [...el.querySelectorAll('[class*="flag-icon-"], [class*="icp-nav-flag-"]')];
  for (const flag of flagElements) {
    const classText = String(flag.getAttribute('class') || '');
    const match = classText.match(/(?:flag-icon|icp-nav-flag)-([a-z]{2})\b/i);
    if (match) return match[1].toLowerCase();
    const hint = norm(flag.getAttribute('title') || flag.getAttribute('aria-label') || '');
    if (hint) return hint;
  }
  return '';
};
const isExcludedContainer = (el) => {
  for (let cur = el; cur && cur !== document.body; cur = cur.parentElement) {
    const id = String(cur.id || '').toLowerCase();
    const cls = String(cur.className || '').toLowerCase();
    if (['nav-belt', 'navbar', 'nav-main', 'navfooter', 'ewc-content', 'nav-flyout-ewc'].includes(id)) return true;
    if (id.includes('ewc') || cls.includes('ewc') || cls.includes('cart')) return true;
  }
  const hrefs = [...el.querySelectorAll('a[href]')].map(a => a.href).join(' ');
  if (/ewc_|\/cart|\/gp\/cart|ref=ewc/i.test(hrefs)) return true;
  return false;
};
const selectors = [
  'article.cellContainer[data-csa-c-item-type="asin"]',
  'article[data-csa-c-item-id*=".asin."]',
  '[data-csa-c-item-id*=".asin."]',
  '#product_grid_container article',
  '.s-result-item[data-asin]:not([data-asin=""])',
  '[data-component-type="s-search-result"][data-asin]:not([data-asin=""])',
  '#gridItemRoot',
  '.zg-grid-general-faceout',
  '.p13n-grid-content',
  '[data-asin]:not([data-asin=""])'
];
// A single selector-list query preserves document order across all card shapes.
// Running one query per selector would group cards by selector priority and can
// change which product becomes the 11th cascade prescreen match.
const elements = [...document.querySelectorAll(selectors.join(','))]
  .filter(el => !isExcludedContainer(el));
const seenAsins = new Set();
const cards = [];
for (const el of elements) {
  const asin = getAsin(el);
  if (!asin || seenAsins.has(asin)) continue;
  const productUrl = getUrl(el, asin);
  if (/ewc_|\/cart|\/gp\/cart|ref=ewc/i.test(productUrl)) continue;
  const imageUrl = getImageUrl(el);
  seenAsins.add(asin);
  cards.push({
    asin,
    title: getTitle(el),
    product_url: productUrl,
    candidate_image_url: imageUrl,
    rank: String(cards.length + 1),
    seller_country_flag_code: getSellerCountryFlagCode(el),
    text: includeText ? norm(el.innerText || el.textContent || '') : ''
  });
}
return cards;
"""
    try:
        cards = list(driver.execute_script(script, include_text) or [])
        if cards:
            return normalize_candidate_cards_image_urls(
                cards,
                normalize_space(str(getattr(driver, "current_url", "") or "")),
            )
    except (JavascriptException, WebDriverException):
        pass
    try:
        cards = extract_lens_candidate_cards_from_html(
            str(driver.page_source or ""),
            include_text=include_text,
        )
        return normalize_candidate_cards_image_urls(
            cards,
            normalize_space(str(getattr(driver, "current_url", "") or "")),
        )
    except WebDriverException:
        return []


def collect_lens_candidate_cards(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    include_text: bool = True,
) -> List[Dict[str, Any]]:
    collected: List[Dict[str, Any]] = []
    seen_asins = set()

    def merge(cards: Sequence[Dict[str, Any]]) -> None:
        for card in cards:
            asin = str(card.get("asin") or "").upper()
            if not asin or asin in seen_asins:
                continue
            copied = dict(card)
            copied["asin"] = asin
            copied["rank"] = copied.get("rank") or str(len(collected) + 1)
            collected.append(copied)
            seen_asins.add(asin)

    merge(extract_lens_candidate_cards(driver, include_text=include_text))
    if len(collected) >= runtime.max_candidates_per_source:
        return collected[: runtime.max_candidates_per_source]

    stable_rounds = 0
    scroll_attempts = min(8, max(2, math.ceil(runtime.max_candidates_per_source / 12)))
    for _ in range(scroll_attempts):
        previous_count = len(collected)
        try:
            driver.execute_script(
                r"""
const scrollables = [document.scrollingElement, document.documentElement, document.body]
  .filter(Boolean);
for (const el of scrollables) {
  try { el.scrollBy(0, Math.max(600, window.innerHeight * 0.85)); } catch (err) {}
}
const grids = [
  '#product_grid_container',
  '[data-testid*="grid"]',
  '[class*="ProductGrid"]',
  '[class*="product-grid"]'
].flatMap(selector => [...document.querySelectorAll(selector)]);
for (const grid of grids) {
  try { grid.scrollBy(0, Math.max(600, window.innerHeight * 0.85)); } catch (err) {}
}
"""
            )
        except (JavascriptException, WebDriverException):
            break
        time.sleep(0.8)
        merge(extract_lens_candidate_cards(driver, include_text=include_text))
        if len(collected) >= runtime.max_candidates_per_source:
            break
        stable_rounds = stable_rounds + 1 if len(collected) == previous_count else 0
        if stable_rounds >= 2:
            break
    return collected[: runtime.max_candidates_per_source]


def merge_lens_product_data(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    plugin_status: str,
) -> List[Dict[str, Any]]:
    include_plugin_data = not runtime.is_count_only
    cards = collect_lens_candidate_cards(driver, runtime, include_text=include_plugin_data)
    table_rows = extract_table_rows(driver) if include_plugin_data else []
    table_by_asin: Dict[str, Dict[str, str]] = {}
    for row in table_rows:
        asin = str(row.get("asin") or "")
        if not asin:
            continue
        parsed = parse_table_row_fields(row)
        target = table_by_asin.setdefault(asin, {})
        for key, value in parsed.items():
            if value and not target.get(key):
                target[key] = value

    records: List[Dict[str, Any]] = []
    for card in cards[: runtime.max_candidates_per_source]:
        asin = str(card.get("asin") or "")
        if not asin:
            continue
        record: Dict[str, Any] = {
            "source_type": "image_search",
            "source_id": current.get("source_id", ""),
            "source_asin": current.get("source_asin", ""),
            "source_product_url": current.get("source_product_url", ""),
            "source_image": current.get("source_image") or current.get("input_image_url") or current.get("input_image_path") or "",
            "input_row": current.get("input_row", ""),
            "asin": asin,
            "title": card.get("title") or "",
            "product_url": card.get("product_url") or "",
            "candidate_image_url": card.get("candidate_image_url") or "",
            "rank": card.get("rank") or "",
            "scraped_at": now_iso(),
            "load_status": plugin_status,
            "note": "",
        }
        text = str(card.get("text") or "")
        for field_name in REQUESTED_DATA_FIELDS:
            record[field_name] = ""
        if include_plugin_data:
            for field_name, value in table_by_asin.get(asin, {}).items():
                if field_name in REQUESTED_DATA_FIELDS and value:
                    record[field_name] = value
            selector_values = extract_by_selectors(driver, card, runtime.field_selectors)
            for field_name, value in selector_values.items():
                if field_name in REQUESTED_DATA_FIELDS and value:
                    record[field_name] = normalize_space(str(value))
            for field_name in REQUESTED_DATA_FIELDS:
                if not record.get(field_name):
                    record[field_name] = parse_field_from_text(field_name, text)
            if not record.get("seller_country"):
                record["seller_country"] = country_from_flag_code_or_text(str(card.get("seller_country_flag_code") or ""))
            missing_count = sum(1 for field_name in REQUESTED_DATA_FIELDS if not record.get(field_name))
            if plugin_status != "ok":
                record["note"] = "插件数据加载超时，已保存页面可见数据"
            elif missing_count == len(REQUESTED_DATA_FIELDS):
                record["note"] = "插件未展示或选择器未匹配"
        records.append(record)
    return records


def image_to_data_url(path: Path) -> str:
    mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def image_ref_to_embedding_input(image_ref: str) -> Dict[str, Any]:
    return {"type": "image_url", "image_url": {"url": image_ref}}


def extract_embedding_vector(data: Dict[str, Any]) -> List[float]:
    items = data.get("data")
    if isinstance(items, dict):
        item = items
    elif isinstance(items, list) and items and isinstance(items[0], dict):
        item = items[0]
    else:
        raise EmbeddingProviderError("多模态向量模型返回格式异常：缺少 data 对象。")

    embedding = item.get("embedding")
    if (
        isinstance(embedding, list)
        and len(embedding) == 1
        and isinstance(embedding[0], list)
    ):
        embedding = embedding[0]
    if not isinstance(embedding, list) or not embedding:
        raise EmbeddingProviderError("多模态向量模型未返回非空 embedding。")

    vector: List[float] = []
    for value in embedding:
        if isinstance(value, bool) or isinstance(value, (dict, list, tuple)):
            raise EmbeddingProviderError("多模态向量模型返回了无效 embedding 数值。")
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise EmbeddingProviderError("多模态向量模型返回了无效 embedding 数值。") from exc
        if not math.isfinite(number):
            raise EmbeddingProviderError("多模态向量模型返回了非有限 embedding 数值。")
        vector.append(number)
    if not any(value != 0 for value in vector):
        raise EmbeddingProviderError("多模态向量模型返回了全零 embedding。")
    return vector


def embedding_http_error_is_fatal(response: requests.Response) -> bool:
    status = int(response.status_code)
    if status in {401, 403, 404}:
        return True

    error_text = str(getattr(response, "text", "") or "")
    try:
        payload = response.json()
    except (TypeError, ValueError):
        payload = None
    if isinstance(payload, dict):
        error = payload.get("error")
        if isinstance(error, dict):
            error_text += " " + " ".join(
                str(error.get(field) or "") for field in ("code", "type", "message")
            )
        error_text += " " + " ".join(
            str(payload.get(field) or "") for field in ("code", "type", "message")
        )
    normalized = error_text.lower()
    fatal_code_markers = (
        "model_not_opened",
        "model_not_found",
        "modelnotfound",
        "invalid_model",
        "endpoint_not_found",
        "endpointnotfound",
        "invalid_endpoint_id",
    )
    model_configuration_error = "model" in normalized and any(
        marker in normalized
        for marker in (
            "not activated",
            "not enabled",
            "not found",
            "no permission",
            "permission denied",
            "access denied",
            "invalid model",
        )
    )
    if any(marker in normalized for marker in fatal_code_markers) or model_configuration_error:
        return True
    if status == 400:
        return False
    if status in {408, 413, 415, 422, 429} or 500 <= status < 600:
        return False
    return 400 <= status < 500


def call_multimodal_embedding(runtime: ImageCompetitorRuntimeConfig, image_ref: str) -> List[float]:
    api_key = runtime.embedding_api_key
    if not api_key:
        raise EmbeddingProviderError("视觉向量服务尚未完成凭据校验。")
    payload = {
        "model": runtime.embedding_model,
        "input": [image_ref_to_embedding_input(image_ref)],
        "encoding_format": runtime.embedding_encoding_format,
    }
    endpoint = _provider_url(runtime.embedding_base_url, runtime.embedding_api_path)
    attempts = max(runtime.embedding_retry_attempts, 1)
    response: Optional[requests.Response] = None
    for attempt in range(attempts):
        try:
            add_provider_metric(runtime, "embedding_api_calls")
            response = requests.post(
                endpoint,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json=payload,
                timeout=runtime.vision_timeout,
                allow_redirects=False,
            )
        except requests.RequestException as exc:
            if attempt + 1 < attempts:
                time.sleep(runtime.embedding_retry_backoff_seconds * (2**attempt))
                continue
            safe_message = redact_sensitive_text(exc, (api_key,))[:300]
            raise EmbeddingProviderError(f"多模态向量模型请求失败：{safe_message}") from exc

        status = int(response.status_code)
        if 300 <= status < 400:
            raise FatalEmbeddingProviderError(
                f"豆包视觉向量端点返回了未接受的重定向（HTTP {status}）；"
                "请直接配置最终 HTTPS 端点。"
            )
        if status < 400:
            break
        retryable = status in RETRYABLE_EMBEDDING_STATUS_CODES or 500 <= status < 600
        if retryable and attempt + 1 < attempts:
            time.sleep(runtime.embedding_retry_backoff_seconds * (2**attempt))
            continue
        if status in {401, 403}:
            raise FatalEmbeddingProviderError(
                f"豆包视觉向量 API 鉴权失败（HTTP {status}）；请检查专用配置文件中的 API Key 和模型权限。"
            )
        if status == 404:
            raise FatalEmbeddingProviderError(
                "豆包视觉向量 API 端点不存在（HTTP 404）；请检查 base_url 和 api_path。"
            )
        if embedding_http_error_is_fatal(response):
            raise FatalEmbeddingProviderError(
                f"豆包视觉向量请求配置错误（HTTP {status}）；"
                "请检查模型是否已开通以及端点配置。"
            )
        if status in {400, 413, 415, 422}:
            raise EmbeddingProviderError(
                f"豆包视觉向量未接受当前图片输入（HTTP {status}）；"
                "可改用本地转码图片重试。"
            )
        if status == 408:
            raise EmbeddingProviderError("豆包视觉向量请求重试后仍超时（HTTP 408）。")
        if status == 429:
            raise EmbeddingProviderError("豆包视觉向量请求重试后仍被限流（HTTP 429）。")
        if 500 <= status < 600:
            raise EmbeddingProviderError(f"豆包视觉向量服务重试后仍不可用（HTTP {status}）。")
        raise FatalEmbeddingProviderError(f"多模态向量模型调用失败（HTTP {status}）。")

    if response is None:  # pragma: no cover - the loop always executes
        raise EmbeddingProviderError("多模态向量模型请求未执行。")
    try:
        data = response.json()
    except ValueError as exc:
        raise EmbeddingProviderError("多模态向量模型返回内容不是有效 JSON。") from exc
    if not isinstance(data, dict):
        raise EmbeddingProviderError("多模态向量模型返回格式异常：根节点不是 JSON 对象。")
    record_provider_usage(runtime, data, "embedding")
    return extract_embedding_vector(data)


def call_multimodal_embedding_cached(runtime: ImageCompetitorRuntimeConfig, image_ref: str) -> List[float]:
    digest = hashlib.sha256(image_ref.encode("utf-8", errors="ignore")).hexdigest()
    cache_key = (
        f"{runtime.embedding_provider}|{runtime.embedding_base_url}|"
        f"{runtime.embedding_api_path}|{runtime.embedding_model}|{digest}"
    )
    cached = EMBEDDING_CACHE.get(cache_key)
    if cached is not None:
        return cached
    embedding = call_multimodal_embedding(runtime, image_ref)
    EMBEDDING_CACHE[cache_key] = embedding
    return embedding


def cosine_similarity(left: Sequence[float], right: Sequence[float]) -> float:
    if not left or not right:
        raise EmbeddingProviderError("无法比较空的图片向量。")
    if len(left) != len(right):
        raise EmbeddingProviderError(
            f"图片向量维度不一致：来源 {len(left)}，候选 {len(right)}。"
        )
    if not all(math.isfinite(value) for value in (*left, *right)):
        raise EmbeddingProviderError("图片向量包含非有限数值。")
    dot = sum(a * b for a, b in zip(left, right))
    norm_left = math.sqrt(sum(a * a for a in left))
    norm_right = math.sqrt(sum(b * b for b in right))
    if not norm_left or not norm_right:
        raise EmbeddingProviderError("图片向量范数为零，无法计算相似度。")
    return dot / (norm_left * norm_right)


def resolve_source_embedding(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    source_image_ref: str,
) -> List[float]:
    source_refs = [image_to_data_url(source_image_path)]
    if source_image_ref and source_image_ref not in source_refs:
        source_refs.append(source_image_ref)
    source_errors: List[str] = []
    for source_ref in source_refs:
        try:
            return call_multimodal_embedding_cached(runtime, source_ref)
        except FatalEmbeddingProviderError:
            raise
        except EmbeddingProviderError as exc:
            source_errors.append(str(exc)[:240])
    raise EmbeddingProviderError("来源图片向量识别失败：" + " | ".join(source_errors))


def resolve_candidate_embedding(
    runtime: ImageCompetitorRuntimeConfig,
    candidate: Dict[str, Any],
) -> tuple[str, Optional[List[float]], str]:
    asin = str(candidate.get("asin") or "").upper()
    image_url = normalize_space(str(candidate.get("candidate_image_url") or ""))
    if not asin or not image_url:
        return asin, None, "候选商品缺少可识别图片。"
    try:
        return asin, call_multimodal_embedding_cached(runtime, image_url), ""
    except FatalEmbeddingProviderError:
        raise
    except EmbeddingProviderError as exc:
        first_error = str(exc)[:180]
    try:
        data_url = image_url_to_data_url(image_url, timeout=min(runtime.vision_timeout, 45))
        return (
            asin,
            call_multimodal_embedding_cached(runtime, data_url),
            "候选图片 URL 识别失败后已改用本地转码图片。",
        )
    except FatalEmbeddingProviderError:
        raise
    except (EmbeddingProviderError, UserFacingError, requests.RequestException) as exc:
        return (
            asin,
            None,
            f"候选图片识别失败：{first_error}；本地转码重试失败：{str(exc)[:180]}",
        )


def extract_response_text(data: Dict[str, Any]) -> str:
    choices = data.get("choices")
    if isinstance(choices, list) and choices:
        message = (choices[0] or {}).get("message") if isinstance(choices[0], dict) else {}
        content = message.get("content") if isinstance(message, dict) else ""
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            parts = []
            for item in content:
                if isinstance(item, dict) and isinstance(item.get("text"), str):
                    parts.append(item["text"])
            return "\n".join(parts)
    if isinstance(data.get("output_text"), str):
        return data["output_text"]
    parts: List[str] = []
    for item in data.get("output") or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content") or []:
            if not isinstance(content, dict):
                continue
            text = content.get("text")
            if isinstance(text, str):
                parts.append(text)
            output_text = content.get("output_text")
            if isinstance(output_text, str):
                parts.append(output_text)
    return "\n".join(parts)


def parse_json_object(text: str) -> Dict[str, Any]:
    text = text.strip()
    if not text:
        return {}
    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, re.S | re.I)
    if fenced:
        text = fenced.group(1).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            return json.loads(text[start : end + 1])
        raise


def build_match_prompt(min_confidence: float, candidates: Sequence[Dict[str, Any]]) -> str:
    candidate_lines = []
    for index, candidate in enumerate(candidates, start=1):
        candidate_lines.append(
            f"{index}. ASIN={candidate.get('asin','')} title={candidate.get('title','')} rank={candidate.get('rank','')}"
        )
    return (
        "你是电商同款竞品视觉审核员。基准图片是来源商品，后面每张候选图片是亚马逊以图搜图结果。\n"
        "核心任务：判断候选图片里的“商品主体”是否和基准图片里的商品主体属于同款/可替代竞品。\n\n"
        "必须遵守的判断规则：\n"
        "1. 只看商品本体，不看整张图片构图、背景、白底、堆叠方式、上方单品+下方集合的版式、光影风格、拍摄角度或模特姿势。\n"
        "2. 先判断硬性门槛：商品类型、用途、核心结构必须一致。只是主题相近、节日氛围相近、构图相近，不能判为同款。\n"
        "3. 重点比较主体细节：形状比例、轮廓、表面纹理/分瓣/凹槽、顶部五金/挂绳/挂环、材质质感、颜色组合、套装构成。\n"
        "4. 允许差异：品牌不同、卖家不同、颜色数量略有不同、展示数量不同、套装/售卖个数不同、配件数量不同、主图排列方式不同。\n"
        "5. 如果候选只是图片排版或构图更像基准图，但商品细节明显不同，必须降低置信度。\n"
        "6. 如果候选构图不一样，但商品本体的核心形态、材质、颜色组合、五金/配件结构高度一致，应该给更高置信度。\n\n"
        "置信度标尺：\n"
        "- 0.90-1.00：几乎同款，商品本体细节高度一致。\n"
        "- 0.80-0.89：高置信同款，只有数量、颜色展示、角度、配件数量等轻微差异。\n"
        "- 0.65-0.79：疑似同款/强竞品，主体一致但关键细节有一些差异。\n"
        "- 0.40-0.64：同类但不同款，不能保留为同款竞品。\n"
        "- 0.00-0.39：明显不同商品。\n\n"
        f"只有 confidence >= {min_confidence:.2f} 且商品主体确实一致时，is_competitor 才能为 true。\n"
        "请只返回 JSON，不要输出解释性正文。格式："
        "{\"matches\":[{\"asin\":\"...\",\"is_competitor\":true,\"confidence\":0.0,"
        "\"reason\":\"简短说明主体商品为何同款或不同款\","
        "\"matched_features\":[\"匹配细节\"],\"different_features\":[\"差异细节\"],"
        "\"composition_similarity_ignored\":true}]}。\n"
        "候选列表：\n"
        + "\n".join(candidate_lines)
    )


def call_openai_vision(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    candidates: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    api_key = os.environ.get(runtime.openai_api_key_env)
    if not api_key:
        raise UserFacingError(f"缺少视觉模型 API Key，请先设置环境变量 {runtime.openai_api_key_env}。")

    content: List[Dict[str, Any]] = [
        {"type": "input_text", "text": build_match_prompt(runtime.min_match_confidence, candidates)},
        {"type": "input_text", "text": "基准图片："},
        {"type": "input_image", "image_url": image_to_data_url(source_image_path)},
    ]
    for index, candidate in enumerate(candidates, start=1):
        image_url = normalize_space(str(candidate.get("candidate_image_url") or ""))
        if not image_url:
            continue
        content.append({"type": "input_text", "text": f"候选 {index} / ASIN {candidate.get('asin','')}："})
        content.append({"type": "input_image", "image_url": image_url})

    if runtime.openai_api_path.endswith("chat/completions"):
        chat_content: List[Dict[str, Any]] = []
        for item in content:
            if item.get("type") == "input_text":
                chat_content.append({"type": "text", "text": item.get("text", "")})
            elif item.get("type") == "input_image":
                chat_content.append({"type": "image_url", "image_url": {"url": item.get("image_url", "")}})
        payload = {
            "model": runtime.vision_model,
            "messages": [{"role": "user", "content": chat_content}],
            "max_tokens": 1800,
            "temperature": 0,
        }
    else:
        payload = {
            "model": runtime.vision_model,
            "input": [{"role": "user", "content": content}],
            "max_output_tokens": 1800,
            "temperature": 0,
        }
    try:
        response = requests.post(
            f"{runtime.openai_base_url}/{runtime.openai_api_path}",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            timeout=runtime.vision_timeout,
        )
    except requests.RequestException as exc:
        safe_message = redact_sensitive_text(exc, (api_key,))[:300]
        raise EmbeddingProviderError(f"视觉模型请求失败：{safe_message}") from exc
    if response.status_code >= 400:
        raise EmbeddingProviderError(f"视觉模型调用失败：HTTP {response.status_code}。")
    try:
        data = response.json()
    except ValueError as exc:
        raise EmbeddingProviderError("视觉模型返回内容不是有效 JSON。") from exc
    text = extract_response_text(data)
    try:
        parsed = parse_json_object(text)
    except json.JSONDecodeError as exc:
        raise EmbeddingProviderError("视觉模型返回内容不是有效 JSON。") from exc
    matches = parsed.get("matches") if isinstance(parsed, dict) else []
    return [dict(item) for item in matches if isinstance(item, dict)]


def build_mini_same_product_prompt(candidates: Sequence[Dict[str, Any]]) -> str:
    candidate_lines = [
        f"{index}. ASIN={candidate.get('asin', '')} title={normalize_space(str(candidate.get('title') or ''))}"
        for index, candidate in enumerate(candidates, start=1)
    ]
    return (
        "你是亚马逊商品同款审核员。第一张图是来源商品，后续图片与候选列表一一对应。\n"
        "只判断商品主体是否为同款，不判断图片构图是否相似。必须遵守：\n"
        "1. 主体产品、核心用途与核心结构相同，即使颜色不同，也视为同款。\n"
        "2. 主体产品相同，配件种类或数量不同、套装件数不同、售卖数量不同，仍视为同款。\n"
        "3. 忽略背景、文字、模特、拍摄角度、排列方式和包装差异。\n"
        "4. 核心品类、核心功能或主体结构不同，必须判为不同款。\n"
        "5. 标题只能辅助识别品类与结构，不能推翻图片中明确可见的主体证据。\n"
        "对每个候选给出 0-1 confidence 和一句简短原因。只返回 JSON 对象，不要输出其他文字。\n"
        "严格格式：{\"matches\":[{\"asin\":\"B000000000\","
        "\"is_same_product\":true,\"confidence\":0.95,\"reason\":\"主体与核心结构一致\"}]}。\n"
        "候选列表：\n" + "\n".join(candidate_lines)
    )


def parse_mini_match_response(
    text: str,
    candidates: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    parsed = json.loads(text.strip())
    if not isinstance(parsed, dict) or set(parsed) != {"matches"}:
        raise ValueError("Mini 返回根对象必须且只能包含 matches。")
    matches = parsed.get("matches")
    if not isinstance(matches, list):
        raise ValueError("Mini 返回缺少 matches 数组。")
    expected = [str(item.get("asin") or "").upper() for item in candidates]
    expected_set = set(expected)
    if not expected or "" in expected_set or len(expected_set) != len(expected):
        raise ValueError("Mini 输入候选 ASIN 无效或重复。")
    normalized: Dict[str, Dict[str, Any]] = {}
    for item in matches:
        if not isinstance(item, dict):
            raise ValueError("Mini matches 包含非对象项。")
        if set(item) != {"asin", "is_same_product", "confidence", "reason"}:
            raise ValueError("Mini match 对象字段不符合固定契约。")
        asin_raw = item.get("asin")
        if not isinstance(asin_raw, str):
            raise ValueError("Mini asin 必须是字符串。")
        asin = asin_raw.upper()
        if asin not in expected_set:
            raise ValueError("Mini 返回了未知 ASIN。")
        if asin in normalized:
            raise ValueError("Mini 返回了重复 ASIN。")
        same_product = item.get("is_same_product")
        if not isinstance(same_product, bool):
            raise ValueError("Mini is_same_product 必须是布尔值。")
        confidence_raw = item.get("confidence")
        if not isinstance(confidence_raw, (int, float)) or isinstance(confidence_raw, bool):
            raise ValueError("Mini confidence 必须是 0-1 数字。")
        confidence = float(confidence_raw)
        if not math.isfinite(confidence) or not 0 <= confidence <= 1:
            raise ValueError("Mini confidence 必须是 0-1 数字。")
        reason_raw = item.get("reason")
        if not isinstance(reason_raw, str):
            raise ValueError("Mini reason 必须是非空字符串。")
        reason = normalize_space(reason_raw)
        if not reason:
            raise ValueError("Mini reason 不能为空。")
        normalized[asin] = {
            "asin": asin,
            "is_same_product": same_product,
            "confidence": round(confidence, 4),
            "reason": reason,
        }
    if set(normalized) != expected_set:
        raise ValueError("Mini 返回未完整覆盖本批候选 ASIN。")
    return [normalized[asin] for asin in expected]


def record_provider_usage(
    runtime: ImageCompetitorRuntimeConfig,
    data: Dict[str, Any],
    prefix: str,
) -> None:
    usage = data.get("usage")
    if not isinstance(usage, dict):
        return
    for source_keys, metric_key in (
        (("prompt_tokens", "input_tokens"), f"{prefix}_prompt_tokens"),
        (("completion_tokens", "output_tokens"), f"{prefix}_completion_tokens"),
        (("total_tokens",), f"{prefix}_total_tokens"),
    ):
        value = next((usage.get(key) for key in source_keys if key in usage), None)
        if isinstance(value, int) and not isinstance(value, bool) and value >= 0:
            add_provider_metric(runtime, metric_key, value)
    for detail_key in ("prompt_tokens_details", "input_tokens_details"):
        details = usage.get(detail_key)
        if not isinstance(details, dict):
            continue
        image_tokens = details.get("image_tokens")
        if isinstance(image_tokens, int) and not isinstance(image_tokens, bool) and image_tokens >= 0:
            add_provider_metric(runtime, f"{prefix}_image_tokens", image_tokens)


def record_mini_usage(runtime: ImageCompetitorRuntimeConfig, data: Dict[str, Any]) -> None:
    record_provider_usage(runtime, data, "mini")


def call_doubao_mini_verifier(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    candidates: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not candidates:
        return []
    if not runtime.mini_api_key:
        raise MiniProviderError("豆包 Mini 同款复核服务尚未完成凭据校验。")
    content: List[Dict[str, Any]] = [
        {"type": "text", "text": build_mini_same_product_prompt(candidates)},
        {"type": "text", "text": "来源商品主图："},
        {"type": "image_url", "image_url": {"url": image_to_data_url(source_image_path)}},
    ]
    for index, candidate in enumerate(candidates, start=1):
        content.extend(
            [
                {
                    "type": "text",
                    "text": f"候选 {index} / ASIN {candidate.get('asin', '')}：",
                },
                {
                    "type": "image_url",
                    "image_url": {
                        "url": normalize_space(str(candidate.get("candidate_image_url") or ""))
                    },
                },
            ]
        )
    payload = {
        "model": runtime.mini_model,
        "messages": [{"role": "user", "content": content}],
        "max_tokens": 1800,
        "temperature": 0,
        "thinking": {"type": "disabled"},
        "response_format": {"type": "json_object"},
    }
    endpoint = _provider_url(runtime.mini_base_url, runtime.mini_api_path)
    attempts = max(runtime.mini_retry_attempts, 1)
    last_error = ""
    for attempt in range(attempts):
        try:
            add_provider_metric(runtime, "mini_api_calls")
            response = requests.post(
                endpoint,
                headers={
                    "Authorization": f"Bearer {runtime.mini_api_key}",
                    "Content-Type": "application/json",
                },
                json=payload,
                timeout=runtime.vision_timeout,
                allow_redirects=False,
            )
        except requests.RequestException as exc:
            last_error = redact_sensitive_text(exc, (runtime.mini_api_key,))[:240]
            if attempt + 1 < attempts:
                time.sleep(runtime.mini_retry_backoff_seconds * (2**attempt))
                continue
            raise MiniProviderError(f"豆包 Mini 请求失败：{last_error}") from exc

        status = int(response.status_code)
        if 300 <= status < 400:
            raise FatalMiniProviderError(
                f"豆包 Mini 端点返回了未接受的重定向（HTTP {status}）；"
                "请直接配置最终 HTTPS 端点。"
            )
        if status >= 400:
            if status in {401, 403}:
                raise FatalMiniProviderError(
                    f"豆包 Mini API 鉴权失败（HTTP {status}）；请检查本地配置文件和模型权限。"
                )
            if status == 404 or embedding_http_error_is_fatal(response):
                raise FatalMiniProviderError(
                    f"豆包 Mini 模型或端点配置错误（HTTP {status}）。"
                )
            retryable = status in {408, 429} or 500 <= status < 600
            last_error = f"HTTP {status}"
            if retryable and attempt + 1 < attempts:
                time.sleep(runtime.mini_retry_backoff_seconds * (2**attempt))
                continue
            raise MiniProviderError(f"豆包 Mini 调用失败：HTTP {status}。")
        try:
            data = response.json()
            if not isinstance(data, dict):
                raise ValueError("响应根节点不是 JSON 对象。")
            record_mini_usage(runtime, data)
            matches = parse_mini_match_response(extract_response_text(data), candidates)
        except (ValueError, json.JSONDecodeError) as exc:
            last_error = str(exc)[:240]
            if attempt + 1 < attempts:
                time.sleep(runtime.mini_retry_backoff_seconds * (2**attempt))
                continue
            raise MiniProviderError(
                f"豆包 Mini 连续 {attempts} 次未返回完整结构化结果：{last_error}"
            ) from exc
        return matches
    raise MiniProviderError(f"豆包 Mini 复核失败：{last_error}")  # pragma: no cover


def run_cascade_match(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    source_image_ref: str,
    records: Sequence[Dict[str, Any]],
) -> MatchEvaluation:
    before_metrics = provider_metric_snapshot(runtime)
    decisions: Dict[str, Dict[str, Any]] = {}
    source_asin = next(
        (str(record.get("source_asin") or "").upper() for record in records if record.get("source_asin")),
        "",
    )
    candidates: List[Dict[str, Any]] = []
    for record in records:
        asin = str(record.get("asin") or "").upper()
        if not asin:
            continue
        candidate = dict(record)
        candidate["asin"] = asin
        if source_asin and asin == source_asin and not runtime.include_source_as_competitor:
            decisions[asin] = {
                "is_competitor": False,
                "match_confidence": "",
                "match_reason": "候选 ASIN 与来源 ASIN 相同，已按配置排除。",
                "prescreen_status": "excluded_source_asin",
            }
            continue
        candidates.append(candidate)

    if not candidates:
        return MatchEvaluation(
            accepted_records=[],
            decisions=decisions,
            prescreen_visual_match_count=0,
            processing_status="verified_zero",
            same_product_count=0,
            same_product_confidence="",
            match_reason="没有需要进入视觉粗筛的候选商品，最终同款数量为 0。",
            provider_metrics=provider_metric_delta(before_metrics, provider_metric_snapshot(runtime)),
        )

    source_embedding = resolve_source_embedding(runtime, source_image_path, source_image_ref)
    prescreen_matches: List[Dict[str, Any]] = []
    processed_asins = set()
    for candidate in candidates:
        asin = str(candidate.get("asin") or "").upper()
        processed_asins.add(asin)
        if not candidate.get("candidate_image_url"):
            raise EmbeddingProviderError(
                f"候选 {asin} 缺少可识别图片；本次不写入同款数量，请修复后重试。"
            )
        _, candidate_embedding, note = resolve_candidate_embedding(runtime, candidate)
        if candidate_embedding is None:
            raise EmbeddingProviderError(
                f"候选 {asin} 无法完成向量识别（{note}）；本次不写入同款数量，请重试。"
            )
        similarity = cosine_similarity(source_embedding, candidate_embedding)
        is_prescreen_match = similarity >= runtime.prescreen_min_similarity
        decision_reason = (
            f"Embedding 粗筛相似度 {similarity:.4f}，阈值 {runtime.prescreen_min_similarity:.2f}。"
        )
        if note:
            decision_reason = f"{decision_reason} {note}"
        decisions[asin] = {
            "is_competitor": False if not is_prescreen_match else "",
            "match_confidence": round(similarity, 4),
            "match_reason": decision_reason,
            "prescreen_similarity": round(similarity, 4),
            "prescreen_match": is_prescreen_match,
            "prescreen_status": "matched" if is_prescreen_match else "rejected",
        }
        if not is_prescreen_match:
            continue
        prescreen_matches.append(candidate)
        if len(prescreen_matches) > runtime.prescreen_max_matches:
            for remaining in candidates:
                remaining_asin = str(remaining.get("asin") or "").upper()
                if remaining_asin in processed_asins:
                    continue
                decisions[remaining_asin] = {
                    "is_competitor": "",
                    "match_confidence": "",
                    "match_reason": "达到粗筛排除阈值后未继续调用向量服务。",
                    "prescreen_status": "not_evaluated_after_limit",
                }
            return MatchEvaluation(
                accepted_records=[],
                decisions=decisions,
                prescreen_visual_match_count=len(prescreen_matches),
                processing_status="prescreen_excluded",
                same_product_count=None,
                same_product_confidence="",
                match_reason=(
                    f"按 Lens 顺序发现至少 {len(prescreen_matches)} 个视觉近似候选，"
                    f"超过上限 {runtime.prescreen_max_matches}；已停止粗筛且未调用 Mini。"
                ),
                provider_metrics=provider_metric_delta(
                    before_metrics,
                    provider_metric_snapshot(runtime),
                ),
            )

    if not prescreen_matches:
        return MatchEvaluation(
            accepted_records=[],
            decisions=decisions,
            prescreen_visual_match_count=0,
            processing_status="verified_zero",
            same_product_count=0,
            same_product_confidence="",
            match_reason="Embedding 粗筛没有命中候选，最终同款数量为 0，未调用 Mini。",
            provider_metrics=provider_metric_delta(before_metrics, provider_metric_snapshot(runtime)),
        )

    accepted: List[Dict[str, Any]] = []
    for start in range(0, len(prescreen_matches), runtime.mini_batch_size):
        batch = prescreen_matches[start : start + runtime.mini_batch_size]
        for match in call_doubao_mini_verifier(runtime, source_image_path, batch):
            asin = str(match["asin"])
            candidate = next(item for item in batch if str(item.get("asin") or "").upper() == asin)
            is_same = bool(match["is_same_product"])
            confidence = float(match["confidence"])
            reason = str(match["reason"])
            decisions[asin].update(
                {
                    "is_competitor": is_same,
                    "match_confidence": confidence,
                    "match_reason": reason,
                    "mini_is_same_product": is_same,
                    "mini_confidence": confidence,
                    "mini_reason": reason,
                }
            )
            if is_same:
                accepted_row = dict(candidate)
                accepted_row.update(
                    {
                        "is_competitor": True,
                        "match_confidence": confidence,
                        "match_reason": reason,
                    }
                )
                accepted.append(accepted_row)

    confidence_value: float | str = ""
    if accepted:
        confidence_value = round(
            min(float(item.get("match_confidence") or 0) for item in accepted),
            4,
        )
    return MatchEvaluation(
        accepted_records=accepted,
        decisions=decisions,
        prescreen_visual_match_count=len(prescreen_matches),
        processing_status="verified",
        same_product_count=len(accepted),
        same_product_confidence=confidence_value,
        match_reason=(
            f"Mini 已复核 {len(prescreen_matches)} 个视觉粗筛候选，"
            f"确认 {len(accepted)} 个同款；最终数量只采用 Mini 结果。"
        ),
        provider_metrics=provider_metric_delta(before_metrics, provider_metric_snapshot(runtime)),
    )


def evaluate_competitor_matches(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    source_image_ref: str,
    records: Sequence[Dict[str, Any]],
) -> MatchEvaluation:
    if runtime.match_mode == "cascade":
        return run_cascade_match(runtime, source_image_path, source_image_ref, records)
    before_metrics = provider_metric_snapshot(runtime)
    accepted, decisions = filter_high_confidence_competitors(
        runtime,
        source_image_path,
        source_image_ref,
        records,
    )
    confidences = [
        float(item.get("match_confidence"))
        for item in accepted
        if isinstance(item.get("match_confidence"), (int, float))
        and not isinstance(item.get("match_confidence"), bool)
    ]
    return MatchEvaluation(
        accepted_records=accepted,
        decisions=decisions,
        prescreen_visual_match_count="",
        processing_status="verified",
        same_product_count=len(accepted),
        same_product_confidence=round(min(confidences), 4) if confidences else "",
        match_reason=f"旧版 {runtime.match_mode} 模式完成，保留原有判断语义。",
        provider_metrics=provider_metric_delta(before_metrics, provider_metric_snapshot(runtime)),
    )


def filter_high_confidence_competitors(
    runtime: ImageCompetitorRuntimeConfig,
    source_image_path: Path,
    source_image_ref: str,
    records: Sequence[Dict[str, Any]],
) -> tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    output: List[Dict[str, Any]] = []
    decisions: Dict[str, Dict[str, Any]] = {}
    source_asin = ""
    for record in records:
        source_asin = str(record.get("source_asin") or "").upper()
        if source_asin:
            break

    by_asin: Dict[str, Dict[str, Any]] = {}
    candidates: List[Dict[str, Any]] = []
    candidate_preparation_errors: List[str] = []
    for record in records:
        asin = str(record.get("asin") or "").upper()
        if not asin:
            continue
        copied = dict(record)
        copied["asin"] = asin
        by_asin[asin] = copied
        if source_asin and asin == source_asin and not runtime.include_source_as_competitor:
            decisions[asin] = {
                "is_competitor": False,
                "match_confidence": "",
                "match_reason": "候选 ASIN 与来源 ASIN 相同，已按配置排除。",
            }
            continue
        if not copied.get("candidate_image_url"):
            candidate_preparation_errors.append(f"{asin}: 缺少可识别图片")
            decisions[asin] = {
                "is_competitor": False,
                "match_confidence": "",
                "match_reason": "候选商品缺少可识别图片。",
            }
            continue
        candidates.append(copied)

    def rank_value(item: Dict[str, Any]) -> int:
        match = re.match(r"\d+", str(item.get("rank") or ""))
        return int(match.group(0)) if match else 9999

    if runtime.match_mode == "embedding":
        if candidate_preparation_errors:
            summary = "；".join(candidate_preparation_errors[:5])
            raise EmbeddingProviderError(
                f"候选图片无法完成向量识别（{summary}）。本次不写入相似竞品数量，请修复后重试。"
            )
        source_embedding = resolve_source_embedding(
            runtime,
            source_image_path,
            source_image_ref,
        )

        max_workers = min(max(runtime.vision_batch_size, 1), 4, max(len(candidates), 1))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(resolve_candidate_embedding, runtime, candidate)
                for candidate in candidates
            ]
            for future in as_completed(futures):
                asin, candidate_embedding, note = future.result()
                if not asin or asin not in by_asin:
                    continue
                candidate = by_asin[asin]
                if candidate_embedding is None:
                    candidate_preparation_errors.append(f"{asin}: {note}")
                    continue
                confidence = cosine_similarity(source_embedding, candidate_embedding)
                is_competitor = confidence >= runtime.min_match_confidence
                candidate["is_competitor"] = is_competitor
                candidate["match_confidence"] = round(confidence, 4)
                candidate["match_reason"] = (
                    f"多模态图片向量相似度 {confidence:.4f}，阈值 {runtime.min_match_confidence:.2f}。"
                )
                if note:
                    candidate["match_reason"] = f"{candidate['match_reason']} {note}"
                decisions[asin] = {
                    "is_competitor": is_competitor,
                    "match_confidence": candidate["match_confidence"],
                    "match_reason": candidate["match_reason"],
                }
                if is_competitor:
                    output.append(candidate)
        if candidate_preparation_errors:
            summary = "；".join(candidate_preparation_errors[:5])
            raise EmbeddingProviderError(
                f"候选图片无法完成向量识别（{summary}）。本次不写入相似竞品数量，请修复后重试。"
            )
        output.sort(key=lambda item: (-float(item.get("match_confidence") or 0), rank_value(item)))
        if runtime.is_count_only:
            return output, decisions
        return output[: runtime.max_competitors_per_source], decisions

    batch_size = max(runtime.vision_batch_size, 1)
    for start in range(0, len(candidates), batch_size):
        batch = candidates[start : start + batch_size]
        matches = call_openai_vision(runtime, source_image_path, batch)
        for match in matches:
            asin = str(match.get("asin") or "").upper()
            if not asin or asin not in by_asin:
                continue
            try:
                confidence = float(match.get("confidence") or 0)
            except (TypeError, ValueError):
                confidence = 0.0
            is_competitor = bool(match.get("is_competitor"))
            candidate = by_asin[asin]
            candidate["is_competitor"] = is_competitor
            candidate["match_confidence"] = confidence
            candidate["match_reason"] = normalize_space(str(match.get("reason") or ""))
            decisions[asin] = {
                "is_competitor": is_competitor,
                "match_confidence": confidence,
                "match_reason": candidate["match_reason"],
            }
            if is_competitor and confidence >= runtime.min_match_confidence:
                output.append(candidate)
    output.sort(key=lambda item: (-float(item.get("match_confidence") or 0), rank_value(item)))
    if runtime.is_count_only:
        return output, decisions
    return output[: runtime.max_competitors_per_source], decisions


def amazon_search_url_for_asin(domain: str, asin: str) -> str:
    return f"https://www.{domain}/s?k={quote_plus(asin)}"


def enrich_accepted_records(
    driver: WebDriver,
    runtime: ImageCompetitorRuntimeConfig,
    records: Sequence[Dict[str, Any]],
    state: Optional[ImageCompetitorStateStore] = None,
) -> List[Dict[str, Any]]:
    if not runtime.enrich_accepted_results or not records:
        return [dict(record) for record in records]

    enriched_records = [dict(record) for record in records]
    old_page_timeout = runtime.page_timeout
    old_plugin_timeout = runtime.plugin_timeout
    runtime.page_timeout = min(runtime.page_timeout, runtime.enrichment_page_timeout)
    runtime.plugin_timeout = runtime.enrichment_plugin_timeout
    try:
        for record in enriched_records:
            asin = str(record.get("asin") or "").upper()
            if not asin:
                continue
            open_image_amazon_page(
                driver,
                amazon_search_url_for_asin(runtime.marketplace_domain, asin),
                runtime,
                state,
            )
            block_reason = detect_block(driver)
            if block_reason:
                handle_image_verification(driver, runtime, state, block_reason)
            try:
                wait_for_amazon_products(driver, runtime)  # type: ignore[arg-type]
            except TimeoutException:
                pass
            plugin_status = wait_for_sellersprite_data(driver, runtime)  # type: ignore[arg-type]
            if plugin_status == "blocked":
                handle_image_sellersprite_block(driver, runtime, state)
                plugin_status = wait_for_sellersprite_data(driver, runtime)  # type: ignore[arg-type]
                if plugin_status == "blocked":
                    raise VerificationUnconfirmedError(
                        verification_unconfirmed_message(
                            sellersprite_block_reason(driver)
                        )
                    )
            if runtime.sellersprite_required and plugin_status != "ok":
                raise UserFacingError(
                    f"卖家精灵数据未达到写入门禁：{plugin_status}。"
                )
            source_context = {
                "source_id": record.get("source_id", ""),
                "source_asin": record.get("source_asin", ""),
                "source_product_url": record.get("source_product_url", ""),
                "source_image": record.get("source_image", ""),
                "input_row": record.get("input_row", ""),
            }
            page_records = merge_lens_product_data(driver, runtime, source_context, plugin_status)
            match = next((item for item in page_records if str(item.get("asin") or "").upper() == asin), None)
            if not match:
                record["load_status"] = f"enrich_{plugin_status}_not_found"
                if not record.get("note"):
                    record["note"] = "补抓页未定位到该 ASIN，保留以图搜图候选信息。"
                continue
            for field_name in REQUESTED_DATA_FIELDS:
                if match.get(field_name):
                    record[field_name] = match[field_name]
            for field_name in ("title", "product_url", "candidate_image_url"):
                if match.get(field_name):
                    record[field_name] = match[field_name]
            record["load_status"] = plugin_status
            missing_count = sum(1 for field_name in REQUESTED_DATA_FIELDS if not record.get(field_name))
            if plugin_status == "ok" and missing_count < len(REQUESTED_DATA_FIELDS):
                if record.get("note") == "插件数据加载超时，已保存页面可见数据":
                    record["note"] = ""
            elif not record.get("note"):
                record["note"] = "补抓页未完整展示卖家精灵字段。"
    finally:
        runtime.page_timeout = old_page_timeout
        runtime.plugin_timeout = old_plugin_timeout
    return enriched_records


def build_candidate_audit_rows(
    records: Sequence[Dict[str, Any]],
    accepted: Sequence[Dict[str, Any]],
    decisions: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    accepted_by_asin = {str(row.get("asin") or ""): row for row in accepted}
    rows: List[Dict[str, Any]] = []
    for record in records:
        asin = str(record.get("asin") or "")
        accepted_row = accepted_by_asin.get(asin)
        decision = decisions.get(asin, {})
        rows.append(
            {
                "source_id": record.get("source_id", ""),
                "source_asin": record.get("source_asin", ""),
                "asin": asin,
                "title": record.get("title", ""),
                "product_url": record.get("product_url", ""),
                "candidate_image_url": record.get("candidate_image_url", ""),
                "rank": record.get("rank", ""),
                "is_competitor": decision.get("is_competitor", accepted_row.get("is_competitor", False) if accepted_row else False),
                "confidence": decision.get("match_confidence", accepted_row.get("match_confidence", "") if accepted_row else ""),
                "reason": decision.get("match_reason", accepted_row.get("match_reason", "") if accepted_row else ""),
                "prescreen_similarity": decision.get("prescreen_similarity", ""),
                "prescreen_status": decision.get("prescreen_status", ""),
                "mini_is_same_product": decision.get("mini_is_same_product", ""),
                "mini_confidence": decision.get("mini_confidence", ""),
                "mini_reason": decision.get("mini_reason", ""),
                "load_status": record.get("load_status", ""),
            }
        )
    return rows


def write_sheet(ws: Any, headers: Sequence[str], rows: Sequence[Dict[str, Any]], field_to_header: Optional[Dict[str, str]] = None) -> None:
    ws.append(list(headers))
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        if field_to_header:
            ws.append([row.get(field, "") for field, header in field_to_header.items() if header in headers])
        else:
            ws.append([row.get(header, "") for header in headers])
    for column_index, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(column_index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def replace_sheet(wb: Any, title: str) -> Any:
    if title in wb.sheetnames:
        wb.remove(wb[title])
    return wb.create_sheet(title)


def write_workbook(
    source_workbook_path: Path,
    records_path: Path,
    candidates_path: Path,
    failures_path: Path,
    output_xlsx: Path,
) -> None:
    if load_workbook is None:
        raise UserFacingError("缺少 openpyxl，无法生成 Excel。")
    records = read_jsonl(records_path)
    candidates = read_jsonl(candidates_path)
    failures = read_jsonl(failures_path)

    wb = load_workbook(source_workbook_path)
    ws = replace_sheet(wb, "同款竞品结果")
    write_sheet(ws, IMAGE_COMPETITOR_HEADERS, records, IMAGE_COMPETITOR_FIELD_TO_HEADER)

    ws_candidates = replace_sheet(wb, "全部候选")
    write_sheet(ws_candidates, CANDIDATE_HEADERS, candidates)

    ws_fail = replace_sheet(wb, "失败记录")
    failure_headers = ["time", "source_id", "source_asin", "source_product_url", "input_row", "page_url", "reason", "message"]
    write_sheet(ws_fail, failure_headers, failures)

    ensure_dir(output_xlsx.parent)
    wb.save(output_xlsx)


def count_display_value(runtime: ImageCompetitorRuntimeConfig, accepted_count: int) -> int | str:
    if accepted_count >= runtime.max_candidates_per_source:
        return f"{runtime.max_candidates_per_source}+"
    return accepted_count


def append_count_result(
    counts_path: Path,
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    search_method: str,
    plugin_status: str,
    candidate_count: int,
    evaluation: MatchEvaluation,
) -> None:
    append_jsonl(
        counts_path,
        build_count_result_row(
            runtime,
            current,
            search_method,
            plugin_status,
            candidate_count,
            evaluation,
        ),
    )


def mini_confirmed_same_product_count_value(result: Dict[str, Any]) -> int | str:
    """Return the final business-facing Mini-only count display.

    A `verified_zero` result is an embedding-only zero and therefore remains
    blank. `prescreen_excluded` is intentionally not a Mini result, so it
    receives the user-facing over-ten marker instead of a numeric count.
    """
    if MINI_CONFIRMED_COUNT_FIELD in result:
        value = result.get(MINI_CONFIRMED_COUNT_FIELD)
        return value if value is not None else ""
    if result.get("match_mode") != "cascade":
        return ""
    processing_status = normalize_space(str(result.get("processing_status") or ""))
    if processing_status == "prescreen_excluded":
        return EMBEDDING_GREATER_THAN_TEN_LABEL
    same_product_count = result.get("same_product_count")
    if processing_status == "verified" and isinstance(same_product_count, int):
        return same_product_count
    return ""


def build_count_result_row(
    runtime: ImageCompetitorRuntimeConfig,
    current: Dict[str, Any],
    search_method: str,
    plugin_status: str,
    candidate_count: int,
    evaluation: MatchEvaluation,
) -> Dict[str, Any]:
    same_product_count = evaluation.same_product_count
    count_display: int | str = ""
    if same_product_count is not None:
        count_display = count_display_value(runtime, same_product_count)
    match_mode = getattr(runtime, "match_mode", "embedding")
    mini_confirmed_count = mini_confirmed_same_product_count_value(
        {
            "match_mode": match_mode,
            "processing_status": evaluation.processing_status,
            "same_product_count": same_product_count,
        }
    )
    return {
        "time": now_iso(),
        "source_id": current.get("source_id", ""),
        "source_asin": current.get("source_asin", ""),
        "source_product_url": current.get("source_product_url", ""),
        "input_row": current.get("input_row", ""),
        "search_method": search_method,
        "plugin_status": plugin_status,
        "candidate_count": candidate_count,
        "match_mode": match_mode,
        "similar_count": same_product_count,
        "count_display": count_display,
        "prescreen_visual_match_count": evaluation.prescreen_visual_match_count,
        "processing_status": evaluation.processing_status,
        "same_product_count": same_product_count,
        "same_product_confidence": evaluation.same_product_confidence,
        "match_reason": evaluation.match_reason,
        MINI_CONFIRMED_COUNT_FIELD: mini_confirmed_count,
        "provider_metrics": dict(evaluation.provider_metrics),
        "max_candidates_per_source": runtime.max_candidates_per_source,
    }


def load_count_results(counts_path: Path) -> Dict[int, Dict[str, Any]]:
    results: Dict[int, Dict[str, Any]] = {}
    for row in read_jsonl(counts_path):
        try:
            input_row = int(row.get("input_row") or 0)
        except (TypeError, ValueError):
            continue
        if input_row >= 2:
            results[input_row] = row
    return results


def validate_committed_count_identity(
    committed_result: Dict[str, Any],
    current: Dict[str, Any],
) -> None:
    """Fail closed before a committed count can be applied to another source."""
    expected = {
        "input_row": str(int(current.get("input_row") or 0)),
        "source_id": normalize_space(str(current.get("source_id") or "")),
        "source_asin": normalize_space(str(current.get("source_asin") or "")).upper(),
        "source_product_url": clean_url(str(current.get("source_product_url") or "")),
    }
    actual = {
        "input_row": str(int(committed_result.get("input_row") or 0)),
        "source_id": normalize_space(str(committed_result.get("source_id") or "")),
        "source_asin": normalize_space(str(committed_result.get("source_asin") or "")).upper(),
        "source_product_url": clean_url(str(committed_result.get("source_product_url") or "")),
    }
    mismatched = [field_name for field_name in expected if expected[field_name] != actual[field_name]]
    if mismatched:
        raise UserFacingError(
            "已提交的同款数量与当前输入来源不一致（"
            + "、".join(mismatched)
            + "）；请保留旧输出并改用新的 job_id。"
        )


def source_result_shard_path(shard_dir: Path, current: Dict[str, Any]) -> Path:
    input_row = int(current.get("input_row") or 0)
    source_id = normalize_space(str(current.get("source_id") or ""))
    identity_hash = hashlib.sha256(source_id.encode("utf-8")).hexdigest()[:16]
    return shard_dir / f"{input_row:08d}-{identity_hash}.json"


def commit_source_result_shard(
    shard_dir: Path,
    current: Dict[str, Any],
    crawl_plan_fingerprint: Dict[str, Any],
    provider_fingerprint: Dict[str, Any],
    candidate_rows: Sequence[Dict[str, Any]],
    accepted_rows: Sequence[Dict[str, Any]],
    count_row: Dict[str, Any],
) -> Path:
    ensure_dir(shard_dir)
    shard_path = source_result_shard_path(shard_dir, current)
    dump_json(
        shard_path,
        {
            "schema": SOURCE_RESULT_SHARD_SEMANTICS,
            "crawl_plan_sha256": crawl_plan_fingerprint.get("sha256", ""),
            "provider_sha256": provider_fingerprint.get("sha256", ""),
            "source": {
                "input_row": current.get("input_row", ""),
                "source_id": current.get("source_id", ""),
                "source_asin": current.get("source_asin", ""),
                "source_product_url": current.get("source_product_url", ""),
            },
            "candidate_rows": [dict(row) for row in candidate_rows],
            "accepted_rows": [dict(row) for row in accepted_rows],
            "count_row": dict(count_row),
        },
    )
    return shard_path


def materialize_source_result_shards(
    shard_dir: Path,
    crawl_plan_fingerprint: Dict[str, Any],
    provider_fingerprint: Dict[str, Any],
    candidates_path: Path,
    records_path: Path,
    counts_path: Path,
) -> Dict[int, Dict[str, Any]]:
    candidate_rows: List[Dict[str, Any]] = []
    accepted_rows: List[Dict[str, Any]] = []
    count_rows: List[Dict[str, Any]] = []
    seen_input_rows: set[int] = set()
    expected_plan = str(crawl_plan_fingerprint.get("sha256") or "")
    expected_provider = str(provider_fingerprint.get("sha256") or "")
    for shard_path in sorted(shard_dir.glob("*.json")):
        payload = load_json(shard_path)
        if payload.get("schema") != SOURCE_RESULT_SHARD_SEMANTICS:
            raise UserFacingError(f"来源结果分片版本不兼容：{shard_path.name}。请使用新的 job_id。")
        if str(payload.get("crawl_plan_sha256") or "") != expected_plan:
            raise UserFacingError(f"来源结果分片的输入指纹不匹配：{shard_path.name}。请使用新的 job_id。")
        if str(payload.get("provider_sha256") or "") != expected_provider:
            raise UserFacingError(f"来源结果分片的模型指纹不匹配：{shard_path.name}。请使用新的 job_id。")
        source = payload.get("source")
        count_row = payload.get("count_row")
        shard_candidates = payload.get("candidate_rows")
        shard_records = payload.get("accepted_rows")
        if (
            not isinstance(source, dict)
            or not isinstance(count_row, dict)
            or not isinstance(shard_candidates, list)
            or not isinstance(shard_records, list)
            or any(not isinstance(row, dict) for row in shard_candidates)
            or any(not isinstance(row, dict) for row in shard_records)
        ):
            raise UserFacingError(f"来源结果分片结构损坏：{shard_path.name}。")
        try:
            input_row = int(source.get("input_row") or 0)
        except (TypeError, ValueError) as exc:
            raise UserFacingError(f"来源结果分片缺少有效 input_row：{shard_path.name}。") from exc
        if input_row < 2 or input_row in seen_input_rows:
            raise UserFacingError(f"来源结果分片 input_row 重复或无效：{shard_path.name}。")
        validate_committed_count_identity(count_row, source)
        seen_input_rows.add(input_row)
        candidate_rows.extend(dict(row) for row in shard_candidates)
        accepted_rows.extend(dict(row) for row in shard_records)
        count_rows.append(dict(count_row))

    write_jsonl_atomic(candidates_path, candidate_rows)
    write_jsonl_atomic(records_path, accepted_rows)
    write_jsonl_atomic(counts_path, count_rows)
    return {int(row["input_row"]): row for row in count_rows}


def copy_header_style(ws: Any, source_col: int, target_col: int) -> None:
    if source_col < 1 or target_col < 1:
        return
    source = ws.cell(row=1, column=source_col)
    target = ws.cell(row=1, column=target_col)
    if source.has_style:
        target._style = copy_style(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy_style(source.alignment)
    if source.fill:
        target.fill = copy_style(source.fill)
    if source.font:
        target.font = copy_style(source.font)


def write_count_only_workbook(
    source_workbook_path: Path,
    counts_path: Path,
    output_xlsx: Path,
) -> None:
    if load_workbook is None or Workbook is None:
        raise UserFacingError("缺少 openpyxl，无法生成 Excel。")
    count_results = load_count_results(counts_path)

    if source_workbook_path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(source_workbook_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rows = read_input_rows(source_workbook_path)
        headers = list(rows[0].keys()) if rows else []
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])

    header_row = 1

    # Count-only workbooks are review workbooks, not a dump of every crawl
    # implementation field. Remove these columns by header (rather than a
    # brittle letter) so input template revisions cannot shift the contract.
    for column_index in range(ws.max_column, 0, -1):
        header = normalize_space(str(ws.cell(row=header_row, column=column_index).value or ""))
        if header in COUNT_ONLY_REMOVED_HEADERS:
            ws.delete_cols(column_index, 1)

    def ensure_column(header: str) -> int:
        for column_index in range(1, ws.max_column + 1):
            existing = normalize_space(
                str(ws.cell(row=header_row, column=column_index).value or "")
            )
            if existing == header:
                return column_index
        column_index = ws.max_column + 1
        if column_index > 1:
            copy_header_style(ws, column_index - 1, column_index)
        ws.cell(row=header_row, column=column_index).value = header
        return column_index

    count_col = ensure_column(COUNT_COLUMN_HEADER)
    existing_product_url_columns = [
        column_index
        for column_index in range(1, count_col)
        if normalize_space(str(ws.cell(row=header_row, column=column_index).value or ""))
        == REVIEW_PRODUCT_URL_HEADER
    ]
    if (
        count_col == 1
        or normalize_space(str(ws.cell(row=header_row, column=count_col - 1).value or ""))
        != REVIEW_PRODUCT_URL_HEADER
        or len(existing_product_url_columns) < 2
    ):
        # Add a dedicated, clickable URL immediately beside the final count.
        # The original product URL remains in the source-data portion of the
        # worksheet; this duplicate is deliberately placed for reviewers.
        ws.insert_cols(count_col, 1)
        copy_header_style(ws, count_col + 1, count_col)
        ws.cell(row=header_row, column=count_col).value = REVIEW_PRODUCT_URL_HEADER
        review_url_col = count_col
        count_col += 1
    else:
        review_url_col = count_col - 1

    source_product_url_col: Optional[int] = None
    for column_index in range(1, review_url_col):
        header = normalize_space(str(ws.cell(row=header_row, column=column_index).value or ""))
        if header == REVIEW_PRODUCT_URL_HEADER:
            source_product_url_col = column_index
            break

    audit_columns = {
        field_name: ensure_column(header)
        for field_name, header in COUNT_RESULT_FIELD_HEADERS.items()
    }

    for row_index, result in count_results.items():
        if result.get("match_mode") == "cascade":
            value = result.get("same_product_count")
        else:
            value = result.get("count_display")
            if value is None or value == "":
                value = result.get("similar_count", "")
        ws.cell(row=row_index, column=count_col).value = value

        review_url = normalize_space(
            str(result.get("source_product_url") or result.get("product_url") or "")
        )
        if not review_url and source_product_url_col is not None:
            review_url = normalize_space(
                str(ws.cell(row=row_index, column=source_product_url_col).value or "")
            )
        review_cell = ws.cell(row=row_index, column=review_url_col)
        review_cell.value = review_url
        if review_url.startswith(("http://", "https://")):
            review_cell.hyperlink = review_url
            review_cell.style = "Hyperlink"
        else:
            review_cell.hyperlink = None

        for field_name, column_index in audit_columns.items():
            value = result.get(field_name, "")
            ws.cell(row=row_index, column=column_index).value = value

    for header, column_index in (
        (REVIEW_PRODUCT_URL_HEADER, review_url_col),
        (COUNT_COLUMN_HEADER, count_col),
        *((header, audit_columns[field_name]) for field_name, header in COUNT_RESULT_FIELD_HEADERS.items()),
    ):
        ws.column_dimensions[get_column_letter(column_index)].width = max(
            ws.column_dimensions[get_column_letter(column_index)].width or 0,
            len(header) + 4,
        )
    ensure_dir(output_xlsx.parent)
    wb.save(output_xlsx)


def run_image_competitor_crawl(runtime: ImageCompetitorRuntimeConfig, dry_run: bool) -> int:
    prepare_vision_provider(runtime)
    initial_queue = load_products(runtime.products_file, runtime.marketplace_domain, dedupe=not runtime.is_count_only)
    job_dir = ensure_dir(runtime.outputs_root / runtime.job_id)
    records_path = job_dir / "records.jsonl"
    candidates_path = job_dir / "candidates.jsonl"
    failures_path = job_dir / "failures.jsonl"
    counts_path = job_dir / "counts.jsonl"
    state_path = job_dir / "state.json"
    source_results_dir = job_dir / "source_results"
    output_suffix = "相似竞品数量" if runtime.is_count_only else "同款竞品结果"
    output_xlsx = job_dir / f"{runtime.products_file.stem}_{output_suffix}.xlsx"
    debug_dir = job_dir / "debug_snapshots"
    image_dir = ensure_dir(job_dir / "source_images")

    print(f"任务目录：{job_dir}")
    print(f"任务模式：image_competitor/{runtime.result_mode}")
    print(f"浏览器后端：{runtime.browser_backend}/{runtime.browser_mode}")
    print(f"卖家精灵门禁：{'required' if runtime.sellersprite_required else 'not_required'}")
    print(f"Amazon 站点：{runtime.marketplace_domain}")
    if runtime.match_mode == "cascade":
        vision_description = (
            f"{runtime.embedding_provider}/{runtime.embedding_model} -> "
            f"{runtime.mini_provider}/{runtime.mini_model}"
        )
    elif runtime.match_mode == "embedding":
        vision_description = f"{runtime.embedding_provider}/{runtime.embedding_model}"
    else:
        vision_description = f"{runtime.embedding_provider}/{runtime.vision_model}"
    print(f"视觉识别：{vision_description}")
    print(f"输入数量：{len(initial_queue)}")
    print(f"输出表格：{output_xlsx}")
    if dry_run:
        print("dry-run：配置、输入表和视觉模型凭据检查完成，未打开浏览器。")
        return 0

    job_lock = JobRunLock(job_dir / ".run.lock")
    job_lock.acquire()
    try:
        if not runtime.resume:
            for old_file in (records_path, candidates_path, failures_path, counts_path, output_xlsx):
                if old_file.exists():
                    old_file.unlink()
            if source_results_dir.exists():
                for old_shard in source_results_dir.glob("*.json"):
                    old_shard.unlink()

        crawl_plan_fingerprint = image_crawl_plan_fingerprint(runtime, initial_queue)
        provider_fingerprint = vision_provider_fingerprint(runtime)
        state = ImageCompetitorStateStore(state_path, runtime, initial_queue)
        state.load_or_create()
        if runtime.is_count_only and runtime.match_mode == "cascade" and source_results_dir.exists():
            committed_count_results = materialize_source_result_shards(
                source_results_dir,
                crawl_plan_fingerprint,
                provider_fingerprint,
                candidates_path,
                records_path,
                counts_path,
            )
        elif runtime.is_count_only and runtime.match_mode == "cascade":
            # Transitional recovery for a count committed by the pre-shard build.
            # New writes always use source shards; identity and plan checks still apply.
            committed_count_results = load_count_results(counts_path)
        else:
            committed_count_results = {}
        driver = start_driver(runtime)
    except BaseException:
        job_lock.release()
        raise
    batch_pause = BatchPauseScheduler(runtime)  # type: ignore[arg-type]
    try:
        if runtime.search_strategy not in {"sellersprite_find_similar_first", "find_similar_first"}:
            ensure_lens_supported(driver, runtime, state, failures_path, debug_dir)
        while True:
            current = state.next_work()
            if not current:
                print("队列已完成。")
                break
            try:
                current_input_row = int(current.get("input_row") or 0)
            except (TypeError, ValueError):
                current_input_row = 0
            committed_result = committed_count_results.get(current_input_row)
            if committed_result is not None:
                validate_committed_count_identity(committed_result, current)
                committed_count = committed_result.get("same_product_count")
                state.finish_current_source(
                    str(committed_result.get("processing_status") or "recovered_count_commit"),
                    count=committed_count if isinstance(committed_count, int) else 0,
                    result={
                        field_name: committed_result.get(field_name, "")
                        for field_name in (
                            "prescreen_visual_match_count",
                            "processing_status",
                            "same_product_count",
                            "same_product_confidence",
                            "match_reason",
                            "provider_metrics",
                        )
                    },
                )
                print(
                    f"恢复已提交数量结果：input_row={current_input_row}；"
                    "跳过重复视觉模型调用。"
                )
                continue
            label = current.get("source_asin") or current.get("source_product_url") or current.get("source_id")
            print(f"处理：{label}")
            source_restore_handles = source_window_restore_order(driver)
            source_claimed_before = claimed_crawler_window_handles(driver)
            try:
                source_image_path = resolve_source_image(
                    driver,
                    runtime,
                    current,
                    image_dir,
                    state,
                )
                state.set_current(current)

                search_method = run_image_search(
                    driver,
                    runtime,
                    current,
                    source_image_path,
                    state,
                )
                current["image_search_method"] = search_method
                block_reason = detect_block_after_navigation(
                    driver,
                    min(runtime.page_timeout, 15),
                )
                if block_reason:
                    try:
                        handle_image_verification(driver, runtime, state, block_reason)
                    except VerificationUnconfirmedError:
                        if runtime.save_debug_snapshots:
                            save_debug_snapshot(driver, debug_dir, block_reason)
                        raise

                try:
                    lens_result_status = wait_for_lens_results(driver, runtime)
                except TimeoutException:
                    if search_method == "sellersprite_find_similar":
                        search_method = "amazon_upload_after_find_similar_timeout"
                        lens_result_status = upload_and_wait_for_lens_results(
                            driver,
                            runtime,
                            source_image_path,
                            state,
                        )
                    elif search_method == "amazon_upload":
                        search_method = "amazon_upload_retry"
                        lens_result_status = upload_and_wait_for_lens_results(
                            driver,
                            runtime,
                            source_image_path,
                            state,
                        )
                    else:
                        raise
                if runtime.sellersprite_on_lens:
                    plugin_status = wait_for_lens_sellersprite_data(driver, runtime)
                    state.mark_sellersprite_readiness(getattr(driver, "_sellersprite_readiness", {}))
                    if plugin_status == "blocked":
                        try:
                            handle_image_sellersprite_block(driver, runtime, state)
                        except VerificationUnconfirmedError:
                            if runtime.save_debug_snapshots:
                                save_debug_snapshot(driver, debug_dir, "verification_timeout")
                            raise
                        plugin_status = wait_for_lens_sellersprite_data(driver, runtime)
                        state.mark_sellersprite_readiness(
                            getattr(driver, "_sellersprite_readiness", {})
                        )
                        if plugin_status == "blocked":
                            raise VerificationUnconfirmedError(
                                verification_unconfirmed_message(
                                    sellersprite_block_reason(driver)
                                )
                            )
                    if runtime.sellersprite_required and plugin_status != "ok":
                        raise UserFacingError(
                            f"卖家精灵数据未达到写入门禁：{plugin_status}。"
                        )
                else:
                    plugin_status = "skipped_on_lens"

                candidate_records = merge_lens_product_data(driver, runtime, current, plugin_status)
                if (
                    runtime.match_mode in {"embedding", "cascade"}
                    and not candidate_records
                    and lens_result_status != "no_results"
                ):
                    raise EmbeddingProviderError(
                        "Amazon Lens 页面曾检测到结果，但候选商品重新提取为空；"
                        "本次不写入零竞品，已保留当前来源供重试。"
                    )
                evaluation = evaluate_competitor_matches(
                    runtime,
                    source_image_path,
                    normalize_space(str(current.get("source_image") or current.get("input_image_url") or "")),
                    candidate_records,
                )
                accepted_records = evaluation.accepted_records
                decisions = evaluation.decisions
                needs_enrichment = runtime.enrich_accepted_results and accepted_records and (
                    plugin_status != "ok"
                    or any(
                        sum(1 for field_name in REQUESTED_DATA_FIELDS if row.get(field_name))
                        < max(3, len(REQUESTED_DATA_FIELDS) // 3)
                        for row in accepted_records
                    )
                )
                if needs_enrichment:
                    accepted_records = enrich_accepted_records(
                        driver,
                        runtime,
                        accepted_records,
                        state,
                    )
                    state.mark_sellersprite_readiness(getattr(driver, "_sellersprite_readiness", {}))
                candidate_audit_rows = build_candidate_audit_rows(candidate_records, accepted_records, decisions)

                if runtime.match_mode == "cascade" and runtime.is_count_only:
                    count_row = build_count_result_row(
                        runtime,
                        current,
                        search_method,
                        plugin_status,
                        len(candidate_records),
                        evaluation,
                    )
                    commit_source_result_shard(
                        source_results_dir,
                        current,
                        crawl_plan_fingerprint,
                        provider_fingerprint,
                        candidate_audit_rows,
                        accepted_records,
                        count_row,
                    )
                    committed_count_results = materialize_source_result_shards(
                        source_results_dir,
                        crawl_plan_fingerprint,
                        provider_fingerprint,
                        candidates_path,
                        records_path,
                        counts_path,
                    )
                else:
                    for row in candidate_audit_rows:
                        append_jsonl(candidates_path, row)
                    for row in accepted_records:
                        append_jsonl(records_path, row)
                if runtime.is_count_only and runtime.match_mode != "cascade":
                    append_count_result(
                        counts_path,
                        runtime,
                        current,
                        search_method,
                        plugin_status,
                        len(candidate_records),
                        evaluation,
                    )
                state.finish_current_source(
                    evaluation.processing_status,
                    count=len(accepted_records),
                    result={
                        "prescreen_visual_match_count": evaluation.prescreen_visual_match_count,
                        "processing_status": evaluation.processing_status,
                        "same_product_count": evaluation.same_product_count,
                        "same_product_confidence": evaluation.same_product_confidence,
                        "match_reason": evaluation.match_reason,
                        "provider_metrics": evaluation.provider_metrics,
                    },
                )
                count_text = (
                    "空（粗筛排除）"
                    if evaluation.same_product_count is None
                    else str(evaluation.same_product_count)
                )
                print(
                    f"搜索方式：{search_method}；候选 {len(candidate_records)} 条，"
                    f"同款数量 {count_text}；状态 {evaluation.processing_status}；"
                    f"插件状态：{plugin_status}；调用统计：{evaluation.provider_metrics}"
                )
                close_claimed_crawler_windows(
                    driver,
                    source_claimed_before,
                    source_restore_handles,
                )
                batch_pause.after_completed_page()
                sleep_between_pages(runtime)
            except EmbeddingProviderError as exc:
                safe_message = redact_sensitive_text(
                    exc,
                    (runtime.embedding_api_key, runtime.mini_api_key),
                )[:500]
                failure_reason = (
                    "embedding_provider_fatal"
                    if isinstance(exc, FatalEmbeddingProviderError)
                    else "embedding_provider_error"
                )
                log_failure(
                    failures_path,
                    state,
                    current,
                    failure_reason,
                    safe_message,
                    safe_driver_current_url(driver),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, failure_reason)
                print(
                    "视觉向量识别失败，当前来源已保留在断点中；"
                    "本次未写入该来源的相似竞品数量。",
                    file=sys.stderr,
                )
                error_type = (
                    FatalEmbeddingProviderError
                    if isinstance(exc, FatalEmbeddingProviderError)
                    else EmbeddingProviderError
                )
                raise error_type(safe_message) from exc
            except DeliveryLocationUnconfirmedError as exc:
                safe_message = redact_sensitive_text(
                    exc,
                    (runtime.embedding_api_key, runtime.mini_api_key),
                )[:500]
                log_failure(
                    failures_path,
                    state,
                    current,
                    "delivery_location_unconfirmed",
                    safe_message,
                    safe_driver_current_url(driver),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, "delivery_location_unconfirmed")
                raise
            except VerificationUnconfirmedError as exc:
                safe_message = redact_sensitive_text(
                    exc,
                    (runtime.embedding_api_key, runtime.mini_api_key),
                )[:500]
                log_failure(
                    failures_path,
                    state,
                    current,
                    "verification_unconfirmed",
                    safe_message,
                    safe_driver_current_url(driver),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, "verification_unconfirmed")
                raise VerificationUnconfirmedError(safe_message) from exc
            except (TimeoutException, requests.RequestException) as exc:
                safe_message = redact_sensitive_text(
                    exc,
                    (runtime.embedding_api_key, runtime.mini_api_key),
                )[:500]
                log_failure(
                    failures_path,
                    state,
                    current,
                    "runtime_error",
                    safe_message,
                    safe_driver_current_url(driver),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, "runtime_error")
                if runtime.match_mode == "cascade":
                    raise UserFacingError(
                        "cascade 来源处理发生暂时性网络/页面超时；"
                        "当前来源已保留，未写入同款数量，请稍后按断点重试。"
                    ) from exc
                state.finish_current_source("runtime_error")
            except (UserFacingError, WebDriverException) as exc:
                safe_message = redact_sensitive_text(
                    exc,
                    (runtime.embedding_api_key, runtime.mini_api_key),
                )[:500]
                log_failure(
                    failures_path,
                    state,
                    current,
                    "crawler_error",
                    safe_message,
                    safe_driver_current_url(driver),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, "crawler_error")
                if runtime.match_mode == "cascade":
                    if isinstance(exc, UserFacingError):
                        raise UserFacingError(safe_message) from exc
                    raise UserFacingError(
                        "cascade 浏览器处理失败；当前来源已保留，未写入同款数量。"
                    ) from exc
                state.finish_current_source("crawler_error")
            finally:
                close_claimed_crawler_windows(
                    driver,
                    source_claimed_before,
                    source_restore_handles,
                )
    finally:
        try:
            try:
                driver.quit()
            except WebDriverException:
                pass
        finally:
            job_lock.release()

    if runtime.is_count_only:
        write_count_only_workbook(runtime.products_file, counts_path, output_xlsx)
        print(f"已生成相似竞品数量表：{output_xlsx}")
    else:
        write_workbook(runtime.products_file, records_path, candidates_path, failures_path, output_xlsx)
        print(f"已生成以图搜图竞品表：{output_xlsx}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Amazon image-search competitor crawler with SellerSprite data.")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="配置文件路径")
    parser.add_argument("--dry-run", action="store_true", help="只检查配置，不打开浏览器")
    parser.add_argument("--no-resume", action="store_true", help="忽略已有断点，重新开始任务")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser()
    if not config_path.is_absolute():
        config_path = ROOT_DIR / config_path
    raw_config = load_json(config_path)
    runtime = build_image_runtime_config(raw_config, args.no_resume)
    return run_image_competitor_crawl(runtime, args.dry_run)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except UserFacingError as exc:
        print(f"运行失败：{exc}", file=sys.stderr)
        raise SystemExit(2)
