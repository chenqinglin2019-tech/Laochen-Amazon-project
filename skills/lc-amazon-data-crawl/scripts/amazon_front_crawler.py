#!/usr/bin/env python3
"""
Unified Amazon front crawler for:
1. BSR category ranking pages
2. keyword search result pages
3. competitor storefront pages

The crawler reads SellerSprite extension data from a visible Chrome session.
It does not solve CAPTCHA or bypass verification; when verification appears,
it saves state and waits for manual handling.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import json
import os
import queue
import random
import re
import shutil
import sys
import threading
import time
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from urllib.parse import parse_qs, quote_plus, urlencode, urlparse, urlunparse

from selenium.common.exceptions import JavascriptException, TimeoutException, WebDriverException
from selenium.webdriver.remote.webdriver import WebDriver

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None

from amazon_category_rank_crawler import (
    BatchPauseScheduler,
    ConcurrentWorkerCancelled,
    JobRunLock,
    ProductFilterConfig,
    REQUESTED_DATA_FIELDS,
    UserFacingError,
    VerificationUnconfirmedError,
    append_jsonl,
    build_delivery_location_config,
    build_product_filter_config,
    build_runtime_config as build_category_runtime_config,
    clean_url,
    config_bool,
    config_float,
    config_int,
    config_text,
    country_from_flag_code_or_text,
    detect_block,
    discover_child_categories,
    dump_json,
    ensure_dir,
    ensure_resume_delivery_fingerprint,
    extract_by_selectors,
    extract_current_category_path,
    extract_node_id,
    extract_table_rows,
    filter_product_records,
    format_subcategory_bsr_ranks,
    find_next_page_url,
    load_json,
    normalize_header,
    normalize_fulfillment_method,
    normalize_subcategory_bsr_ranks,
    normalize_space,
    now_iso,
    now_ts,
    open_amazon_page,
    parse_field_from_text,
    parse_fulfillment_evidence,
    parse_subcategory_bsr_ranks,
    parse_table_row_fields,
    record_contract_fingerprint,
    read_jsonl,
    resolve_path,
    run_crawl as run_category_crawl,
    save_debug_snapshot,
    safe_sellersprite_readiness,
    select_fulfillment_evidence,
    slugify,
    start_driver,
    validate_amazon_url,
    wait_for_amazon_products,
    wait_for_manual_clear,
    wait_for_sellersprite_data,
    wait_for_sellersprite_data_or_prompt,
)


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CONFIG = ROOT_DIR / "config" / "amazon_front_crawler.json"
SUPPORTED_MODES = {"bsr_category", "keyword_search", "storefront"}
SEARCH_SORT_OPTIONS = {
    "Featured": "featured-rank",
    "Price: Low to High": "price-asc-rank",
    "Price: High to Low": "price-desc-rank",
    "Avg. Customer Review": "review-rank",
    "Newest Arrivals": "date-desc-rank",
    "Best Sellers": "exact-aware-popularity-rank",
}


FRONT_DEDUP_HEADERS = [
    "来源类型",
    "来源关键词列表",
    "来源搜索排序规则列表",
    "来源店铺URL列表",
    "来源店铺名称列表",
    "来源店铺排序规则列表",
    "来源类目路径列表",
    "来源页面URL列表",
    "出现次数",
    "最佳页码",
    "最佳排名",
    "ASIN",
    "商品标题",
    "商品URL",
    "评论数量",
    "评分值",
    "卖家名称",
    "品牌名称",
    "卖家所处国家",
    "子类目节点排名",
    "近30天销量（子体）",
    "近30天销量（父体）",
    "FBA费用",
    "毛利率",
    "配送方式",
    "配送时长",
    "上架时间",
    "自然搜索词数量",
    "广告搜索词数量",
    "抓取时间",
    "加载状态",
    "备注",
]

FRONT_FIELD_TO_HEADER = {
    "source_types": "来源类型",
    "source_keywords": "来源关键词列表",
    "source_search_sort_orders": "来源搜索排序规则列表",
    "source_store_urls": "来源店铺URL列表",
    "source_store_names": "来源店铺名称列表",
    "source_store_sort_orders": "来源店铺排序规则列表",
    "source_category_paths": "来源类目路径列表",
    "source_page_urls": "来源页面URL列表",
    "appearance_count": "出现次数",
    "best_page_number": "最佳页码",
    "best_rank": "最佳排名",
    "asin": "ASIN",
    "title": "商品标题",
    "product_url": "商品URL",
    "review_count": "评论数量",
    "rating_value": "评分值",
    "seller_name": "卖家名称",
    "brand_name": "品牌名称",
    "seller_country": "卖家所处国家",
    "subcategory_bsr_ranks": "子类目节点排名",
    "sales_30_days_child": "近30天销量（子体）",
    "sales_30_days_parent": "近30天销量（父体）",
    "fba_fee": "FBA费用",
    "gross_margin": "毛利率",
    "fulfillment_method": "配送方式",
    "delivery_duration": "配送时长",
    "launch_date": "上架时间",
    "organic_keywords_count": "自然搜索词数量",
    "ad_keywords_count": "广告搜索词数量",
    "scraped_at": "抓取时间",
    "load_status": "加载状态",
    "note": "备注",
}


@dataclass
class FrontRuntimeConfig:
    mode: str
    job_id: str
    outputs_root: Path
    keywords_file: Path
    store_urls_file: Path
    start_url: str
    max_pages_per_keyword: int
    include_sponsored: bool
    keyword_sort_orders: List[str]
    store_page_limit: Optional[int]
    store_sort_orders: List[str]
    browser_tab_concurrency: int
    product_filters: ProductFilterConfig
    record_contract_fingerprint: str
    resume: bool
    browser_backend: str
    browser_mode: str
    chrome_binary: str
    chrome_user_data_dir: Path
    chrome_profile_directory: str
    debugger_address: str
    extension_path: Path
    activate_plugin: bool
    page_timeout: int
    plugin_timeout: int
    plugin_retry_attempts: int
    plugin_retry_wait_seconds: float
    plugin_retry_wait_seconds_max: float
    plugin_relaunch_retry_attempts: int
    plugin_relaunch_wait_seconds: float
    plugin_second_relaunch_retry_attempts: int
    plugin_second_relaunch_wait_seconds: float
    manual_pause_timeout: int
    delivery_location_enabled: bool
    delivery_locations_file: Path
    delivery_location_timeout: int
    delivery_locations: Dict[str, Dict[str, str]]
    delivery_location_fingerprint: str
    delay_seconds_min: float
    delay_seconds_max: float
    batch_pause_pages_min: int
    batch_pause_pages_max: int
    batch_pause_seconds_min: float
    batch_pause_seconds_max: float
    page_scroll_before_extract: bool
    page_scroll_max_rounds: int
    page_scroll_step_ratio: float
    page_scroll_wait_seconds: float
    page_scroll_stable_rounds: int
    sellersprite_required: bool
    sellersprite_min_enriched_records: int
    sellersprite_min_fields_per_record: int
    sellersprite_stable_checks: int
    save_debug_snapshots: bool
    field_selectors: Dict[str, List[str]] = field(default_factory=dict)


FRONT_STATE_SCHEMA_VERSION = 2
FRONT_RETRYABLE_ERRORS = {"webdriver_error", "page_timeout", "no_product_cards"}
FRONT_MAX_TASK_RETRIES = 2


def prepare_front_retry_task(task: Dict[str, Any]) -> Tuple[Dict[str, Any], bool, int]:
    retry_task = copy.deepcopy(task)
    previous_count = int(retry_task.get("worker_retry_count") or 0)
    next_count = min(previous_count + 1, FRONT_MAX_TASK_RETRIES)
    retry_task["worker_retry_count"] = next_count
    return retry_task, previous_count < FRONT_MAX_TASK_RETRIES, next_count


def front_source_id(task: Dict[str, Any]) -> str:
    return str(task.get("source_id") or task.get("keyword") or task.get("store_url") or "")


def front_task_identity(task: Dict[str, Any]) -> str:
    return "|".join(
        [
            str(task.get("source_type") or ""),
            front_source_id(task),
            f"page:{task.get('page_number') or 1}",
            clean_url(str(task.get("page_url") or "")),
        ]
    )


def front_crawl_plan_fingerprint(
    runtime: FrontRuntimeConfig, initial_queue: Sequence[Dict[str, Any]]
) -> str:
    """Fingerprint crawl semantics while intentionally excluding tab concurrency."""
    canonical_tasks = sorted(
        (copy.deepcopy(task) for task in initial_queue),
        key=lambda task: (
            front_task_identity(task),
            json.dumps(task, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
        ),
    )
    payload = {
        "mode": runtime.mode,
        "initial_queue": canonical_tasks,
        "include_sponsored": runtime.include_sponsored,
        "max_pages_per_keyword": runtime.max_pages_per_keyword,
        "store_page_limit": runtime.store_page_limit,
        "sellersprite_required": runtime.sellersprite_required,
        "field_selectors": runtime.field_selectors,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


class FrontStateStore:
    """Versioned, single-writer state for resumable multi-tab crawling."""

    def __init__(self, path: Path, runtime: FrontRuntimeConfig, initial_queue: Sequence[Dict[str, Any]]) -> None:
        self.path = path
        self.runtime = runtime
        self.initial_queue = [copy.deepcopy(item) for item in initial_queue]
        self.crawl_plan_fingerprint = front_crawl_plan_fingerprint(
            runtime, self.initial_queue
        )
        self.page_results_dir = self.path.parent / "page_results"
        self.data: Dict[str, Any] = {}

    @property
    def records_path(self) -> Path:
        return self.path.with_name("records.jsonl")

    def _new_data(self) -> Dict[str, Any]:
        return {
            "schema_version": FRONT_STATE_SCHEMA_VERSION,
            "job_id": self.runtime.job_id,
            "mode": self.runtime.mode,
            "created_at": now_iso(),
            "pending": [copy.deepcopy(item) for item in self.initial_queue],
            "in_flight": {},
            "completed_pages": [],
            "completed_page_order": [],
            "completed_sources": [],
            "completed_source_reasons": {},
            "records_count": 0,
            "scanned_count": 0,
            "filtered_count": 0,
            "filter_reason_counts": {},
            "failures_count": 0,
            "manual_pauses": {},
            "sellersprite_readiness": {},
            "delivery_location_fingerprint": self.runtime.delivery_location_fingerprint,
            "record_contract_fingerprint": self.runtime.record_contract_fingerprint,
            "crawl_plan_fingerprint": self.crawl_plan_fingerprint,
        }

    @staticmethod
    def _has_progress(data: Dict[str, Any], records_path: Path) -> bool:
        return bool(
            int(data.get("records_count") or 0) > 0
            or data.get("completed_pages")
            or data.get("in_flight")
            or data.get("current")
            or (records_path.exists() and records_path.stat().st_size > 0)
        )

    def load_or_create(self) -> None:
        ensure_dir(self.page_results_dir)
        if not self.runtime.resume:
            self.data = self._new_data()
            self.flush()
            return
        if not self.path.exists():
            has_page_results = any(self.page_results_dir.glob("*.json"))
            if self.records_path.exists() and self.records_path.stat().st_size > 0 and not has_page_results:
                raise UserFacingError(
                    "现有 records.jsonl 没有兼容的状态/page shard，不能安全续传；请更换 job_id。"
                )
            self.data = self._new_data()
            if has_page_results:
                self._recover_from_page_results()
            self.flush()
            return

        loaded = load_json(self.path)
        if int(loaded.get("schema_version") or 0) != FRONT_STATE_SCHEMA_VERSION:
            if self._has_progress(loaded, self.records_path):
                raise UserFacingError(
                    "现有断点属于旧数据结构，不能与子类目排名/过滤结果混写；请更换 job_id。"
                )
            self.data = self._new_data()
            self.flush()
            return

        previous_contract = str(loaded.get("record_contract_fingerprint") or "")
        if previous_contract != self.runtime.record_contract_fingerprint:
            if self._has_progress(loaded, self.records_path) or any(self.page_results_dir.glob("*.json")):
                raise UserFacingError(
                    "当前抓取字段或过滤条件与已有断点不一致；请更换 job_id。"
                )
            loaded["record_contract_fingerprint"] = self.runtime.record_contract_fingerprint

        previous_plan = str(loaded.get("crawl_plan_fingerprint") or "")
        if previous_plan != self.crawl_plan_fingerprint:
            if self._has_progress(loaded, self.records_path) or any(
                self.page_results_dir.glob("*.json")
            ):
                raise UserFacingError(
                    "当前抓取模式、输入来源或抓取参数与已有断点不一致；请更换 job_id。"
                )
            self.data = self._new_data()
            self.flush()
            return

        self.data = loaded
        ensure_resume_delivery_fingerprint(self.data, self.runtime, self.records_path)
        self._recover_from_page_results()
        self.flush()

    def flush(self) -> None:
        self.data["updated_at"] = now_iso()
        dump_json(self.path, self.data)

    def _read_page_result(self, path: Path) -> Dict[str, Any]:
        try:
            payload = load_json(path)
        except Exception as exc:
            raise UserFacingError(f"页面提交文件损坏，无法安全恢复：{path.name}: {exc}") from exc
        if int(payload.get("schema_version") or 0) != FRONT_STATE_SCHEMA_VERSION:
            raise UserFacingError(f"页面提交文件结构版本不兼容：{path.name}；请更换 job_id。")
        if str(payload.get("record_contract_fingerprint") or "") != self.runtime.record_contract_fingerprint:
            raise UserFacingError(f"页面提交文件与当前数据契约不一致：{path.name}；请更换 job_id。")
        if str(payload.get("crawl_plan_fingerprint") or "") != self.crawl_plan_fingerprint:
            raise UserFacingError(f"页面提交文件与当前抓取计划不一致：{path.name}；请更换 job_id。")
        return payload

    def page_result_path(self, page_key: str) -> Path:
        digest = hashlib.sha256(page_key.encode("utf-8")).hexdigest()
        return self.page_results_dir / f"{digest}.json"

    def iter_page_results(self) -> List[Dict[str, Any]]:
        by_key: Dict[str, Dict[str, Any]] = {}
        for path in sorted(self.page_results_dir.glob("*.json")):
            payload = self._read_page_result(path)
            page_key = str(payload.get("page_key") or "")
            if page_key:
                by_key[page_key] = payload
        order = [str(item) for item in self.data.get("completed_page_order") or []]
        completed_pages = [str(item) for item in self.data.get("completed_pages") or []]
        if set(order) != set(completed_pages) or len(order) != len(completed_pages):
            raise UserFacingError(
                "断点中的 completed_pages 与提交顺序不一致，无法安全恢复；请更换 job_id。"
            )
        ordered: List[Dict[str, Any]] = []
        missing_keys: List[str] = []
        for page_key in order:
            payload = by_key.pop(page_key, None)
            if payload:
                ordered.append(payload)
            else:
                missing_keys.append(page_key)
        if missing_keys:
            raise UserFacingError(
                "断点引用的页面提交文件缺失，无法保证 records 完整性；"
                "请恢复 page_results 或更换 job_id。"
            )
        if int(self.data.get("records_count") or 0) > 0 and not ordered and not by_key:
            raise UserFacingError(
                "断点已有 records_count，但 page_results 为空；请恢复提交文件或更换 job_id。"
            )
        ordered.extend(
            sorted(
                by_key.values(),
                key=lambda item: (str(item.get("committed_at") or ""), str(item.get("page_key") or "")),
            )
        )
        return ordered

    def _recover_from_page_results(self) -> None:
        payloads = self.iter_page_results()
        completed_pages: List[str] = []
        completed_task_ids = set()
        expected_completed_sources = set(
            str(item) for item in self.data.get("completed_sources") or []
        )
        completed_sources: set[str] = set()
        completed_reasons: Dict[str, str] = {}
        scanned_count = 0
        records_count = 0
        filter_reason_counts: Dict[str, int] = {}
        next_tasks: List[Dict[str, Any]] = []

        for payload in payloads:
            page_key = str(payload.get("page_key") or "")
            if not page_key:
                continue
            completed_pages.append(page_key)
            completed_task_ids.add(str(payload.get("task_identity") or ""))
            scanned_count += int(payload.get("scanned_count") or 0)
            records_count += len(payload.get("records") or [])
            for reason, count in dict(payload.get("rejection_counts") or {}).items():
                filter_reason_counts[str(reason)] = filter_reason_counts.get(str(reason), 0) + int(count or 0)
            source_id = str(payload.get("source_id") or "")
            finish_reason = str(payload.get("finish_reason") or "")
            if source_id and finish_reason:
                completed_sources.add(source_id)
                completed_reasons[source_id] = finish_reason
            next_task = payload.get("next_task")
            if isinstance(next_task, dict):
                next_tasks.append(copy.deepcopy(next_task))

        missing_completed_sources = expected_completed_sources - completed_sources
        if missing_completed_sources:
            raise UserFacingError(
                "断点中的已完成来源缺少最终页面提交文件，无法保证 records 完整性；"
                "请恢复 page_results 或更换 job_id。"
            )

        pending: List[Dict[str, Any]] = []
        for item in list(self.data.get("pending") or []) + list((self.data.get("in_flight") or {}).values()) + next_tasks:
            if not isinstance(item, dict):
                continue
            source_id = front_source_id(item)
            task_id = front_task_identity(item)
            if source_id in completed_sources or task_id in completed_task_ids:
                continue
            if any(front_task_identity(existing) == task_id for existing in pending):
                continue
            pending.append(copy.deepcopy(item))

        self.data["pending"] = pending
        self.data["in_flight"] = {}
        self.data["completed_pages"] = completed_pages
        self.data["completed_page_order"] = completed_pages
        self.data["completed_sources"] = sorted(completed_sources)
        self.data["completed_source_reasons"] = completed_reasons
        self.data["scanned_count"] = scanned_count
        self.data["records_count"] = records_count
        self.data["filtered_count"] = max(scanned_count - records_count, 0)
        self.data["filter_reason_counts"] = filter_reason_counts
        self.data["manual_pauses"] = {}
        self.data["record_contract_fingerprint"] = self.runtime.record_contract_fingerprint
        self.data["crawl_plan_fingerprint"] = self.crawl_plan_fingerprint

    def _prune_pending(self) -> bool:
        completed_sources = set(self.data.get("completed_sources") or [])
        pending = [
            copy.deepcopy(task)
            for task in self.data.get("pending") or []
            if isinstance(task, dict) and front_source_id(task) not in completed_sources
        ]
        if pending == list(self.data.get("pending") or []):
            return False
        self.data["pending"] = pending
        return True

    def lease_next(self, worker_id: str) -> Optional[Dict[str, Any]]:
        pruned = self._prune_pending()
        pending = list(self.data.get("pending") or [])
        active_sources = {
            front_source_id(item)
            for item in (self.data.get("in_flight") or {}).values()
            if isinstance(item, dict)
        }
        completed_sources = set(self.data.get("completed_sources") or [])
        selected_index: Optional[int] = None
        for index, task in enumerate(pending):
            source_id = front_source_id(task)
            if source_id and source_id not in active_sources and source_id not in completed_sources:
                selected_index = index
                break
        if selected_index is None:
            if pruned:
                self.flush()
            return None
        task = pending.pop(selected_index)
        self.data["pending"] = pending
        self.data.setdefault("in_flight", {})[worker_id] = copy.deepcopy(task)
        self.flush()
        return copy.deepcopy(task)

    def has_work(self) -> bool:
        if self._prune_pending():
            self.flush()
        return bool(self.data.get("pending") or self.data.get("in_flight"))

    def is_page_completed(self, key: str) -> bool:
        return key in set(self.data.get("completed_pages") or [])

    def commit_page_result(self, result: Any) -> bool:
        page_key = str(result.page_key or "")
        worker_id = str(result.worker_id)
        if not page_key:
            raise UserFacingError("页面结果缺少 page_key，拒绝提交。")
        if self.is_page_completed(page_key):
            self.data.setdefault("in_flight", {}).pop(worker_id, None)
            self.flush()
            return False

        payload = {
            "schema_version": FRONT_STATE_SCHEMA_VERSION,
            "record_contract_fingerprint": self.runtime.record_contract_fingerprint,
            "crawl_plan_fingerprint": self.crawl_plan_fingerprint,
            "page_key": page_key,
            "task_identity": front_task_identity(result.task),
            "source_id": front_source_id(result.task),
            "task": result.task,
            "page_url": result.page_url,
            "plugin_status": result.plugin_status,
            "scanned_count": len(result.raw_records),
            "records": result.accepted_records,
            "rejection_counts": result.rejection_counts,
            "next_task": result.next_task,
            "finish_reason": result.finish_reason,
            "committed_at": now_iso(),
        }
        dump_json(self.page_result_path(page_key), payload)

        completed = list(self.data.get("completed_pages") or [])
        completed.append(page_key)
        self.data["completed_pages"] = completed
        self.data.setdefault("completed_page_order", []).append(page_key)
        self.data.setdefault("in_flight", {}).pop(worker_id, None)
        self.data["scanned_count"] = int(self.data.get("scanned_count") or 0) + len(result.raw_records)
        self.data["records_count"] = int(self.data.get("records_count") or 0) + len(result.accepted_records)
        self.data["filtered_count"] = int(self.data.get("filtered_count") or 0) + max(
            len(result.raw_records) - len(result.accepted_records), 0
        )
        totals = self.data.setdefault("filter_reason_counts", {})
        for reason, count in result.rejection_counts.items():
            totals[reason] = int(totals.get(reason) or 0) + int(count or 0)

        if result.next_task:
            task_id = front_task_identity(result.next_task)
            if not any(front_task_identity(item) == task_id for item in self.data.get("pending") or []):
                self.data.setdefault("pending", []).append(copy.deepcopy(result.next_task))
        else:
            source_id = front_source_id(result.task)
            completed_sources = set(self.data.get("completed_sources") or [])
            completed_sources.add(source_id)
            self.data["completed_sources"] = sorted(completed_sources)
            self.data.setdefault("completed_source_reasons", {})[source_id] = result.finish_reason
        self.flush()
        return True

    def requeue_task(self, worker_id: str, task: Dict[str, Any]) -> None:
        self.data.setdefault("in_flight", {}).pop(worker_id, None)
        task_id = front_task_identity(task)
        if not any(front_task_identity(item) == task_id for item in self.data.get("pending") or []):
            self.data.setdefault("pending", []).insert(0, copy.deepcopy(task))
        self.flush()

    def log_failure(self) -> None:
        self.data["failures_count"] = int(self.data.get("failures_count") or 0) + 1
        self.flush()

    def mark_sellersprite_readiness(self, report: Dict[str, Any], worker_id: str = "main") -> None:
        self.data.setdefault("sellersprite_readiness", {})[worker_id] = safe_sellersprite_readiness(report)
        self.flush()

    def set_manual_snapshot(self, snapshot: Optional[Dict[str, Any]]) -> None:
        normalized = snapshot or {}
        if self.data.get("manual_pauses") != normalized:
            self.data["manual_pauses"] = normalized
            self.flush()

    def mark_manual_pause(self, reason: str, page_url: str) -> None:
        self.set_manual_snapshot(
            {"main": {"paused_at": now_iso(), "reason": reason, "page_url": page_url}}
        )

    def clear_manual_pause(self) -> None:
        self.set_manual_snapshot(None)


def parse_sort_orders(raw_value: Any, field_name: str, subject_name: str) -> List[str]:
    if raw_value in ("", None):
        return []
    if isinstance(raw_value, str):
        candidates = [part.strip() for part in re.split(r"[,;，；]+", raw_value) if part.strip()]
    elif isinstance(raw_value, list):
        candidates = [str(item).strip() for item in raw_value if str(item).strip()]
    else:
        raise UserFacingError(f"配置项 `{field_name}` 必须是数组、逗号分隔文本或空值。")
    normalized: List[str] = []
    alias_map = {normalize_header(key): key for key in SEARCH_SORT_OPTIONS}
    for candidate in candidates:
        key = alias_map.get(normalize_header(candidate))
        if not key:
            raise UserFacingError(f"不支持的{subject_name}排序规则：{candidate}")
        if key not in normalized:
            normalized.append(key)
    return normalized


def parse_store_sort_orders(raw_value: Any) -> List[str]:
    return parse_sort_orders(raw_value, "store_sort_orders", "店铺")


def parse_keyword_sort_orders(raw_value: Any) -> List[str]:
    return parse_sort_orders(raw_value, "keyword_sort_orders", "关键词")


def prompt_sort_orders(subject_name: str) -> List[str]:
    labels = list(SEARCH_SORT_OPTIONS.keys())
    print(f"请选择{subject_name}排序规则，可多选。输入序号，用逗号分隔；直接回车默认 Featured。")
    for index, label in enumerate(labels, start=1):
        print(f"  {index}. {label}")
    try:
        raw = input("排序规则：").strip()
    except EOFError:
        raw = ""
    if not raw:
        return ["Featured"]
    selected: List[str] = []
    for part in re.split(r"[,;，；\\s]+", raw):
        if not part:
            continue
        if not part.isdigit():
            raise UserFacingError("排序规则请输入序号，例如：1,6")
        index = int(part)
        if index < 1 or index > len(labels):
            raise UserFacingError(f"排序规则序号超出范围：{index}")
        label = labels[index - 1]
        if label not in selected:
            selected.append(label)
    return selected or ["Featured"]


def prompt_store_sort_orders() -> List[str]:
    return prompt_sort_orders("店铺商品")


def prompt_keyword_sort_orders() -> List[str]:
    return prompt_sort_orders("关键词搜索结果")


def prompt_store_page_limit() -> int:
    print("请输入每个店铺、每个排序规则要抓取前多少页商品，最多 20 页；直接回车默认 20。")
    try:
        raw = input("抓取页数：").strip()
    except EOFError:
        raw = ""
    if not raw:
        return 20
    if not raw.isdigit():
        raise UserFacingError("店铺抓取页数请输入 1-20 的整数。")
    value = int(raw)
    if value < 1 or value > 20:
        raise UserFacingError("店铺抓取页数必须在 1-20 之间。")
    return value


def build_front_runtime_config(config: Dict[str, Any], no_resume: bool) -> FrontRuntimeConfig:
    mode = config_text(config, "mode", "keyword_search").lower()
    if mode not in SUPPORTED_MODES:
        raise UserFacingError("配置项 `mode` 只支持 bsr_category、keyword_search 或 storefront。")

    job_id = config_text(config, "job_id") or f"amazon-front-{mode}-{now_ts()}"
    outputs_root = resolve_path(config_text(config, "outputs_root", "outputs"))
    keywords_file = resolve_path(config_text(config, "keywords_file", "inputs/keywords.csv"))
    store_urls_file = resolve_path(config_text(config, "store_urls_file", "inputs/storefronts.csv"))
    start_url = config_text(config, "start_url")

    if mode == "bsr_category":
        if not start_url:
            raise UserFacingError("bsr_category 模式下 `start_url` 不能为空。")
        validate_amazon_url(start_url)
    if mode == "keyword_search" and not keywords_file.exists():
        raise UserFacingError(f"没有找到关键词表格：{keywords_file}")
    if mode == "storefront" and not store_urls_file.exists():
        raise UserFacingError(f"没有找到店铺 URL 表格：{store_urls_file}")

    browser_backend = config_text(config, "browser_backend", "cdp").lower()
    if browser_backend not in {"cdp", "selenium"}:
        raise UserFacingError("配置项 `browser_backend` 只支持 cdp 或 selenium。")
    browser_mode = config_text(config, "browser_mode", "launch").lower()
    if browser_mode not in {"launch", "attach", "reuse", "applescript"}:
        raise UserFacingError("配置项 `browser_mode` 只支持 launch、attach、reuse 或 applescript。")
    browser_tab_concurrency = config_int(config, "browser_tab_concurrency", 1)
    browser_tab_concurrency = 1 if browser_tab_concurrency is None else browser_tab_concurrency
    if browser_tab_concurrency < 1 or browser_tab_concurrency > 3:
        raise UserFacingError("配置项 `browser_tab_concurrency` 必须是 1-3 的整数。")
    if browser_tab_concurrency > 1 and (
        browser_backend != "cdp" or browser_mode not in {"reuse", "attach"}
    ):
        raise UserFacingError(
            "多标签并发仅支持 browser_backend=cdp 且 browser_mode=reuse/attach。"
        )

    product_filters = build_product_filter_config(config)
    sellersprite_required = config_bool(config, "sellersprite_required", True)
    if product_filters.enabled and not sellersprite_required:
        raise UserFacingError(
            "启用 product_filters 时必须设置 sellersprite_required=true。"
        )

    extension_path_text = config_text(config, "extension_path")
    extension_path = resolve_path(extension_path_text) if extension_path_text else Path("")
    if browser_mode == "launch" and extension_path_text and not extension_path.exists():
        raise UserFacingError(f"没有找到卖家精灵扩展目录：{extension_path}")

    min_delay = config_float(config, "delay_seconds_min", 4)
    max_delay = config_float(config, "delay_seconds_max", 9)
    if max_delay < min_delay:
        max_delay = min_delay
    batch_pages_min = config_int(config, "batch_pause_pages_min", 20) or 0
    batch_pages_max = config_int(config, "batch_pause_pages_max", 30) or 0
    if batch_pages_min < 0 or batch_pages_max < 0:
        raise UserFacingError("配置项 batch_pause_pages_min / batch_pause_pages_max 不能小于 0。")
    if batch_pages_max and batch_pages_max < batch_pages_min:
        batch_pages_max = batch_pages_min
    batch_seconds_min = config_float(config, "batch_pause_seconds_min", 60)
    batch_seconds_max = config_float(config, "batch_pause_seconds_max", 180)
    if batch_seconds_min < 0 or batch_seconds_max < 0:
        raise UserFacingError("配置项 batch_pause_seconds_min / batch_pause_seconds_max 不能小于 0。")
    if batch_seconds_max < batch_seconds_min:
        batch_seconds_max = batch_seconds_min
    plugin_retry_wait_min = config_float(
        config,
        "plugin_retry_wait_seconds_min",
        config_float(config, "plugin_retry_wait_seconds", 10),
    )
    plugin_retry_wait_max = config_float(config, "plugin_retry_wait_seconds_max", 20)
    if plugin_retry_wait_min < 0 or plugin_retry_wait_max < 0:
        raise UserFacingError("配置项 plugin_retry_wait_seconds_min / plugin_retry_wait_seconds_max 不能小于 0。")
    if plugin_retry_wait_max < plugin_retry_wait_min:
        plugin_retry_wait_max = plugin_retry_wait_min
    page_scroll_max_rounds = max(config_int(config, "page_scroll_max_rounds", 18) or 0, 0)
    page_scroll_step_ratio = config_float(config, "page_scroll_step_ratio", 0.85) or 0.85
    if page_scroll_step_ratio <= 0:
        raise UserFacingError("配置项 page_scroll_step_ratio 必须大于 0。")
    page_scroll_wait_seconds = max(config_float(config, "page_scroll_wait_seconds", 1.0) or 0, 0)
    page_scroll_stable_rounds = max(config_int(config, "page_scroll_stable_rounds", 2) or 1, 1)
    delivery_config = build_delivery_location_config(config)

    raw_selectors = config.get("field_selectors") or {}
    field_selectors: Dict[str, List[str]] = {}
    if isinstance(raw_selectors, dict):
        for key, value in raw_selectors.items():
            if isinstance(value, list):
                field_selectors[key] = [str(item).strip() for item in value if str(item).strip()]

    max_pages_per_keyword = config_int(config, "max_pages_per_keyword", 7) or 7
    if max_pages_per_keyword < 1:
        raise UserFacingError("配置项 `max_pages_per_keyword` 必须大于 0。")
    keyword_sort_orders = parse_keyword_sort_orders(config.get("keyword_sort_orders"))
    store_sort_orders = parse_store_sort_orders(config.get("store_sort_orders"))
    store_page_limit = config_int(config, "store_page_limit")
    if mode == "keyword_search" and not keyword_sort_orders:
        keyword_sort_orders = ["Featured"]
    if mode == "storefront":
        if not store_sort_orders:
            store_sort_orders = prompt_store_sort_orders()
        if store_page_limit is None:
            store_page_limit = prompt_store_page_limit()
        if store_page_limit < 1:
            raise UserFacingError("店铺抓取页数必须大于 0。")
        if store_page_limit > 20:
            raise UserFacingError("店铺抓取最多支持 20 页，请把 store_page_limit 调整到 20 以内。")

    return FrontRuntimeConfig(
        mode=mode,
        job_id=slugify(job_id),
        outputs_root=outputs_root,
        keywords_file=keywords_file,
        store_urls_file=store_urls_file,
        start_url=start_url,
        max_pages_per_keyword=max_pages_per_keyword,
        include_sponsored=config_bool(config, "include_sponsored", False),
        keyword_sort_orders=keyword_sort_orders,
        store_page_limit=store_page_limit,
        store_sort_orders=store_sort_orders,
        browser_tab_concurrency=browser_tab_concurrency,
        product_filters=product_filters,
        record_contract_fingerprint=record_contract_fingerprint(product_filters),
        resume=False if no_resume else config_bool(config, "resume", True),
        browser_backend=browser_backend,
        browser_mode=browser_mode,
        chrome_binary=config_text(config, "chrome_binary"),
        chrome_user_data_dir=resolve_path(config_text(config, "chrome_user_data_dir", "chrome_profiles/category-rank-sellersprite")),
        chrome_profile_directory=config_text(config, "chrome_profile_directory", "Default") or "Default",
        debugger_address=config_text(config, "debugger_address", "127.0.0.1:9222"),
        extension_path=extension_path,
        activate_plugin=config_bool(config, "activate_plugin", True),
        page_timeout=config_int(config, "page_timeout", 90) or 90,
        plugin_timeout=config_int(config, "plugin_timeout", 120) or 120,
        plugin_retry_attempts=max(config_int(config, "plugin_retry_attempts", 5) or 0, 0),
        plugin_retry_wait_seconds=plugin_retry_wait_min,
        plugin_retry_wait_seconds_max=plugin_retry_wait_max,
        plugin_relaunch_retry_attempts=max(config_int(config, "plugin_relaunch_retry_attempts", 3) or 0, 0),
        plugin_relaunch_wait_seconds=max(config_float(config, "plugin_relaunch_wait_seconds", 300) or 0, 0),
        plugin_second_relaunch_retry_attempts=max(config_int(config, "plugin_second_relaunch_retry_attempts", 3) or 0, 0),
        plugin_second_relaunch_wait_seconds=max(config_float(config, "plugin_second_relaunch_wait_seconds", 600) or 0, 0),
        manual_pause_timeout=config_int(config, "manual_pause_timeout", 900) or 900,
        **delivery_config,
        delay_seconds_min=min_delay,
        delay_seconds_max=max_delay,
        batch_pause_pages_min=batch_pages_min,
        batch_pause_pages_max=batch_pages_max,
        batch_pause_seconds_min=batch_seconds_min,
        batch_pause_seconds_max=batch_seconds_max,
        page_scroll_before_extract=config_bool(config, "page_scroll_before_extract", True),
        page_scroll_max_rounds=page_scroll_max_rounds,
        page_scroll_step_ratio=page_scroll_step_ratio,
        page_scroll_wait_seconds=page_scroll_wait_seconds,
        page_scroll_stable_rounds=page_scroll_stable_rounds,
        sellersprite_required=sellersprite_required,
        sellersprite_min_enriched_records=max(
            config_int(config, "sellersprite_min_enriched_records", 1) or 1,
            1,
        ),
        sellersprite_min_fields_per_record=max(
            config_int(config, "sellersprite_min_fields_per_record", 2) or 2,
            1,
        ),
        sellersprite_stable_checks=max(
            config_int(config, "sellersprite_stable_checks", 3) or 3,
            1,
        ),
        save_debug_snapshots=config_bool(config, "save_debug_snapshots", True),
        field_selectors=field_selectors,
    )


def read_input_rows(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            return [{str(k or "").strip(): normalize_space(str(v or "")) for k, v in row.items()} for row in reader]
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise UserFacingError("缺少 openpyxl，无法读取 Excel 输入表。")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_space(str(value or "")) for value in rows[0]]
        output: List[Dict[str, str]] = []
        for raw_row in rows[1:]:
            row: Dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = raw_row[index] if index < len(raw_row) else ""
                row[header] = normalize_space(str(value or ""))
            output.append(row)
        return output
    raise UserFacingError(f"暂不支持的输入表格式：{path.suffix}")


def pick_column(row: Dict[str, str], aliases: Sequence[str]) -> str:
    normalized_to_key = {normalize_header(key): key for key in row.keys()}
    for alias in aliases:
        key = normalized_to_key.get(normalize_header(alias))
        if key:
            return normalize_space(row.get(key) or "")
    return ""


def load_keywords(path: Path) -> List[str]:
    keywords: List[str] = []
    seen = set()
    for row in read_input_rows(path):
        keyword = pick_column(row, ["keyword", "关键词"])
        if keyword and keyword not in seen:
            keywords.append(keyword)
            seen.add(keyword)
    if not keywords:
        raise UserFacingError("关键词表必须包含 `keyword` 或 `关键词` 列，且至少有一个关键词。")
    return keywords


def load_storefronts(path: Path) -> List[Dict[str, str]]:
    stores: List[Dict[str, str]] = []
    seen = set()
    for row in read_input_rows(path):
        store_url = pick_column(row, ["store_url", "店铺URL", "店铺url"])
        store_name = pick_column(row, ["store_name", "店铺名称", "店铺名"])
        if not store_url or store_url in seen:
            continue
        validate_amazon_url(store_url)
        stores.append({"store_url": store_url, "store_name": store_name})
        seen.add(store_url)
    if not stores:
        raise UserFacingError("店铺 URL 表必须包含 `store_url` 或 `店铺URL` 列，且至少有一个 Amazon 店铺链接。")
    return stores


def build_initial_queue(runtime: FrontRuntimeConfig) -> List[Dict[str, Any]]:
    if runtime.mode == "keyword_search":
        queue: List[Dict[str, Any]] = []
        for keyword in load_keywords(runtime.keywords_file):
            for sort_order in runtime.keyword_sort_orders:
                queue.append(
                    {
                        "source_type": "keyword_search",
                        "source_id": f"{keyword}|sort:{sort_order}",
                        "keyword": keyword,
                        "search_sort_order": sort_order,
                        "page_number": 1,
                        "page_url": build_keyword_search_url(keyword, sort_order),
                        "seen_page_urls": [],
                        "previous_page_asins": [],
                    }
                )
        return queue
    if runtime.mode == "storefront":
        queue: List[Dict[str, Any]] = []
        for store in load_storefronts(runtime.store_urls_file):
            for sort_order in runtime.store_sort_orders:
                queue.append(
                    {
                        "source_type": "storefront",
                        "source_id": f"{store['store_url']}|sort:{sort_order}",
                        "store_url": store["store_url"],
                        "store_name": store.get("store_name", ""),
                        "store_sort_order": sort_order,
                        "page_number": 1,
                        "page_url": store["store_url"],
                        "seen_page_urls": [],
                        "previous_page_asins": [],
                        "prepared_storefront": False,
                    }
                )
        return queue
    return []


def build_keyword_search_url(keyword: str, sort_order: str = "Featured") -> str:
    base_url = f"https://www.amazon.com/s?k={quote_plus(keyword)}"
    return apply_sort_to_url(base_url, sort_order)


def apply_sort_to_url(url: str, sort_order: str) -> str:
    sort_code = SEARCH_SORT_OPTIONS.get(sort_order)
    if not sort_code:
        return url
    parsed = urlparse(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query["s"] = [sort_code]
    flat_query = urlencode([(key, value) for key, values in query.items() for value in values])
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, flat_query, parsed.fragment))


def wait_for_product_cards(
    driver: WebDriver,
    runtime: FrontRuntimeConfig,
    timeout: Optional[int] = None,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    effective_runtime = runtime if timeout is None else replace(runtime, page_timeout=timeout)
    try:
        wait_for_amazon_products(  # type: ignore[arg-type]
            driver,
            effective_runtime,
            stop_event=stop_event,
        )
        return True
    except TimeoutException:
        return False


def wait_for_page_or_manual_front(
    driver: WebDriver,
    runtime: FrontRuntimeConfig,
    state: FrontStateStore,
    failures_path: Path,
    debug_dir: Path,
    current: Dict[str, Any],
) -> bool:
    block_reason = detect_block(driver)
    if block_reason:
        state.mark_manual_pause(block_reason, driver.current_url)
        cleared = wait_for_manual_clear(driver, block_reason, runtime.manual_pause_timeout)
        if cleared:
            state.clear_manual_pause()
        if not cleared:
            log_front_failure(failures_path, state, current, block_reason, "人工处理超时", driver.current_url)
            if runtime.save_debug_snapshots:
                save_debug_snapshot(driver, debug_dir, block_reason)
            raise VerificationUnconfirmedError(
                f"{block_reason}_unconfirmed: 人工处理超时，任务已停止且未提取当前页数据。"
            )
    if wait_for_product_cards(driver, runtime):
        return True
    return False


def find_store_products_url(driver: WebDriver) -> str:
    script = r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const phrases = [
  'products', 'shop all', 'all products', 'see all products', 'visit the store',
  '商品', '全部商品', '所有商品', '店铺商品'
];
const links = [...document.querySelectorAll('a[href]')];
for (const link of links) {
  const text = norm(link.innerText || link.textContent || link.getAttribute('aria-label') || '').toLowerCase();
  const href = absUrl(link.getAttribute('href') || link.href || '');
  if (!href || !/amazon\./i.test(href)) continue;
  if (phrases.some(phrase => text.includes(phrase))) return href;
}
for (const link of links) {
  const href = absUrl(link.getAttribute('href') || link.href || '');
  if (/\/s\?/.test(href) && /(?:me=|rh=|i=merchant-items)/.test(href)) return href;
}
return '';
"""
    try:
        return str(driver.execute_script(script) or "")
    except (JavascriptException, WebDriverException):
        return ""


def prepare_storefront_page(
    driver: WebDriver,
    runtime: FrontRuntimeConfig,
    current: Dict[str, Any],
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
) -> None:
    if current.get("prepared_storefront"):
        return
    sort_order = str(current.get("store_sort_order") or "Featured")
    if wait_for_product_cards(
        driver,
        runtime,
        timeout=12,
        stop_event=stop_event,
    ):
        sorted_url = apply_sort_to_url(driver.current_url, sort_order)
        if clean_url(sorted_url) != clean_url(driver.current_url):
            open_amazon_page(
                driver,
                sorted_url,
                runtime,
                on_manual_pause,
                on_manual_resume,
                stop_event=stop_event,
            )
        current["prepared_storefront"] = True
        current["page_url"] = driver.current_url
        return
    products_url = find_store_products_url(driver)
    if products_url and clean_url(products_url) != clean_url(driver.current_url):
        sorted_url = apply_sort_to_url(products_url, sort_order)
        open_amazon_page(
            driver,
            sorted_url,
            runtime,
            on_manual_pause,
            on_manual_resume,
            stop_event=stop_event,
        )
        current["page_url"] = sorted_url
    current["prepared_storefront"] = True


def extract_front_product_cards(driver: WebDriver, include_sponsored: bool) -> List[Dict[str, Any]]:
    script = r"""
const includeSponsored = arguments[0];
const asinRe = /\b([A-Z0-9]{10})\b/;
const asinUrlRe = /\/(?:dp|gp\/product)\/([A-Z0-9]{10})(?:[/?#]|$)/i;
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const isSponsored = (el) => {
  if (el.matches('[data-component-type*="sp-sponsored" i], [class*="AdHolder" i]')) return true;
  const explicit = el.querySelector('[aria-label*="Sponsored" i], .puis-sponsored-label-text, [class*="sponsored-label" i]');
  if (explicit) return true;
  const shortLabels = [...el.querySelectorAll('span, a, i')].map(item => norm(item.innerText || item.textContent || item.getAttribute('aria-label') || ''));
  if (shortLabels.some(text => /^(Sponsored|赞助)$/i.test(text))) return true;
  return false;
};
const getAsin = (el) => {
  for (const attr of ['data-asin', 'asin', 'data-csa-c-asin']) {
    const value = el.getAttribute(attr);
    if (value && /^[A-Z0-9]{10}$/.test(value.trim())) return value.trim();
  }
  for (const a of el.querySelectorAll('a[href]')) {
    const match = a.href.match(asinUrlRe);
    if (match) return match[1];
  }
  const textMatch = norm(el.innerText || el.textContent || '').match(asinRe);
  return textMatch ? textMatch[1] : '';
};
const getUrl = (el, asin) => {
  const links = [...el.querySelectorAll('a[href]')];
  for (const a of links) {
    if (asin && a.href.includes(asin)) return absUrl(a.href);
  }
  for (const a of links) {
    if (/\/(?:dp|gp\/product)\//i.test(a.href)) return absUrl(a.href);
  }
  return '';
};
const getTitle = (el) => {
  const selectors = ['h2 a span', 'h2 span', 'a.a-link-normal span', 'img[alt]', '.p13n-sc-truncate'];
  for (const selector of selectors) {
    const item = el.querySelector(selector);
    if (!item) continue;
    const text = norm(item.getAttribute('alt') || item.innerText || item.textContent || '');
    if (text && text.length > 5) return text;
  }
  return '';
};
const getSellerCountryFlagCode = (el) => {
  const flagElements = [...el.querySelectorAll('[class*="flag-icon-"], [class*="icp-nav-flag-"]')];
  for (const flag of flagElements) {
    const classText = String(flag.getAttribute('class') || '');
    const match = classText.match(/(?:flag-icon|icp-nav-flag)-([a-z]{2})\b/i);
    if (match) return match[1].toLowerCase();
    const hint = norm(flag.getAttribute('title') || flag.getAttribute('aria-label') || '');
    if (hint) return hint;
  }
  return '';
};
const getBsrText = (el) => [...el.querySelectorAll('.rank-number-box .bsr-list-item')]
  .map(item => norm(item.innerText || item.textContent || ''))
  .filter(Boolean)
  .join(' ');
const selectors = [
  '.s-result-item[data-asin]:not([data-asin=""])',
  '[data-component-type="s-search-result"][data-asin]:not([data-asin=""])',
  '#gridItemRoot',
  '.zg-grid-general-faceout',
  '.p13n-grid-content',
  '[data-asin]:not([data-asin=""])'
];
const elements = [];
const seenElements = new Set();
for (const selector of selectors) {
  for (const el of document.querySelectorAll(selector)) {
    if (seenElements.has(el)) continue;
    seenElements.add(el);
    elements.push(el);
  }
}
const seenAsins = new Set();
const cards = [];
for (const el of elements) {
  const asin = getAsin(el);
  if (!asin || seenAsins.has(asin)) continue;
  const sponsored = isSponsored(el);
  if (sponsored && !includeSponsored) continue;
  seenAsins.add(asin);
  cards.push({
    asin,
    title: getTitle(el),
    product_url: getUrl(el, asin),
    rank: String(cards.length + 1),
    is_sponsored: sponsored ? 'yes' : 'no',
    seller_country_flag_code: getSellerCountryFlagCode(el),
    bsr_text: getBsrText(el),
    text: norm(el.innerText || el.textContent || '')
  });
}
return cards;
"""
    result = driver.execute_script(script, include_sponsored)
    if result is None:
        return []
    if not isinstance(result, list):
        raise WebDriverException("商品卡片提取脚本返回了非数组结果。")
    return list(result)


def merge_front_product_data(
    driver: WebDriver,
    runtime: FrontRuntimeConfig,
    current: Dict[str, Any],
    plugin_status: str,
) -> List[Dict[str, Any]]:
    # Always retain sponsored cards in the raw page evidence used by pagination.
    # They are excluded from written records below when include_sponsored=false.
    cards = extract_front_product_cards(driver, True)
    table_rows = extract_table_rows(
        driver,
        strict=bool(getattr(runtime, "sellersprite_required", True)),
    )
    table_by_asin: Dict[str, Dict[str, Any]] = {}
    for row in table_rows:
        asin = str(row.get("asin") or "")
        if not asin:
            continue
        parsed = parse_table_row_fields(row)
        target = table_by_asin.setdefault(asin, {})
        for key, value in parsed.items():
            if key in {"fulfillment_method", "fulfillment_method_raw"}:
                continue
            if value and not target.get(key):
                target[key] = value
        table_method, table_raw = select_fulfillment_evidence(
            (
                target.get("fulfillment_method"),
                target.get("fulfillment_method_raw"),
            ),
            (
                parsed.get("fulfillment_method"),
                parsed.get("fulfillment_method_raw"),
            ),
        )
        target["fulfillment_method"] = table_method
        target["fulfillment_method_raw"] = table_raw
        row_ranks = parse_subcategory_bsr_ranks(
            str(row.get("bsr_text") or row.get("text") or "")
        )
        if row_ranks:
            target["subcategory_bsr_ranks"] = normalize_subcategory_bsr_ranks(
                [*(target.get("subcategory_bsr_ranks") or []), *row_ranks]
            )

    records: List[Dict[str, Any]] = []
    page_number = int(current.get("page_number") or 1)
    for card in cards:
        asin = str(card.get("asin") or "")
        if not asin:
            continue
        record: Dict[str, Any] = {
            "source_type": current.get("source_type", ""),
            "source_id": current.get("source_id", ""),
            "keyword": current.get("keyword", ""),
            "search_sort_order": current.get("search_sort_order", ""),
            "store_url": current.get("store_url", ""),
            "store_name": current.get("store_name", ""),
            "store_sort_order": current.get("store_sort_order", ""),
            "category_path": current.get("category_path", ""),
            "category_name": current.get("category_name", ""),
            "category_node_id": current.get("category_node_id", ""),
            "source_url": current.get("source_url", current.get("page_url", "")),
            "page_url": driver.current_url,
            "page_number": page_number,
            "rank": card.get("rank") or "",
            "is_sponsored": card.get("is_sponsored") or "unknown",
            "asin": asin,
            "title": card.get("title") or "",
            "product_url": card.get("product_url") or "",
            "scraped_at": now_iso(),
            "load_status": plugin_status,
            "note": "",
            "fulfillment_method_raw": "",
        }
        text = str(card.get("text") or "")
        bsr_text = str(card.get("bsr_text") or "")
        for field_name in REQUESTED_DATA_FIELDS:
            record[field_name] = ""
        record["subcategory_bsr_ranks"] = []
        for field_name, value in table_by_asin.get(asin, {}).items():
            if field_name == "subcategory_bsr_ranks" and value:
                record[field_name] = normalize_subcategory_bsr_ranks(value)
            elif field_name == "fulfillment_method":
                continue
            elif field_name == "fulfillment_method_raw" and value:
                record[field_name] = normalize_space(str(value))[:120]
            elif field_name in REQUESTED_DATA_FIELDS and value:
                record[field_name] = value
        selector_values = extract_by_selectors(driver, card, runtime.field_selectors)
        for field_name, value in selector_values.items():
            if field_name == "subcategory_bsr_ranks" and value:
                record[field_name] = parse_subcategory_bsr_ranks(str(value))
            elif field_name == "fulfillment_method" and value:
                continue
            elif field_name in REQUESTED_DATA_FIELDS and value:
                record[field_name] = normalize_space(str(value))
        if not record["subcategory_bsr_ranks"]:
            record["subcategory_bsr_ranks"] = parse_subcategory_bsr_ranks(
                bsr_text or text
            )
        for field_name in REQUESTED_DATA_FIELDS:
            if field_name == "fulfillment_method":
                continue
            if not record.get(field_name):
                record[field_name] = parse_field_from_text(field_name, text)
        selector_fulfillment = parse_fulfillment_evidence(
            str(selector_values.get("fulfillment_method") or ""),
            explicit_value=True,
        )
        table_fulfillment = (
            table_by_asin.get(asin, {}).get("fulfillment_method"),
            table_by_asin.get(asin, {}).get("fulfillment_method_raw"),
        )
        card_fulfillment = parse_fulfillment_evidence(text)
        method, raw = select_fulfillment_evidence(
            selector_fulfillment,
            table_fulfillment,
            card_fulfillment,
        )
        record["fulfillment_method"] = method
        record["fulfillment_method_raw"] = raw
        record["fulfillment_method"] = normalize_fulfillment_method(
            record.get("fulfillment_method")
        )
        if not record.get("seller_country"):
            record["seller_country"] = country_from_flag_code_or_text(str(card.get("seller_country_flag_code") or ""))
        missing_count = sum(1 for field_name in REQUESTED_DATA_FIELDS if not record.get(field_name))
        if plugin_status != "ok":
            record["note"] = "插件数据加载超时，已保存页面可见数据"
        elif missing_count == len(REQUESTED_DATA_FIELDS):
            record["note"] = "插件未展示或选择器未匹配"
        records.append(record)
    return records


def front_page_key(current: Dict[str, Any], page_url: str) -> str:
    return "|".join(
        [
            str(current.get("source_type") or ""),
            str(current.get("source_id") or ""),
            f"page:{current.get('page_number') or 1}",
            clean_url(page_url),
        ]
    )


def log_front_failure(
    failures_path: Path,
    state: FrontStateStore,
    current: Dict[str, Any],
    reason: str,
    message: str,
    page_url: str,
) -> None:
    append_jsonl(
        failures_path,
        {
            "time": now_iso(),
            "source_type": current.get("source_type", ""),
            "keyword": current.get("keyword", ""),
            "search_sort_order": current.get("search_sort_order", ""),
            "store_url": current.get("store_url", ""),
            "store_name": current.get("store_name", ""),
            "store_sort_order": current.get("store_sort_order", ""),
            "category_path": current.get("category_path", ""),
            "page_number": current.get("page_number", ""),
            "page_url": page_url,
            "reason": reason,
            "message": message,
        },
    )
    state.log_failure()


def should_stop_on_repeated_store_page(current: Dict[str, Any], records: Sequence[Dict[str, Any]]) -> bool:
    if current.get("source_type") != "storefront":
        return False
    previous = set(current.get("previous_page_asins") or [])
    current_asins = {str(record.get("asin") or "") for record in records if record.get("asin")}
    return bool(previous and current_asins and previous == current_asins)


def build_next_front_task(
    driver: WebDriver,
    runtime: FrontRuntimeConfig,
    current: Dict[str, Any],
    raw_records: Sequence[Dict[str, Any]],
) -> Tuple[Optional[Dict[str, Any]], str]:
    source_type = current.get("source_type")
    page_number = int(current.get("page_number") or 1)
    if not raw_records:
        return None, "empty_page"
    if should_stop_on_repeated_store_page(current, raw_records):
        return None, "repeated_asins"

    next_url = find_next_page_url(driver, strict=True)
    if not next_url:
        return None, "no_next_page"

    seen_urls = [clean_url(url) for url in current.get("seen_page_urls") or []]
    current_clean_url = clean_url(driver.current_url)
    if current_clean_url not in seen_urls:
        seen_urls.append(current_clean_url)
    next_clean_url = clean_url(next_url)
    if next_clean_url in seen_urls:
        return None, "repeated_next_url"

    if source_type == "keyword_search" and page_number >= runtime.max_pages_per_keyword:
        return None, "keyword_page_limit"
    if source_type == "storefront" and runtime.store_page_limit is not None and page_number >= runtime.store_page_limit:
        return None, "store_page_limit"

    next_task = copy.deepcopy(current)
    next_task.pop("worker_retry_count", None)
    next_task["page_number"] = page_number + 1
    next_task["page_url"] = next_url
    next_task["seen_page_urls"] = seen_urls
    next_task["previous_page_asins"] = [
        str(record.get("asin") or "") for record in raw_records if record.get("asin")
    ]
    return next_task, ""


def add_unique(values: List[str], value: Any) -> None:
    text = normalize_space(str(value or ""))
    if text and text not in values:
        values.append(text)


def numeric_sort_value(value: Any, fallback: int = 1_000_000) -> int:
    try:
        text = normalize_space(str(value or ""))
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else fallback
    except (TypeError, ValueError):
        return fallback


def merge_subcategory_bsr_ranks(existing: Any, incoming: Any) -> List[Dict[str, Any]]:
    """Merge rank/name pairs atomically, preserving the first rank per category."""
    merged: List[Dict[str, Any]] = []
    seen_categories = set()
    for item in normalize_subcategory_bsr_ranks(existing) + normalize_subcategory_bsr_ranks(incoming):
        category_key = str(item["category_name"]).casefold()
        if category_key in seen_categories:
            continue
        seen_categories.add(category_key)
        merged.append(dict(item))
    return merged


def build_front_dedup_rows(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    accumulators: Dict[str, Dict[str, List[str]]] = {}
    best_keys: Dict[str, tuple[int, int]] = {}

    for row in records:
        asin = str(row.get("asin") or "")
        if not asin:
            continue
        target = grouped.setdefault(asin, {})
        acc = accumulators.setdefault(
            asin,
            {
                "source_types": [],
                "source_keywords": [],
                "source_search_sort_orders": [],
                "source_store_urls": [],
                "source_store_names": [],
                "source_store_sort_orders": [],
                "source_category_paths": [],
                "source_page_urls": [],
                "notes": [],
                "statuses": [],
            },
        )
        add_unique(acc["source_types"], row.get("source_type") or "bsr_category")
        add_unique(acc["source_keywords"], row.get("keyword"))
        add_unique(acc["source_search_sort_orders"], row.get("search_sort_order"))
        add_unique(acc["source_store_urls"], row.get("store_url"))
        add_unique(acc["source_store_names"], row.get("store_name"))
        add_unique(acc["source_store_sort_orders"], row.get("store_sort_order"))
        add_unique(acc["source_category_paths"], row.get("category_path"))
        add_unique(acc["source_page_urls"], row.get("page_url") or row.get("category_url"))
        add_unique(acc["notes"], row.get("note"))
        add_unique(acc["statuses"], row.get("load_status"))

        for key in ["asin", "title", "product_url", "scraped_at"]:
            if row.get(key) and not target.get(key):
                target[key] = row[key]
        for field_name in REQUESTED_DATA_FIELDS:
            if row.get(field_name) and not target.get(field_name):
                target[field_name] = row[field_name]
        target["subcategory_bsr_ranks"] = merge_subcategory_bsr_ranks(
            target.get("subcategory_bsr_ranks"), row.get("subcategory_bsr_ranks")
        )

        page_number = numeric_sort_value(row.get("page_number"))
        rank = numeric_sort_value(row.get("rank"))
        best_key = (page_number, rank)
        if asin not in best_keys or best_key < best_keys[asin]:
            best_keys[asin] = best_key
            target["best_page_number"] = row.get("page_number", "")
            target["best_rank"] = row.get("rank", "")

        target["appearance_count"] = int(target.get("appearance_count") or 0) + 1

    output: List[Dict[str, Any]] = []
    for asin, row in grouped.items():
        acc = accumulators.get(asin, {})
        merged = dict(row)
        merged["source_types"] = " ; ".join(acc.get("source_types", []))
        merged["source_keywords"] = " ; ".join(acc.get("source_keywords", []))
        merged["source_search_sort_orders"] = " ; ".join(acc.get("source_search_sort_orders", []))
        merged["source_store_urls"] = " ; ".join(acc.get("source_store_urls", []))
        merged["source_store_names"] = " ; ".join(acc.get("source_store_names", []))
        merged["source_store_sort_orders"] = " ; ".join(acc.get("source_store_sort_orders", []))
        merged["source_category_paths"] = " ; ".join(acc.get("source_category_paths", []))
        merged["source_page_urls"] = " ; ".join(acc.get("source_page_urls", []))
        merged["load_status"] = "ok" if "ok" in acc.get("statuses", []) else " ; ".join(acc.get("statuses", []))
        merged["note"] = " ; ".join(acc.get("notes", []))
        output.append(merged)
    output.sort(key=lambda item: (numeric_sort_value(item.get("best_page_number")), numeric_sort_value(item.get("best_rank")), item.get("asin", "")))
    return output


def write_front_sheet(ws: Any, headers: Sequence[str], rows: Sequence[Dict[str, Any]], raw_headers: bool = False) -> None:
    ws.append(list(headers))
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        if raw_headers:
            ws.append([row.get(header, "") for header in headers])
        else:
            values: List[Any] = []
            for field, header in FRONT_FIELD_TO_HEADER.items():
                if header not in headers:
                    continue
                value = row.get(field, "")
                if field == "subcategory_bsr_ranks":
                    value = format_subcategory_bsr_ranks(value)
                values.append(value)
            ws.append(values)
    for column_index, header in enumerate(headers, start=1):
        width = 42 if header == "子类目节点排名" else min(max(len(str(header)) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(column_index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_front_workbook(records_path: Path, failures_path: Path, output_xlsx: Path) -> None:
    if Workbook is None:
        raise UserFacingError("缺少 openpyxl，无法生成 Excel。")
    records = read_jsonl(records_path)
    failures = read_jsonl(failures_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "ASIN去重总表"
    write_front_sheet(ws, FRONT_DEDUP_HEADERS, build_front_dedup_rows(records))

    ws_fail = wb.create_sheet("失败页面")
    failure_headers = ["time", "source_type", "keyword", "search_sort_order", "store_url", "store_name", "store_sort_order", "category_path", "page_number", "page_url", "reason", "message"]
    write_front_sheet(ws_fail, failure_headers, failures, raw_headers=True)

    ensure_dir(output_xlsx.parent)
    wb.save(output_xlsx)


def write_jsonl_atomic(path: Path, records: Sequence[Dict[str, Any]]) -> None:
    ensure_dir(path.parent)
    temp_path = path.with_name(
        f".{path.name}.{os.getpid()}.{threading.get_ident()}.{time.time_ns()}.tmp"
    )
    try:
        with temp_path.open("w", encoding="utf-8") as fh:
            for record in records:
                fh.write(json.dumps(record, ensure_ascii=False) + "\n")
            fh.flush()
            os.fsync(fh.fileno())
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def materialize_front_records(state: FrontStateStore, records_path: Path) -> int:
    records: List[Dict[str, Any]] = []
    seen = set()
    for payload in state.iter_page_results():
        page_key = str(payload.get("page_key") or "")
        for record in payload.get("records") or []:
            if not isinstance(record, dict):
                continue
            identity = (page_key, str(record.get("asin") or ""))
            if identity in seen:
                continue
            seen.add(identity)
            records.append(record)
    write_jsonl_atomic(records_path, records)
    return len(records)


@dataclass
class FrontPageResult:
    worker_id: str
    task: Dict[str, Any]
    page_key: str = ""
    page_url: str = ""
    raw_records: List[Dict[str, Any]] = field(default_factory=list)
    accepted_records: List[Dict[str, Any]] = field(default_factory=list)
    rejection_counts: Dict[str, int] = field(default_factory=dict)
    plugin_status: str = ""
    next_task: Optional[Dict[str, Any]] = None
    finish_reason: str = ""
    error_reason: str = ""
    error_message: str = ""
    fatal: bool = False
    readiness: Dict[str, Any] = field(default_factory=dict)


class ManualActionCoordinator:
    """Allow only one browser tab to request human action at a time."""

    def __init__(self) -> None:
        self._condition = threading.Condition()
        self._owner = ""
        self._snapshot: Dict[str, Any] = {}
        self._snapshot_stack: List[Dict[str, Any]] = []

    def begin(
        self,
        worker_id: str,
        task: Dict[str, Any],
        reason: str,
        page_url: str,
        stop_event: Optional[threading.Event] = None,
    ) -> None:
        with self._condition:
            while self._owner and self._owner != worker_id:
                if stop_event is not None and stop_event.is_set():
                    raise ConcurrentWorkerCancelled(
                        "并发任务正在停止，已取消等待人工操作。"
                    )
                self._condition.wait(timeout=0.25)
            self._owner = worker_id
            snapshot = {
                "paused_at": now_iso(),
                "worker_id": worker_id,
                "task_id": front_task_identity(task),
                "source_id": front_source_id(task),
                "page_url": page_url,
                "reason": reason,
                "tab": worker_id,
            }
            self._snapshot_stack.append(snapshot)
            self._snapshot = {worker_id: copy.deepcopy(snapshot)}
            self._condition.notify_all()

    def end(self, worker_id: str) -> None:
        with self._condition:
            if self._owner == worker_id:
                if self._snapshot_stack:
                    self._snapshot_stack.pop()
                if self._snapshot_stack:
                    self._snapshot = {
                        worker_id: copy.deepcopy(self._snapshot_stack[-1])
                    }
                else:
                    self._owner = ""
                    self._snapshot = {}
                self._condition.notify_all()

    def release_all(self, worker_id: str) -> None:
        """Release every nested pause held by a worker after its task exits."""
        with self._condition:
            if self._owner == worker_id:
                self._owner = ""
                self._snapshot = {}
                self._snapshot_stack = []
                self._condition.notify_all()

    def wait_if_paused(
        self,
        worker_id: str,
        stop_event: Optional[threading.Event] = None,
    ) -> None:
        with self._condition:
            while self._owner and self._owner != worker_id:
                if stop_event is not None and stop_event.is_set():
                    raise ConcurrentWorkerCancelled(
                        "并发任务正在停止，已取消等待人工操作。"
                    )
                self._condition.wait(timeout=0.25)

    def is_paused(self) -> bool:
        with self._condition:
            return bool(self._owner)

    def snapshot(self) -> Dict[str, Any]:
        with self._condition:
            return copy.deepcopy(self._snapshot)


class NavigationThrottle:
    """Globally stagger page navigations across workers."""

    def __init__(self, minimum: float, maximum: float) -> None:
        self.minimum = max(float(minimum), 0.0)
        self.maximum = max(float(maximum), self.minimum)
        self._lock = threading.Lock()
        self._next_navigation_at = 0.0

    def wait(self, stop_event: Optional[threading.Event] = None) -> None:
        with self._lock:
            now = time.monotonic()
            navigation_at = max(now, self._next_navigation_at)
            self._next_navigation_at = navigation_at + random.uniform(self.minimum, self.maximum)
        delay = navigation_at - time.monotonic()
        if delay > 0:
            if stop_event is not None:
                if stop_event.wait(delay):
                    raise ConcurrentWorkerCancelled(
                        "并发任务正在停止，已取消导航等待。"
                    )
            else:
                time.sleep(delay)


class DeliveryDomainLocks:
    def __init__(self) -> None:
        self._guard = threading.Lock()
        self._locks: Dict[str, threading.Lock] = {}

    def for_url(self, url: str) -> threading.Lock:
        domain = (urlparse(url).hostname or "unknown").lower()
        with self._guard:
            return self._locks.setdefault(domain, threading.Lock())


def worker_debug_label(worker_id: str, task: Dict[str, Any], reason: str) -> str:
    task_digest = hashlib.sha256(front_task_identity(task).encode("utf-8")).hexdigest()[:10]
    safe_reason = re.sub(r"[^A-Za-z0-9_-]+", "_", reason)[:40]
    return f"{int(time.time() * 1000)}_{worker_id}_{task_digest}_{safe_reason}"


class FrontWorker:
    """A dedicated thread with a private Playwright connection and owned tab."""

    def __init__(
        self,
        worker_id: str,
        runtime: FrontRuntimeConfig,
        results: "queue.Queue[FrontPageResult]",
        manual: ManualActionCoordinator,
        throttle: NavigationThrottle,
        delivery_locks: DeliveryDomainLocks,
        debug_dir: Path,
    ) -> None:
        self.worker_id = worker_id
        self.runtime = runtime
        self.results = results
        self.manual = manual
        self.throttle = throttle
        self.delivery_locks = delivery_locks
        self.debug_dir = debug_dir
        self.tasks: "queue.Queue[Optional[Dict[str, Any]]]" = queue.Queue(maxsize=1)
        self.thread = threading.Thread(target=self._run, name=f"amazon-{worker_id}", daemon=True)
        self.stop_event = threading.Event()
        self.driver: Optional[WebDriver] = None
        self._confirmed_domains: set[str] = set()
        self._active_task: Dict[str, Any] = {}
        self._readiness: Dict[str, Any] = {}

    def start(self) -> None:
        self.thread.start()

    def submit(self, task: Dict[str, Any]) -> None:
        self.tasks.put(copy.deepcopy(task))

    def stop(self) -> None:
        self.stop_event.set()
        try:
            self.tasks.put_nowait(None)
        except queue.Full:
            pass

    def join(self, timeout: Optional[float] = None) -> bool:
        self.thread.join(timeout=timeout)
        return not self.thread.is_alive()

    def is_alive(self) -> bool:
        return self.thread.is_alive()

    def _raise_if_stopping(self) -> None:
        if self.stop_event.is_set():
            raise ConcurrentWorkerCancelled("并发任务正在停止，当前标签页已取消。")

    def _ensure_driver(self) -> WebDriver:
        if self.driver is None:
            self.driver = start_driver(self.runtime)  # type: ignore[arg-type]
            self._confirmed_domains.clear()
        return self.driver

    def _close_driver(self) -> None:
        current = self.driver
        self.driver = None
        self._confirmed_domains.clear()
        if current is None:
            return
        try:
            current.quit()
        except Exception:
            pass

    def _manual_pause(self, reason: str, page_url: str) -> None:
        self.manual.begin(
            self.worker_id,
            self._active_task,
            reason,
            page_url,
            stop_event=self.stop_event,
        )

    def _manual_resume(self) -> None:
        self.manual.end(self.worker_id)

    def _publish_readiness(self, report: Dict[str, Any]) -> None:
        self._readiness = safe_sellersprite_readiness(report)

    def _before_navigation(self) -> None:
        self._raise_if_stopping()
        self.manual.wait_if_paused(self.worker_id, self.stop_event)
        self.throttle.wait(self.stop_event)

    def _open_page(self, url: str) -> None:
        self._before_navigation()
        driver = self._ensure_driver()
        domain = (urlparse(url).hostname or "unknown").lower()
        if domain not in self._confirmed_domains:
            domain_lock = self.delivery_locks.for_url(url)
            while not domain_lock.acquire(timeout=0.25):
                self._raise_if_stopping()
            try:
                open_amazon_page(
                    driver,
                    url,
                    self.runtime,
                    on_manual_pause=self._manual_pause,
                    on_manual_resume=self._manual_resume,
                    stop_event=self.stop_event,
                )
            finally:
                domain_lock.release()
            self._confirmed_domains.add(domain)
            return
        open_amazon_page(
            driver,
            url,
            self.runtime,
            on_manual_pause=self._manual_pause,
            on_manual_resume=self._manual_resume,
            stop_event=self.stop_event,
        )

    def _restart_plugin_driver(self, current_driver: WebDriver, page_url: str, wait_seconds: float) -> WebDriver:
        if self.driver is current_driver:
            self._close_driver()
        else:
            try:
                current_driver.quit()
            except Exception:
                pass
        if wait_seconds > 0 and self.stop_event.wait(wait_seconds):
            raise ConcurrentWorkerCancelled("并发任务正在停止，已取消插件重启等待。")
        self._raise_if_stopping()
        driver = self._ensure_driver()
        self._open_page(page_url)
        try:
            wait_for_amazon_products(  # type: ignore[arg-type]
                driver,
                self.runtime,
                stop_event=self.stop_event,
            )
        except TimeoutException:
            pass
        return driver

    def _wait_for_page_or_manual(self, current: Dict[str, Any]) -> bool:
        driver = self._ensure_driver()
        block_reason = detect_block(driver)
        if block_reason:
            self.manual.begin(
                self.worker_id,
                current,
                block_reason,
                driver.current_url,
                stop_event=self.stop_event,
            )
            try:
                cleared = wait_for_manual_clear(
                    driver,
                    block_reason,
                    self.runtime.manual_pause_timeout,
                    stop_event=self.stop_event,
                )
            finally:
                self.manual.end(self.worker_id)
            if not cleared:
                raise VerificationUnconfirmedError(
                    f"{block_reason}_unconfirmed: 人工处理超时，任务已停止且未提取当前页数据。"
                )
        return wait_for_product_cards(
            driver,
            self.runtime,
            stop_event=self.stop_event,
        )

    def _save_debug(self, task: Dict[str, Any], reason: str) -> None:
        if not self.runtime.save_debug_snapshots or self.driver is None:
            return
        save_debug_snapshot(
            self.driver,
            self.debug_dir,
            worker_debug_label(self.worker_id, task, reason),
        )

    def _process(self, task: Dict[str, Any]) -> FrontPageResult:
        leased_task = copy.deepcopy(task)
        current = copy.deepcopy(task)
        self._active_task = leased_task
        self._readiness = {}
        page_url = str(current.get("page_url") or "")
        try:
            self._open_page(page_url)
            driver = self._ensure_driver()
            if current.get("source_type") == "storefront":
                prepare_storefront_page(
                    driver,
                    self.runtime,
                    current,
                    on_manual_pause=self._manual_pause,
                    on_manual_resume=self._manual_resume,
                    stop_event=self.stop_event,
                )
            self.manual.wait_if_paused(self.worker_id, self.stop_event)
            if not self._wait_for_page_or_manual(current):
                self._save_debug(current, "no_product_cards")
                self._close_driver()
                return FrontPageResult(
                    worker_id=self.worker_id,
                    task=leased_task,
                    page_url=str(getattr(driver, "current_url", "") or page_url),
                    error_reason="no_product_cards",
                    error_message="页面未检测到商品卡片",
                    finish_reason="page_wait_failed",
                )

            plugin_status = wait_for_sellersprite_data_or_prompt(
                driver,
                self.runtime,  # type: ignore[arg-type]
                on_manual_pause=self._manual_pause,
                on_manual_resume=self._manual_resume,
                restart_driver=self._restart_plugin_driver,
                on_readiness=self._publish_readiness,
                before_navigation=self._before_navigation,
                stop_event=self.stop_event,
            )
            self.manual.wait_if_paused(self.worker_id, self.stop_event)
            driver = self._ensure_driver()
            if plugin_status == "blocked":
                self._save_debug(current, "verification_timeout")
                raise VerificationUnconfirmedError(
                    "sellersprite_verification_unconfirmed: 人工处理超时，任务已停止且未提取当前页数据。"
                )

            actual_url = str(driver.current_url or page_url)
            page_key = front_page_key(current, actual_url)
            raw_records = merge_front_product_data(driver, self.runtime, current, plugin_status)
            if not raw_records:
                raise WebDriverException(
                    "页面已检测到商品卡片，但提取结果为空；为避免静默漏页，保留断点并重试。"
                )
            eligible_records = (
                raw_records
                if self.runtime.include_sponsored
                else [
                    record
                    for record in raw_records
                    if str(record.get("is_sponsored") or "").lower() != "yes"
                ]
            )
            accepted_records, rejection_counts = filter_product_records(
                eligible_records, self.runtime.product_filters
            )
            sponsored_excluded = len(raw_records) - len(eligible_records)
            if sponsored_excluded:
                rejection_counts["sponsored_excluded"] = sponsored_excluded
            next_task, finish_reason = build_next_front_task(
                driver, self.runtime, current, raw_records
            )
            if plugin_status not in {"ok", "not_required"}:
                self._save_debug(current, f"plugin_{plugin_status}")
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_key=page_key,
                page_url=actual_url,
                raw_records=raw_records,
                accepted_records=accepted_records,
                rejection_counts=rejection_counts,
                plugin_status=plugin_status,
                next_task=next_task,
                finish_reason=finish_reason,
                readiness=self._readiness,
            )
        except VerificationUnconfirmedError as exc:
            self._save_debug(current, "verification_timeout")
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_url=str(getattr(self.driver, "current_url", "") or page_url),
                error_reason="verification_timeout",
                error_message=str(exc),
                fatal=True,
                readiness=self._readiness,
            )
        except TimeoutException as exc:
            self._save_debug(current, "page_timeout")
            self._close_driver()
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_url=str(getattr(self.driver, "current_url", "") or page_url),
                error_reason="page_timeout",
                error_message=str(exc),
                finish_reason="page_timeout",
                readiness=self._readiness,
            )
        except WebDriverException as exc:
            self._save_debug(current, "webdriver_error")
            self._close_driver()
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_url=page_url,
                error_reason="webdriver_error",
                error_message=str(exc)[:500],
                finish_reason="webdriver_error",
                readiness=self._readiness,
            )
        except UserFacingError as exc:
            self._save_debug(current, "user_facing_error")
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_url=str(getattr(self.driver, "current_url", "") or page_url),
                error_reason="crawl_blocked",
                error_message=str(exc),
                fatal=True,
                readiness=self._readiness,
            )
        except Exception as exc:
            self._save_debug(current, "unexpected_error")
            return FrontPageResult(
                worker_id=self.worker_id,
                task=leased_task,
                page_url=str(getattr(self.driver, "current_url", "") or page_url),
                error_reason="unexpected_error",
                error_message=f"{type(exc).__name__}: {exc}"[:500],
                fatal=True,
                readiness=self._readiness,
            )
        finally:
            self.manual.release_all(self.worker_id)

    def _run(self) -> None:
        try:
            while True:
                task = self.tasks.get()
                if task is None:
                    return
                try:
                    result = self._process(task)
                except Exception as exc:
                    self.manual.release_all(self.worker_id)
                    result = FrontPageResult(
                        worker_id=self.worker_id,
                        task=copy.deepcopy(task),
                        page_url=str(task.get("page_url") or ""),
                        error_reason="worker_crashed",
                        error_message=f"{type(exc).__name__}: {exc}"[:500],
                        fatal=True,
                    )
                self.results.put(result)
                if self.stop_event.is_set():
                    return
        finally:
            self.manual.release_all(self.worker_id)
            self._close_driver()


def preflight_front_delivery(
    runtime: FrontRuntimeConfig,
    initial_queue: Sequence[Dict[str, Any]],
    state: FrontStateStore,
) -> None:
    if runtime.browser_tab_concurrency <= 1 or not runtime.delivery_location_enabled:
        return
    first_url_by_domain: Dict[str, str] = {}
    for task in initial_queue:
        url = str(task.get("page_url") or "")
        domain = (urlparse(url).hostname or "").lower()
        if domain and domain not in first_url_by_domain:
            first_url_by_domain[domain] = url
    if not first_url_by_domain:
        return

    print(f"并发预检：串行确认 {len(first_url_by_domain)} 个 Amazon 域名的配送地址。")
    driver = start_driver(runtime)  # type: ignore[arg-type]
    try:
        for domain, url in first_url_by_domain.items():
            open_amazon_page(
                driver,
                url,
                runtime,
                on_manual_pause=lambda reason, page_url: state.mark_manual_pause(reason, page_url),
                on_manual_resume=state.clear_manual_pause,
            )
            print(f"配送地址已确认：{domain}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def run_bsr_category_mode(raw_config: Dict[str, Any], runtime: FrontRuntimeConfig, dry_run: bool, no_resume: bool) -> int:
    category_runtime = build_category_runtime_config(raw_config, DEFAULT_CONFIG, no_resume)
    result = run_category_crawl(category_runtime, dry_run)
    if dry_run:
        return result
    job_dir = category_runtime.outputs_root / category_runtime.job_id
    records_path = job_dir / "records.jsonl"
    failures_path = job_dir / "failures.jsonl"
    output_xlsx = job_dir / "dedup_total.xlsx"
    write_front_workbook(records_path, failures_path, output_xlsx)
    print(f"已生成 ASIN 去重总表：{output_xlsx}")
    return result


def _run_front_modes_unlocked(
    raw_config: Dict[str, Any],
    runtime: FrontRuntimeConfig,
    dry_run: bool,
) -> int:
    initial_queue = build_initial_queue(runtime)
    job_dir = ensure_dir(runtime.outputs_root / runtime.job_id)
    records_path = job_dir / "records.jsonl"
    failures_path = job_dir / "failures.jsonl"
    state_path = job_dir / "state.json"
    output_xlsx = job_dir / "dedup_total.xlsx"
    debug_dir = job_dir / "debug_snapshots"

    print(f"任务目录：{job_dir}")
    print(f"任务模式：{runtime.mode}")
    print(f"浏览器后端：{runtime.browser_backend}/{runtime.browser_mode}")
    print(f"浏览器标签页并发：{runtime.browser_tab_concurrency}")
    print(f"卖家精灵门禁：{'required' if runtime.sellersprite_required else 'not_required'}")
    print(f"商品过滤：{json.dumps(runtime.product_filters.as_dict(), ensure_ascii=False)}")
    print(f"输入数量：{len(initial_queue)}")
    effective_concurrency = min(
        runtime.browser_tab_concurrency,
        len({front_source_id(item) for item in initial_queue}),
    )
    print(f"有效来源并发：{effective_concurrency}（同一来源分页保持串行）")
    print(f"输出表格：{output_xlsx}")
    if dry_run:
        print("dry-run：配置和输入表检查完成，未打开浏览器。")
        return 0
    if not runtime.resume:
        for old_file in (records_path, failures_path, state_path, output_xlsx):
            if old_file.exists():
                old_file.unlink()
        page_results_dir = job_dir / "page_results"
        if page_results_dir.exists():
            shutil.rmtree(page_results_dir)

    state = FrontStateStore(state_path, runtime, initial_queue)
    state.load_or_create()
    materialize_front_records(state, records_path)
    preflight_front_delivery(
        runtime,
        [item for item in state.data.get("pending") or [] if isinstance(item, dict)],
        state,
    )

    result_queue: "queue.Queue[FrontPageResult]" = queue.Queue()
    manual = ManualActionCoordinator()
    throttle = NavigationThrottle(runtime.delay_seconds_min, runtime.delay_seconds_max)
    delivery_locks = DeliveryDomainLocks()
    workers = [
        FrontWorker(
            worker_id=f"tab-{index}",
            runtime=runtime,
            results=result_queue,
            manual=manual,
            throttle=throttle,
            delivery_locks=delivery_locks,
            debug_dir=debug_dir,
        )
        for index in range(1, runtime.browser_tab_concurrency + 1)
    ]
    worker_by_id = {worker.worker_id: worker for worker in workers}
    for worker in workers:
        worker.start()

    batch_pause = BatchPauseScheduler(runtime)  # type: ignore[arg-type]
    idle_workers = set(worker_by_id)
    busy_workers: set[str] = set()
    fatal_message = ""
    try:
        while True:
            state.set_manual_snapshot(manual.snapshot())
            if not fatal_message and not manual.is_paused():
                for worker_id in sorted(list(idle_workers)):
                    if not worker_by_id[worker_id].is_alive():
                        idle_workers.discard(worker_id)
                        fatal_message = (
                            fatal_message
                            or f"{worker_id} 已意外退出；未派发任务仍保留在断点中。"
                        )
                        continue
                    task = state.lease_next(worker_id)
                    if task is None:
                        break
                    label = (
                        task.get("keyword")
                        or task.get("store_name")
                        or task.get("store_url")
                        or task.get("page_url")
                    )
                    print(f"[{worker_id}] 处理：{label} / 第 {task.get('page_number') or 1} 页")
                    worker_by_id[worker_id].submit(task)
                    idle_workers.remove(worker_id)
                    busy_workers.add(worker_id)

            if fatal_message and not busy_workers:
                break
            if not state.has_work() and not busy_workers:
                print("队列已完成。")
                break
            try:
                result = result_queue.get(timeout=0.25)
            except queue.Empty:
                for worker_id in sorted(list(busy_workers)):
                    if worker_by_id[worker_id].is_alive():
                        continue
                    task = (state.data.get("in_flight") or {}).get(worker_id)
                    if isinstance(task, dict):
                        state.requeue_task(worker_id, task)
                    busy_workers.discard(worker_id)
                    idle_workers.discard(worker_id)
                    fatal_message = (
                        fatal_message
                        or f"{worker_id} 已意外退出；当前任务已保留在断点中。"
                    )
                continue

            worker_id = result.worker_id
            busy_workers.discard(worker_id)
            idle_workers.add(worker_id)
            if result.readiness:
                state.mark_sellersprite_readiness(result.readiness, worker_id)

            if result.error_reason:
                log_front_failure(
                    failures_path,
                    state,
                    result.task,
                    result.error_reason,
                    result.error_message,
                    result.page_url or str(result.task.get("page_url") or ""),
                )
                if result.error_reason in FRONT_RETRYABLE_ERRORS:
                    retry_task, should_retry, retry_number = prepare_front_retry_task(
                        result.task
                    )
                    state.requeue_task(worker_id, retry_task)
                    if should_retry:
                        print(
                            f"[{worker_id}] 页面暂时失败（{result.error_reason}），"
                            f"保留断点并重试当前任务（{retry_number}/"
                            f"{FRONT_MAX_TASK_RETRIES}）。"
                        )
                        continue
                    fatal_message = fatal_message or (
                        f"页面连续 {FRONT_MAX_TASK_RETRIES + 1} 次失败"
                        f"（{result.error_reason}），当前任务已保留在断点中："
                        f"{result.error_message}"
                    )
                    print(f"[{worker_id}] 页面重试仍失败，任务保留在断点中。")
                    continue
                if result.fatal:
                    state.requeue_task(worker_id, result.task)
                    fatal_message = fatal_message or result.error_message
                else:
                    state.finish_failed_task(
                        worker_id,
                        result.task,
                        result.finish_reason or result.error_reason,
                    )
                print(f"[{worker_id}] 页面失败：{result.error_reason}")
                continue

            committed = state.commit_page_result(result)
            if committed:
                materialize_front_records(state, records_path)
                print(
                    f"[{worker_id}] 扫描 {len(result.raw_records)} 条，保留 "
                    f"{len(result.accepted_records)} 条，过滤 "
                    f"{len(result.raw_records) - len(result.accepted_records)} 条，"
                    f"插件状态：{result.plugin_status}"
                )
                batch_pause.after_completed_page()
            else:
                print(f"[{worker_id}] 页面已提交，跳过重复写入。")
    finally:
        for worker in workers:
            worker.stop()
        stuck_workers: List[str] = []
        for worker in workers:
            if not worker.join(timeout=5.0):
                stuck_workers.append(worker.worker_id)
        if stuck_workers:
            state.set_manual_snapshot(manual.snapshot())
            print(
                "以下标签页仍在等待不可中断的浏览器调用，主流程不再无限等待："
                + ", ".join(stuck_workers)
            )
        else:
            state.set_manual_snapshot(None)

    materialize_front_records(state, records_path)
    write_front_workbook(records_path, failures_path, output_xlsx)
    print(f"已生成 ASIN 去重总表：{output_xlsx}")
    if fatal_message:
        raise UserFacingError(fatal_message)
    return 0


def run_front_modes(
    raw_config: Dict[str, Any],
    runtime: FrontRuntimeConfig,
    dry_run: bool,
) -> int:
    if dry_run:
        return _run_front_modes_unlocked(raw_config, runtime, dry_run=True)
    job_dir = runtime.outputs_root / runtime.job_id
    with JobRunLock(job_dir / ".run.lock"):
        return _run_front_modes_unlocked(raw_config, runtime, dry_run=False)


def main() -> int:
    parser = argparse.ArgumentParser(description="Unified Amazon front crawler with SellerSprite extension data.")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="配置文件路径")
    parser.add_argument("--dry-run", action="store_true", help="只检查配置，不打开浏览器")
    parser.add_argument("--no-resume", action="store_true", help="忽略已有断点，重新开始任务")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser()
    if not config_path.is_absolute():
        config_path = ROOT_DIR / config_path
    raw_config = load_json(config_path)
    runtime = build_front_runtime_config(raw_config, args.no_resume)
    if runtime.mode == "bsr_category":
        return run_bsr_category_mode(raw_config, runtime, args.dry_run, args.no_resume)
    return run_front_modes(raw_config, runtime, args.dry_run)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except UserFacingError as exc:
        print(f"运行失败：{exc}", file=sys.stderr)
        raise SystemExit(2)
