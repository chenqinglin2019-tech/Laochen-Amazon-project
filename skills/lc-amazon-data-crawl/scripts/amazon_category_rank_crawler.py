#!/usr/bin/env python3
"""
Amazon category ranking crawler that reads SellerSprite extension data from a
visible Chrome session.

The crawler intentionally does not solve CAPTCHA or bypass verification. When
Amazon or SellerSprite verification appears, it saves a checkpoint and waits for
manual handling in the visible browser.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import copy
import datetime as dt
import hashlib
import json
import os
import queue
import random
import re
import select
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, FrozenSet, Iterable, List, Mapping, Optional, Sequence, Tuple
from urllib.parse import parse_qs, urljoin, urlparse, urlunparse
from urllib.request import urlopen

from selenium import webdriver
from selenium.common.exceptions import (
    JavascriptException,
    NoSuchElementException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from browser_runtime import CdpWebDriver
from amazon_page_recovery import (
    AmazonPageRetryController,
    AmazonPageRetryExhausted,
    DEFAULT_RETRY_SCHEDULE_SECONDS,
    DomainCooldownRegistry,
    PageHealthAssessment,
    PageHealthStatus,
    PageSnapshot,
    RetryCallbacks,
    RetryConfigurationError,
    RetrySchedule,
    TransientAmazonPageUnavailable,
    classify_page_snapshot,
    retry_schedule_from_config,
)

try:  # POSIX process lock
    import fcntl as _fcntl
except ImportError:  # pragma: no cover - Windows
    _fcntl = None

try:  # Windows process lock
    import msvcrt as _msvcrt
except ImportError:  # pragma: no cover - POSIX
    _msvcrt = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CONFIG = ROOT_DIR / "config" / "category_rank_crawler.json"
DEFAULT_DELIVERY_LOCATIONS_FILE = "config/amazon_delivery_locations.json"
SELLERSPRITE_EXTENSION_ID = "lnbmbgocenenhhhdojdielgnmeflbnfb"
ASIN_RE = re.compile(r"\b([A-Z0-9]{10})\b")
URL_ASIN_RE = re.compile(r"/(?:dp|gp/product)/([A-Z0-9]{10})(?:[/?#]|$)", re.I)
FULFILLMENT_METHODS = {"FBA", "FBM", "AMZ"}
RECORD_SCHEMA_VERSION = 2
CATEGORY_STATE_SCHEMA_VERSION = 2
CRAWL_PLAN_SCHEMA_VERSION = 1
SUBCATEGORY_BSR_SEMANTICS = {
    "single_row_is_child": True,
    "multiple_rows_skip_first": True,
    "preserve_all_children": True,
}
FULFILLMENT_SEMANTICS = {
    "version": 3,
    "known_methods": sorted(FULFILLMENT_METHODS),
    "explicit_known_prefix_accepts_any_suffix": True,
    "unknown_nonempty_is_missing": False,
}

AMAZON_SIGN_IN_STOP_MESSAGE = (
    "检测到 Amazon 登录页。本工具只采集公开页面，不使用也不需要 Amazon 买家账号。"
    "本条采集已停止且未写入当前页数据；请稍后重试，仍出现时可更换公开页面入口。"
    "不要登录买家账号。"
)


OUTPUT_HEADERS = [
    "根类目URL",
    "类目路径",
    "类目名称",
    "类目节点ID",
    "类目URL",
    "页码",
    "排名",
    "ASIN",
    "商品标题",
    "商品URL",
    "评论数量",
    "评分值",
    "卖家名称",
    "品牌名称",
    "卖家所处国家",
    "子类目节点排名",
    "近30天销量（子体）",
    "近30天销量（父体）",
    "FBA费用",
    "毛利率",
    "配送方式",
    "配送时长",
    "上架时间",
    "自然搜索词数量",
    "广告搜索词数量",
    "抓取时间",
    "加载状态",
    "备注",
]

FIELD_TO_HEADER = {
    "root_url": "根类目URL",
    "category_path": "类目路径",
    "category_name": "类目名称",
    "category_node_id": "类目节点ID",
    "category_url": "类目URL",
    "page_number": "页码",
    "rank": "排名",
    "asin": "ASIN",
    "title": "商品标题",
    "product_url": "商品URL",
    "review_count": "评论数量",
    "rating_value": "评分值",
    "seller_name": "卖家名称",
    "brand_name": "品牌名称",
    "seller_country": "卖家所处国家",
    "subcategory_bsr_ranks": "子类目节点排名",
    "sales_30_days_child": "近30天销量（子体）",
    "sales_30_days_parent": "近30天销量（父体）",
    "fba_fee": "FBA费用",
    "gross_margin": "毛利率",
    "fulfillment_method": "配送方式",
    "delivery_duration": "配送时长",
    "launch_date": "上架时间",
    "organic_keywords_count": "自然搜索词数量",
    "ad_keywords_count": "广告搜索词数量",
    "scraped_at": "抓取时间",
    "load_status": "加载状态",
    "note": "备注",
}

REQUESTED_DATA_FIELDS = [
    "review_count",
    "rating_value",
    "seller_name",
    "brand_name",
    "seller_country",
    "sales_30_days_child",
    "sales_30_days_parent",
    "fba_fee",
    "gross_margin",
    "fulfillment_method",
    "delivery_duration",
    "launch_date",
    "organic_keywords_count",
    "ad_keywords_count",
]

SELLERSPRITE_EVIDENCE_FIELDS = [
    "sales_30_days_child",
    "sales_30_days_parent",
    "fba_fee",
    "gross_margin",
    "fulfillment_method",
    "delivery_duration",
    "launch_date",
    "organic_keywords_count",
    "ad_keywords_count",
]

HEADER_ALIASES = {
    "review_count": ["review", "ratings", "rating count", "评论", "评价", "ratings units sold", "number of ratings"],
    "rating_value": ["latest rating", "star", "评分", "星级", "rating"],
    "seller_name": ["seller", "卖家", "店铺"],
    "brand_name": ["brand", "品牌"],
    "seller_country": ["country", "国家", "所在地", "location"],
    "sales_30_days_child": ["units sold", "30 days", "30-day", "近30", "月销量", "销量", "子体"],
    "sales_30_days_parent": ["units sold", "30 days", "30-day", "近30", "月销量", "销量", "父体"],
    "fba_fee": ["fba fee", "fba fees", "fba费用", "配送费"],
    "gross_margin": ["gross margin", "margin", "毛利率", "利润率"],
    "fulfillment_method": ["fulfillment", "配送方式", "fba/fbm", "fulfilment"],
    "delivery_duration": ["delivery", "配送时长", "estimated delivery", "到达"],
    "launch_date": ["launch", "上架", "available since", "date first available"],
    "organic_keywords_count": ["organic", "自然搜索词", "自然词"],
    "ad_keywords_count": ["ad keyword", "sponsored", "广告搜索词", "广告词", "ppc"],
}

COUNTRY_BY_FLAG_CODE = {
    "ae": "阿联酋",
    "au": "澳大利亚",
    "be": "比利时",
    "br": "巴西",
    "ca": "加拿大",
    "ch": "瑞士",
    "cn": "中国",
    "de": "德国",
    "es": "西班牙",
    "fr": "法国",
    "gb": "英国",
    "hk": "中国香港",
    "ie": "爱尔兰",
    "in": "印度",
    "it": "意大利",
    "jp": "日本",
    "kr": "韩国",
    "mx": "墨西哥",
    "my": "马来西亚",
    "nl": "荷兰",
    "pl": "波兰",
    "sa": "沙特阿拉伯",
    "se": "瑞典",
    "sg": "新加坡",
    "th": "泰国",
    "tr": "土耳其",
    "tw": "中国台湾",
    "uk": "英国",
    "us": "美国",
    "vn": "越南",
}


class UserFacingError(RuntimeError):
    pass


class DeliveryLocationUnconfirmedError(UserFacingError):
    """Stop the whole crawl when the requested marketplace delivery location is not confirmed."""

    pass


class VerificationUnconfirmedError(UserFacingError):
    """Stop the whole crawl when Amazon or SellerSprite verification times out."""

    pass


class ConcurrentWorkerCancelled(UserFacingError):
    """Internal cooperative cancellation used after another worker fails fatally."""

    pass


@dataclass(frozen=True)
class ProductFilterConfig:
    allowed_fulfillment_methods: Tuple[str, ...] = ()
    allow_missing_fulfillment: bool = False
    require_subcategory_rank: bool = False
    excluded_fulfillment_methods: Tuple[str, ...] = ()

    @property
    def enabled(self) -> bool:
        return bool(
            self.allowed_fulfillment_methods
            or self.excluded_fulfillment_methods
            or self.allow_missing_fulfillment
            or self.require_subcategory_rank
        )

    def as_dict(self) -> Dict[str, Any]:
        return {
            "allowed_fulfillment_methods": list(self.allowed_fulfillment_methods),
            "excluded_fulfillment_methods": list(self.excluded_fulfillment_methods),
            "allow_missing_fulfillment": self.allow_missing_fulfillment,
            "require_subcategory_rank": self.require_subcategory_rank,
        }


def _strict_json_bool(value: Any, field_name: str, default: bool = False) -> bool:
    if value is None:
        return default
    if not isinstance(value, bool):
        raise UserFacingError(f"配置项 `{field_name}` 必须是 JSON 布尔值 true 或 false。")
    return value


def _fulfillment_method_config_list(value: Any, field_name: str) -> Tuple[str, ...]:
    if not isinstance(value, list):
        raise UserFacingError(f"配置项 `product_filters.{field_name}` 必须是数组。")
    methods: List[str] = []
    for raw_method in value:
        if not isinstance(raw_method, str):
            raise UserFacingError(
                f"配置项 `product_filters.{field_name}` 只能包含字符串。"
            )
        method = raw_method.strip().upper()
        if method not in FULFILLMENT_METHODS:
            raise UserFacingError(
                f"配置项 `product_filters.{field_name}` 只支持 FBA、FBM 或 AMZ。"
            )
        if method not in methods:
            methods.append(method)
    return tuple(sorted(methods))


def build_product_filter_config(config: Dict[str, Any]) -> ProductFilterConfig:
    raw = config.get("product_filters")
    if raw is None:
        return ProductFilterConfig()
    if not isinstance(raw, dict):
        raise UserFacingError("配置项 `product_filters` 必须是对象。")

    supported_keys = {
        "allowed_fulfillment_methods",
        "excluded_fulfillment_methods",
        "allow_missing_fulfillment",
        "require_subcategory_rank",
    }
    unknown_keys = sorted(str(key) for key in raw if key not in supported_keys)
    if unknown_keys:
        raise UserFacingError(
            "配置项 `product_filters` 包含不支持的字段：" + ", ".join(unknown_keys)
        )

    allowed_methods = _fulfillment_method_config_list(
        raw.get("allowed_fulfillment_methods", []),
        "allowed_fulfillment_methods",
    )
    excluded_methods = _fulfillment_method_config_list(
        raw.get("excluded_fulfillment_methods", []),
        "excluded_fulfillment_methods",
    )
    if allowed_methods and excluded_methods:
        raise UserFacingError(
            "配置项 `product_filters.allowed_fulfillment_methods` 与 "
            "`product_filters.excluded_fulfillment_methods` 不能同时设置。"
        )

    return ProductFilterConfig(
        allowed_fulfillment_methods=allowed_methods,
        allow_missing_fulfillment=_strict_json_bool(
            raw.get("allow_missing_fulfillment"),
            "product_filters.allow_missing_fulfillment",
            False,
        ),
        require_subcategory_rank=_strict_json_bool(
            raw.get("require_subcategory_rank"),
            "product_filters.require_subcategory_rank",
            False,
        ),
        excluded_fulfillment_methods=excluded_methods,
    )


def record_contract_fingerprint(product_filters: ProductFilterConfig) -> str:
    payload = {
        "schema_version": RECORD_SCHEMA_VERSION,
        "product_filters": product_filters.as_dict(),
        "subcategory_bsr_semantics": SUBCATEGORY_BSR_SEMANTICS,
        "fulfillment_semantics": FULFILLMENT_SEMANTICS,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def crawl_plan_fingerprint(
    start_url: str,
    include_root: bool,
    max_depth: Optional[int],
    max_pages_per_category: Optional[int],
    field_selectors: Dict[str, List[str]],
) -> str:
    normalized_selectors = {
        str(key): [normalize_space(str(value)) for value in values]
        for key, values in sorted(field_selectors.items())
    }
    payload = {
        "schema_version": CRAWL_PLAN_SCHEMA_VERSION,
        "start_url": normalize_space(start_url),
        "include_root": bool(include_root),
        "max_depth": max_depth,
        "max_pages_per_category": max_pages_per_category,
        "field_selectors": normalized_selectors,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def now_ts() -> str:
    return dt.datetime.now().strftime("%Y%m%d-%H%M%S")


def now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


class JobRunLock:
    """Prevent two crawler processes from mutating one job checkpoint."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self._handle: Optional[Any] = None

    def acquire(self) -> None:
        if self._handle is not None:
            return
        ensure_dir(self.path.parent)
        handle = self.path.open("a+", encoding="utf-8")
        try:
            if _fcntl is not None:
                _fcntl.flock(handle.fileno(), _fcntl.LOCK_EX | _fcntl.LOCK_NB)
            elif _msvcrt is not None:  # pragma: no cover - Windows
                handle.seek(0, os.SEEK_END)
                if handle.tell() == 0:
                    handle.write(" ")
                    handle.flush()
                handle.seek(0)
                _msvcrt.locking(handle.fileno(), _msvcrt.LK_NBLCK, 1)
            else:  # pragma: no cover - unsupported platform
                raise UserFacingError("当前系统不支持任务级文件锁。")
        except (BlockingIOError, OSError) as exc:
            handle.close()
            raise UserFacingError(
                f"同一 job_id 已有抓取进程运行：{self.path.parent.name}。"
                "请等待其结束或使用新的 job_id。"
            ) from exc

        handle.seek(0)
        handle.truncate()
        handle.write(
            json.dumps(
                {"pid": os.getpid(), "acquired_at": now_iso()},
                ensure_ascii=False,
            )
        )
        handle.flush()
        os.fsync(handle.fileno())
        self._handle = handle

    def release(self) -> None:
        handle = self._handle
        self._handle = None
        if handle is None:
            return
        try:
            handle.seek(0)
            if _fcntl is not None:
                _fcntl.flock(handle.fileno(), _fcntl.LOCK_UN)
            elif _msvcrt is not None:  # pragma: no cover - Windows
                _msvcrt.locking(handle.fileno(), _msvcrt.LK_UNLCK, 1)
        finally:
            handle.close()

    def __enter__(self) -> "JobRunLock":
        self.acquire()
        return self

    def __exit__(self, _exc_type: Any, _exc: Any, _traceback: Any) -> None:
        self.release()


def load_json(path: Path) -> Dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise UserFacingError(f"没有找到配置文件：{path}") from exc
    except json.JSONDecodeError as exc:
        raise UserFacingError(f"配置文件不是有效 JSON：{path}") from exc


def dump_json(path: Path, data: Dict[str, Any]) -> None:
    ensure_dir(path.parent)
    file_descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=str(path.parent),
    )
    temp_path = Path(temp_name)
    try:
        with os.fdopen(file_descriptor, "w", encoding="utf-8") as handle:
            json.dump(data, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    except BaseException:
        try:
            temp_path.unlink()
        except FileNotFoundError:
            pass
        raise


def append_jsonl(path: Path, record: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    records: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return records


def resolve_path(value: str, base: Path = ROOT_DIR) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = base / path
    return path


def config_text(config: Dict[str, Any], key: str, default: str = "") -> str:
    value = config.get(key, default)
    return "" if value is None else str(value).strip()


def config_int(config: Dict[str, Any], key: str, default: Optional[int] = None) -> Optional[int]:
    value = config.get(key, default)
    if value in ("", None):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise UserFacingError(f"配置项 `{key}` 必须是数字或空值。") from exc


def config_float(config: Dict[str, Any], key: str, default: float) -> float:
    value = config.get(key, default)
    if value in ("", None):
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise UserFacingError(f"配置项 `{key}` 必须是数字。") from exc


def config_bool(config: Dict[str, Any], key: str, default: bool = False) -> bool:
    value = config.get(key, default)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on"}
    return bool(value)


def slugify(value: str) -> str:
    value = value.strip() or "category-rank"
    value = re.sub(r"[\\/:*?\"<>|]+", " ", value)
    value = re.sub(r"\s+", "-", value)
    return value.strip("-_")[:90] or "category-rank"


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\u00a0", " ")).strip()


def normalize_header(value: str) -> str:
    value = normalize_space(value).lower()
    value = value.replace("（", "(").replace("）", ")")
    value = re.sub(r"[\s_:/|]+", " ", value)
    return value


_BSR_RANK_START_RE = re.compile(r"#\s*([\d][\d,]*)\s+in\s+", re.I)
_BSR_CATEGORY_TRAILING_FIELDS_RE = re.compile(
    r"\s+(?=(?:"
    r"近30天销量(?:\([^)]*\))?|销售额|FBA费用|毛利率|变体数|"
    r"评分(?:\(评分数\))?|价格|全部流量词|搜索推荐词|自然搜索词|"
    r"广告(?:搜索|流量)词|上架时间|配送(?:时长)?|卖家|品牌|ASIN|"
    r"Color|Size"
    r")\s*[:：]|加入产品库(?:\s|$))",
    re.I,
)


def normalize_subcategory_bsr_ranks(value: Any) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return []
    normalized: List[Dict[str, Any]] = []
    seen: set[Tuple[int, str]] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        raw_rank = item.get("rank")
        if isinstance(raw_rank, bool):
            continue
        try:
            rank = int(str(raw_rank).replace(",", "").strip())
        except (TypeError, ValueError):
            continue
        category_name = normalize_space(str(item.get("category_name") or ""))
        if rank <= 0 or not category_name:
            continue
        key = (rank, category_name.casefold())
        if key in seen:
            continue
        seen.add(key)
        normalized.append({"rank": rank, "category_name": category_name})
    return normalized


def parse_subcategory_bsr_ranks(text: str) -> List[Dict[str, Any]]:
    normalized_text = normalize_space(text)
    matches = list(_BSR_RANK_START_RE.finditer(normalized_text))
    if not matches:
        return []

    parsed: List[Dict[str, Any]] = []
    start_index = 0 if len(matches) == 1 else 1
    for index, match in enumerate(matches[start_index:], start=start_index):
        category_end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized_text)
        category_name = normalized_text[match.end() : category_end]
        trailing = _BSR_CATEGORY_TRAILING_FIELDS_RE.search(category_name)
        if trailing:
            category_name = category_name[: trailing.start()]
        category_name = normalize_space(category_name).strip(" -|;；,，。")
        try:
            rank = int(match.group(1).replace(",", ""))
        except (TypeError, ValueError):
            continue
        if rank > 0 and category_name:
            parsed.append({"rank": rank, "category_name": category_name})
    return normalize_subcategory_bsr_ranks(parsed)


def format_subcategory_bsr_ranks(value: Any) -> str:
    return " ; ".join(
        f"#{item['rank']:,} in {item['category_name']}"
        for item in normalize_subcategory_bsr_ranks(value)
    )


def normalize_fulfillment_method(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    method = normalize_space(value).upper()
    return method if method in FULFILLMENT_METHODS else ""


_FULFILLMENT_LABEL_PATTERN = (
    r"(?:配送(?!\s*(?:时长|费))\s*(?:方式)?|"
    r"fulfillment(?:\s+method)?|fulfilment(?:\s+method)?)"
)
_FULFILLMENT_LABEL_RE = re.compile(
    _FULFILLMENT_LABEL_PATTERN
    + r"\s*(?:[:：]\s*|\s+)([^:：|;,，；]{1,120})",
    re.I,
)
_FULFILLMENT_KNOWN_PREFIX_RE = re.compile(
    r"^(FBA|FBM|AMZ)",
    re.I,
)


def _known_fulfillment_prefix(value: str) -> str:
    match = _FULFILLMENT_KNOWN_PREFIX_RE.match(normalize_space(value))
    return match.group(1).upper() if match else ""


def parse_fulfillment_evidence(text: Any, *, explicit_value: bool = False) -> Tuple[str, str]:
    """Return (canonical_method, raw_evidence) without hiding unknown values.

    Free-form product text only counts as evidence when it has an explicit
    fulfillment label. Values originating from a mapped table column or an
    explicit selector use ``explicit_value=True`` and therefore remain evidence
    even when SellerSprite introduces a value that is not FBA/FBM/AMZ.
    """

    if not isinstance(text, str):
        return "", ""
    normalized_text = normalize_space(text)
    if not normalized_text:
        return "", ""
    match = _FULFILLMENT_LABEL_RE.search(normalized_text)
    if match:
        raw = normalize_space(match.group(1))[:120]
        method = normalize_fulfillment_method(raw) or _known_fulfillment_prefix(raw)
        return method, raw
    if explicit_value:
        raw = normalized_text[:120]
        return _known_fulfillment_prefix(raw), raw
    return "", ""


def select_fulfillment_evidence(
    *evidence: Tuple[Any, Any],
) -> Tuple[str, str]:
    """Select evidence in source-priority order without hiding unknown values.

    A canonical value from any source is stronger than unknown raw evidence.
    Among equally strong values the first source wins, so callers can pass
    explicit selector, structured table and labelled card text in that order.
    """

    prepared: List[Tuple[str, str]] = []
    for method_value, raw_value in evidence:
        method = normalize_fulfillment_method(method_value)
        raw = normalize_space(str(raw_value or ""))[:120]
        prepared.append((method, raw))
    for method, raw in prepared:
        if method:
            return method, raw or method
    for _method, raw in prepared:
        if raw:
            return "", raw
    return "", ""


def parse_fulfillment_method(text: str) -> str:
    method, _raw = parse_fulfillment_evidence(text)
    return method


def filter_product_records(
    records: Sequence[Dict[str, Any]],
    product_filters: ProductFilterConfig,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    if not product_filters.enabled:
        return list(records), {}

    accepted: List[Dict[str, Any]] = []
    rejection_counts: Dict[str, int] = {}
    allowed = set(product_filters.allowed_fulfillment_methods)
    excluded = set(product_filters.excluded_fulfillment_methods)
    fulfillment_filter_enabled = bool(
        allowed or excluded or product_filters.allow_missing_fulfillment
    )
    for record in records:
        rejected_reasons: List[str] = []
        if fulfillment_filter_enabled:
            fulfillment_value = record.get("fulfillment_method")
            raw_evidence = record.get("fulfillment_method_raw")
            method = normalize_fulfillment_method(fulfillment_value)
            if not method and raw_evidence:
                method, _parsed_raw = parse_fulfillment_evidence(
                    str(raw_evidence),
                    explicit_value=True,
                )
            evidence_text = normalize_space(str(raw_evidence or "")) or normalize_space(
                str(fulfillment_value or "")
            )
            truly_missing = not evidence_text
            if excluded:
                fulfillment_allowed = method not in excluded
            else:
                fulfillment_allowed = method in allowed or (
                    truly_missing and product_filters.allow_missing_fulfillment
                )
            if not fulfillment_allowed:
                rejected_reasons.append("fulfillment_method_not_allowed")
        if product_filters.require_subcategory_rank and not normalize_subcategory_bsr_ranks(
            record.get("subcategory_bsr_ranks")
        ):
            rejected_reasons.append("subcategory_bsr_rank_missing")

        if rejected_reasons:
            for reason in rejected_reasons:
                rejection_counts[reason] = rejection_counts.get(reason, 0) + 1
            continue
        accepted.append(record)
    return accepted, rejection_counts


def country_from_flag_code_or_text(value: str) -> str:
    text = normalize_space(value).strip().lower()
    if not text:
        return ""
    code_match = re.search(r"(?:flag-icon|icp-nav-flag)-([a-z]{2})\b", text)
    if code_match:
        text = code_match.group(1)
    text = text.removeprefix("flag-icon-").removeprefix("icp-nav-flag-")
    return COUNTRY_BY_FLAG_CODE.get(text, text.upper() if re.fullmatch(r"[a-z]{2}", text) else "")


def clean_url(url: str) -> str:
    parsed = urlparse(url)
    clean_query = []
    for key, values in parse_qs(parsed.query, keep_blank_values=True).items():
        if key.lower() in {"ref", "qid", "sprefix", "crid", "rnid"}:
            continue
        for value in values:
            clean_query.append((key, value))
    query = "&".join(f"{key}={value}" for key, value in clean_query)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", query, ""))


def validate_amazon_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise UserFacingError("start_url 必须是 http 或 https 地址。")
    host = parsed.netloc.lower()
    if "amazon." not in host:
        raise UserFacingError("start_url 必须是 Amazon 类目页面地址。")


def load_delivery_locations(path: Path) -> Dict[str, Dict[str, str]]:
    raw = load_json(path)
    raw_locations = raw.get("locations", raw)
    if not isinstance(raw_locations, dict) or not raw_locations:
        raise UserFacingError(f"配送地址配置必须包含非空对象 `locations`：{path}")

    locations: Dict[str, Dict[str, str]] = {}
    for raw_domain, raw_location in raw_locations.items():
        domain = str(raw_domain or "").strip().lower().removeprefix("www.").rstrip(".")
        if not re.fullmatch(r"amazon(?:\.[a-z0-9-]+)+", domain):
            raise UserFacingError(f"配送地址配置包含无效 Amazon 域名：{raw_domain}")
        if not isinstance(raw_location, dict):
            raise UserFacingError(f"配送地址 `{domain}` 必须是对象。")
        city = normalize_space(str(raw_location.get("city") or ""))
        postal_code = str(raw_location.get("postal_code") or "").strip()
        strategy = str(raw_location.get("strategy") or "postal").strip().lower()
        if not city or not postal_code:
            raise UserFacingError(f"配送地址 `{domain}` 的 city 和 postal_code 不能为空。")
        if strategy not in {"postal", "postal_then_city"}:
            raise UserFacingError(
                f"配送地址 `{domain}` 的 strategy 只支持 postal 或 postal_then_city。"
            )
        locations[domain] = {
            "city": city,
            "postal_code": postal_code,
            "strategy": strategy,
        }
    return dict(sorted(locations.items()))


def delivery_location_fingerprint(enabled: bool, locations: Dict[str, Dict[str, str]]) -> str:
    payload = {
        "enabled": bool(enabled),
        "locations": locations if enabled else {},
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def build_delivery_location_config(config: Dict[str, Any]) -> Dict[str, Any]:
    enabled = config_bool(config, "delivery_location_enabled", True)
    configured_file = (
        config_text(config, "delivery_locations_file", DEFAULT_DELIVERY_LOCATIONS_FILE)
        or DEFAULT_DELIVERY_LOCATIONS_FILE
    )
    locations_file = resolve_path(configured_file)
    bundled_file = ROOT_DIR / "assets" / "config" / "amazon_delivery_locations.json"
    if enabled and configured_file == DEFAULT_DELIVERY_LOCATIONS_FILE and not locations_file.exists():
        locations_file = bundled_file
    timeout = config_int(config, "delivery_location_timeout", 20) or 20
    if timeout <= 0:
        raise UserFacingError("配置项 `delivery_location_timeout` 必须大于 0。")
    locations = load_delivery_locations(locations_file) if enabled else {}
    return {
        "delivery_location_enabled": enabled,
        "delivery_locations_file": locations_file,
        "delivery_location_timeout": timeout,
        "delivery_locations": locations,
        "delivery_location_fingerprint": delivery_location_fingerprint(enabled, locations),
    }


def ensure_resume_delivery_fingerprint(
    state: Dict[str, Any],
    runtime: Any,
    records_path: Optional[Path] = None,
) -> bool:
    expected = str(getattr(runtime, "delivery_location_fingerprint", "") or "")
    previous = str(state.get("delivery_location_fingerprint") or "")
    has_records = bool(
        int(state.get("records_count") or 0) > 0
        or state.get("completed_pages")
        or (records_path is not None and records_path.exists() and records_path.stat().st_size > 0)
    )
    if has_records and previous != expected:
        raise UserFacingError(
            "配送地址配置与已有断点不一致，拒绝混合续跑。请保留旧输出并改用新的 `job_id`。"
        )
    changed = previous != expected
    state["delivery_location_fingerprint"] = expected
    return changed


def ensure_resume_record_contract_fingerprint(
    state: Dict[str, Any],
    runtime: Any,
    records_path: Optional[Path] = None,
) -> bool:
    expected = str(getattr(runtime, "record_contract_fingerprint", "") or "")
    previous = str(state.get("record_contract_fingerprint") or "")
    has_records = bool(
        int(state.get("records_count") or 0) > 0
        or state.get("completed_pages")
        or (records_path is not None and records_path.exists() and records_path.stat().st_size > 0)
    )
    if has_records and previous != expected:
        raise UserFacingError(
            "产品过滤条件、子类目排名或配送方式输出语义与已有断点不一致，拒绝混合续跑。"
            "请保留旧输出并改用新的 `job_id`。"
        )
    changed = previous != expected
    state["record_contract_fingerprint"] = expected
    return changed


def ensure_resume_crawl_plan_fingerprint(
    state: Dict[str, Any],
    runtime: Any,
    records_path: Optional[Path] = None,
    page_results_dir: Optional[Path] = None,
) -> bool:
    expected = str(getattr(runtime, "crawl_plan_fingerprint", "") or "")
    previous = str(state.get("crawl_plan_fingerprint") or "")
    has_page_results = bool(
        page_results_dir is not None
        and page_results_dir.exists()
        and any(page_results_dir.glob("*.json"))
    )
    has_progress = bool(
        int(state.get("processed_categories_count") or 0) > 0
        or state.get("completed_pages")
        or state.get("done_categories")
        or state.get("current")
        or state.get("in_flight_categories")
        or has_page_results
        or (records_path is not None and records_path.exists() and records_path.stat().st_size > 0)
    )
    if has_progress and previous != expected:
        raise UserFacingError(
            "起始类目、递归深度、分页上限或字段选择器与已有断点不一致，拒绝混合续跑。"
            "请保留旧输出并改用新的 `job_id`。"
        )
    changed = previous != expected
    state["crawl_plan_fingerprint"] = expected
    return changed


def extract_node_id(url: str) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    for key in ("node", "rh"):
        values = query.get(key)
        if values:
            match = re.search(r"(\d{4,})", values[-1])
            if match:
                return match.group(1)
    matches = re.findall(r"/(\d{4,})(?:[/?#]|$)", parsed.path)
    if matches:
        return matches[-1]
    matches = re.findall(r"_(\d{4,})(?:[/?#_]|$)", url)
    return matches[-1] if matches else ""


def category_key(node: Dict[str, Any]) -> str:
    node_id = str(node.get("node_id") or "").strip()
    if node_id:
        return f"node:{node_id}"
    return f"url:{clean_url(str(node.get('url') or ''))}"


def page_key(node: Dict[str, Any], page_number: int, url: str) -> str:
    return f"{category_key(node)}|page:{page_number}|{clean_url(url)}"


def infer_category_name_from_url(url: str) -> str:
    path = urlparse(url).path.rstrip("/")
    parts = [part for part in path.split("/") if part]
    for part in reversed(parts):
        if not re.fullmatch(r"\d{4,}", part):
            return normalize_space(part.replace("-", " ").title())
    return "Amazon Category"


@dataclass
class RuntimeConfig:
    start_url: str
    job_id: str
    outputs_root: Path
    browser_backend: str
    browser_mode: str
    chrome_binary: str
    chrome_user_data_dir: Path
    chrome_profile_directory: str
    debugger_address: str
    browser_tab_concurrency: int
    extension_path: Path
    include_root: bool
    max_depth: Optional[int]
    max_pages_per_category: Optional[int]
    max_categories: Optional[int]
    resume: bool
    activate_plugin: bool
    page_timeout: int
    amazon_page_retry_schedule: RetrySchedule
    plugin_timeout: int
    plugin_retry_attempts: int
    plugin_retry_wait_seconds: float
    plugin_retry_wait_seconds_max: float
    plugin_relaunch_retry_attempts: int
    plugin_relaunch_wait_seconds: float
    plugin_second_relaunch_retry_attempts: int
    plugin_second_relaunch_wait_seconds: float
    manual_pause_timeout: int
    delivery_location_enabled: bool
    delivery_locations_file: Path
    delivery_location_timeout: int
    delivery_locations: Dict[str, Dict[str, str]]
    delivery_location_fingerprint: str
    product_filters: ProductFilterConfig
    record_contract_fingerprint: str
    crawl_plan_fingerprint: str
    delay_seconds_min: float
    delay_seconds_max: float
    batch_pause_pages_min: int
    batch_pause_pages_max: int
    batch_pause_seconds_min: float
    batch_pause_seconds_max: float
    page_scroll_before_extract: bool
    page_scroll_max_rounds: int
    page_scroll_step_ratio: float
    page_scroll_wait_seconds: float
    page_scroll_stable_rounds: int
    sellersprite_required: bool
    sellersprite_min_enriched_records: int
    sellersprite_min_fields_per_record: int
    sellersprite_stable_checks: int
    save_debug_snapshots: bool
    field_selectors: Dict[str, List[str]] = field(default_factory=dict)


@dataclass
class CategoryPageBatch:
    key: str
    page_number: int
    page_url: str
    plugin_status: str
    extracted_count: int
    records: List[Dict[str, Any]]
    rejection_counts: Dict[str, int]


@dataclass
class CategoryCrawlBatch:
    node: Dict[str, Any]
    pages: List[CategoryPageBatch] = field(default_factory=list)
    children: List[Dict[str, Any]] = field(default_factory=list)
    skipped_intermediate: bool = False
    failures: List[Dict[str, Any]] = field(default_factory=list)
    terminal_error_type: str = ""
    terminal_error_message: str = ""


@dataclass
class CategoryPageWorkResult:
    """Uncommitted result of one fully validated category page attempt."""

    node: Dict[str, Any]
    next_url: str = ""
    page: Optional[CategoryPageBatch] = None
    children: List[Dict[str, Any]] = field(default_factory=list)
    skipped_intermediate: bool = False


MANUAL_INTERACTION_LOCK = threading.Lock()


def build_runtime_config(config: Dict[str, Any], config_path: Path, no_resume: bool) -> RuntimeConfig:
    start_url = config_text(config, "start_url")
    if not start_url:
        raise UserFacingError("配置项 `start_url` 不能为空。")
    validate_amazon_url(start_url)

    job_id = config_text(config, "job_id") or f"category-rank-{now_ts()}"
    outputs_root = resolve_path(config_text(config, "outputs_root", "outputs"))
    browser_backend = config_text(config, "browser_backend", "cdp").lower()
    if browser_backend != "cdp":
        raise UserFacingError(
            "配置项 `browser_backend` 必须为 cdp；旧 Selenium 后端无法证明标签页归属，已停用。"
        )
    browser_mode = config_text(config, "browser_mode", "reuse").lower()
    if browser_mode not in {"launch", "attach", "reuse"}:
        raise UserFacingError(
            "配置项 `browser_mode` 只支持 launch、attach 或 reuse；"
            "AppleScript 无法证明标签页归属，已停用。"
        )
    browser_tab_concurrency = config_int(config, "browser_tab_concurrency", 1)
    browser_tab_concurrency = 1 if browser_tab_concurrency is None else browser_tab_concurrency
    if browser_tab_concurrency < 1 or browser_tab_concurrency > 3:
        raise UserFacingError("配置项 `browser_tab_concurrency` 必须是 1-3 的整数。")
    if browser_tab_concurrency > 1 and not (
        browser_backend == "cdp" and browser_mode in {"attach", "reuse"}
    ):
        raise UserFacingError(
            "browser_tab_concurrency 大于 1 时只支持 browser_backend=cdp，"
            "且 browser_mode 必须是 attach 或 reuse。"
        )

    try:
        amazon_page_retry_schedule = retry_schedule_from_config(config)
    except RetryConfigurationError as exc:
        raise UserFacingError(str(exc)) from exc

    chrome_binary = config_text(config, "chrome_binary")
    chrome_user_data_dir = resolve_path(config_text(config, "chrome_user_data_dir", "chrome_profiles/category-rank-sellersprite"))
    chrome_profile_directory = config_text(config, "chrome_profile_directory", "Default") or "Default"
    debugger_address = config_text(config, "debugger_address", "127.0.0.1:9222")
    extension_path_text = config_text(config, "extension_path")
    extension_path = resolve_path(extension_path_text) if extension_path_text else Path("")
    if browser_mode == "launch" and extension_path_text and not extension_path.exists():
        raise UserFacingError(f"没有找到卖家精灵扩展目录：{extension_path}")

    min_delay = config_float(config, "delay_seconds_min", 4)
    max_delay = config_float(config, "delay_seconds_max", 9)
    if max_delay < min_delay:
        max_delay = min_delay
    batch_pages_min = config_int(config, "batch_pause_pages_min", 20) or 0
    batch_pages_max = config_int(config, "batch_pause_pages_max", 30) or 0
    if batch_pages_min < 0 or batch_pages_max < 0:
        raise UserFacingError("配置项 batch_pause_pages_min / batch_pause_pages_max 不能小于 0。")
    if batch_pages_max and batch_pages_max < batch_pages_min:
        batch_pages_max = batch_pages_min
    batch_seconds_min = config_float(config, "batch_pause_seconds_min", 60)
    batch_seconds_max = config_float(config, "batch_pause_seconds_max", 180)
    if batch_seconds_min < 0 or batch_seconds_max < 0:
        raise UserFacingError("配置项 batch_pause_seconds_min / batch_pause_seconds_max 不能小于 0。")
    if batch_seconds_max < batch_seconds_min:
        batch_seconds_max = batch_seconds_min
    plugin_retry_wait_min = config_float(
        config,
        "plugin_retry_wait_seconds_min",
        config_float(config, "plugin_retry_wait_seconds", 10),
    )
    plugin_retry_wait_max = config_float(config, "plugin_retry_wait_seconds_max", 20)
    if plugin_retry_wait_min < 0 or plugin_retry_wait_max < 0:
        raise UserFacingError("配置项 plugin_retry_wait_seconds_min / plugin_retry_wait_seconds_max 不能小于 0。")
    if plugin_retry_wait_max < plugin_retry_wait_min:
        plugin_retry_wait_max = plugin_retry_wait_min
    page_scroll_max_rounds = max(config_int(config, "page_scroll_max_rounds", 18) or 0, 0)
    page_scroll_step_ratio = config_float(config, "page_scroll_step_ratio", 0.85) or 0.85
    if page_scroll_step_ratio <= 0:
        raise UserFacingError("配置项 page_scroll_step_ratio 必须大于 0。")
    page_scroll_wait_seconds = max(config_float(config, "page_scroll_wait_seconds", 1.0) or 0, 0)
    page_scroll_stable_rounds = max(config_int(config, "page_scroll_stable_rounds", 2) or 1, 1)
    delivery_config = build_delivery_location_config(config)
    product_filters = build_product_filter_config(config)
    sellersprite_required = config_bool(config, "sellersprite_required", True)
    if product_filters.enabled and not sellersprite_required:
        raise UserFacingError(
            "启用 product_filters 时必须设置 `sellersprite_required: true`，"
            "否则无法可靠判断配送方式和子类目节点排名。"
        )

    raw_selectors = config.get("field_selectors") or {}
    field_selectors: Dict[str, List[str]] = {}
    if isinstance(raw_selectors, dict):
        for key, value in raw_selectors.items():
            if isinstance(value, list):
                field_selectors[key] = [str(item).strip() for item in value if str(item).strip()]

    include_root = config_bool(config, "include_root", False)
    max_depth = config_int(config, "max_depth")
    max_pages_per_category = config_int(config, "max_pages_per_category")

    return RuntimeConfig(
        start_url=start_url,
        job_id=slugify(job_id),
        outputs_root=outputs_root,
        browser_backend=browser_backend,
        browser_mode=browser_mode,
        chrome_binary=chrome_binary,
        chrome_user_data_dir=chrome_user_data_dir,
        chrome_profile_directory=chrome_profile_directory,
        debugger_address=debugger_address,
        browser_tab_concurrency=browser_tab_concurrency,
        extension_path=extension_path,
        include_root=include_root,
        max_depth=max_depth,
        max_pages_per_category=max_pages_per_category,
        max_categories=config_int(config, "max_categories"),
        resume=False if no_resume else config_bool(config, "resume", True),
        activate_plugin=config_bool(config, "activate_plugin", True),
        page_timeout=config_int(config, "page_timeout", 90) or 90,
        amazon_page_retry_schedule=amazon_page_retry_schedule,
        plugin_timeout=config_int(config, "plugin_timeout", 120) or 120,
        plugin_retry_attempts=max(config_int(config, "plugin_retry_attempts", 5) or 0, 0),
        plugin_retry_wait_seconds=plugin_retry_wait_min,
        plugin_retry_wait_seconds_max=plugin_retry_wait_max,
        plugin_relaunch_retry_attempts=max(config_int(config, "plugin_relaunch_retry_attempts", 3) or 0, 0),
        plugin_relaunch_wait_seconds=max(config_float(config, "plugin_relaunch_wait_seconds", 300) or 0, 0),
        plugin_second_relaunch_retry_attempts=max(config_int(config, "plugin_second_relaunch_retry_attempts", 3) or 0, 0),
        plugin_second_relaunch_wait_seconds=max(config_float(config, "plugin_second_relaunch_wait_seconds", 600) or 0, 0),
        manual_pause_timeout=config_int(config, "manual_pause_timeout", 900) or 900,
        **delivery_config,
        product_filters=product_filters,
        record_contract_fingerprint=record_contract_fingerprint(product_filters),
        crawl_plan_fingerprint=crawl_plan_fingerprint(
            start_url,
            include_root,
            max_depth,
            max_pages_per_category,
            field_selectors,
        ),
        delay_seconds_min=min_delay,
        delay_seconds_max=max_delay,
        batch_pause_pages_min=batch_pages_min,
        batch_pause_pages_max=batch_pages_max,
        batch_pause_seconds_min=batch_seconds_min,
        batch_pause_seconds_max=batch_seconds_max,
        page_scroll_before_extract=config_bool(config, "page_scroll_before_extract", True),
        page_scroll_max_rounds=page_scroll_max_rounds,
        page_scroll_step_ratio=page_scroll_step_ratio,
        page_scroll_wait_seconds=page_scroll_wait_seconds,
        page_scroll_stable_rounds=page_scroll_stable_rounds,
        sellersprite_required=sellersprite_required,
        sellersprite_min_enriched_records=max(
            config_int(config, "sellersprite_min_enriched_records", 1) or 1,
            1,
        ),
        sellersprite_min_fields_per_record=max(
            config_int(config, "sellersprite_min_fields_per_record", 2) or 2,
            1,
        ),
        sellersprite_stable_checks=max(
            config_int(config, "sellersprite_stable_checks", 3) or 3,
            1,
        ),
        save_debug_snapshots=config_bool(config, "save_debug_snapshots", True),
        field_selectors=field_selectors,
    )


class StateStore:
    def __init__(self, path: Path, runtime: RuntimeConfig) -> None:
        self.path = path
        self.runtime = runtime
        self.page_results_dir = self.path.parent / "page_results"
        self.data: Dict[str, Any] = {}

    def _new_data(self) -> Dict[str, Any]:
        root_name = infer_category_name_from_url(self.runtime.start_url)
        root_node = {
            "url": self.runtime.start_url,
            "name": root_name,
            "path": [root_name],
            "node_id": extract_node_id(self.runtime.start_url),
            "depth": 0,
        }
        return {
            "state_version": CATEGORY_STATE_SCHEMA_VERSION,
            "job_id": self.runtime.job_id,
            "start_url": self.runtime.start_url,
            "created_at": now_iso(),
            "queue": [root_node],
            "current": None,
            "in_flight_categories": {},
            "seen_categories": [category_key(root_node)],
            "done_categories": [],
            "completed_pages": [],
            "completed_page_order": [],
            "processed_categories_count": 0,
            "records_count": 0,
            "filtered_out_count": 0,
            "filter_rejection_counts": {},
            "failures_count": 0,
            "delivery_location_fingerprint": self.runtime.delivery_location_fingerprint,
            "record_contract_fingerprint": self.runtime.record_contract_fingerprint,
            "crawl_plan_fingerprint": self.runtime.crawl_plan_fingerprint,
        }

    def load_or_create(self) -> None:
        ensure_dir(self.page_results_dir)
        if self.runtime.resume and self.path.exists():
            self.data = load_json(self.path)
            delivery_changed = ensure_resume_delivery_fingerprint(
                self.data,
                self.runtime,
                self.path.with_name("records.jsonl"),
            )
            contract_changed = ensure_resume_record_contract_fingerprint(
                self.data,
                self.runtime,
                self.path.with_name("records.jsonl"),
            )
            plan_changed = ensure_resume_crawl_plan_fingerprint(
                self.data,
                self.runtime,
                self.path.with_name("records.jsonl"),
                self.page_results_dir,
            )
            if plan_changed:
                # A pending-only checkpoint has no committed work and can be
                # safely rebuilt from the new start URL/plan. Keeping its old
                # queue would relabel stale work with the new fingerprint.
                self.data = self._new_data()
                self.flush()
                return
            self._recover_from_page_results()
            if delivery_changed or contract_changed or plan_changed:
                self.data["checkpoint_contract_updated_at"] = now_iso()
            self.flush()
            return
        if self.runtime.resume:
            records_path = self.path.with_name("records.jsonl")
            has_records = records_path.exists() and records_path.stat().st_size > 0
            has_page_results = any(self.page_results_dir.glob("*.json"))
            if not has_page_results:
                ensure_resume_delivery_fingerprint({}, self.runtime, records_path)
                ensure_resume_record_contract_fingerprint({}, self.runtime, records_path)
                ensure_resume_crawl_plan_fingerprint(
                    {},
                    self.runtime,
                    records_path,
                    self.page_results_dir,
                )
            if has_records and not has_page_results:
                raise UserFacingError(
                    "现有 records.jsonl 没有原子 page shard，不能安全续跑；请更换 job_id。"
                )
        self.data = self._new_data()
        if self.runtime.resume:
            self._recover_from_page_results()
        self.flush()

    def flush(self) -> None:
        # Keep the generic state-v2 vocabulary alongside the recursive
        # category crawler's legacy queue/done names for operators and tooling.
        self.data["schema_version"] = CATEGORY_STATE_SCHEMA_VERSION
        self.data["pending"] = list(self.data.get("queue") or [])
        self.data["in_flight"] = dict(self.data.get("in_flight_categories") or {})
        self.data["completed_sources"] = list(self.data.get("done_categories") or [])
        self.data["updated_at"] = now_iso()
        dump_json(self.path, self.data)

    def page_result_path(self, key: str) -> Path:
        digest = hashlib.sha256(key.encode("utf-8")).hexdigest()
        return self.page_results_dir / f"{digest}.json"

    def _read_page_result(self, path: Path) -> Dict[str, Any]:
        try:
            payload = load_json(path)
        except Exception as exc:
            raise UserFacingError(f"页面提交文件损坏，无法安全恢复：{path.name}: {exc}") from exc
        if int(payload.get("schema_version") or 0) != CATEGORY_STATE_SCHEMA_VERSION:
            raise UserFacingError(f"页面提交文件版本不兼容：{path.name}；请更换 job_id。")
        if str(payload.get("record_contract_fingerprint") or "") != self.runtime.record_contract_fingerprint:
            raise UserFacingError(f"页面提交文件与当前数据契约不一致：{path.name}；请更换 job_id。")
        if str(payload.get("delivery_location_fingerprint") or "") != self.runtime.delivery_location_fingerprint:
            raise UserFacingError(f"页面提交文件与当前配送地址配置不一致：{path.name}；请更换 job_id。")
        if str(payload.get("crawl_plan_fingerprint") or "") != self.runtime.crawl_plan_fingerprint:
            raise UserFacingError(f"页面提交文件与当前抓取计划不一致：{path.name}；请更换 job_id。")
        return payload

    def iter_page_results(self) -> List[Dict[str, Any]]:
        by_key: Dict[str, Dict[str, Any]] = {}
        for path in sorted(self.page_results_dir.glob("*.json")):
            payload = self._read_page_result(path)
            key = str(payload.get("page_key") or "")
            if key:
                by_key[key] = payload
        ordered: List[Dict[str, Any]] = []
        for key in [str(item) for item in self.data.get("completed_page_order") or []]:
            payload = by_key.pop(key, None)
            if payload:
                ordered.append(payload)
        ordered.extend(
            sorted(
                by_key.values(),
                key=lambda item: (
                    str(item.get("committed_at") or ""),
                    str(item.get("page_key") or ""),
                ),
            )
        )
        return ordered

    def _recover_from_page_results(self) -> None:
        payloads = self.iter_page_results()
        shard_keys = [str(payload.get("page_key") or "") for payload in payloads]
        shard_key_set = {key for key in shard_keys if key}
        completed_in_state = set(self.data.get("completed_pages") or [])
        missing_shards = sorted(completed_in_state - shard_key_set)
        if missing_shards:
            raise UserFacingError(
                "现有类目断点缺少原子 page shard，不能保证 records 去重恢复；"
                "请保留旧输出并更换 job_id。"
            )

        records_count = 0
        scanned_count = 0
        rejection_counts: Dict[str, int] = {}
        for payload in payloads:
            records = payload.get("records") or []
            records_count += len(records) if isinstance(records, list) else 0
            scanned_count += max(int(payload.get("extracted_count") or 0), 0)
            for reason, count in dict(payload.get("rejection_counts") or {}).items():
                rejection_counts[str(reason)] = rejection_counts.get(str(reason), 0) + max(
                    int(count or 0), 0
                )
        self.data["state_version"] = CATEGORY_STATE_SCHEMA_VERSION
        self.data["completed_pages"] = shard_keys
        self.data["completed_page_order"] = shard_keys
        self.data["records_count"] = records_count
        self.data["filtered_out_count"] = max(scanned_count - records_count, 0)
        self.data["filter_rejection_counts"] = rejection_counts
        self.data["record_contract_fingerprint"] = self.runtime.record_contract_fingerprint
        self.data["delivery_location_fingerprint"] = self.runtime.delivery_location_fingerprint
        self.data["crawl_plan_fingerprint"] = self.runtime.crawl_plan_fingerprint

    def commit_page_batch(self, page: CategoryPageBatch) -> bool:
        if self.is_page_completed(page.key):
            return False
        normalized_records: List[Dict[str, Any]] = []
        seen_asins: set[str] = set()
        for record in page.records:
            if not isinstance(record, dict):
                continue
            asin = str(record.get("asin") or "")
            if asin in seen_asins:
                continue
            seen_asins.add(asin)
            normalized_records.append(record)
        payload = {
            "schema_version": CATEGORY_STATE_SCHEMA_VERSION,
            "record_contract_fingerprint": self.runtime.record_contract_fingerprint,
            "delivery_location_fingerprint": self.runtime.delivery_location_fingerprint,
            "crawl_plan_fingerprint": self.runtime.crawl_plan_fingerprint,
            "page_key": page.key,
            "page_number": page.page_number,
            "page_url": page.page_url,
            "plugin_status": page.plugin_status,
            "extracted_count": page.extracted_count,
            "records": normalized_records,
            "rejection_counts": page.rejection_counts,
            "committed_at": now_iso(),
        }
        dump_json(self.page_result_path(page.key), payload)
        self.mark_page_completed(
            page.key,
            len(normalized_records),
            page.extracted_count - len(normalized_records),
            page.rejection_counts,
        )
        return True

    def next_work(self) -> Optional[Dict[str, Any]]:
        current = self.data.get("current")
        if current:
            return current
        queue = self.data.get("queue") or []
        if not queue:
            return None
        node = queue.pop(0)
        current = {
            "node": node,
            "page_number": 1,
            "page_url": node["url"],
            "children_enqueued": False,
            "children": [],
        }
        self.data["queue"] = queue
        self.data["current"] = current
        self.flush()
        return current

    def prepare_concurrent_resume(self) -> None:
        """Recover stale work and migrate the legacy single-current checkpoint."""

        recovered: List[Dict[str, Any]] = []
        current = self.data.get("current")
        if isinstance(current, dict) and isinstance(current.get("node"), dict):
            recovered.append(dict(current["node"]))
        in_flight = self.data.get("in_flight_categories") or {}
        if isinstance(in_flight, dict):
            for item in in_flight.values():
                if isinstance(item, dict) and isinstance(item.get("node"), dict):
                    recovered.append(dict(item["node"]))

        done = set(self.data.setdefault("done_categories", []))
        pending: List[Dict[str, Any]] = []
        pending_keys: set[str] = set()
        for node in [*recovered, *(self.data.get("queue") or [])]:
            if not isinstance(node, dict) or not node.get("url"):
                continue
            key = category_key(node)
            if key in done or key in pending_keys:
                continue
            pending_keys.add(key)
            pending.append(node)
        self.data["state_version"] = 2
        self.data["queue"] = pending
        self.data["current"] = None
        self.data["in_flight_categories"] = {}
        self.flush()

    def recover_stale_in_flight(self) -> None:
        """Make a crashed concurrent run resumable in sequential mode too."""

        in_flight = self.data.get("in_flight_categories") or {}
        if not isinstance(in_flight, dict) or not in_flight:
            self.data.setdefault("in_flight_categories", {})
            self.data.setdefault("state_version", 2)
            return
        done = set(self.data.setdefault("done_categories", []))
        queue_items = list(self.data.get("queue") or [])
        queued_keys = {category_key(node) for node in queue_items if isinstance(node, dict)}
        recovered: List[Dict[str, Any]] = []
        for item in in_flight.values():
            node = item.get("node") if isinstance(item, dict) else None
            if not isinstance(node, dict) or not node.get("url"):
                continue
            key = category_key(node)
            if key in done or key in queued_keys:
                continue
            queued_keys.add(key)
            recovered.append(node)
        self.data["state_version"] = 2
        self.data["queue"] = [*recovered, *queue_items]
        self.data["in_flight_categories"] = {}
        self.flush()

    def claim_next_category(self) -> Optional[Tuple[str, Dict[str, Any]]]:
        queue_items = list(self.data.get("queue") or [])
        done = set(self.data.setdefault("done_categories", []))
        in_flight = self.data.setdefault("in_flight_categories", {})
        while queue_items:
            node = queue_items.pop(0)
            if not isinstance(node, dict) or not node.get("url"):
                continue
            key = category_key(node)
            if key in done or key in in_flight:
                continue
            self.data["queue"] = queue_items
            in_flight[key] = {
                "node": node,
                "claimed_at": now_iso(),
            }
            self.flush()
            return key, node
        self.data["queue"] = []
        self.flush()
        return None

    def complete_claimed_category(self, claim_key: str, node: Dict[str, Any]) -> None:
        in_flight = self.data.setdefault("in_flight_categories", {})
        in_flight.pop(claim_key, None)
        done = set(self.data.setdefault("done_categories", []))
        done.add(claim_key)
        done.add(category_key(node))
        self.data["done_categories"] = sorted(done)
        seen = set(self.data.setdefault("seen_categories", []))
        seen.add(claim_key)
        seen.add(category_key(node))
        self.data["seen_categories"] = sorted(seen)
        self.data["processed_categories_count"] = int(
            self.data.get("processed_categories_count") or 0
        ) + 1
        self.flush()

    def requeue_claimed_categories(self, claim_keys: Sequence[str]) -> None:
        in_flight = self.data.setdefault("in_flight_categories", {})
        queue_items = list(self.data.get("queue") or [])
        done = set(self.data.setdefault("done_categories", []))
        queued_keys = {category_key(node) for node in queue_items if isinstance(node, dict)}
        recovered: List[Dict[str, Any]] = []
        for claim_key in claim_keys:
            item = in_flight.pop(claim_key, None)
            node = item.get("node") if isinstance(item, dict) else None
            if not isinstance(node, dict) or not node.get("url"):
                continue
            key = category_key(node)
            if claim_key in done or key in done or key in queued_keys:
                continue
            queued_keys.add(key)
            recovered.append(node)
        self.data["queue"] = [*recovered, *queue_items]
        self.flush()

    def requeue_claimed_category(
        self,
        claim_key: str,
        updated_node: Optional[Dict[str, Any]] = None,
    ) -> None:
        in_flight = self.data.setdefault("in_flight_categories", {})
        item = in_flight.pop(claim_key, None)
        original_node = item.get("node") if isinstance(item, dict) else None
        node = updated_node if isinstance(updated_node, dict) else original_node
        queue_items = list(self.data.get("queue") or [])
        if isinstance(node, dict) and node.get("url"):
            key = category_key(node)
            done = set(self.data.setdefault("done_categories", []))
            queued_keys = {
                category_key(pending)
                for pending in queue_items
                if isinstance(pending, dict)
            }
            if claim_key not in done and key not in done and key not in queued_keys:
                queue_items.insert(0, node)
        self.data["queue"] = queue_items
        self.flush()

    def set_current(self, current: Optional[Dict[str, Any]]) -> None:
        self.data["current"] = current
        self.flush()

    def enqueue_children(self, children: Sequence[Dict[str, Any]]) -> int:
        queue = self.data.setdefault("queue", [])
        seen = set(self.data.setdefault("seen_categories", []))
        added = 0
        for child in children:
            key = category_key(child)
            if key in seen:
                continue
            seen.add(key)
            queue.append(child)
            added += 1
        self.data["seen_categories"] = sorted(seen)
        self.data["queue"] = queue
        self.flush()
        return added

    def mark_page_completed(
        self,
        key: str,
        count: int,
        filtered_out_count: int = 0,
        rejection_counts: Optional[Dict[str, int]] = None,
    ) -> None:
        completed = set(self.data.setdefault("completed_pages", []))
        completed.add(key)
        self.data["completed_pages"] = sorted(completed)
        order = self.data.setdefault("completed_page_order", [])
        if key not in order:
            order.append(key)
        self.data["records_count"] = int(self.data.get("records_count") or 0) + count
        self.data["filtered_out_count"] = (
            int(self.data.get("filtered_out_count") or 0) + max(filtered_out_count, 0)
        )
        totals = self.data.setdefault("filter_rejection_counts", {})
        for reason, rejected_count in (rejection_counts or {}).items():
            totals[reason] = int(totals.get(reason) or 0) + max(int(rejected_count), 0)
        self.flush()

    def is_page_completed(self, key: str) -> bool:
        return key in set(self.data.get("completed_pages") or [])

    def finish_current_category(self) -> None:
        current = self.data.get("current")
        if current:
            node = current.get("node") or {}
            done = set(self.data.setdefault("done_categories", []))
            done.add(category_key(node))
            self.data["done_categories"] = sorted(done)
            self.data["processed_categories_count"] = int(self.data.get("processed_categories_count") or 0) + 1
        self.data["current"] = None
        self.flush()

    def log_failure(self, count: int = 1) -> None:
        self.data["failures_count"] = int(self.data.get("failures_count") or 0) + max(
            int(count), 0
        )
        self.flush()

    def mark_sellersprite_readiness(self, report: Dict[str, Any]) -> None:
        self.data["sellersprite_readiness"] = safe_sellersprite_readiness(report)
        self.flush()

    def mark_manual_pause(self, reason: str, page_url: str, work_key: str = "") -> None:
        if work_key:
            pauses = self.data.setdefault("manual_pauses", {})
            pauses[work_key] = {
                "paused_at": now_iso(),
                "reason": reason,
                "page_url": page_url,
                "in_flight": (self.data.get("in_flight_categories") or {}).get(work_key),
            }
            self.flush()
            return
        self.data["manual_pause"] = {
            "paused_at": now_iso(),
            "reason": reason,
            "page_url": page_url,
            "current": self.data.get("current"),
        }
        self.flush()

    def clear_manual_pause(self, work_key: str = "") -> None:
        if work_key:
            pauses = self.data.get("manual_pauses") or {}
            if pauses.pop(work_key, None) is not None:
                if pauses:
                    self.data["manual_pauses"] = pauses
                else:
                    self.data.pop("manual_pauses", None)
                self.flush()
            return
        if self.data.pop("manual_pause", None) is not None:
            self.flush()

    @staticmethod
    def amazon_page_retry_entry_key(retry_state: Mapping[str, Any]) -> str:
        work_key = str(retry_state.get("work_key") or "")
        stage = str(retry_state.get("stage") or "")
        return f"{work_key}|stage:{stage}"

    def _amazon_page_retry_entries(self) -> Dict[str, Dict[str, Any]]:
        current = self.data.get("amazon_page_retry")
        entries: Dict[str, Dict[str, Any]] = {}
        if not isinstance(current, dict):
            return entries
        raw_entries = current.get("entries")
        if isinstance(raw_entries, dict):
            for key, value in raw_entries.items():
                if isinstance(value, dict):
                    entries[str(key)] = copy.deepcopy(value)
        elif current.get("work_key") and current.get("stage"):
            key = self.amazon_page_retry_entry_key(current)
            entries[key] = copy.deepcopy(current)
        return entries

    @staticmethod
    def _representative_amazon_page_retry(
        entries: Mapping[str, Mapping[str, Any]],
    ) -> Dict[str, Any]:
        priorities = {
            "manual_resume_required": 3,
            "waiting": 2,
            "attempting": 1,
        }
        def updated_at(item: Tuple[str, Mapping[str, Any]]) -> float:
            try:
                return float(item[1].get("updated_at") or 0)
            except (TypeError, ValueError, OverflowError):
                return 0.0

        _key, representative = max(
            entries.items(),
            key=lambda item: (
                priorities.get(str(item[1].get("status") or ""), 0),
                updated_at(item),
                str(item[0]),
            ),
        )
        payload = copy.deepcopy(dict(representative))
        payload["entries"] = {
            str(key): copy.deepcopy(dict(value))
            for key, value in entries.items()
        }
        return payload

    def load_amazon_page_retry(self, retry_key: str) -> Optional[Dict[str, Any]]:
        value = self._amazon_page_retry_entries().get(str(retry_key))
        return copy.deepcopy(value) if value is not None else None

    def write_amazon_page_retry(
        self,
        retry_key: str,
        retry_state: Mapping[str, Any],
    ) -> None:
        state_value = copy.deepcopy(dict(retry_state))
        entries = self._amazon_page_retry_entries()
        entries[str(retry_key)] = state_value
        self.data["amazon_page_retry"] = self._representative_amazon_page_retry(
            entries
        )
        self.flush()

    def clear_amazon_page_retry(self, retry_key: str) -> None:
        entries = self._amazon_page_retry_entries()
        if entries.pop(str(retry_key), None) is None:
            return
        if entries:
            self.data["amazon_page_retry"] = self._representative_amazon_page_retry(
                entries
            )
        else:
            self.data.pop("amazon_page_retry", None)
        self.flush()


def write_jsonl_atomic(path: Path, records: Sequence[Dict[str, Any]]) -> None:
    ensure_dir(path.parent)
    file_descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=str(path.parent),
    )
    temp_path = Path(temp_name)
    try:
        with os.fdopen(file_descriptor, "w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json.dumps(record, ensure_ascii=False) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    except BaseException:
        try:
            temp_path.unlink()
        except FileNotFoundError:
            pass
        raise


def materialize_category_records(state: StateStore, records_path: Path) -> int:
    records: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str]] = set()
    for payload in state.iter_page_results():
        key = str(payload.get("page_key") or "")
        for record in payload.get("records") or []:
            if not isinstance(record, dict):
                continue
            identity = (key, str(record.get("asin") or ""))
            if identity in seen:
                continue
            seen.add(identity)
            records.append(record)
    write_jsonl_atomic(records_path, records)
    return len(records)


def start_driver(runtime: RuntimeConfig) -> WebDriver:
    if runtime.browser_mode == "applescript":
        return AppleScriptChromeDriver(runtime.page_timeout)  # type: ignore[return-value]
    if runtime.browser_backend == "cdp":
        owned_process = None
        if runtime.browser_mode == "launch":
            owned_process = launch_debug_chrome(runtime)
        elif not wait_for_debugger(runtime.debugger_address, timeout=2):
            raise UserFacingError(
                "没有找到可连接的 Chrome CDP 调试窗口。"
                f"当前配置的调试地址是：{runtime.debugger_address}"
            )
        try:
            return CdpWebDriver(  # type: ignore[return-value]
                debugger_address=runtime.debugger_address,
                page_timeout=runtime.page_timeout,
                expected_user_data_dir=runtime.chrome_user_data_dir,
                profile_directory=runtime.chrome_profile_directory,
                owns_browser=owned_process is not None,
                owned_process=owned_process,
            )
        except WebDriverException as exc:
            raise UserFacingError(str(exc)) from exc

    options = Options()
    if runtime.browser_mode == "launch":
        launch_debug_chrome(runtime)
        options.add_experimental_option("debuggerAddress", runtime.debugger_address)
        return webdriver.Chrome(options=options)
    if runtime.browser_mode in {"attach", "reuse"}:
        if runtime.browser_mode == "reuse" and not wait_for_debugger(runtime.debugger_address, timeout=2):
            raise UserFacingError(
                "没有找到可复用的 Chrome 调试窗口。请先打开带调试端口的 Chrome，"
                f"或把 browser_mode 改回 launch。当前配置的调试地址是：{runtime.debugger_address}"
            )
        options.add_experimental_option("debuggerAddress", runtime.debugger_address)
        return webdriver.Chrome(options=options)

    raise UserFacingError(f"不支持的浏览器模式：{runtime.browser_mode}")


def launch_debug_chrome(runtime: RuntimeConfig) -> Optional[subprocess.Popen]:
    if wait_for_debugger(runtime.debugger_address, timeout=2):
        return None
    chrome_binary = Path(runtime.chrome_binary)
    if runtime.chrome_binary and not chrome_binary.exists():
        raise UserFacingError(f"没有找到 Chrome：{runtime.chrome_binary}")
    ensure_dir(runtime.chrome_user_data_dir)
    port = parse_debugger_port(runtime.debugger_address)
    command = [
        runtime.chrome_binary,
        f"--remote-debugging-port={port}",
        f"--user-data-dir={runtime.chrome_user_data_dir}",
        f"--profile-directory={runtime.chrome_profile_directory}",
        "--no-first-run",
        "--new-window",
        "about:blank",
    ]
    if runtime.extension_path and runtime.extension_path.exists():
        command.insert(-2, f"--load-extension={runtime.extension_path}")
    process = subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    if not wait_for_debugger(runtime.debugger_address, timeout=90):
        raise UserFacingError(
            "Chrome 调试端口没有启动成功。请关闭刚打开的专用 Chrome 后重试，"
            "或把 browser_mode 改成 attach 并手动打开调试端口。"
        )
    return process


def parse_debugger_port(address: str) -> int:
    address = address.strip()
    if ":" not in address:
        raise UserFacingError("debugger_address 需要形如 127.0.0.1:9222。")
    try:
        return int(address.rsplit(":", 1)[-1])
    except ValueError as exc:
        raise UserFacingError("debugger_address 端口不是有效数字。") from exc


def wait_for_debugger(address: str, timeout: int) -> bool:
    deadline = time.time() + timeout
    endpoint = f"http://{address}/json/version"
    while time.time() < deadline:
        try:
            with urlopen(endpoint, timeout=2) as response:
                return response.status == 200
        except Exception:
            time.sleep(1)
    return False


def start_driver_legacy(runtime: RuntimeConfig) -> WebDriver:
    options = Options()
    if runtime.chrome_binary:
        options.binary_location = runtime.chrome_binary
    ensure_dir(runtime.chrome_user_data_dir)
    options.add_argument(f"--user-data-dir={runtime.chrome_user_data_dir}")
    options.add_argument(f"--profile-directory={runtime.chrome_profile_directory}")
    options.add_argument("--start-maximized")
    if runtime.extension_path and runtime.extension_path.exists():
        options.add_argument(f"--disable-extensions-except={runtime.extension_path}")
        options.add_argument(f"--load-extension={runtime.extension_path}")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(runtime.page_timeout)
    return driver


class AppleScriptElement:
    def __init__(
        self,
        text: str = "",
        driver: Optional["AppleScriptChromeDriver"] = None,
        selector: str = "",
    ) -> None:
        self.text = text
        self._driver = driver
        self._selector = selector

    def clear(self) -> None:
        if self._driver is None or not self._selector:
            raise WebDriverException("AppleScript 元素缺少可操作的 CSS 选择器。")
        changed = self._driver.execute_script(
            r"""
const selector = arguments[0];
const input = [...document.querySelectorAll(selector)].find(el => el.offsetParent !== null)
  || document.querySelector(selector);
if (!input) return false;
const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
if (descriptor && descriptor.set) descriptor.set.call(input, ''); else input.value = '';
for (const eventName of ['input', 'change']) input.dispatchEvent(new Event(eventName, {bubbles: true}));
return true;
""",
            self._selector,
        )
        if not changed:
            raise WebDriverException(f"没有找到可清空的 AppleScript 元素：{self._selector}")

    def send_keys(self, value: str) -> None:
        if self._driver is None or not self._selector:
            raise WebDriverException("AppleScript 元素缺少可操作的 CSS 选择器。")
        typed = self._driver.execute_script(
            r"""
const selector = arguments[0];
const text = String(arguments[1] || '');
const input = [...document.querySelectorAll(selector)].find(el => el.offsetParent !== null)
  || document.querySelector(selector);
if (!input) return false;
const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
let current = String(input.value || '');
for (const character of text) {
  input.dispatchEvent(new KeyboardEvent('keydown', {key: character, bubbles: true}));
  current += character;
  if (descriptor && descriptor.set) descriptor.set.call(input, current); else input.value = current;
  input.dispatchEvent(new InputEvent('input', {data: character, inputType: 'insertText', bubbles: true}));
  input.dispatchEvent(new KeyboardEvent('keyup', {key: character, bubbles: true}));
}
input.dispatchEvent(new Event('change', {bubbles: true}));
return true;
""",
            self._selector,
            value,
        )
        if not typed:
            raise WebDriverException(f"没有找到可输入的 AppleScript 元素：{self._selector}")

    def type_text(self, value: str) -> None:
        self.send_keys(value)

    def click(self) -> None:
        if self._driver is None or not self._selector:
            raise WebDriverException("AppleScript 元素缺少可操作的 CSS 选择器。")
        clicked = self._driver.execute_script(
            r"""
const selector = arguments[0];
const element = [...document.querySelectorAll(selector)].find(el => el.offsetParent !== null)
  || document.querySelector(selector);
if (!element) return false;
element.click();
return true;
""",
            self._selector,
        )
        if not clicked:
            raise WebDriverException(f"没有找到可点击的 AppleScript 元素：{self._selector}")


class AppleScriptChromeDriver:
    is_applescript_driver = True

    def __init__(self, page_timeout: int) -> None:
        self.page_timeout = page_timeout
        self.set_page_load_timeout(page_timeout)
        try:
            self.execute_script("return String(123);")
        except WebDriverException as exc:
            raise UserFacingError(
                "当前 Chrome 还没有打开“允许 Apple 事件中的 JavaScript”。"
                "请在 Chrome 菜单栏选择：显示 > 开发者 > 允许 Apple 事件中的 JavaScript。"
            ) from exc

    def _run_osascript(self, script: str, *args: str) -> str:
        result = subprocess.run(
            ["osascript", "-e", script, *args],
            text=True,
            capture_output=True,
            timeout=max(self.page_timeout, 30),
        )
        if result.returncode != 0:
            raise WebDriverException((result.stderr or result.stdout or "").strip())
        return (result.stdout or "").rstrip("\n")

    def set_page_load_timeout(self, timeout: int) -> None:
        self.page_timeout = timeout

    def get(self, url: str) -> None:
        script = r'''
on run argv
  set targetUrl to item 1 of argv
  tell application "Google Chrome"
    activate
    if (count of windows) = 0 then make new window
    set URL of active tab of front window to targetUrl
  end tell
end run
'''
        self._run_osascript(script, url)

    @property
    def current_url(self) -> str:
        script = r'''
tell application "Google Chrome"
  if (count of windows) = 0 then return ""
  return URL of active tab of front window
end tell
'''
        return self._run_osascript(script)

    @property
    def title(self) -> str:
        script = r'''
tell application "Google Chrome"
  if (count of windows) = 0 then return ""
  return title of active tab of front window
end tell
'''
        return self._run_osascript(script)

    @property
    def page_source(self) -> str:
        return str(self.execute_script("return document.documentElement ? document.documentElement.outerHTML : '';") or "")

    def refresh(self) -> None:
        script = r'''
tell application "Google Chrome"
  if (count of windows) > 0 then reload active tab of front window
end tell
'''
        self._run_osascript(script)

    def quit(self) -> None:
        return None

    def save_screenshot(self, path: str) -> bool:
        return False

    def find_element(self, by: str = By.ID, value: Optional[str] = None) -> AppleScriptElement:
        if by == By.TAG_NAME and value and value.lower() == "body":
            text = str(self.execute_script("return document.body ? document.body.innerText : '';") or "")
            return AppleScriptElement(text, self, "body")
        if by == By.CSS_SELECTOR and value:
            exists = bool(self.execute_script("return !!document.querySelector(arguments[0]);", value))
            if exists:
                text = str(self.execute_script("const el = document.querySelector(arguments[0]); return el ? (el.innerText || el.textContent || '') : '';", value) or "")
                return AppleScriptElement(text, self, value)
        raise NoSuchElementException(value or "")

    def execute_script(self, script: str, *args: Any) -> Any:
        js_args = json.dumps(list(args), ensure_ascii=False)
        wrapped = f"""
(function() {{
  const __codexArgs = {js_args};
  const __codexUserFn = function() {{
{script}
  }};
  const __codexResult = __codexUserFn.apply(window, __codexArgs);
  return JSON.stringify(__codexResult === undefined ? null : __codexResult);
}})();
"""
        apple_script = r'''
on run argv
  set jsPath to item 1 of argv
  set jsCode to read (POSIX file jsPath) as «class utf8»
  tell application "Google Chrome"
    if (count of windows) = 0 then make new window
    return execute active tab of front window javascript jsCode
  end tell
end run
'''
        with tempfile.NamedTemporaryFile("w", suffix=".js", encoding="utf-8", delete=True) as fh:
            fh.write(wrapped)
            fh.flush()
            raw = self._run_osascript(apple_script, fh.name)
        if raw == "":
            return None
        try:
            return json.loads(raw)
        except json.JSONDecodeError as exc:
            raise JavascriptException(raw) from exc


def sleep_between_pages(runtime: RuntimeConfig) -> None:
    delay = random.uniform(runtime.delay_seconds_min, runtime.delay_seconds_max)
    time.sleep(delay)


class BatchPauseScheduler:
    def __init__(self, runtime: RuntimeConfig) -> None:
        self.runtime = runtime
        self.pages_since_pause = 0
        self.next_pause_after = self._pick_next_pause_after()
        self._lock = threading.Lock()

    def _pick_next_pause_after(self) -> int:
        min_pages = self.runtime.batch_pause_pages_min
        max_pages = self.runtime.batch_pause_pages_max
        if min_pages <= 0 or max_pages <= 0:
            return 0
        return random.randint(min_pages, max_pages)

    def after_completed_page(self) -> None:
        with self._lock:
            if not self.next_pause_after:
                return
            self.pages_since_pause += 1
            if self.pages_since_pause < self.next_pause_after:
                return
            duration = random.uniform(self.runtime.batch_pause_seconds_min, self.runtime.batch_pause_seconds_max)
            if duration > 0:
                print(f"已连续抓取 {self.pages_since_pause} 页，自动休息 {duration / 60:.1f} 分钟以降低风控风险。")
                time.sleep(duration)
            self.pages_since_pause = 0
            self.next_pause_after = self._pick_next_pause_after()


class NavigationThrottle:
    """Globally stagger browser navigations across category workers."""

    def __init__(self, minimum: float, maximum: float) -> None:
        self.minimum = max(float(minimum), 0.0)
        self.maximum = max(float(maximum), self.minimum)
        self._lock = threading.Lock()
        self._next_navigation_at = 0.0

    def wait(self) -> None:
        with self._lock:
            now = time.monotonic()
            navigation_at = max(now, self._next_navigation_at)
            self._next_navigation_at = navigation_at + random.uniform(
                self.minimum,
                self.maximum,
            )
        delay = navigation_at - time.monotonic()
        if delay > 0:
            time.sleep(delay)


class DeliveryDomainLocks:
    def __init__(self) -> None:
        self._guard = threading.Lock()
        self._locks: Dict[str, threading.Lock] = {}

    def for_url(self, url: str) -> threading.Lock:
        domain = (urlparse(url).hostname or "unknown").lower()
        with self._guard:
            return self._locks.setdefault(domain, threading.Lock())


def safe_find_text(driver: WebDriver) -> str:
    try:
        return normalize_space(driver.find_element(By.TAG_NAME, "body").text)
    except (WebDriverException, AttributeError):
        return ""


def detect_block(driver: WebDriver) -> Optional[str]:
    title = ""
    try:
        title = driver.title or ""
    except WebDriverException:
        pass
    text = safe_find_text(driver)
    haystack = f"{title}\n{text}".lower()
    if "sellersprite" in haystack or "卖家精灵" in haystack:
        sellersprite_verify_markers = [
            "slide to verify",
            "complete the verification",
            "enter the characters you see",
            "validatecaptcha",
            "captcha",
            "robot check",
            "机器人检测",
            "我不是机器人",
            "验证码",
        ]
        if any(marker in haystack for marker in sellersprite_verify_markers):
            return "sellersprite_verification"
    checks = [
        ("amazon_robot_check", ["robot check", "enter the characters you see", "validatecaptcha", "captcha", "机器人检测", "我不是机器人"]),
        ("amazon_sign_in", ["sign in", "authentication required", "enter your password", "请先登录亚马逊", "登录亚马逊"]),
    ]
    for reason, markers in checks:
        if any(marker in haystack for marker in markers):
            if reason == "amazon_sign_in":
                sign_in_url = urlparse(str(getattr(driver, "current_url", "") or ""))
                if (
                    "/ap/signin" not in sign_in_url.path.lower()
                    or "amazon." not in sign_in_url.netloc.lower()
                ):
                    continue
            return reason
    return None


def _raise_if_stop_requested(stop_event: Optional[threading.Event]) -> None:
    if stop_event is not None and stop_event.is_set():
        raise ConcurrentWorkerCancelled("并发任务正在停止，已取消等待操作。")


def _sleep_with_stop(seconds: float, stop_event: Optional[threading.Event]) -> None:
    duration = max(float(seconds), 0.0)
    if duration <= 0:
        _raise_if_stop_requested(stop_event)
        return
    if stop_event is None:
        time.sleep(duration)
        return
    if stop_event.wait(duration):
        raise ConcurrentWorkerCancelled("并发任务正在停止，已取消等待操作。")


def wait_for_manual_continue(
    timeout: int,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    started_at = time.time()
    deadline = started_at + timeout
    next_status_at = started_at + 30
    prompt = "人工处理完成后按 Enter 继续；在 Codex 中请告诉我“继续”："
    print(prompt, flush=True)
    while time.time() < deadline:
        _raise_if_stop_requested(stop_event)
        remaining = max(0, deadline - time.time())
        try:
            poll_seconds = 0.25 if stop_event is not None else 30
            ready, _, _ = select.select([sys.stdin], [], [], min(remaining, poll_seconds))
        except (OSError, ValueError):
            ready = []
        if ready:
            try:
                line = sys.stdin.readline()
            except EOFError:
                return False
            if line == "":
                return False
            return True
        now = time.time()
        if now >= next_status_at:
            print("仍在等待人工确认继续。", flush=True)
            next_status_at = now + 30
    return False


def wait_for_manual_clear(
    driver: WebDriver,
    reason: str,
    timeout: int,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    if reason == "amazon_sign_in":
        print(AMAZON_SIGN_IN_STOP_MESSAGE)
        return False
    print(f"检测到需要人工处理：{reason}")
    print("脚本已暂停并保留当前页面。请在 Chrome 中手动完成页面验证。")
    deadline = time.time() + max(int(timeout), 1)
    while True:
        remaining = deadline - time.time()
        continued = False
        if remaining > 0:
            wait_seconds = max(int(remaining), 1)
            continued = (
                wait_for_manual_continue(wait_seconds)
                if stop_event is None
                else wait_for_manual_continue(wait_seconds, stop_event=stop_event)
            )
        if not continued:
            return False
        if not detect_block(driver):
            return True
        print("页面仍显示验证或限制，请继续处理后再确认。")


def verification_unconfirmed_message(reason: str) -> str:
    if reason == "amazon_sign_in":
        return f"amazon_sign_in_terminal: {AMAZON_SIGN_IN_STOP_MESSAGE}"
    return f"{reason}_unconfirmed: 人工处理超时，任务已停止且未提取当前页数据。"


def sellersprite_block_reason(driver: WebDriver) -> str:
    readiness = get_sellersprite_readiness(driver)
    return normalize_space(str(readiness.get("blocked_reason") or "")) or (
        detect_block(driver) or "sellersprite_verification"
    )


def amazon_marketplace_domain(url: str, locations: Dict[str, Dict[str, str]]) -> str:
    host = (urlparse(url).hostname or "").lower().rstrip(".")
    for domain in sorted(locations, key=len, reverse=True):
        if host == domain or host.endswith(f".{domain}"):
            return domain
    return ""


def normalize_delivery_value(value: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", str(value or "").lower())


def delivery_value_is_present(snapshot: str, expected: str) -> bool:
    """Match a normalized value without allowing a longer alphanumeric value."""
    characters = re.findall(r"[0-9a-z]", str(expected or "").lower())
    if not characters:
        return False
    pattern = (
        r"(?<![0-9a-z])"
        + r"[^0-9a-z]*".join(re.escape(character) for character in characters)
        + r"(?![0-9a-z])"
    )
    return re.search(pattern, str(snapshot or "").lower()) is not None


def delivery_postal_candidates(postal_code: str) -> List[str]:
    original = str(postal_code or "").strip()
    compact = re.sub(r"[\s-]+", "", original)
    return [value for index, value in enumerate((original, compact)) if value and value not in (original, compact)[:index]]


def _runtime_delivery_locations(runtime: Any) -> Dict[str, Dict[str, str]]:
    locations = getattr(runtime, "delivery_locations", None)
    if isinstance(locations, dict) and locations:
        return locations
    path = getattr(runtime, "delivery_locations_file", None)
    if path:
        return load_delivery_locations(Path(path))
    return build_delivery_location_config({})["delivery_locations"]


def _delivery_location_snapshot(driver: WebDriver) -> str:
    script = r"""
/* lc_delivery_snapshot */
const selectors = [
  '#glow-ingress-block', '#glow-ingress-line1', '#glow-ingress-line2',
  '#nav-global-location-popover-link', '#nav-global-location-slot',
  '#nav-global-location-data-modal-action', '#GLUXHiddenSuccessMessage',
  '#GLUXHiddenSuccessDialog', '#GLUXHiddenSuccessSelectedAddressPlaceholder',
  '#GLUXHiddenSuccessSubTextAisEgress',
  '[data-action="GLUXPostalInputAction"] .a-alert-content',
  '[data-action="GLUXConfirmAction"]'
];
const values = [];
for (const selector of selectors) {
  for (const el of document.querySelectorAll(selector)) {
    values.push(el.innerText || el.textContent || el.getAttribute('aria-label') || el.getAttribute('title') || '');
  }
}
return values.join(' ').replace(/\s+/g, ' ').trim();
"""
    try:
        return normalize_space(str(driver.execute_script(script) or ""))
    except (JavascriptException, WebDriverException):
        return ""


def delivery_location_is_confirmed(
    driver: WebDriver,
    location: Dict[str, str],
) -> bool:
    snapshot = _delivery_location_snapshot(driver)
    if not snapshot:
        return False
    expected = location["city"] if location.get("strategy") == "postal_then_city" else location["postal_code"]
    return delivery_value_is_present(snapshot, expected)


def _submitted_delivery_location_is_confirmed(
    driver: WebDriver,
    location: Dict[str, str],
    allow_reloaded_city_input: bool = False,
) -> bool:
    """Confirm an exact submit using only persisted post-navigation evidence.

    Some marketplaces, including amazon.ca, replace the last postal character in
    the reopened header with a zero-width mask. UAE confirmation may read the full
    city restored in the modal only after the transient value was cleared and the
    target URL reopened. A fresh driver still has to submit and confirm again
    because the initial check remains exact.
    """
    if delivery_location_is_confirmed(driver, location):
        return True
    if location.get("strategy") == "postal_then_city":
        if not allow_reloaded_city_input:
            return False
        script = r"""
/* lc_delivery_city_input_snapshot */
const input = [...document.querySelectorAll('#GLUXCityWithDistrictCityInput')]
  .find(el => el.offsetParent !== null);
return input ? input.value : '';
"""
        try:
            selected_city = str(driver.execute_script(script) or "")
        except (JavascriptException, WebDriverException):
            return False
        if normalize_delivery_value(selected_city) != normalize_delivery_value(location["city"]):
            return False
        return _close_delivery_city_dialog(driver)
    script = r"""
/* lc_delivery_header_snapshot */
const el = document.querySelector('#glow-ingress-line2');
return el ? (el.innerText || el.textContent || '') : '';
"""
    try:
        header = str(driver.execute_script(script) or "")
    except (JavascriptException, WebDriverException):
        return False
    privacy_masks = ("\u200b", "\u200c", "\u200d", "\u2060", "\ufeff")
    expected = normalize_delivery_value(location["postal_code"])
    if len(expected) <= 1:
        return False
    for index, character in enumerate(header):
        if character not in privacy_masks:
            continue
        before_mask = normalize_delivery_value(header[:index])
        after_mask = normalize_delivery_value(header[index + 1 :])
        if before_mask == expected[:-1] and not after_mask:
            return True
    return False


def _dismiss_delivery_obstructions(driver: WebDriver) -> None:
    """Close only SellerSprite extension dialogs that cover Amazon's location UI."""
    script = r"""
/* lc_delivery_dismiss_obstructions */
const roots = [...document.querySelectorAll('#seller-sprite-extension-app [role="dialog"]')]
  .filter(el => el.offsetParent !== null);
for (const root of roots) {
  const selectors = [
    '.el-dialog__headerbtn', '.el-message-box__headerbtn',
    'button[aria-label*="close" i]', '[role="button"][aria-label*="close" i]'
  ];
  for (const selector of selectors) {
    const button = root.querySelector(selector);
    if (button && button.offsetParent !== null) { button.click(); return true; }
  }
}
return false;
"""
    try:
        driver.execute_script(script)
    except (JavascriptException, WebDriverException):
        pass


def _close_delivery_city_dialog(driver: WebDriver) -> bool:
    script = r"""
/* lc_delivery_city_dialog_close */
const cityInput = [...document.querySelectorAll('#GLUXCityWithDistrictCityInput')]
  .find(el => el.offsetParent !== null);
const root = cityInput && cityInput.closest('.a-popover-modal, [role="dialog"]');
if (root && !root.closest('#seller-sprite-extension-app')) {
  const selectors = [
    '.a-button-close', '[data-action="a-popover-close"]',
    'button[aria-label*="close" i]', '[role="button"][aria-label*="close" i]'
  ];
  for (const selector of selectors) {
    const button = root.querySelector(selector);
    if (button && button.offsetParent !== null) { button.click(); return true; }
  }
}
return false;
"""
    try:
        clicked = bool(driver.execute_script(script))
    except (JavascriptException, WebDriverException):
        return False
    if not clicked:
        return False
    hidden_deadline = time.time() + 1.5
    visible_script = r"""
/* lc_delivery_city_dialog_visible */
return [...document.querySelectorAll('#GLUXCityWithDistrictCityInput')]
  .some(el => el.offsetParent !== null);
"""
    while time.time() < hidden_deadline:
        try:
            if not bool(driver.execute_script(visible_script)):
                return True
        except (JavascriptException, WebDriverException):
            return False
        time.sleep(0.1)
    try:
        return not bool(driver.execute_script(visible_script))
    except (JavascriptException, WebDriverException):
        return False


def _clear_transient_delivery_city_input(driver: WebDriver) -> bool:
    """Clear any current UAE form value before a navigation-based persistence check."""
    script = r"""
/* lc_delivery_city_input_reset */
const input = document.querySelector('#GLUXCityWithDistrictCityInput');
if (!input) return 'absent';
const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
if (descriptor && descriptor.set) descriptor.set.call(input, ''); else input.value = '';
return input.value === '' ? 'cleared' : 'failed';
"""
    try:
        return str(driver.execute_script(script) or "") in {"absent", "cleared"}
    except (JavascriptException, WebDriverException):
        return False


def _click_delivery_trigger(driver: WebDriver) -> bool:
    _dismiss_delivery_obstructions(driver)
    visible_input_script = r"""
/* lc_delivery_open_visible */
const visibleInputs = [
  '#GLUXZipUpdateInput', 'input[name="zipCode"]', 'input[name="postalCode"]',
  '#GLUXPostalCodeWithCity_PostalCodeInput',
  '#GLUXCityWithDistrictCityInput',
  '[data-action="GLUXPostalInputAction"] input[type="text"]'
];
const inputVisible = visibleInputs.some(selector => {
  return [...document.querySelectorAll(selector)].some(input => input.offsetParent !== null);
});
const deliverySurfaceVisible = [
  '#GLUXAddressBlock', '#GLUXSpecifyLocationDiv', '#GLUXHiddenSuccessDialog',
  '#GLUXHiddenSuccessSelectedAddressPlaceholder'
].some(selector => [...document.querySelectorAll(selector)].some(el => el.offsetParent !== null));
return inputVisible || deliverySurfaceVisible;
"""
    try:
        if bool(driver.execute_script(visible_input_script)):
            return True
    except (JavascriptException, WebDriverException):
        pass

    current_domain = (urlparse(str(getattr(driver, "current_url", "") or "")).hostname or "").lower()
    try:
        opened_domain, opened_at = getattr(driver, "_lc_delivery_trigger_opened", ("", 0.0))
        if opened_domain == current_domain and time.time() - float(opened_at) < 5:
            return True
    except (TypeError, ValueError):
        pass

    selectors = [
        "#nav-global-location-popover-link",
        "#nav-global-location-slot",
        "#nav-global-location-data-modal-action",
        '[data-csa-c-content-id="nav-global-location-slot"]',
    ]
    for selector in selectors:
        try:
            driver.find_element(By.CSS_SELECTOR, selector).click()
            setattr(driver, "_lc_delivery_trigger_opened", (current_domain, time.time()))
            return True
        except (AttributeError, NoSuchElementException, WebDriverException):
            continue

    script = r"""
/* lc_delivery_open */
const selectors = [
  '#nav-global-location-popover-link', '#nav-global-location-slot',
  '#nav-global-location-data-modal-action', '[data-csa-c-content-id="nav-global-location-slot"]'
];
for (const selector of selectors) {
  const el = document.querySelector(selector);
  if (el) { el.click(); return true; }
}
return false;
"""
    try:
        clicked = bool(driver.execute_script(script))
        if clicked:
            setattr(driver, "_lc_delivery_trigger_opened", (current_domain, time.time()))
        return clicked
    except (JavascriptException, WebDriverException):
        return False


def _submit_delivery_postal(driver: WebDriver, postal_code: str, city: str = "") -> Optional[bool]:
    city_only_script = r"""
/* lc_delivery_city_only_form */
return [...document.querySelectorAll('#GLUXCityWithDistrictCityInput')]
  .some(el => el.offsetParent !== null);
"""
    try:
        if bool(driver.execute_script(city_only_script)):
            return None
    except (JavascriptException, WebDriverException):
        pass

    postal_with_city_script = r"""
/* lc_delivery_postal_with_city_submit */
const postal = String(arguments[0] || '');
const city = String(arguments[1] || '');
const input = [...document.querySelectorAll('#GLUXPostalCodeWithCity_PostalCodeInput')]
  .find(el => el.offsetParent !== null);
if (!input) return 'not_applicable';
if (input.value !== postal) {
  return 'needs_typing';
}
const norm = value => String(value || '').replace(/\s+/g, ' ').trim().toLowerCase();
const select = [...document.querySelectorAll('#GLUXPostalCodeWithCity_DropdownList')]
  .find(el => el.offsetParent !== null);
if (!select) return 'waiting';
const option = [...select.options].find(item => item.value && norm(item.textContent).includes(norm(city)));
if (!option) return 'waiting';
if (select.value !== option.value) {
  select.value = option.value;
  for (const eventName of ['input', 'change']) select.dispatchEvent(new Event(eventName, {bubbles: true}));
}
for (const button of document.querySelectorAll('#GLUXPostalCodeWithCityApplyButton')) {
  if (button.offsetParent !== null) { button.click(); return 'submitted'; }
}
return 'waiting';
"""
    if city:
        try:
            postal_with_city_status = str(
                driver.execute_script(postal_with_city_script, postal_code, city) or ""
            )
        except (JavascriptException, WebDriverException):
            postal_with_city_status = ""
        if postal_with_city_status == "submitted":
            return True
        if postal_with_city_status == "needs_typing":
            try:
                postal_input = driver.find_element(
                    By.CSS_SELECTOR,
                    "#GLUXPostalCodeWithCity_PostalCodeInput",
                )
                postal_input.clear()
                type_text = getattr(postal_input, "type_text", None)
                if callable(type_text):
                    type_text(postal_code)
                else:
                    postal_input.send_keys(postal_code)
            except (AttributeError, NoSuchElementException, WebDriverException):
                pass
            return False
        if postal_with_city_status == "waiting":
            return False

    input_selectors = [
        "#GLUXZipUpdateInput",
        'input[name="zipCode"]',
        'input[name="postalCode"]',
        '[data-action="GLUXPostalInputAction"] input[type="text"]',
        'input[autocomplete="postal-code"]',
    ]
    submit_selectors = [
        "#GLUXZipUpdate",
        '[data-action="GLUXPostalInputAction"] input[type="submit"]',
        '[data-action="GLUXPostalInputAction"] button',
        'button[name="glowDoneButton"]',
    ]
    split_fields_script = r"""
/* lc_delivery_split_postal_fields */
return [...document.querySelectorAll('input[id^="GLUXZipUpdateInput_"]')]
  .filter(input => input.offsetParent !== null)
  .sort((a, b) => String(a.id).localeCompare(String(b.id)))
  .map(input => ({id: input.id, maxLength: Number(input.maxLength || 0), value: input.value || ''}));
"""
    try:
        split_fields = driver.execute_script(split_fields_script) or []
    except (JavascriptException, WebDriverException):
        split_fields = []
    if isinstance(split_fields, list) and len(split_fields) >= 2:
        compact = re.sub(r"[\s-]+", "", postal_code)
        offset = 0
        chunks: List[str] = []
        for index, field in enumerate(split_fields):
            remaining_inputs = len(split_fields) - index
            configured_length = int(field.get("maxLength") or 0) if isinstance(field, dict) else 0
            remaining_characters = max(len(compact) - offset, 0)
            chunk_length = configured_length if configured_length > 0 else (
                (remaining_characters + remaining_inputs - 1) // remaining_inputs
            )
            chunks.append(compact[offset : offset + chunk_length])
            offset += chunk_length
        typed = False
        for field, chunk in zip(split_fields, chunks):
            if not isinstance(field, dict) or str(field.get("value") or "") == chunk:
                continue
            try:
                split_input = driver.find_element(By.CSS_SELECTOR, f"#{field['id']}")
                split_input.clear()
                type_text = getattr(split_input, "type_text", None)
                if callable(type_text):
                    type_text(chunk)
                else:
                    split_input.send_keys(chunk)
                typed = True
            except (AttributeError, KeyError, NoSuchElementException, WebDriverException):
                split_fields = []
                break
        if split_fields and typed:
            return False

    split_input_script = r"""
/* lc_delivery_split_postal_fill */
const compact = String(arguments[0] || '').replace(/[\s-]+/g, '');
const inputs = [...document.querySelectorAll('input[id^="GLUXZipUpdateInput_"]')]
  .filter(input => input.offsetParent !== null)
  .sort((a, b) => String(a.id).localeCompare(String(b.id)));
if (inputs.length < 2 || !compact) return false;
let offset = 0;
for (let index = 0; index < inputs.length; index += 1) {
  const input = inputs[index];
  const remainingInputs = inputs.length - index;
  const configuredLength = Number(input.maxLength || 0);
  const chunkLength = configuredLength > 0
    ? configuredLength
    : Math.ceil((compact.length - offset) / remainingInputs);
  const value = compact.slice(offset, offset + chunkLength);
  offset += chunkLength;
  const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
  if (descriptor && descriptor.set) descriptor.set.call(input, value); else input.value = value;
  for (const eventName of ['input', 'change']) {
    input.dispatchEvent(new Event(eventName, {bubbles: true}));
  }
}
return offset >= compact.length;
"""
    if split_fields:
        split_ready = True
    else:
        try:
            split_ready = bool(driver.execute_script(split_input_script, postal_code))
        except (JavascriptException, WebDriverException):
            split_ready = False
    if split_ready:
        split_submit_script = r"""
/* lc_delivery_split_postal_submit */
const selectors = [
        '#GLUXZipUpdate', '[data-action="GLUXPostalInputAction"] input[type="submit"]',
        '[data-action="GLUXPostalInputAction"] button', 'button[name="glowDoneButton"]'
];
for (const selector of selectors) {
  for (const button of document.querySelectorAll(selector)) {
    if (button.offsetParent !== null) { button.click(); return true; }
  }
}
return false;
"""
        try:
            if bool(driver.execute_script(split_submit_script)):
                return True
        except (JavascriptException, WebDriverException):
            pass
        for selector in submit_selectors:
            try:
                driver.find_element(By.CSS_SELECTOR, selector).click()
                return True
            except (AttributeError, NoSuchElementException, WebDriverException):
                continue

    native_input = None
    for selector in input_selectors:
        try:
            native_input = driver.find_element(By.CSS_SELECTOR, selector)
            native_input.clear()
            type_text = getattr(native_input, "type_text", None)
            if callable(type_text):
                type_text(postal_code)
            else:
                native_input.send_keys(postal_code)
            break
        except (AttributeError, NoSuchElementException, WebDriverException):
            native_input = None
    if native_input is not None:
        for selector in submit_selectors:
            try:
                driver.find_element(By.CSS_SELECTOR, selector).click()
                return True
            except (AttributeError, NoSuchElementException, WebDriverException):
                continue

    script = r"""
/* lc_delivery_postal_submit */
const postal = arguments[0];
const inputSelectors = [
  '#GLUXZipUpdateInput', 'input[name="zipCode"]', 'input[name="postalCode"]',
  '[data-action="GLUXPostalInputAction"] input[type="text"]',
  'input[autocomplete="postal-code"]'
];
let input = null;
for (const selector of inputSelectors) {
  for (const candidate of document.querySelectorAll(selector)) {
    if (candidate.offsetParent !== null) { input = candidate; break; }
  }
  if (input) break;
}
if (!input) return false;
const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
if (descriptor && descriptor.set) descriptor.set.call(input, postal); else input.value = postal;
for (const eventName of ['input', 'change']) input.dispatchEvent(new Event(eventName, {bubbles: true}));
const submitSelectors = [
  '#GLUXZipUpdate', '[data-action="GLUXPostalInputAction"] input[type="submit"]',
  '[data-action="GLUXPostalInputAction"] button', 'button[name="glowDoneButton"]'
];
for (const selector of submitSelectors) {
  for (const button of document.querySelectorAll(selector)) {
    if (button.offsetParent !== null) { button.click(); return true; }
  }
}
if (input.form) { input.form.requestSubmit ? input.form.requestSubmit() : input.form.submit(); return true; }
return false;
"""
    try:
        return bool(driver.execute_script(script, postal_code))
    except (JavascriptException, WebDriverException):
        return False


def _submit_delivery_city(driver: WebDriver, city: str) -> bool:
    city_with_district_script = r"""
/* lc_delivery_city_with_district_submit */
const city = String(arguments[0] || '');
const norm = value => String(value || '').replace(/\s+/g, ' ').trim().toLowerCase();
for (const selector of [
  '#GLUXConfirmClose', '[data-action="GLUXConfirmAction"] button',
  'button[name="glowDoneButton"]'
]) {
  for (const button of document.querySelectorAll(selector)) {
    if (button.offsetParent !== null) { button.click(); return 'submitted'; }
  }
}
const input = [...document.querySelectorAll('#GLUXCityWithDistrictCityInput')]
  .find(el => el.offsetParent !== null);
if (!input) return 'not_applicable';
if (norm(input.value) !== norm(city)) return 'needs_typing';
const suggestion = [...document.querySelectorAll('#GLUXCityWithDistrictCityList li')]
  .find(el => el.offsetParent !== null && norm(el.textContent) === norm(city));
if (suggestion) { suggestion.click(); return 'waiting'; }
for (const button of document.querySelectorAll('#GLUXCityWithDistrictApplyButton')) {
  if (button.offsetParent !== null) { button.click(); return 'submitted'; }
}
return 'waiting';
"""
    try:
        city_with_district_status = str(
            driver.execute_script(city_with_district_script, city) or ""
        )
    except (JavascriptException, WebDriverException):
        city_with_district_status = ""
    if city_with_district_status == "submitted":
        return True
    if city_with_district_status == "needs_typing":
        try:
            city_input = driver.find_element(
                By.CSS_SELECTOR,
                "#GLUXCityWithDistrictCityInput",
            )
            city_input.clear()
            type_text = getattr(city_input, "type_text", None)
            if callable(type_text):
                type_text(city)
            else:
                city_input.send_keys(city)
        except (AttributeError, NoSuchElementException, WebDriverException):
            pass
        return False
    if city_with_district_status == "waiting":
        return False

    script = r"""
/* lc_delivery_city_submit */
const city = String(arguments[0] || '');
const norm = value => String(value || '').replace(/\s+/g, ' ').trim().toLowerCase();
const target = norm(city);
const clickable = [...document.querySelectorAll('button, a, [role="button"], [role="option"], label')]
  .find(el => el.offsetParent !== null && norm(el.innerText || el.textContent || el.getAttribute('aria-label')).includes(target));
if (clickable) { clickable.click(); return true; }
const inputSelectors = [
  'input[name*="city" i]', 'input[placeholder*="city" i]', 'input[aria-label*="city" i]',
  '[role="dialog"] input[type="text"]'
];
for (const selector of inputSelectors) {
  for (const input of document.querySelectorAll(selector)) {
    if (input.offsetParent === null) continue;
    const descriptor = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
    if (descriptor && descriptor.set) descriptor.set.call(input, city); else input.value = city;
    for (const eventName of ['input', 'change']) input.dispatchEvent(new Event(eventName, {bubbles: true}));
    input.dispatchEvent(new KeyboardEvent('keydown', {key: 'Enter', code: 'Enter', bubbles: true}));
    input.dispatchEvent(new KeyboardEvent('keyup', {key: 'Enter', code: 'Enter', bubbles: true}));
    return true;
  }
}
return false;
"""
    try:
        return bool(driver.execute_script(script, city))
    except (JavascriptException, WebDriverException):
        return False


def _dismiss_delivery_dialog(driver: WebDriver) -> None:
    script = r"""
/* lc_delivery_dismiss */
for (const selector of ['#GLUXConfirmClose', 'button[name="glowDoneButton"]', '[data-action="GLUXConfirmAction"] button']) {
  for (const button of document.querySelectorAll(selector)) {
    if (button.offsetParent !== null) { button.click(); return true; }
  }
}
return false;
"""
    try:
        driver.execute_script(script)
    except (JavascriptException, WebDriverException):
        pass


def handle_amazon_verification(
    driver: WebDriver,
    runtime: Any,
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
) -> None:
    _raise_if_stop_requested(stop_event)
    reason = detect_block(driver)
    if not reason:
        return
    if reason in {"amazon_rate_limit", "access_denied"}:
        # These are transient page-health failures. The shared recovery
        # controller applies the configured long backoff; they are not manual
        # CAPTCHA tasks.
        return
    if on_manual_pause:
        on_manual_pause(reason, str(getattr(driver, "current_url", "") or ""))
    timeout = int(getattr(runtime, "manual_pause_timeout", 900) or 900)
    cleared = (
        wait_for_manual_clear(driver, reason, timeout)
        if stop_event is None
        else wait_for_manual_clear(driver, reason, timeout, stop_event=stop_event)
    )
    if not cleared:
        raise VerificationUnconfirmedError(verification_unconfirmed_message(reason))
    if on_manual_resume:
        on_manual_resume()


def _delivery_cache(runtime: Any, driver: WebDriver) -> set[str]:
    try:
        cache = getattr(driver, "_lc_delivery_location_cache", None)
        if not isinstance(cache, set):
            cache = set()
            setattr(driver, "_lc_delivery_location_cache", cache)
        return cache
    except Exception:
        cache = getattr(runtime, "_lc_delivery_location_cache", None)
        if not isinstance(cache, dict):
            cache = {}
            setattr(runtime, "_lc_delivery_location_cache", cache)
        return cache.setdefault(id(driver), set())


def _wait_for_delivery_confirmation(
    driver: WebDriver,
    location: Dict[str, str],
    deadline: float,
    allow_reloaded_city_input: bool = False,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    while time.time() < deadline:
        _raise_if_stop_requested(stop_event)
        if _submitted_delivery_location_is_confirmed(
            driver,
            location,
            allow_reloaded_city_input=allow_reloaded_city_input,
        ):
            return True
        if allow_reloaded_city_input and location.get("strategy") == "postal_then_city":
            _click_delivery_trigger(driver)
        _sleep_with_stop(0.25, stop_event)
    _raise_if_stop_requested(stop_event)
    return _submitted_delivery_location_is_confirmed(
        driver,
        location,
        allow_reloaded_city_input=allow_reloaded_city_input,
    )


def _attempt_delivery_value(
    driver: WebDriver,
    value: str,
    deadline: float,
    submitter: Any,
    stop_event: Optional[threading.Event] = None,
) -> Optional[bool]:
    while time.time() < deadline:
        _raise_if_stop_requested(stop_event)
        _click_delivery_trigger(driver)
        submitted = submitter(driver, value)
        if submitted is None:
            return None
        if submitted:
            return True
        _sleep_with_stop(0.25, stop_event)
    return False


def _reopen_amazon_target(
    driver: WebDriver,
    target_url: str,
    runtime: Any,
    on_manual_pause: Optional[Any],
    on_manual_resume: Optional[Any],
    clear_transient_city_input: bool = False,
    stop_event: Optional[threading.Event] = None,
    before_navigation: Optional[Callable[[], None]] = None,
    page_health_validator: Optional[Callable[[], Any]] = None,
) -> bool:
    city_input_reset = (
        _clear_transient_delivery_city_input(driver)
        if clear_transient_city_input
        else False
    )
    last_error: Optional[WebDriverException] = None
    for attempt in range(3):
        _raise_if_stop_requested(stop_event)
        try:
            if before_navigation is not None:
                before_navigation()
            driver.get(target_url)
            last_error = None
            break
        except WebDriverException as exc:
            if "ERR_ABORTED" not in str(exc).upper():
                assessment = category_page_assessment(
                    driver,
                    navigation_error=str(exc),
                )
                raise TransientAmazonPageUnavailable.from_assessment(
                    assessment,
                    url=safe_driver_current_url(driver, target_url),
                ) from exc
            last_error = exc
            _sleep_with_stop(0.75 * (attempt + 1), stop_event)
    if last_error is not None:
        assessment = category_page_assessment(
            driver,
            navigation_error=str(last_error),
        )
        raise TransientAmazonPageUnavailable.from_assessment(
            assessment,
            url=safe_driver_current_url(driver, target_url),
        ) from last_error
    try:
        delattr(driver, "_lc_delivery_trigger_opened")
    except (AttributeError, TypeError):
        pass
    handle_amazon_verification(
        driver,
        runtime,
        on_manual_pause,
        on_manual_resume,
        stop_event=stop_event,
    )
    if page_health_validator is not None:
        page_health_validator()
    else:
        assessment = category_page_assessment(driver)
        if (
            assessment.status is PageHealthStatus.TRANSIENT_UNAVAILABLE
            and assessment.reason != "expected_content_missing"
        ):
            raise TransientAmazonPageUnavailable.from_assessment(
                assessment,
                url=safe_driver_current_url(driver, target_url),
            )
    return city_input_reset


def ensure_amazon_delivery_location(
    driver: WebDriver,
    runtime: Any,
    original_url: str = "",
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    before_navigation: Optional[Callable[[], None]] = None,
    page_health_validator: Optional[Callable[[], Any]] = None,
) -> None:
    _raise_if_stop_requested(stop_event)
    if not bool(getattr(runtime, "delivery_location_enabled", True)):
        return
    locations = _runtime_delivery_locations(runtime)
    current_url = safe_driver_current_url(driver, original_url)
    domain = amazon_marketplace_domain(current_url or original_url, locations)
    if not domain:
        raise DeliveryLocationUnconfirmedError(
            f"delivery_location_unsupported: 当前 Amazon 站点没有配送地址映射：{current_url or original_url}"
        )
    location = locations[domain]
    fingerprint = str(getattr(runtime, "delivery_location_fingerprint", "") or "")
    cache_key = f"{fingerprint}:{domain}"
    cache = _delivery_cache(runtime, driver)
    if cache_key in cache:
        return
    if delivery_location_is_confirmed(driver, location):
        cache.add(cache_key)
        return

    target_url = original_url or current_url
    timeout = max(int(getattr(runtime, "delivery_location_timeout", 20) or 20), 1)

    def attempt_value(value: str, deadline: float, submitter: Any) -> Optional[bool]:
        if stop_event is None:
            return _attempt_delivery_value(driver, value, deadline, submitter)
        return _attempt_delivery_value(
            driver,
            value,
            deadline,
            submitter,
            stop_event=stop_event,
        )

    def wait_confirmation(
        deadline: float,
        allow_reloaded_city_input: bool = False,
    ) -> bool:
        if stop_event is None:
            return _wait_for_delivery_confirmation(
                driver,
                location,
                deadline,
                allow_reloaded_city_input=allow_reloaded_city_input,
            )
        return _wait_for_delivery_confirmation(
            driver,
            location,
            deadline,
            allow_reloaded_city_input=allow_reloaded_city_input,
            stop_event=stop_event,
        )

    def reopen_target(clear_transient_city_input: bool) -> bool:
        if not target_url:
            return False
        if stop_event is None:
            return _reopen_amazon_target(
                driver,
                target_url,
                runtime,
                on_manual_pause,
                on_manual_resume,
                clear_transient_city_input=clear_transient_city_input,
                before_navigation=before_navigation,
                page_health_validator=page_health_validator,
            )
        return _reopen_amazon_target(
            driver,
            target_url,
            runtime,
            on_manual_pause,
            on_manual_resume,
            clear_transient_city_input=clear_transient_city_input,
            stop_event=stop_event,
            before_navigation=before_navigation,
            page_health_validator=page_health_validator,
        )

    def validate_page_before_manual_delivery() -> None:
        if page_health_validator is not None:
            page_health_validator()
            return
        assessment = category_page_assessment(driver)
        if (
            assessment.status is PageHealthStatus.TRANSIENT_UNAVAILABLE
            and assessment.reason != "expected_content_missing"
        ):
            raise TransientAmazonPageUnavailable.from_assessment(
                assessment,
                url=safe_driver_current_url(driver, target_url),
            )

    for postal_code in delivery_postal_candidates(location["postal_code"]):
        deadline = time.time() + timeout
        submitted = attempt_value(
            postal_code,
            deadline,
            lambda current_driver, value: _submit_delivery_postal(
                current_driver,
                value,
                location["city"],
            ),
        )
        if submitted is None:
            break
        if not submitted:
            continue
        ready_deadline = min(deadline, time.time() + 5)
        wait_confirmation(ready_deadline)
        _dismiss_delivery_dialog(driver)
        time.sleep(min(0.75, max(deadline - time.time(), 0)))
        city_input_reloaded = False
        if target_url:
            city_input_reloaded = reopen_target(
                location.get("strategy") == "postal_then_city"
            )
        if wait_confirmation(
            time.time() + timeout,
            allow_reloaded_city_input=city_input_reloaded,
        ):
            cache.add(cache_key)
            return

    if location.get("strategy") == "postal_then_city":
        deadline = time.time() + timeout
        submitted = attempt_value(
            location["city"],
            deadline,
            _submit_delivery_city,
        )
        if submitted:
            ready_deadline = min(deadline, time.time() + 5)
            wait_confirmation(ready_deadline)
            _dismiss_delivery_dialog(driver)
            time.sleep(min(0.75, max(deadline - time.time(), 0)))
            city_input_reloaded = False
            if target_url:
                city_input_reloaded = reopen_target(True)
            if wait_confirmation(
                time.time() + timeout,
                allow_reloaded_city_input=city_input_reloaded,
            ):
                cache.add(cache_key)
                return

    validate_page_before_manual_delivery()
    reason = "delivery_location_unconfirmed"
    if on_manual_pause:
        on_manual_pause(reason, safe_driver_current_url(driver, target_url))
    print(
        f"自动设置配送地址失败。请在当前 Chrome 页面手动设置 {domain} 配送地址为 "
        f"{location['city']} / {location['postal_code']}。"
    )
    manual_deadline = time.time() + max(int(getattr(runtime, "manual_pause_timeout", 900) or 900), 1)
    while time.time() < manual_deadline:
        remaining = max(int(manual_deadline - time.time()), 1)
        continued = (
            wait_for_manual_continue(remaining)
            if stop_event is None
            else wait_for_manual_continue(remaining, stop_event=stop_event)
        )
        if not continued:
            break
        if delivery_location_is_confirmed(driver, location):
            if on_manual_resume:
                on_manual_resume()
            cache.add(cache_key)
            return
        if target_url:
            city_input_reloaded = reopen_target(
                location.get("strategy") == "postal_then_city"
            )
            confirmation_deadline = min(manual_deadline, time.time() + timeout)
            if wait_confirmation(
                confirmation_deadline,
                allow_reloaded_city_input=city_input_reloaded,
            ):
                if on_manual_resume:
                    on_manual_resume()
                cache.add(cache_key)
                return
        validate_page_before_manual_delivery()
        print("配送地址仍未确认，请继续在当前 Chrome 页面处理。")
    raise DeliveryLocationUnconfirmedError(
        "delivery_location_unconfirmed: 配送地址人工确认超时，任务已停止且未提取当前页数据。"
    )


def open_amazon_page(
    driver: WebDriver,
    url: str,
    runtime: Any,
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    defer_delivery: bool = False,
) -> None:
    _raise_if_stop_requested(stop_event)
    driver.get(url)
    handle_amazon_verification(
        driver,
        runtime,
        on_manual_pause,
        on_manual_resume,
        stop_event=stop_event,
    )
    if defer_delivery:
        return
    ensure_amazon_delivery_location(
        driver,
        runtime,
        original_url=url,
        on_manual_pause=on_manual_pause,
        on_manual_resume=on_manual_resume,
        stop_event=stop_event,
    )


def wait_for_amazon_products(
    driver: WebDriver,
    runtime: RuntimeConfig,
    stop_event: Optional[threading.Event] = None,
) -> None:
    selectors = [
        "#gridItemRoot",
        ".zg-grid-general-faceout",
        ".p13n-grid-content",
        "[data-asin]:not([data-asin=''])",
        ".s-result-item[data-asin]:not([data-asin=''])",
    ]
    combined_selector = ",".join(selectors)
    if stop_event is None:
        condition = EC.presence_of_element_located((By.CSS_SELECTOR, combined_selector))
        WebDriverWait(driver, runtime.page_timeout).until(condition)
        return

    deadline = time.monotonic() + max(float(runtime.page_timeout), 0.1)
    while time.monotonic() < deadline:
        _raise_if_stop_requested(stop_event)
        try:
            driver.find_element(By.CSS_SELECTOR, combined_selector)
            return
        except NoSuchElementException:
            pass
        _sleep_with_stop(0.25, stop_event)
    raise TimeoutException("等待 Amazon 商品卡片超时。")


CATEGORY_PRODUCT_SELECTORS = (
    "#gridItemRoot",
    ".zg-grid-general-faceout",
    "[data-asin]:not([data-asin=''])",
    ".s-result-item[data-asin]:not([data-asin=''])",
)

CATEGORY_EXPLICIT_EMPTY_MARKERS = (
    "there are no products available in this category",
    "there are no products in this category",
    "no results for",
    "did not match any products",
    "we couldn't find any results",
    "we couldn’t find any results",
    "currently no listings available",
    "此分类中没有可用的商品",
    "没有找到任何商品",
    "没有符合条件的商品",
)


def category_expected_content_present(driver: WebDriver) -> bool:
    try:
        driver.find_element(By.CSS_SELECTOR, ",".join(CATEGORY_PRODUCT_SELECTORS))
        return True
    except (NoSuchElementException, WebDriverException, AttributeError):
        return False


def category_explicit_empty_present(driver: WebDriver) -> bool:
    text = safe_find_text(driver).lower()
    return any(marker in text for marker in CATEGORY_EXPLICIT_EMPTY_MARKERS)


def _category_http_status(driver: WebDriver, title: str, body_text: str) -> Optional[int]:
    for attribute in ("last_http_status", "_last_http_status", "http_status"):
        try:
            value = getattr(driver, attribute)
        except (AttributeError, WebDriverException):
            continue
        if isinstance(value, int) and not isinstance(value, bool) and 100 <= value <= 599:
            return value
    haystack = normalize_space(f"{title} {body_text}").lower()
    if "too many requests" in haystack:
        return 429
    status_markers = {
        500: ("internal server error", "500 server error"),
        502: ("bad gateway",),
        503: ("service unavailable",),
        504: ("gateway timeout",),
    }
    for status, markers in status_markers.items():
        if any(marker in haystack for marker in markers):
            return status
    return None


def category_page_assessment(
    driver: WebDriver,
    *,
    navigation_error: str = "",
) -> PageHealthAssessment:
    """Capture and classify a category page without mutating browser state."""

    try:
        current_url = str(getattr(driver, "current_url", "") or "")
    except WebDriverException:
        current_url = ""
    try:
        title = str(getattr(driver, "title", "") or "")
    except WebDriverException:
        title = ""
    body_text = safe_find_text(driver)
    expected_content_present = category_expected_content_present(driver)
    explicit_empty = (
        False
        if expected_content_present
        else category_explicit_empty_present(driver)
    )
    return classify_page_snapshot(
        PageSnapshot(
            page_kind="search_category",
            url=current_url,
            title=title,
            body_text=body_text,
            http_status=(
                None
                if expected_content_present
                else _category_http_status(driver, title, body_text)
            ),
            navigation_error=navigation_error,
            expected_content_present=expected_content_present,
            explicit_empty=explicit_empty,
        )
    )


def safe_driver_current_url(driver: WebDriver, fallback: str = "") -> str:
    """Read a remote browser URL without masking the exception being handled."""

    try:
        return str(getattr(driver, "current_url", "") or fallback)
    except Exception:
        return str(fallback or "")


def _handle_assessment_interaction(
    driver: WebDriver,
    runtime: RuntimeConfig,
    assessment: PageHealthAssessment,
    on_manual_pause: Optional[Any],
    on_manual_resume: Optional[Any],
    stop_event: Optional[threading.Event],
) -> None:
    if assessment.status is PageHealthStatus.AMAZON_SIGN_IN:
        raise VerificationUnconfirmedError(
            verification_unconfirmed_message("amazon_sign_in")
        )
    if assessment.status is not PageHealthStatus.INTERACTIVE_VERIFICATION:
        return
    handle_amazon_verification(
        driver,
        runtime,
        on_manual_pause,
        on_manual_resume,
        stop_event=stop_event,
    )


def wait_for_category_page_health(
    driver: WebDriver,
    runtime: RuntimeConfig,
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
) -> PageHealthAssessment:
    """Wait for a product DOM or an explicit empty marker, then classify."""

    assessment = category_page_assessment(driver)
    _handle_assessment_interaction(
        driver,
        runtime,
        assessment,
        on_manual_pause,
        on_manual_resume,
        stop_event,
    )
    if assessment.status is PageHealthStatus.INTERACTIVE_VERIFICATION:
        assessment = category_page_assessment(driver)
    if assessment.status in {
        PageHealthStatus.HEALTHY,
        PageHealthStatus.VERIFIED_EMPTY,
    }:
        return assessment
    # Strong dog/error/rate-limit/access-denied signatures should back off
    # immediately. A blank or still-building DOM gets the configured timeout.
    if assessment.reason not in {"blank_page", "expected_content_missing"}:
        raise TransientAmazonPageUnavailable.from_assessment(
            assessment,
            url=safe_driver_current_url(driver),
        )
    try:
        wait_for_amazon_products(driver, runtime, stop_event=stop_event)
    except TimeoutException:
        pass
    assessment = category_page_assessment(driver)
    _handle_assessment_interaction(
        driver,
        runtime,
        assessment,
        on_manual_pause,
        on_manual_resume,
        stop_event,
    )
    if assessment.status is PageHealthStatus.INTERACTIVE_VERIFICATION:
        assessment = category_page_assessment(driver)
    if assessment.status in {
        PageHealthStatus.HEALTHY,
        PageHealthStatus.VERIFIED_EMPTY,
    }:
        return assessment
    raise TransientAmazonPageUnavailable.from_assessment(
        assessment,
        url=safe_driver_current_url(driver),
    )


def category_owned_handle_snapshot(driver: WebDriver) -> Optional[FrozenSet[str]]:
    snapshot = getattr(driver, "owned_handle_snapshot", None)
    if not callable(snapshot):
        return None
    try:
        return frozenset(str(handle) for handle in snapshot())
    except WebDriverException:
        return None


def close_category_owned_since(
    driver: WebDriver,
    snapshot: Optional[FrozenSet[str]],
) -> None:
    """Close only event-owned popups and always restore the dedicated worker."""

    if snapshot is None:
        restore_worker = getattr(driver, "restore_worker_page", None)
        if callable(restore_worker):
            try:
                restore_worker()
            except WebDriverException as exc:
                print(f"恢复类目 worker 标签页失败：{exc}", file=sys.stderr)
        return
    close_owned = getattr(driver, "close_owned_since", None)
    restore_worker = getattr(driver, "restore_worker_page", None)
    try:
        if callable(close_owned):
            close_owned(snapshot)
    except WebDriverException as exc:
        print(f"清理 crawler-owned 弹窗失败，已保留诊断记录：{exc}", file=sys.stderr)
    finally:
        if callable(restore_worker):
            try:
                restore_worker()
            except WebDriverException as exc:
                print(f"恢复类目 worker 标签页失败：{exc}", file=sys.stderr)


def _runtime_recovery_clock(runtime: Any) -> Callable[[], float]:
    value = getattr(runtime, "_amazon_page_retry_clock", None)
    return value if callable(value) else time.time


def _runtime_recovery_rng(runtime: Any) -> Any:
    return getattr(runtime, "_amazon_page_retry_rng", None)


def _runtime_recovery_waiter(
    runtime: Any,
    stop_event: Optional[threading.Event],
) -> Callable[[float], Any]:
    value = getattr(runtime, "_amazon_page_retry_waiter", None)
    if callable(value):
        return value
    return lambda seconds: _sleep_with_stop(seconds, stop_event)


def _retry_entry_key(work_key: str, stage: str) -> str:
    return f"{work_key}|stage:{stage}"


def state_retry_callbacks(
    state: StateStore,
    retry_key: str,
) -> RetryCallbacks:
    return RetryCallbacks(
        load_state=lambda: state.load_amazon_page_retry(retry_key),
        write_state=lambda value: state.write_amazon_page_retry(retry_key, value),
        clear_state=lambda: state.clear_amazon_page_retry(retry_key),
    )


def run_category_page_work_with_recovery(
    runtime: RuntimeConfig,
    page_url: str,
    work_key: str,
    *,
    retry_callbacks: RetryCallbacks,
    driver_provider: Callable[[], WebDriver],
    operation: Callable[[Any], CategoryPageWorkResult],
    stop_event: Optional[threading.Event] = None,
    domain_cooldowns: Optional[DomainCooldownRegistry] = None,
    stage: str = "category_page_work",
) -> CategoryPageWorkResult:
    """Retry navigation, plugin inspection, extraction and pagination as one stage.

    The operation must not commit state or output files. Its result is committed
    by the caller only after this controller succeeds, so a retry can never
    duplicate a page shard.
    """

    domain = (urlparse(page_url).hostname or "unknown").lower()
    attempt_snapshot: List[Optional[FrozenSet[str]]] = [None]
    base_callbacks = retry_callbacks

    def cleanup() -> None:
        try:
            current_driver = driver_provider()
        except BaseException:
            current_driver = None
        if current_driver is not None:
            close_category_owned_since(current_driver, attempt_snapshot[0])
        attempt_snapshot[0] = None
        base_callbacks.cleanup()

    def begin_domain_cooldown(current_domain: str, deadline: float) -> None:
        if domain_cooldowns is not None:
            domain_cooldowns.extend(current_domain, deadline)
        base_callbacks.begin_domain_cooldown(current_domain, deadline)

    def end_domain_cooldown(current_domain: str, deadline: float) -> None:
        if domain_cooldowns is not None:
            domain_cooldowns.release(current_domain, deadline)
        base_callbacks.end_domain_cooldown(current_domain, deadline)

    def heartbeat(value: Mapping[str, Any]) -> None:
        base_callbacks.heartbeat(value)
        remaining = max(float(value.get("remaining_wait_seconds") or 0), 0.0)
        print(
            f"Amazon 页面处理暂不可用；{work_key} 将在约 {remaining:.0f} 秒后重试。",
            flush=True,
        )

    callbacks = RetryCallbacks(
        load_state=base_callbacks.load_state,
        write_state=base_callbacks.write_state,
        clear_state=base_callbacks.clear_state,
        cleanup=cleanup,
        begin_domain_cooldown=begin_domain_cooldown,
        end_domain_cooldown=end_domain_cooldown,
        heartbeat=heartbeat,
    )

    def guarded(attempt: Any) -> CategoryPageWorkResult:
        _raise_if_stop_requested(stop_event)
        current_driver = driver_provider()
        attempt_snapshot[0] = category_owned_handle_snapshot(current_driver)
        try:
            return operation(attempt)
        except TransientAmazonPageUnavailable:
            raise
        except (TimeoutException, WebDriverException) as exc:
            reason = (
                "page_timeout"
                if isinstance(exc, TimeoutException)
                else "webdriver_error"
            )
            raise TransientAmazonPageUnavailable(
                str(exc) or reason,
                reason=reason,
                url=safe_driver_current_url(current_driver, page_url),
            ) from exc

    controller = AmazonPageRetryController(
        domain=domain,
        work_key=work_key,
        stage=stage,
        url=page_url,
        schedule=getattr(
            runtime,
            "amazon_page_retry_schedule",
            DEFAULT_RETRY_SCHEDULE_SECONDS,
        ),
        callbacks=callbacks,
        clock=_runtime_recovery_clock(runtime),
        rng=_runtime_recovery_rng(runtime),
        waiter=_runtime_recovery_waiter(runtime, stop_event),
    )
    return controller.run(guarded)


def load_category_page_attempt(
    driver: WebDriver,
    page_url: str,
    runtime: RuntimeConfig,
    *,
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    before_navigation: Optional[Callable[[], None]] = None,
    delivery_lock: Optional[threading.Lock] = None,
    domain_cooldowns: Optional[DomainCooldownRegistry] = None,
) -> PageHealthAssessment:
    """Perform one uncommitted category navigation/health/delivery attempt."""

    domain = (urlparse(page_url).hostname or "unknown").lower()
    if domain_cooldowns is not None:
        domain_cooldowns.wait(domain)
    _raise_if_stop_requested(stop_event)
    if before_navigation is not None:
        before_navigation()

    def before_delivery_navigation() -> None:
        if domain_cooldowns is not None:
            domain_cooldowns.wait(domain)
        _raise_if_stop_requested(stop_event)
        if before_navigation is not None:
            before_navigation()

    def navigate_with_delivery() -> PageHealthAssessment:
        try:
            driver.get(page_url)
        except (TimeoutException, WebDriverException) as exc:
            assessment = category_page_assessment(
                driver,
                navigation_error=str(exc),
            )
            _handle_assessment_interaction(
                driver,
                runtime,
                assessment,
                on_manual_pause,
                on_manual_resume,
                stop_event,
            )
            if assessment.status is PageHealthStatus.INTERACTIVE_VERIFICATION:
                assessment = category_page_assessment(
                    driver,
                    navigation_error=str(exc),
                )
            raise TransientAmazonPageUnavailable.from_assessment(
                assessment,
                url=safe_driver_current_url(driver, page_url),
            ) from exc

        handle_amazon_verification(
            driver,
            runtime,
            on_manual_pause,
            on_manual_resume,
            stop_event=stop_event,
        )

        def validate_category_page() -> PageHealthAssessment:
            return wait_for_category_page_health(
                driver,
                runtime,
                on_manual_pause,
                on_manual_resume,
                stop_event,
            )

        assessment = validate_category_page()
        ensure_amazon_delivery_location(
            driver,
            runtime,
            original_url=page_url,
            on_manual_pause=on_manual_pause,
            on_manual_resume=on_manual_resume,
            stop_event=stop_event,
            before_navigation=before_delivery_navigation,
            page_health_validator=validate_category_page,
        )
        # Delivery selection can reload the target. Revalidate before any
        # discovery, plugin wait, extraction, or empty-page commit.
        return validate_category_page()

    if delivery_lock is None:
        return navigate_with_delivery()
    with delivery_lock:
        return navigate_with_delivery()


def load_category_page_with_recovery(
    driver: WebDriver,
    page_url: str,
    runtime: RuntimeConfig,
    work_key: str,
    *,
    retry_callbacks: RetryCallbacks,
    stage: str = "category_page",
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
    before_navigation: Optional[Callable[[], None]] = None,
    delivery_lock: Optional[threading.Lock] = None,
    domain_cooldowns: Optional[DomainCooldownRegistry] = None,
) -> PageHealthAssessment:
    """Navigate and validate one category page with the shared five-attempt policy."""

    domain = (urlparse(page_url).hostname or "unknown").lower()
    attempt_snapshot: List[Optional[FrozenSet[str]]] = [None]
    base_callbacks = retry_callbacks

    def cleanup() -> None:
        close_category_owned_since(driver, attempt_snapshot[0])
        attempt_snapshot[0] = None
        base_callbacks.cleanup()

    def begin_domain_cooldown(current_domain: str, deadline: float) -> None:
        if domain_cooldowns is not None:
            domain_cooldowns.extend(current_domain, deadline)
        base_callbacks.begin_domain_cooldown(current_domain, deadline)

    def end_domain_cooldown(current_domain: str, deadline: float) -> None:
        if domain_cooldowns is not None:
            domain_cooldowns.release(current_domain, deadline)
        base_callbacks.end_domain_cooldown(current_domain, deadline)

    def heartbeat(value: Mapping[str, Any]) -> None:
        base_callbacks.heartbeat(value)
        remaining = max(float(value.get("remaining_wait_seconds") or 0), 0.0)
        print(
            f"Amazon 页面暂不可用；{work_key} 将在约 {remaining:.0f} 秒后重试。",
            flush=True,
        )

    callbacks = RetryCallbacks(
        load_state=base_callbacks.load_state,
        write_state=base_callbacks.write_state,
        clear_state=base_callbacks.clear_state,
        cleanup=cleanup,
        begin_domain_cooldown=begin_domain_cooldown,
        end_domain_cooldown=end_domain_cooldown,
        heartbeat=heartbeat,
    )

    def navigate_and_validate(_attempt: Any) -> PageHealthAssessment:
        _raise_if_stop_requested(stop_event)
        attempt_snapshot[0] = category_owned_handle_snapshot(driver)
        return load_category_page_attempt(
            driver,
            page_url,
            runtime,
            on_manual_pause=on_manual_pause,
            on_manual_resume=on_manual_resume,
            stop_event=stop_event,
            before_navigation=before_navigation,
            delivery_lock=delivery_lock,
            domain_cooldowns=domain_cooldowns,
        )

    schedule = getattr(
        runtime,
        "amazon_page_retry_schedule",
        DEFAULT_RETRY_SCHEDULE_SECONDS,
    )
    controller = AmazonPageRetryController(
        domain=domain,
        work_key=work_key,
        stage=stage,
        url=page_url,
        schedule=schedule,
        callbacks=callbacks,
        clock=_runtime_recovery_clock(runtime),
        rng=_runtime_recovery_rng(runtime),
        waiter=_runtime_recovery_waiter(runtime, stop_event),
    )
    return controller.run(navigate_and_validate)


def try_activate_plugin(driver: WebDriver) -> bool:
    script = r"""
const words = ['Product Research', 'SellerSprite', '卖家精灵', '产品调研', 'BSR'];
const pluginRootSelector = [
  '[id*="sellersprite" i]',
  '[class*="sellersprite" i]',
  '[id*="seller-sprite" i]',
  '[class*="seller-sprite" i]',
  '[id*="__ss" i]',
  '[class*="ss-" i]'
].join(',');
const isVisible = (el) => {
  const rect = el.getBoundingClientRect();
  const style = window.getComputedStyle(el);
  return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
};
const roots = [...document.querySelectorAll(pluginRootSelector)];
const candidates = roots.flatMap(root => [
  ...(root.matches('button,[role="button"]') ? [root] : []),
  ...root.querySelectorAll('button,[role="button"]')
]).filter(el => {
  const text = (el.innerText || el.textContent || '').trim();
  if (!text || text.length > 80 || !isVisible(el)) return false;
  return words.some(word => text.includes(word));
});
for (const el of candidates.slice(0, 5)) {
  try {
    el.click();
    return true;
  } catch (err) {}
}
return false;
"""
    try:
        return bool(driver.execute_script(script))
    except JavascriptException:
        return False


def plugin_node_count(driver: WebDriver) -> int:
    script = r"""
const selector = [
  '[id*="sellersprite" i]',
  '[class*="sellersprite" i]',
  '[id*="seller-sprite" i]',
  '[class*="seller-sprite" i]',
  '[id*="__ss" i]',
  '[class*="vxe-" i]',
  '[class*="ss-" i]',
  '[class*="sprite" i]'
].join(',');
return document.querySelectorAll(selector).length;
"""
    try:
        return int(driver.execute_script(script) or 0)
    except (JavascriptException, WebDriverException):
        return 0


def sellersprite_login_required(driver: WebDriver) -> bool:
    script = r"""
const pluginSelector = [
  '[id*="sellersprite" i]',
  '[class*="sellersprite" i]',
  '[id*="seller-sprite" i]',
  '[class*="seller-sprite" i]',
  '[id*="__ss" i]',
  '[class*="vxe-" i]',
  '[class*="ss-" i]',
  '[class*="sprite" i]'
].join(',');
const phrases = [
  '请登录', '立即登录', '登录卖家精灵', '重新登录', '授权已过期',
  'login required', 'sign in to sellersprite', 'session expired',
  'please login', 'please sign in'
];
const text = [...document.querySelectorAll(pluginSelector)]
  .map(el => el.innerText || el.textContent || '')
  .join(' ')
  .replace(/\s+/g, ' ')
  .toLowerCase();
return phrases.some(phrase => text.includes(phrase.toLowerCase()));
"""
    try:
        return bool(driver.execute_script(script))
    except (JavascriptException, WebDriverException):
        return False


def classify_sellersprite_snapshot(
    snapshot: Dict[str, Any],
    min_enriched_records: int,
    min_fields_per_record: int,
) -> str:
    if snapshot.get("blocked"):
        return "blocked"
    if int(snapshot.get("plugin_nodes") or 0) <= 0:
        return "plugin_absent"
    if bool(snapshot.get("login_required")):
        return "login_required"
    if int(snapshot.get("product_count") or 0) <= 0:
        return "data_loading"
    if (
        int(snapshot.get("enriched_records") or 0) >= max(min_enriched_records, 1)
        and int(snapshot.get("max_fields_per_record") or 0) >= max(min_fields_per_record, 1)
    ):
        return "ready_candidate"
    return "data_loading"


def set_sellersprite_readiness(driver: WebDriver, report: Dict[str, Any]) -> None:
    try:
        setattr(driver, "_sellersprite_readiness", dict(report))
    except Exception:
        pass


def get_sellersprite_readiness(driver: WebDriver) -> Dict[str, Any]:
    try:
        value = getattr(driver, "_sellersprite_readiness", {})
    except Exception:
        value = {}
    return dict(value) if isinstance(value, dict) else {}


def safe_sellersprite_readiness(report: Dict[str, Any]) -> Dict[str, Any]:
    allowed = {
        "status",
        "checked_at",
        "page_url",
        "plugin_nodes",
        "login_required",
        "product_count",
        "enriched_records",
        "max_fields_per_record",
        "stable_checks",
        "blocked",
        "blocked_reason",
    }
    return {key: report.get(key) for key in allowed if key in report}


def inspect_sellersprite_readiness(driver: WebDriver, runtime: RuntimeConfig) -> Dict[str, Any]:
    blocked_reason = detect_block(driver)
    cards = extract_product_cards(driver)
    rows = extract_table_rows(driver)
    product_asins = {
        normalize_space(str(item.get("asin") or "")).upper()
        for item in [*cards, *rows]
        if normalize_space(str(item.get("asin") or ""))
    }
    field_counts_by_asin: Dict[str, int] = {}
    bsr_ranks_by_asin: Dict[str, List[Dict[str, Any]]] = {}
    fulfillment_by_asin: Dict[str, str] = {}
    table_fulfillment_by_asin: Dict[str, Tuple[str, str]] = {}
    for row in rows:
        asin = normalize_space(str(row.get("asin") or "")).upper()
        if not asin:
            continue
        parsed = parse_table_row_fields(row)
        table_evidence = (
            parsed.get("fulfillment_method"),
            parsed.get("fulfillment_method_raw"),
        )
        table_fulfillment_by_asin[asin] = select_fulfillment_evidence(
            table_fulfillment_by_asin.get(asin, ("", "")),
            table_evidence,
        )
        count = sum(
            1
            for field_name in SELLERSPRITE_EVIDENCE_FIELDS
            if field_name != "fulfillment_method" and parsed.get(field_name)
        ) + int(bool(table_evidence[0] or table_evidence[1]))
        field_counts_by_asin[asin] = max(field_counts_by_asin.get(asin, 0), count)
        selected_method, selected_raw = table_fulfillment_by_asin[asin]
        if selected_method or selected_raw:
            fulfillment_by_asin[asin] = selected_method or f"raw:{selected_raw}"
        row_bsr_text = str(row.get("bsr_text") or "")
        row_ranks = parse_subcategory_bsr_ranks(
            row_bsr_text or str(row.get("text") or "")
        )
        if row_ranks:
            bsr_ranks_by_asin[asin] = normalize_subcategory_bsr_ranks(
                [*(bsr_ranks_by_asin.get(asin) or []), *row_ranks]
            )
    runtime_field_selectors = getattr(runtime, "field_selectors", {}) or {}
    fulfillment_selector_map = {}
    if runtime_field_selectors.get("fulfillment_method"):
        fulfillment_selector_map["fulfillment_method"] = runtime_field_selectors[
            "fulfillment_method"
        ]
    for card in cards:
        asin = normalize_space(str(card.get("asin") or "")).upper()
        if not asin:
            continue
        text = str(card.get("text") or "")
        selector_values = (
            extract_by_selectors(driver, card, fulfillment_selector_map)
            if fulfillment_selector_map
            else {}
        )
        selector_evidence = parse_fulfillment_evidence(
            str(selector_values.get("fulfillment_method") or ""),
            explicit_value=True,
        )
        card_evidence = parse_fulfillment_evidence(text)
        fulfillment_evidence = select_fulfillment_evidence(
            selector_evidence,
            table_fulfillment_by_asin.get(asin, ("", "")),
            card_evidence,
        )
        count = sum(
            1
            for field_name in SELLERSPRITE_EVIDENCE_FIELDS
            if field_name != "fulfillment_method"
            and parse_sellersprite_inline_field(field_name, text)
        ) + int(bool(fulfillment_evidence[0] or fulfillment_evidence[1]))
        field_counts_by_asin[asin] = max(field_counts_by_asin.get(asin, 0), count)
        fulfillment_method, fulfillment_raw = fulfillment_evidence
        if fulfillment_method or fulfillment_raw:
            fulfillment_by_asin[asin] = fulfillment_method or f"raw:{fulfillment_raw}"
        card_bsr_text = str(card.get("bsr_text") or "")
        card_ranks = parse_subcategory_bsr_ranks(card_bsr_text or text)
        if card_ranks:
            bsr_ranks_by_asin[asin] = normalize_subcategory_bsr_ranks(
                [*(bsr_ranks_by_asin.get(asin) or []), *card_ranks]
            )

    min_fields = max(int(getattr(runtime, "sellersprite_min_fields_per_record", 2) or 2), 1)
    enriched_records = sum(1 for count in field_counts_by_asin.values() if count >= min_fields)
    signature = "|".join(
        [
            str(plugin_node_count(driver)),
            str(len(product_asins)),
            str(len(rows)),
            ",".join(f"{asin}:{count}" for asin, count in sorted(field_counts_by_asin.items())),
            json.dumps(fulfillment_by_asin, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
            json.dumps(bsr_ranks_by_asin, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
        ]
    )
    report: Dict[str, Any] = {
        "status": "data_loading",
        "checked_at": now_iso(),
        "page_url": safe_driver_current_url(driver),
        "plugin_nodes": plugin_node_count(driver),
        "login_required": sellersprite_login_required(driver),
        "product_count": len(product_asins),
        "enriched_records": enriched_records,
        "max_fields_per_record": max(field_counts_by_asin.values(), default=0),
        "stable_checks": 0,
        "signature": signature,
        "blocked": bool(blocked_reason),
        "blocked_reason": blocked_reason or "",
    }
    report["status"] = classify_sellersprite_snapshot(
        report,
        int(getattr(runtime, "sellersprite_min_enriched_records", 1) or 1),
        min_fields,
    )
    return report


def wait_for_user_plugin_action(
    driver: WebDriver,
    reason: str,
    before_refresh: Optional[Any] = None,
    manual_pause_timeout: int = 900,
    stop_event: Optional[threading.Event] = None,
) -> bool:
    print(f"卖家精灵插件需要人工处理：{reason}")
    print("请在当前 Chrome 窗口安装/启用卖家精灵插件，并完成登录。")
    print("这里仅指卖家精灵账号，不是 Amazon 买家账号；不要登录 Amazon 买家账号。")
    print("处理完成后回到这里按 Enter；如果是 Codex 正在运行，请告诉我“已登录插件，继续”。")
    wait_seconds = max(int(manual_pause_timeout), 1)
    continued = (
        wait_for_manual_continue(wait_seconds)
        if stop_event is None
        else wait_for_manual_continue(wait_seconds, stop_event=stop_event)
    )
    if not continued:
        return False
    try:
        if before_refresh:
            before_refresh()
        driver.refresh()
    except WebDriverException:
        pass
    return True


def wait_for_sellersprite_data_or_prompt(
    driver: WebDriver,
    runtime: RuntimeConfig,
    on_manual_pause: Optional[Any] = None,
    on_manual_resume: Optional[Any] = None,
    restart_driver: Optional[Any] = None,
    on_readiness: Optional[Any] = None,
    before_navigation: Optional[Any] = None,
    recover_amazon_page: Optional[Any] = None,
    stop_event: Optional[threading.Event] = None,
) -> str:
    _raise_if_stop_requested(stop_event)
    if not bool(getattr(runtime, "sellersprite_required", True)):
        report = {
            "status": "not_required",
            "checked_at": now_iso(),
            "page_url": safe_driver_current_url(driver),
        }
        set_sellersprite_readiness(driver, report)
        if on_readiness:
            on_readiness(report)
        return "not_required"

    def publish(current_driver: WebDriver) -> None:
        if on_readiness:
            on_readiness(get_sellersprite_readiness(current_driver))

    def ensure_page_health(current_driver: WebDriver) -> WebDriver:
        for _check in range(2):
            assessment = category_page_assessment(current_driver)
            if assessment.status is PageHealthStatus.AMAZON_SIGN_IN:
                if on_manual_pause:
                    on_manual_pause(
                        "amazon_sign_in",
                        safe_driver_current_url(current_driver),
                    )
                raise VerificationUnconfirmedError(
                    verification_unconfirmed_message("amazon_sign_in")
                )
            if assessment.status is PageHealthStatus.INTERACTIVE_VERIFICATION:
                handle_amazon_verification(
                    current_driver,
                    runtime,
                    on_manual_pause,
                    on_manual_resume,
                    stop_event=stop_event,
                )
                continue
            if assessment.status is PageHealthStatus.TRANSIENT_UNAVAILABLE:
                if recover_amazon_page is None:
                    raise TransientAmazonPageUnavailable.from_assessment(
                        assessment,
                        url=safe_driver_current_url(current_driver),
                    )
                recovered = recover_amazon_page(
                    current_driver,
                    safe_driver_current_url(current_driver),
                )
                if recovered is not None:
                    current_driver = recovered
                continue
            return current_driver
        assessment = category_page_assessment(current_driver)
        if assessment.status is PageHealthStatus.AMAZON_SIGN_IN:
            if on_manual_pause:
                on_manual_pause(
                    "amazon_sign_in",
                    safe_driver_current_url(current_driver),
                )
            raise VerificationUnconfirmedError(
                verification_unconfirmed_message("amazon_sign_in")
            )
        if assessment.status is PageHealthStatus.INTERACTIVE_VERIFICATION:
            handle_amazon_verification(
                current_driver,
                runtime,
                on_manual_pause,
                on_manual_resume,
                stop_event=stop_event,
            )
            return current_driver
        if assessment.status is PageHealthStatus.TRANSIENT_UNAVAILABLE:
            raise TransientAmazonPageUnavailable.from_assessment(
                assessment,
                url=safe_driver_current_url(current_driver),
            )
        return current_driver

    def handle_block(current_driver: WebDriver) -> Optional[str]:
        reason = sellersprite_block_reason(current_driver)
        if on_manual_pause:
            on_manual_pause(reason, safe_driver_current_url(current_driver))
        cleared = (
            wait_for_manual_clear(
                current_driver,
                reason,
                runtime.manual_pause_timeout,
            )
            if stop_event is None
            else wait_for_manual_clear(
                current_driver,
                reason,
                runtime.manual_pause_timeout,
                stop_event=stop_event,
            )
        )
        if cleared and on_manual_resume:
            on_manual_resume()
        if not cleared and reason == "amazon_sign_in":
            raise VerificationUnconfirmedError(verification_unconfirmed_message(reason))
        return None if cleared else "blocked"

    def retry_with_refresh(current_driver: WebDriver, attempts: int, label: str) -> tuple[str, WebDriver]:
        last_status = "plugin_timeout"
        for attempt in range(1, attempts + 1):
            _raise_if_stop_requested(stop_event)
            wait_seconds = random_plugin_retry_wait(runtime)
            print(
                f"{label} {attempt}/{attempts}：刷新当前页后最多等待 {wait_seconds:.1f} 秒，重新检测卖家精灵数据。"
            )
            try:
                if before_navigation:
                    before_navigation()
                current_driver.refresh()
            except WebDriverException:
                pass
            current_driver = ensure_page_health(current_driver)
            try:
                wait_for_amazon_products(
                    current_driver,
                    runtime,
                    stop_event=stop_event,
                )
            except TimeoutException:
                pass
            last_status = wait_for_sellersprite_data(
                current_driver,
                runtime,
                wait_seconds,
                stop_event=stop_event,
            )
            publish(current_driver)
            if last_status == "ok":
                return last_status, current_driver
            if last_status == "blocked":
                blocked_result = handle_block(current_driver)
                if blocked_result:
                    return blocked_result, current_driver
        return last_status, current_driver

    while True:
        _raise_if_stop_requested(stop_event)
        driver = ensure_page_health(driver)
        status = wait_for_sellersprite_data(driver, runtime, stop_event=stop_event)
        publish(driver)
        if status == "ok":
            return status
        if status == "blocked":
            blocked_result = handle_block(driver)
            if blocked_result:
                return blocked_result
            continue

        status, driver = retry_with_refresh(driver, runtime.plugin_retry_attempts, "卖家精灵数据未成功加载，自动重试")
        if status == "ok":
            return status
        if status == "blocked":
            return status

        if restart_driver is not None and runtime.plugin_relaunch_retry_attempts > 0:
            page_url = driver.current_url
            print(
                f"连续 {runtime.plugin_retry_attempts} 次仍未加载卖家精灵数据，关闭窗口并等待 "
                f"{runtime.plugin_relaunch_wait_seconds / 60:.1f} 分钟后重新拉起。"
            )
            driver = restart_driver(driver, page_url, runtime.plugin_relaunch_wait_seconds)
            status, driver = retry_with_refresh(driver, runtime.plugin_relaunch_retry_attempts, "重启浏览器后自动重试")
            if status == "ok":
                return status
            if status == "blocked":
                return status

        if restart_driver is not None and runtime.plugin_second_relaunch_retry_attempts > 0:
            page_url = driver.current_url
            print(
                f"重启后仍未加载卖家精灵数据，再次关闭窗口并等待 "
                f"{runtime.plugin_second_relaunch_wait_seconds / 60:.1f} 分钟后重新拉起。"
            )
            driver = restart_driver(driver, page_url, runtime.plugin_second_relaunch_wait_seconds)
            status, driver = retry_with_refresh(driver, runtime.plugin_second_relaunch_retry_attempts, "第二次重启浏览器后自动重试")
            if status == "ok":
                return status
            if status == "blocked":
                return status

        report = get_sellersprite_readiness(driver)
        status = str(report.get("status") or status)
        manual_reason = "sellersprite_manual_action"
        if on_manual_pause:
            on_manual_pause(manual_reason, str(getattr(driver, "current_url", "") or ""))
        if status == "plugin_absent":
            action_reason = "当前页面未检测到卖家精灵插件注入，可能未安装、未启用或当前 Chrome 不是安装插件的用户资料。"
        elif status == "login_required":
            action_reason = "卖家精灵插件需要登录、重新授权或确认登录状态。"
        else:
            action_reason = "已检测到卖家精灵插件，但真实字段尚未稳定显示。"
        try:
            continued = wait_for_user_plugin_action(
                driver,
                action_reason,
                before_refresh=before_navigation,
                manual_pause_timeout=int(
                    getattr(runtime, "manual_pause_timeout", 900) or 900
                ),
                stop_event=stop_event,
            )
        finally:
            if on_manual_resume:
                on_manual_resume()
        if not continued:
            return "blocked"
        try:
            wait_for_amazon_products(driver, runtime, stop_event=stop_event)
        except TimeoutException:
            pass


def extract_product_cards(
    driver: WebDriver,
    *,
    strict: bool = False,
) -> List[Dict[str, Any]]:
    script = r"""
const asinRe = /\b([A-Z0-9]{10})\b/;
const asinUrlRe = /\/(?:dp|gp\/product)\/([A-Z0-9]{10})(?:[/?#]|$)/i;
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const getAsin = (el) => {
  const attrs = ['data-asin', 'asin', 'data-csa-c-asin'];
  for (const attr of attrs) {
    const value = el.getAttribute(attr);
    if (value && /^[A-Z0-9]{10}$/.test(value.trim())) return value.trim();
  }
  const links = [...el.querySelectorAll('a[href]')].map(a => a.href);
  for (const href of links) {
    const match = href.match(asinUrlRe);
    if (match) return match[1];
  }
  const text = norm(el.innerText || el.textContent || '');
  const match = text.match(asinRe);
  return match ? match[1] : '';
};
const getUrl = (el, asin) => {
  const links = [...el.querySelectorAll('a[href]')];
  for (const a of links) {
    if (asin && a.href.includes(asin)) return absUrl(a.href);
  }
  for (const a of links) {
    if (/\/(?:dp|gp\/product)\//i.test(a.href)) return absUrl(a.href);
  }
  return '';
};
const getTitle = (el) => {
  const selectors = ['h2 a span', 'h2 span', 'a.a-link-normal span', 'img[alt]', '.p13n-sc-truncate'];
  for (const selector of selectors) {
    const item = el.querySelector(selector);
    if (!item) continue;
    const text = norm(item.getAttribute('alt') || item.innerText || item.textContent || '');
    if (text && text.length > 5) return text;
  }
  return '';
};
const getRank = (el) => {
  const selectors = ['.zg-bdg-text', '.zg-badge-text', '[aria-label^="#"]'];
  for (const selector of selectors) {
    const item = el.querySelector(selector);
    if (!item) continue;
    const text = norm(item.getAttribute('aria-label') || item.innerText || item.textContent || '');
    const match = text.match(/#?\s*(\d{1,5})/);
    if (match) return match[1];
  }
  const text = norm(el.innerText || '');
  const match = text.match(/#\s*(\d{1,5})/);
  return match ? match[1] : '';
};
const getSellerCountryFlagCode = (el) => {
  const flagElements = [...el.querySelectorAll('[class*="flag-icon-"], [class*="icp-nav-flag-"]')];
  for (const flag of flagElements) {
    const classText = String(flag.getAttribute('class') || '');
    const match = classText.match(/(?:flag-icon|icp-nav-flag)-([a-z]{2})\b/i);
    if (match) return match[1].toLowerCase();
    const hint = norm(flag.getAttribute('title') || flag.getAttribute('aria-label') || '');
    if (hint) return hint;
  }
  return '';
};
const getBsrText = (el) => [...el.querySelectorAll('.rank-number-box .bsr-list-item')]
  .map(item => norm(item.innerText || item.textContent || ''))
  .filter(Boolean)
  .join('\n');
const selectors = [
  '#gridItemRoot',
  '.zg-grid-general-faceout',
  '.p13n-grid-content',
  '.s-result-item[data-asin]:not([data-asin=""])',
  '[data-asin]:not([data-asin=""])'
];
const seenElements = new Set();
const cards = [];
for (const selector of selectors) {
  for (const el of document.querySelectorAll(selector)) {
    if (seenElements.has(el)) continue;
    seenElements.add(el);
    const asin = getAsin(el);
    if (!asin) continue;
    cards.push({
      asin,
      title: getTitle(el),
      product_url: getUrl(el, asin),
      rank: getRank(el),
      seller_country_flag_code: getSellerCountryFlagCode(el),
      bsr_text: getBsrText(el),
      text: norm(el.innerText || el.textContent || '')
    });
  }
}
const byAsin = new Map();
for (const card of cards) {
  if (!byAsin.has(card.asin) || card.text.length > byAsin.get(card.asin).text.length) {
    byAsin.set(card.asin, card);
  }
}
return [...byAsin.values()];
"""
    try:
        return list(driver.execute_script(script) or [])
    except (JavascriptException, WebDriverException):
        if strict:
            raise
        return []


def extract_table_rows(
    driver: WebDriver,
    *,
    strict: bool = False,
) -> List[Dict[str, Any]]:
    script = r"""
const asinRe = /\b([A-Z0-9]{10})\b/;
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const directCellText = (row) => {
  let cells = [...row.querySelectorAll(':scope > td, :scope > th, :scope > [role="cell"], :scope > [class*="cell"], :scope > div')];
  if (!cells.length) cells = [...row.children];
  return cells.map(cell => norm(cell.innerText || cell.textContent || ''));
};
const headerTexts = (row) => {
  const table = row.closest('table,[role="table"],[class*="table"],[class*="vxe"]');
  if (!table) return [];
  const headers = [...table.querySelectorAll('thead th, [role="columnheader"], .vxe-header--column')];
  return headers.map(cell => norm(cell.innerText || cell.textContent || ''));
};
const rows = [];
const selectors = ['tr', '[role="row"]', '.vxe-body--row', '[class*="body--row"]', '[class*="table-row"]'];
const seen = new Set();
for (const selector of selectors) {
  for (const row of document.querySelectorAll(selector)) {
    if (seen.has(row)) continue;
    seen.add(row);
    const text = norm(row.innerText || row.textContent || '');
    const match = text.match(asinRe);
    if (!match) continue;
    rows.push({
      asin: match[1],
      text,
      bsr_text: [...row.querySelectorAll('.rank-number-box .bsr-list-item')]
        .map(item => norm(item.innerText || item.textContent || ''))
        .filter(Boolean)
        .join('\n'),
      cells: directCellText(row),
      headers: headerTexts(row)
    });
  }
}
return rows;
"""
    try:
        return list(driver.execute_script(script) or [])
    except (JavascriptException, WebDriverException):
        if strict:
            raise
        return []


def extract_by_selectors(driver: WebDriver, card: Dict[str, Any], selectors: Dict[str, List[str]]) -> Dict[str, str]:
    if not selectors:
        return {}
    asin = card.get("asin") or ""
    script = r"""
const asin = arguments[0];
const selectorMap = arguments[1] || {};
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const cards = [...document.querySelectorAll('#gridItemRoot,.zg-grid-general-faceout,.p13n-grid-content,.s-result-item,[data-asin]')];
const card = cards.find(el => {
  if (el.getAttribute('data-asin') === asin) return true;
  return (el.innerText || el.textContent || '').includes(asin) || [...el.querySelectorAll('a[href]')].some(a => a.href.includes(asin));
});
const output = {};
for (const [field, selectors] of Object.entries(selectorMap)) {
  for (const selector of selectors || []) {
    let found = null;
    try {
      found = card ? card.querySelector(selector) : document.querySelector(selector);
    } catch (err) {
      found = null;
    }
    if (!found) continue;
    const text = norm(found.innerText || found.textContent || found.getAttribute('title') || found.getAttribute('aria-label') || '');
    if (text) {
      output[field] = text;
      break;
    }
  }
}
return output;
"""
    try:
        return dict(driver.execute_script(script, asin, selectors) or {})
    except (JavascriptException, WebDriverException):
        return {}


def split_lines(text: str) -> List[str]:
    lines = []
    for raw in re.split(r"[\n\r]+| {2,}", text or ""):
        line = normalize_space(raw)
        if line:
            lines.append(line)
    return lines


def value_near_labels(text: str, labels: Sequence[str]) -> str:
    lines = split_lines(text)
    lower_labels = [label.lower() for label in labels]
    for index, line in enumerate(lines):
        line_lower = line.lower()
        for label in lower_labels:
            if label not in line_lower:
                continue
            pattern = re.compile(re.escape(label), re.I)
            value = pattern.split(line, maxsplit=1)[-1]
            value = value.strip(" :：|-")
            if value and value.lower() != label:
                return value[:120]
            if index + 1 < len(lines):
                return lines[index + 1][:120]
    return ""


def first_regex(text: str, patterns: Sequence[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return normalize_space(match.group(1))
    return ""


def parse_field_from_text(field_name: str, text: str) -> str:
    text = normalize_space(text)
    if not text:
        return ""

    zh_value = parse_sellersprite_inline_field(field_name, text)
    if zh_value:
        return zh_value
    if field_name == "sales_30_days_parent":
        return ""

    if field_name == "review_count":
        return first_regex(text, [r"评分\(评分数\)[:：]\s*(?:[0-9.]+|N/A)\(([\d,]+|N/A)\)", r"([\d,]+)\s+(?:ratings?|reviews?)", r"(?:评论|评价)\D{0,8}([\d,]+)"])
    if field_name == "rating_value":
        return first_regex(text, [r"评分\(评分数\)[:：]\s*([0-9.]+|N/A)\((?:[\d,]+|N/A)\)", r"([1-5](?:\.\d)?)\s+out of\s+5", r"([1-5](?:\.\d)?)\s*(?:stars?|星)"])
    if field_name in {"sales_30_days", "sales_30_days_child", "sales_30_days_parent"}:
        return first_regex(text, [r"([\d,.]+[Kk万]?)\s*(?:units sold|sold|月销量|销量)", r"(?:30 days?|近30天)\D{0,12}([\d,.]+[Kk万]?)"])
    if field_name == "fba_fee":
        return first_regex(text, [r"FBA费用[:：]\s*(N/A|\$?\s*[\d,.]+)", r"FBA\D{0,20}(N/A|\$?\s*[\d,.]+)", r"(\$?\s*[\d,.]+)\s*FBA"])
    if field_name == "gross_margin":
        return first_regex(text, [r"毛利率[:：]\s*(N/A|[\d,.]+%)", r"([\d,.]+%)\s*(?:gross margin|margin|毛利率)", r"(?:gross margin|margin|毛利率)\D{0,12}([\d,.]+%)"])
    if field_name == "fulfillment_method":
        return parse_fulfillment_method(text)
    if field_name in {"organic_keywords_count", "ad_keywords_count"}:
        return first_regex(text, [r"([\d,]+)\s*(?:keywords?|词)"])

    labels = HEADER_ALIASES.get(field_name, [])
    labelled = value_near_labels(text, labels)
    if labelled:
        return labelled
    return ""


def parse_sellersprite_inline_field(field_name: str, text: str) -> str:
    if field_name == "fulfillment_method":
        method, _raw = parse_fulfillment_evidence(text)
        return method
    patterns = {
        "brand_name": [r"品牌[:：]\s*(.*?)\s+卖家[:：]"],
        "seller_name": [r"卖家[:：]\s*(.*?)\s+配送[:：]"],
        "seller_country": [r"卖家所处国家[:：]\s*([^\s]+)", r"卖家国家[:：]\s*([^\s]+)"],
        "sales_30_days": [r"近30天销量\(子体\)[:：]\s*([^\s]+)", r"近30天销量[:：]\s*([^\s]+)"],
        "sales_30_days_child": [r"近30天销量\(子体\)[:：]\s*([^\s]+)", r"近30天销量[:：]\s*([^\s]+)"],
        "sales_30_days_parent": [r"近30天销量\(父体\)[:：]\s*([^\s]+)"],
        "fba_fee": [r"FBA费用[:：]\s*(N/A|\$?\s*[\d,.]+)"],
        "gross_margin": [r"毛利率[:：]\s*(N/A|[\d,.]+%)"],
        "delivery_duration": [r"(?<!Prime)配送时长[:：]\s*([^\s]+)"],
        "launch_date": [r"上架时间[:：]\s*(\d{4}-\d{2}-\d{2}(?:\s*\([^)]+\))?)"],
        "organic_keywords_count": [r"自然搜索词[:：]\s*([\d,]+)"],
        "ad_keywords_count": [r"广告(?:搜索|流量)词[:：]\s*([\d,]+)"],
    }
    if field_name in {"rating_value", "review_count"}:
        match = re.search(r"评分\(评分数\)[:：]\s*([0-9.]+|N/A)\(([\d,]+|N/A)\)", text)
        if match:
            return match.group(1) if field_name == "rating_value" else match.group(2)
    for pattern in patterns.get(field_name, []):
        match = re.search(pattern, text, re.I)
        if match:
            return normalize_space(match.group(1))
    return ""


def map_headers_to_fields(headers: Sequence[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if not normalized:
            continue
        if "近30" in normalized and "父体" in normalized:
            mapping[index] = "sales_30_days_parent"
            continue
        if "近30" in normalized and "子体" in normalized:
            mapping[index] = "sales_30_days_child"
            continue
        for field_name, aliases in HEADER_ALIASES.items():
            if any(alias.lower() in normalized for alias in aliases):
                if field_name == "rating_value" and "count" in normalized:
                    continue
                mapping[index] = field_name
                break
    return mapping


def parse_table_row_fields(row: Dict[str, Any]) -> Dict[str, str]:
    output: Dict[str, str] = {}
    cells = [normalize_space(str(cell)) for cell in row.get("cells") or []]
    headers = [normalize_space(str(header)) for header in row.get("headers") or []]
    header_mapping = map_headers_to_fields(headers)
    for index, field_name in header_mapping.items():
        if index < len(cells) and cells[index]:
            if field_name == "fulfillment_method":
                method, raw = parse_fulfillment_evidence(cells[index], explicit_value=True)
                output[field_name] = method
                output["fulfillment_method_raw"] = raw
            else:
                output[field_name] = cells[index]
    text = str(row.get("text") or "")
    for field_name in REQUESTED_DATA_FIELDS:
        if field_name == "fulfillment_method":
            continue
        if not output.get(field_name):
            output[field_name] = parse_field_from_text(field_name, text)
    if not output.get("fulfillment_method_raw"):
        method, raw = parse_fulfillment_evidence(text)
        output["fulfillment_method"] = method
        output["fulfillment_method_raw"] = raw
    return output


def merge_product_data(
    driver: WebDriver,
    runtime: RuntimeConfig,
    node: Dict[str, Any],
    page_number: int,
    plugin_status: str,
) -> List[Dict[str, Any]]:
    cards = extract_product_cards(driver, strict=True)
    table_rows = extract_table_rows(
        driver,
        strict=bool(getattr(runtime, "sellersprite_required", True)),
    )
    table_by_asin: Dict[str, Dict[str, Any]] = {}
    for row in table_rows:
        asin = str(row.get("asin") or "")
        if not asin:
            continue
        parsed = parse_table_row_fields(row)
        current = table_by_asin.setdefault(asin, {})
        for key, value in parsed.items():
            if key in {"fulfillment_method", "fulfillment_method_raw"}:
                continue
            if value and not current.get(key):
                current[key] = value
        table_method, table_raw = select_fulfillment_evidence(
            (
                current.get("fulfillment_method"),
                current.get("fulfillment_method_raw"),
            ),
            (
                parsed.get("fulfillment_method"),
                parsed.get("fulfillment_method_raw"),
            ),
        )
        current["fulfillment_method"] = table_method
        current["fulfillment_method_raw"] = table_raw
        row_bsr_text = str(row.get("bsr_text") or "")
        row_ranks = parse_subcategory_bsr_ranks(
            row_bsr_text or str(row.get("text") or "")
        )
        if row_ranks:
            current["subcategory_bsr_ranks"] = normalize_subcategory_bsr_ranks(
                [*(current.get("subcategory_bsr_ranks") or []), *row_ranks]
            )

    records: List[Dict[str, Any]] = []
    for card in cards:
        asin = str(card.get("asin") or "")
        if not asin:
            continue
        record: Dict[str, Any] = {
            "root_url": runtime.start_url,
            "category_path": " > ".join(str(part) for part in node.get("path") or []),
            "category_name": node.get("name") or "",
            "category_node_id": node.get("node_id") or "",
            "category_url": node.get("url") or "",
            "page_number": page_number,
            "rank": card.get("rank") or "",
            "asin": asin,
            "title": card.get("title") or "",
            "product_url": card.get("product_url") or "",
            "scraped_at": now_iso(),
            "load_status": plugin_status,
            "note": "",
            "subcategory_bsr_ranks": [],
            "fulfillment_method_raw": "",
        }
        text = str(card.get("text") or "")
        for field_name in REQUESTED_DATA_FIELDS:
            record[field_name] = ""
        for field_name, value in table_by_asin.get(asin, {}).items():
            if field_name == "fulfillment_method":
                continue
            if field_name in REQUESTED_DATA_FIELDS and value:
                record[field_name] = value
            elif field_name == "fulfillment_method_raw" and value:
                record[field_name] = normalize_space(str(value))[:120]
        record["subcategory_bsr_ranks"] = normalize_subcategory_bsr_ranks(
            table_by_asin.get(asin, {}).get("subcategory_bsr_ranks")
        )
        selector_values = extract_by_selectors(driver, card, runtime.field_selectors)
        for field_name, value in selector_values.items():
            if field_name in REQUESTED_DATA_FIELDS and value:
                if field_name == "fulfillment_method":
                    continue
                else:
                    record[field_name] = normalize_space(str(value))
        selector_bsr = selector_values.get("subcategory_bsr_ranks")
        if selector_bsr:
            record["subcategory_bsr_ranks"] = parse_subcategory_bsr_ranks(
                str(selector_bsr)
            )
        for field_name in REQUESTED_DATA_FIELDS:
            if field_name == "fulfillment_method":
                continue
            if not record.get(field_name):
                record[field_name] = parse_field_from_text(field_name, text)
        selector_fulfillment = parse_fulfillment_evidence(
            str(selector_values.get("fulfillment_method") or ""),
            explicit_value=True,
        )
        table_fulfillment = (
            table_by_asin.get(asin, {}).get("fulfillment_method"),
            table_by_asin.get(asin, {}).get("fulfillment_method_raw"),
        )
        card_fulfillment = parse_fulfillment_evidence(text)
        method, raw = select_fulfillment_evidence(
            selector_fulfillment,
            table_fulfillment,
            card_fulfillment,
        )
        record["fulfillment_method"] = method
        record["fulfillment_method_raw"] = raw
        if not record["subcategory_bsr_ranks"]:
            card_bsr_text = str(card.get("bsr_text") or "")
            record["subcategory_bsr_ranks"] = parse_subcategory_bsr_ranks(
                card_bsr_text or text
            )
        record["fulfillment_method"] = normalize_fulfillment_method(
            record.get("fulfillment_method")
        )
        if not record.get("seller_country"):
            record["seller_country"] = country_from_flag_code_or_text(str(card.get("seller_country_flag_code") or ""))
        missing_count = sum(1 for field_name in REQUESTED_DATA_FIELDS if not record.get(field_name))
        if plugin_status != "ok":
            record["note"] = "插件数据加载超时，已保存页面可见数据"
        elif missing_count == len(REQUESTED_DATA_FIELDS):
            record["note"] = "插件未展示或选择器未匹配"
        records.append(record)
    return records


def preload_page_data_with_scroll(driver: WebDriver, runtime: Any) -> None:
    if not bool(getattr(runtime, "page_scroll_before_extract", True)):
        return
    max_rounds = max(int(getattr(runtime, "page_scroll_max_rounds", 18) or 0), 0)
    if max_rounds <= 0:
        return
    step_ratio = float(getattr(runtime, "page_scroll_step_ratio", 0.85) or 0.85)
    if step_ratio <= 0:
        step_ratio = 0.85
    wait_seconds = max(float(getattr(runtime, "page_scroll_wait_seconds", 1.0) or 0), 0)
    stable_target = max(int(getattr(runtime, "page_scroll_stable_rounds", 2) or 1), 1)

    print("向下滚动页面，等待 Amazon 商品和卖家精灵插件数据加载。", flush=True)
    try:
        driver.execute_script("window.scrollTo(0, 0);")
    except (JavascriptException, WebDriverException):
        return
    if wait_seconds:
        time.sleep(min(wait_seconds, 0.5))

    last_signature = ""
    stable_rounds = 0
    for round_index in range(max_rounds):
        if detect_block(driver):
            return
        if round_index == 0 and bool(getattr(runtime, "activate_plugin", True)):
            try_activate_plugin(driver)
        try:
            metrics = driver.execute_script(
                r"""
const stepRatio = Number(arguments[0]) || 0.85;
const asinRe = /\b[A-Z0-9]{10}\b/;
const asinGlobalRe = /\b[A-Z0-9]{10}\b/g;
const pluginSelector = [
  '[id*="sellersprite" i]',
  '[class*="sellersprite" i]',
  '[id*="seller-sprite" i]',
  '[class*="seller-sprite" i]',
  '[id*="__ss" i]',
  '[class*="vxe-" i]',
  '[class*="ss-" i]',
  '[class*="sprite" i]'
].join(',');
const productSelector = [
  '#gridItemRoot',
  '.zg-grid-general-faceout',
  '.p13n-grid-content',
  '.s-result-item[data-asin]:not([data-asin=""])',
  '[data-asin]:not([data-asin=""])'
].join(',');
const tableRowSelector = [
  'tr',
  '[role="row"]',
  '.vxe-body--row',
  '[class*="body--row"]',
  '[class*="table-row"]'
].join(',');
const scrollSelector = [
  pluginSelector,
  '.vxe-table--body-wrapper',
  '[class*="vxe-table--body-wrapper"]',
  '[class*="el-table__body-wrapper"]',
  '[class*="table-body"]',
  '[role="grid"]',
  '[role="table"]'
].join(',');
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const isScrollable = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (style.display === 'none' || style.visibility === 'hidden') return false;
  return (el.scrollHeight || 0) > (el.clientHeight || 0) + 40;
};
const root = document.scrollingElement || document.documentElement || document.body;
const scrollables = [root, document.documentElement, document.body, ...document.querySelectorAll(scrollSelector)];
const seen = new Set();
for (const el of scrollables) {
  if (!el || seen.has(el)) continue;
  seen.add(el);
  if (el !== root && !isScrollable(el)) continue;
  try { el.scrollBy(0, Math.max(500, window.innerHeight * stepRatio)); } catch (err) {}
}
const bodyText = document.body ? (document.body.innerText || document.body.textContent || '') : '';
const asins = new Set((bodyText.match(asinGlobalRe) || []));
for (const el of document.querySelectorAll(productSelector)) {
  const asin = el.getAttribute('data-asin') || '';
  if (asin) asins.add(asin);
}
const pluginNodes = document.querySelectorAll(pluginSelector).length;
const tableRows = [...document.querySelectorAll(tableRowSelector)].filter(row => asinRe.test(norm(row.innerText || row.textContent || ''))).length;
const scrollTop = root ? (root.scrollTop || window.scrollY || 0) : window.scrollY;
const scrollHeight = root ? root.scrollHeight : document.documentElement.scrollHeight;
const viewport = window.innerHeight || document.documentElement.clientHeight || 0;
return {
  asinCount: asins.size,
  pluginNodes,
  tableRows,
  textLength: bodyText.length,
  scrollTop,
  scrollHeight,
  atBottom: scrollTop + viewport >= scrollHeight - 20
};
""",
                step_ratio,
            )
        except (JavascriptException, WebDriverException):
            return
        if not isinstance(metrics, dict):
            return
        signature = (
            f"{metrics.get('scrollHeight')}:{metrics.get('asinCount')}:"
            f"{metrics.get('pluginNodes')}:{metrics.get('tableRows')}:{metrics.get('textLength')}"
        )
        if bool(metrics.get("atBottom")) and signature == last_signature:
            stable_rounds += 1
        else:
            stable_rounds = 0
        last_signature = signature
        if stable_rounds >= stable_target:
            break
        if wait_seconds:
            time.sleep(wait_seconds)


def wait_for_sellersprite_data(
    driver: WebDriver,
    runtime: RuntimeConfig,
    timeout_seconds: Optional[float] = None,
    stop_event: Optional[threading.Event] = None,
) -> str:
    _raise_if_stop_requested(stop_event)
    if not bool(getattr(runtime, "sellersprite_required", True)):
        report = {
            "status": "not_required",
            "checked_at": now_iso(),
            "page_url": str(getattr(driver, "current_url", "") or ""),
        }
        set_sellersprite_readiness(driver, report)
        return "not_required"
    if runtime.activate_plugin:
        try_activate_plugin(driver)
    preload_page_data_with_scroll(driver, runtime)
    timeout = runtime.plugin_timeout if timeout_seconds is None else max(float(timeout_seconds), 1)
    deadline = time.time() + timeout
    stable_seen = 0
    last_signature = ""
    last_status = "data_loading"
    while time.time() < deadline:
        _raise_if_stop_requested(stop_event)
        report = inspect_sellersprite_readiness(driver, runtime)
        status = str(report.get("status") or "data_loading")
        if status == "blocked":
            set_sellersprite_readiness(driver, report)
            return "blocked"
        signature = str(report.get("signature") or "")
        if status == "ready_candidate":
            stable_seen = stable_seen + 1 if signature == last_signature else 1
            report["stable_checks"] = stable_seen
            required_checks = max(int(getattr(runtime, "sellersprite_stable_checks", 3) or 3), 1)
            if stable_seen >= required_checks:
                report["status"] = "ready"
                set_sellersprite_readiness(driver, report)
                return "ok"
            report["status"] = "data_loading"
        else:
            stable_seen = 0
        last_status = str(report.get("status") or status)
        set_sellersprite_readiness(driver, report)
        last_signature = signature
        _sleep_with_stop(2, stop_event)
    return last_status


def random_plugin_retry_wait(runtime: RuntimeConfig) -> float:
    return random.uniform(runtime.plugin_retry_wait_seconds, runtime.plugin_retry_wait_seconds_max)


def discover_child_categories(
    driver: WebDriver,
    node: Dict[str, Any],
    *,
    strict: bool = False,
) -> List[Dict[str, Any]]:
    script = r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim();
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const nodeIdFromUrl = (url) => {
  const decoded = decodeURIComponent(url || '');
  const nodeMatch = decoded.match(/[?&]node=(\d{4,})/);
  if (nodeMatch) return nodeMatch[1];
  const pathMatches = [...decoded.matchAll(/\/(\d{4,})(?:[/?#]|$)/g)].map(m => m[1]);
  if (pathMatches.length) return pathMatches[pathMatches.length - 1];
  const refMatches = [...decoded.matchAll(/_(\d{4,})(?:[/?#_]|$)/g)].map(m => m[1]);
  return refMatches.length ? refMatches[refMatches.length - 1] : '';
};
const currentNodeId = nodeIdFromUrl(location.href);
const browseRoot = document.querySelector('#zg_browseRoot,#zg-left-col,[data-testid="departments"],#departments,#s-refinements') || document.body;
let anchors = [];
const selected = browseRoot.querySelector('.zg_selected,.a-color-state,[aria-current="page"],span[aria-current],strong');
const selectedLi = selected ? selected.closest('li') : null;
if (selectedLi) {
  const childLists = [...selectedLi.children].filter(el => el.tagName && el.tagName.toLowerCase() === 'ul');
  for (const list of childLists) {
    anchors.push(...[...list.querySelectorAll('a[href]')].filter(a => a.closest('ul') === list || a.closest('li')?.parentElement === list));
  }
  const nextLi = selectedLi.nextElementSibling && selectedLi.nextElementSibling.matches('li') ? selectedLi.nextElementSibling : null;
  if (nextLi) {
    anchors.push(...nextLi.querySelectorAll('a[href]'));
  }
}
if (!anchors.length) {
  const selectedParent = selectedLi || browseRoot;
  anchors = [...selectedParent.querySelectorAll(':scope > ul a[href], :scope > div > ul a[href]')];
}
const seen = new Set();
const output = [];
for (const a of anchors) {
  const name = norm(a.innerText || a.textContent || a.getAttribute('aria-label') || '');
  const href = absUrl(a.getAttribute('href') || a.href || '');
  const childNodeId = nodeIdFromUrl(href);
  if (!name || !href || !childNodeId || childNodeId === currentNodeId) continue;
  if (!/amazon\./i.test(href)) continue;
  const key = childNodeId || href;
  if (seen.has(key)) continue;
  seen.add(key);
  output.push({name, url: href, node_id: childNodeId});
}
return output;
"""
    try:
        raw_children = list(driver.execute_script(script) or [])
    except (JavascriptException, WebDriverException):
        if strict:
            raise
        raw_children = []

    parent_path = [str(part) for part in node.get("path") or []]
    depth = int(node.get("depth") or 0) + 1
    children: List[Dict[str, Any]] = []
    for child in raw_children:
        name = normalize_space(str(child.get("name") or ""))
        url = str(child.get("url") or "")
        if not name or not url:
            continue
        children.append(
            {
                "url": url,
                "name": name,
                "path": parent_path + [name],
                "node_id": str(child.get("node_id") or extract_node_id(url)),
                "depth": depth,
            }
        )
    return children


def extract_current_category_path(driver: WebDriver) -> List[str]:
    script = r"""
const norm = (text) => (text || '').replace(/\s+/g, ' ').trim().replace(/\(Current\)$/i, '').trim();
const root = document.querySelector('#zg-left-col,#zg_browseRoot,#departments,#s-refinements');
if (!root) return [];
const output = [];
for (const a of root.querySelectorAll('li[class*="browse-up"] a[href], li[class*="browse-up"] span a[href]')) {
  const text = norm(a.innerText || a.textContent || '');
  if (text && !/^any department$/i.test(text)) output.push(text);
}
const selected = root.querySelector('[aria-current="page"], [class*="selected"]');
if (selected) {
  const text = norm(selected.innerText || selected.textContent || '');
  if (text && !output.includes(text)) output.push(text);
}
return output;
"""
    try:
        values = list(driver.execute_script(script) or [])
    except (JavascriptException, WebDriverException):
        values = []
    return [normalize_space(str(value)) for value in values if normalize_space(str(value))]


def should_descend(runtime: RuntimeConfig, node: Dict[str, Any], children: Sequence[Dict[str, Any]]) -> bool:
    if not children:
        return False
    depth = int(node.get("depth") or 0)
    if runtime.max_depth is not None and depth >= runtime.max_depth:
        return False
    return True


def find_next_page_url(driver: WebDriver, *, strict: bool = False) -> str:
    script = r"""
const absUrl = (href) => {
  try { return new URL(href, location.href).href; } catch (err) { return href || ''; }
};
const selectors = [
  'li.a-last a[href]',
  'ul.a-pagination a[aria-label*="Next" i]',
  'a[aria-label*="Next" i]',
  'a.s-pagination-next[href]'
];
for (const selector of selectors) {
  const el = document.querySelector(selector);
  if (el && !el.closest('.a-disabled')) return absUrl(el.getAttribute('href') || el.href || '');
}
const links = [...document.querySelectorAll('a[href]')].filter(a => /next|下一页|后一页/i.test(a.innerText || a.getAttribute('aria-label') || ''));
for (const link of links) return absUrl(link.getAttribute('href') || link.href || '');
return '';
"""
    try:
        return str(driver.execute_script(script) or "")
    except (JavascriptException, WebDriverException):
        if strict:
            raise
        return ""


def save_debug_snapshot(driver: WebDriver, debug_dir: Path, label: str) -> None:
    ensure_dir(debug_dir)
    safe_label = slugify(label)
    html_path = debug_dir / f"{now_ts()}_{safe_label}.html"
    png_path = debug_dir / f"{now_ts()}_{safe_label}.png"
    try:
        html_path.write_text(driver.page_source, encoding="utf-8")
    except WebDriverException:
        pass
    try:
        driver.save_screenshot(str(png_path))
    except WebDriverException:
        pass


def build_failure_record(
    runtime: RuntimeConfig,
    node: Dict[str, Any],
    page_number: int,
    url: str,
    reason: str,
    message: str,
) -> Dict[str, Any]:
    return {
        "time": now_iso(),
        "root_url": runtime.start_url,
        "category_path": " > ".join(str(part) for part in node.get("path") or []),
        "category_name": node.get("name") or "",
        "category_node_id": node.get("node_id") or "",
        "category_url": node.get("url") or "",
        "page_number": page_number,
        "page_url": url,
        "reason": reason,
        "message": message,
    }


def build_amazon_retry_exhausted_failure_record(
    runtime: RuntimeConfig,
    node: Dict[str, Any],
    page_number: int,
    url: str,
    exhausted: AmazonPageRetryExhausted,
) -> Dict[str, Any]:
    retry_state = exhausted.state
    work_key = str(retry_state.get("work_key") or page_key(node, page_number, url))
    stage = str(retry_state.get("stage") or "category_page_work")
    cycle = max(int(retry_state.get("cycle") or 1), 1)
    record = build_failure_record(
        runtime,
        node,
        page_number,
        url,
        AmazonPageRetryExhausted.failure_code,
        str(exhausted),
    )
    record["recovery_failure_key"] = f"{work_key}|stage:{stage}|cycle:{cycle}"
    record["recovery_stage"] = stage
    record["recovery_cycle"] = cycle
    return record


def append_failure_record_once(path: Path, record: Dict[str, Any]) -> bool:
    identity = str(record.get("recovery_failure_key") or "")
    if identity and any(
        str(existing.get("recovery_failure_key") or "") == identity
        for existing in read_jsonl(path)
    ):
        return False
    append_jsonl(path, record)
    return True


def log_amazon_retry_exhausted_once(
    failures_path: Path,
    state: StateStore,
    runtime: RuntimeConfig,
    node: Dict[str, Any],
    page_number: int,
    url: str,
    exhausted: AmazonPageRetryExhausted,
) -> bool:
    record = build_amazon_retry_exhausted_failure_record(
        runtime,
        node,
        page_number,
        url,
        exhausted,
    )
    if not append_failure_record_once(failures_path, record):
        return False
    state.log_failure()
    return True


def log_failure(
    failures_path: Path,
    state: StateStore,
    runtime: RuntimeConfig,
    node: Dict[str, Any],
    page_number: int,
    url: str,
    reason: str,
    message: str,
) -> None:
    append_jsonl(
        failures_path,
        build_failure_record(runtime, node, page_number, url, reason, message),
    )
    state.log_failure()


def write_workbook(records_path: Path, failures_path: Path, output_xlsx: Path) -> None:
    if Workbook is None:
        raise UserFacingError("缺少 openpyxl，无法生成 Excel。")
    records = read_jsonl(records_path)
    failures = read_jsonl(failures_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    write_sheet(ws, OUTPUT_HEADERS, records)

    dedup_rows = build_dedup_rows(records)
    ws_dedup = wb.create_sheet("去重ASIN表")
    write_sheet(ws_dedup, OUTPUT_HEADERS, dedup_rows)

    ws_fail = wb.create_sheet("失败页面")
    failure_headers = ["time", "category_path", "category_name", "category_node_id", "page_number", "page_url", "reason", "message"]
    write_sheet(ws_fail, failure_headers, failures, raw_headers=True)

    ensure_dir(output_xlsx.parent)
    wb.save(output_xlsx)


def write_sheet(ws: Any, headers: Sequence[str], rows: Sequence[Dict[str, Any]], raw_headers: bool = False) -> None:
    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in rows:
        if raw_headers:
            ws.append([row.get(header, "") for header in headers])
        else:
            ws.append(
                [
                    format_subcategory_bsr_ranks(row.get(field))
                    if field == "subcategory_bsr_ranks"
                    else row.get(field, "")
                    for field, header in FIELD_TO_HEADER.items()
                    if header in headers
                ]
            )
    for column_index, header in enumerate(headers, start=1):
        width = 42 if header == "子类目节点排名" else min(max(len(str(header)) + 2, 12), 36)
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_dedup_rows(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    paths: Dict[str, List[str]] = {}
    for row in records:
        asin = str(row.get("asin") or "")
        if not asin:
            continue
        target = grouped.setdefault(asin, dict(row))
        for key, value in row.items():
            if key == "subcategory_bsr_ranks":
                merged: List[Dict[str, Any]] = []
                seen_categories: set[str] = set()
                for item in [
                    *normalize_subcategory_bsr_ranks(target.get(key)),
                    *normalize_subcategory_bsr_ranks(value),
                ]:
                    category_name_key = str(item["category_name"]).casefold()
                    if category_name_key in seen_categories:
                        continue
                    seen_categories.add(category_name_key)
                    merged.append(dict(item))
                target[key] = merged
                continue
            if value and not target.get(key):
                target[key] = value
        path = str(row.get("category_path") or "")
        if path:
            paths.setdefault(asin, [])
            if path not in paths[asin]:
                paths[asin].append(path)
    output = []
    for asin, row in grouped.items():
        merged = dict(row)
        merged["category_path"] = " ; ".join(paths.get(asin) or [])
        output.append(merged)
    return output


class WorkerManualGate:
    """Serialize interactive prompts while reporting state changes to the main thread."""

    def __init__(
        self,
        claim_key: str,
        events: "queue.Queue[Dict[str, Any]]",
        stop_event: threading.Event,
    ) -> None:
        self.claim_key = claim_key
        self.events = events
        self.stop_event = stop_event
        self._held = False

    def pause(self, reason: str, page_url: str) -> None:
        if not self._held:
            while not MANUAL_INTERACTION_LOCK.acquire(timeout=0.25):
                if self.stop_event.is_set():
                    raise ConcurrentWorkerCancelled("并发任务正在停止，已取消等待人工操作。")
            if self.stop_event.is_set():
                MANUAL_INTERACTION_LOCK.release()
                raise ConcurrentWorkerCancelled("并发任务正在停止，已取消等待人工操作。")
            self._held = True
        self.events.put(
            {
                "type": "manual_pause",
                "claim_key": self.claim_key,
                "reason": reason,
                "page_url": page_url,
            }
        )

    def resume(self) -> None:
        if not self._held:
            return
        self.events.put({"type": "manual_resume", "claim_key": self.claim_key})
        self._held = False
        MANUAL_INTERACTION_LOCK.release()

    def close(self) -> None:
        self.resume()


def _worker_retry_state_request(
    events: "queue.Queue[Dict[str, Any]]",
    stop_event: threading.Event,
    event_type: str,
    retry_key: str,
    retry_state: Optional[Mapping[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    acknowledgment = threading.Event()
    event: Dict[str, Any] = {
        "type": event_type,
        "retry_key": retry_key,
        "acknowledgment": acknowledgment,
    }
    if retry_state is not None:
        event["retry_state"] = copy.deepcopy(dict(retry_state))
    events.put(event)
    while not acknowledgment.wait(0.25):
        if stop_event.is_set():
            if event_type == "amazon_page_retry_load":
                raise ConcurrentWorkerCancelled(
                    "并发任务正在停止，已取消读取 Amazon 页面恢复断点。"
                )
            # The main-thread shutdown path drains queued writes/clears after
            # workers exit. Never deadlock executor shutdown waiting for ack.
            return None
    error = event.get("error")
    if isinstance(error, BaseException):
        raise error
    result = event.get("result")
    return copy.deepcopy(result) if isinstance(result, dict) else None


def worker_retry_callbacks(
    events: "queue.Queue[Dict[str, Any]]",
    stop_event: threading.Event,
    retry_key: str,
) -> RetryCallbacks:
    return RetryCallbacks(
        load_state=lambda: _worker_retry_state_request(
            events,
            stop_event,
            "amazon_page_retry_load",
            retry_key,
        ),
        write_state=lambda value: _worker_retry_state_request(
            events,
            stop_event,
            "amazon_page_retry_write",
            retry_key,
            value,
        ),
        clear_state=lambda: _worker_retry_state_request(
            events,
            stop_event,
            "amazon_page_retry_clear",
            retry_key,
        ),
    )


def _drain_worker_events(
    events: "queue.Queue[Dict[str, Any]]",
    state: StateStore,
) -> None:
    while True:
        try:
            event = events.get_nowait()
        except queue.Empty:
            return
        event_type = str(event.get("type") or "")
        claim_key = str(event.get("claim_key") or "")
        acknowledgment = event.get("acknowledgment")
        try:
            if event_type == "manual_pause":
                state.mark_manual_pause(
                    str(event.get("reason") or "manual_action"),
                    str(event.get("page_url") or ""),
                    claim_key,
                )
            elif event_type == "manual_resume":
                state.clear_manual_pause(claim_key)
            elif event_type == "readiness":
                report = event.get("report")
                if isinstance(report, dict):
                    state.mark_sellersprite_readiness(report)
            elif event_type == "amazon_page_retry_load":
                event["result"] = state.load_amazon_page_retry(
                    str(event.get("retry_key") or "")
                )
            elif event_type == "amazon_page_retry_write":
                retry_state = event.get("retry_state")
                if not isinstance(retry_state, dict):
                    raise UserFacingError("Amazon 页面恢复状态不是 JSON 对象。")
                state.write_amazon_page_retry(
                    str(event.get("retry_key") or ""),
                    retry_state,
                )
            elif event_type == "amazon_page_retry_clear":
                state.clear_amazon_page_retry(
                    str(event.get("retry_key") or "")
                )
        except BaseException as exc:
            event["error"] = exc
        finally:
            if isinstance(acknowledgment, threading.Event):
                acknowledgment.set()


def _wait_for_page_without_worker_writes(
    driver: WebDriver,
    runtime: RuntimeConfig,
    manual_gate: WorkerManualGate,
    stop_event: threading.Event,
) -> None:
    block_reason = detect_block(driver)
    if block_reason:
        manual_gate.pause(block_reason, str(getattr(driver, "current_url", "") or ""))
        cleared = wait_for_manual_clear(
            driver,
            block_reason,
            runtime.manual_pause_timeout,
            stop_event=stop_event,
        )
        if cleared:
            manual_gate.resume()
        if not cleared:
            raise VerificationUnconfirmedError(
                verification_unconfirmed_message(block_reason)
            )
    wait_for_amazon_products(driver, runtime, stop_event=stop_event)


def crawl_category_source(
    runtime: RuntimeConfig,
    claim_key: str,
    source_node: Dict[str, Any],
    completed_page_keys: set[str],
    debug_dir: Path,
    events: "queue.Queue[Dict[str, Any]]",
    stop_event: threading.Event,
    navigation_throttle: NavigationThrottle,
    delivery_locks: DeliveryDomainLocks,
) -> CategoryCrawlBatch:
    """Crawl one category source serially; never write shared state or JSONL files."""

    node = dict(source_node)
    node["path"] = list(source_node.get("path") or [])
    result = CategoryCrawlBatch(node=node)
    manual_gate = WorkerManualGate(claim_key, events, stop_event)
    driver: Optional[WebDriver] = None
    confirmed_domains: set[str] = set()
    page_number = 1
    page_url = str(node.get("url") or "")
    domain_cooldowns = getattr(runtime, "_amazon_domain_cooldowns", None)
    if not isinstance(domain_cooldowns, DomainCooldownRegistry):
        domain_cooldowns = DomainCooldownRegistry(
            clock=_runtime_recovery_clock(runtime),
            waiter=_runtime_recovery_waiter(runtime, stop_event),
        )
        setattr(runtime, "_amazon_domain_cooldowns", domain_cooldowns)

    def open_worker_page(
        current_driver: WebDriver,
        target_url: str,
        current_page_number: int,
        stage: str = "category_page",
    ) -> PageHealthAssessment:
        if stop_event.is_set():
            raise ConcurrentWorkerCancelled("并发任务正在停止，已取消页面导航。")
        domain = (urlparse(target_url).hostname or "unknown").lower()
        def before_navigation() -> None:
            navigation_throttle.wait()
            _raise_if_stop_requested(stop_event)

        assessment = load_category_page_attempt(
            current_driver,
            target_url,
            runtime,
            on_manual_pause=manual_gate.pause,
            on_manual_resume=manual_gate.resume,
            stop_event=stop_event,
            before_navigation=before_navigation,
            delivery_lock=(
                delivery_locks.for_url(target_url)
                if domain not in confirmed_domains
                else None
            ),
            domain_cooldowns=domain_cooldowns,
        )
        confirmed_domains.add(domain)
        return assessment

    def restart_plugin_driver(
        current_driver: WebDriver,
        target_url: str,
        wait_seconds: float,
    ) -> WebDriver:
        nonlocal driver
        try:
            current_driver.quit()
        except WebDriverException:
            pass
        if wait_seconds > 0:
            _sleep_with_stop(wait_seconds, stop_event)
        if stop_event.is_set():
            raise ConcurrentWorkerCancelled("并发任务正在停止，已取消浏览器重启。")
        driver = start_driver(runtime)
        confirmed_domains.clear()
        open_worker_page(
            driver,
            target_url,
            page_number,
            stage="category_plugin_restart",
        )
        return driver

    def recover_plugin_page(
        current_driver: WebDriver,
        target_url: str,
    ) -> WebDriver:
        assessment = open_worker_page(
            current_driver,
            target_url or page_url,
            page_number,
            stage="category_plugin_refresh",
        )
        if assessment.status is PageHealthStatus.VERIFIED_EMPTY:
            raise TransientAmazonPageUnavailable(
                "页面在插件刷新期间变为明确空结果；重新执行整页事务。",
                reason="verified_empty_during_plugin_refresh",
                url=safe_driver_current_url(current_driver, target_url or page_url),
            )
        return current_driver

    try:
        driver = start_driver(runtime)
        while not stop_event.is_set():
            print(
                f"并发处理类目：{' > '.join(node.get('path') or [])} / 第 {page_number} 页",
                flush=True,
            )
            try:
                work_key = page_key(node, page_number, page_url)
                retry_key = _retry_entry_key(work_key, "category_page_work")

                def current_driver() -> WebDriver:
                    if driver is None:
                        raise WebDriverException("类目 worker 浏览器尚未启动。")
                    return driver

                def process_page_attempt(attempt: Any) -> CategoryPageWorkResult:
                    nonlocal driver
                    if attempt.attempt_number > 1:
                        try:
                            current_driver().quit()
                        except Exception:
                            pass
                        driver = start_driver(runtime)
                        confirmed_domains.clear()

                    active_driver = current_driver()
                    page_assessment = open_worker_page(
                        active_driver,
                        page_url,
                        page_number,
                    )
                    attempt_node = copy.deepcopy(node)
                    page_path = extract_current_category_path(active_driver)
                    if page_path:
                        attempt_node["path"] = page_path
                        attempt_node["name"] = page_path[-1]
                        attempt_node["node_id"] = (
                            attempt_node.get("node_id")
                            or extract_node_id(
                                safe_driver_current_url(active_driver, page_url)
                            )
                        )

                    children: List[Dict[str, Any]] = []
                    if page_number == 1:
                        children = discover_child_categories(
                            active_driver,
                            attempt_node,
                            strict=True,
                        )
                        if (
                            should_descend(runtime, attempt_node, children)
                            and not runtime.include_root
                        ):
                            return CategoryPageWorkResult(
                                node=attempt_node,
                                children=children,
                                skipped_intermediate=True,
                            )

                    key = page_key(attempt_node, page_number, page_url)
                    page_batch: Optional[CategoryPageBatch] = None
                    if key in completed_page_keys:
                        print(f"断点已完成，跳过并发页写入：{key}", flush=True)
                    elif page_assessment.status is PageHealthStatus.VERIFIED_EMPTY:
                        page_batch = CategoryPageBatch(
                            key=key,
                            page_number=page_number,
                            page_url=page_url,
                            plugin_status="verified_empty",
                            extracted_count=0,
                            records=[],
                            rejection_counts={},
                        )
                    else:
                        plugin_status = wait_for_sellersprite_data_or_prompt(
                            active_driver,
                            runtime,
                            on_manual_pause=manual_gate.pause,
                            on_manual_resume=manual_gate.resume,
                            restart_driver=restart_plugin_driver,
                            on_readiness=lambda report: events.put(
                                {
                                    "type": "readiness",
                                    "claim_key": claim_key,
                                    "report": dict(report),
                                }
                            ),
                            before_navigation=lambda: (
                                domain_cooldowns.wait(
                                    (urlparse(page_url).hostname or "unknown").lower()
                                ),
                                navigation_throttle.wait(),
                                _raise_if_stop_requested(stop_event),
                            ),
                            recover_amazon_page=recover_plugin_page,
                            stop_event=stop_event,
                        )
                        if plugin_status == "blocked":
                            raise VerificationUnconfirmedError(
                                "sellersprite_verification_unconfirmed: "
                                "人工处理超时，任务已停止且未提取当前页数据。"
                            )
                        active_driver = current_driver()
                        post_plugin_health = wait_for_category_page_health(
                            active_driver,
                            runtime,
                            manual_gate.pause,
                            manual_gate.resume,
                            stop_event,
                        )
                        if post_plugin_health.status is PageHealthStatus.VERIFIED_EMPTY:
                            raise TransientAmazonPageUnavailable(
                                "页面在插件等待期间变为明确空结果；重新执行整页事务。",
                                reason="verified_empty_after_plugin_wait",
                                url=safe_driver_current_url(active_driver, page_url),
                            )
                        extracted_records = merge_product_data(
                            active_driver,
                            runtime,
                            attempt_node,
                            page_number,
                            plugin_status,
                        )
                        if not extracted_records:
                            raise TransientAmazonPageUnavailable(
                                "页面已检测到商品卡片，但提取结果为空；"
                                "为避免静默漏页，保留断点并重试。",
                                reason="empty_extraction_after_expected_content",
                                url=safe_driver_current_url(active_driver, page_url),
                            )
                        accepted_records, rejection_counts = filter_product_records(
                            extracted_records,
                            runtime.product_filters,
                        )
                        page_batch = CategoryPageBatch(
                            key=key,
                            page_number=page_number,
                            page_url=page_url,
                            plugin_status=plugin_status,
                            extracted_count=len(extracted_records),
                            records=accepted_records,
                            rejection_counts=rejection_counts,
                        )
                        if plugin_status != "ok" and runtime.save_debug_snapshots:
                            save_debug_snapshot(
                                active_driver,
                                debug_dir,
                                f"plugin_{plugin_status}_{slugify(claim_key)}_{page_number}",
                            )

                    active_driver = current_driver()
                    return CategoryPageWorkResult(
                        node=attempt_node,
                        next_url=find_next_page_url(active_driver, strict=True),
                        page=page_batch,
                        children=children,
                    )

                page_result = run_category_page_work_with_recovery(
                    runtime,
                    page_url,
                    work_key,
                    retry_callbacks=worker_retry_callbacks(
                        events,
                        stop_event,
                        retry_key,
                    ),
                    driver_provider=current_driver,
                    operation=process_page_attempt,
                    stop_event=stop_event,
                    domain_cooldowns=domain_cooldowns,
                )
                node = page_result.node
                result.node = node
                if page_result.children:
                    result.children = page_result.children
                if page_result.skipped_intermediate:
                    result.skipped_intermediate = True
                    return result
                if page_result.page is not None:
                    result.pages.append(page_result.page)

                next_url = page_result.next_url
                page_limit = runtime.max_pages_per_category
                if not next_url or (page_limit is not None and page_number >= page_limit):
                    return result
                page_number += 1
                page_url = next_url
            except AmazonPageRetryExhausted as exc:
                result.failures.append(
                    build_amazon_retry_exhausted_failure_record(
                        runtime,
                        node,
                        page_number,
                        page_url,
                        exc,
                    )
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(
                        driver,
                        debug_dir,
                        f"amazon_page_retry_exhausted_{slugify(claim_key)}_{page_number}",
                    )
                result.terminal_error_type = AmazonPageRetryExhausted.failure_code
                result.terminal_error_message = str(exc)
                stop_event.set()
                return result
            except DeliveryLocationUnconfirmedError as exc:
                result.failures.append(
                    build_failure_record(
                        runtime,
                        node,
                        page_number,
                        page_url,
                        "delivery_location_unconfirmed",
                        str(exc),
                    )
                )
                result.terminal_error_type = "delivery_location_unconfirmed"
                result.terminal_error_message = str(exc)
                stop_event.set()
                return result
            except VerificationUnconfirmedError as exc:
                result.failures.append(
                    build_failure_record(
                        runtime,
                        node,
                        page_number,
                        page_url,
                        "verification_timeout",
                        str(exc),
                    )
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(
                        driver,
                        debug_dir,
                        f"verification_timeout_{slugify(claim_key)}_{page_number}",
                    )
                result.terminal_error_type = "verification_unconfirmed"
                result.terminal_error_message = str(exc)
                stop_event.set()
                return result
        raise ConcurrentWorkerCancelled(
            "并发任务已收到停止信号，当前类目保留在断点中。"
        )
    finally:
        manual_gate.close()
        if driver is not None:
            try:
                driver.quit()
            except WebDriverException:
                pass


def _commit_category_batch(
    batch: CategoryCrawlBatch,
    claim_key: str,
    runtime: RuntimeConfig,
    state: StateStore,
    records_path: Path,
    failures_path: Path,
) -> int:
    committed_pages = 0
    for page in batch.pages:
        if not state.commit_page_batch(page):
            print(f"主线程提交时发现页面已完成，跳过：{page.key}")
            continue
        committed_pages += 1
        materialize_category_records(state, records_path)
        written_count = len(
            {
                str(record.get("asin") or "")
                for record in page.records
                if isinstance(record, dict)
            }
        )
        filtered_out_count = page.extracted_count - written_count
        print(
            f"提取商品 {page.extracted_count} 条，写入 {written_count} 条，"
            f"过滤 {filtered_out_count} 条，插件状态：{page.plugin_status}"
        )

    committed_failures = 0
    for failure in batch.failures:
        if append_failure_record_once(failures_path, failure):
            committed_failures += 1
    if committed_failures:
        state.log_failure(committed_failures)

    if batch.children:
        added = state.enqueue_children(batch.children)
        if batch.skipped_intermediate:
            print(f"发现下级类目 {len(batch.children)} 个，新增 {added} 个；跳过当前中间节点。")
        else:
            print(f"当前节点已抓取，同时新增下级类目 {added} 个。")

    state.clear_manual_pause(claim_key)
    if not batch.terminal_error_type:
        state.complete_claimed_category(claim_key, batch.node)
    return committed_pages


def _raise_concurrent_terminal(batch: CategoryCrawlBatch) -> None:
    if batch.terminal_error_type == "delivery_location_unconfirmed":
        raise DeliveryLocationUnconfirmedError(batch.terminal_error_message)
    if batch.terminal_error_type == "verification_unconfirmed":
        raise VerificationUnconfirmedError(batch.terminal_error_message)
    if batch.terminal_error_type:
        raise UserFacingError(batch.terminal_error_message or batch.terminal_error_type)


def preflight_category_delivery(
    runtime: RuntimeConfig,
    state: StateStore,
    failures_path: Optional[Path] = None,
    domain_cooldowns: Optional[DomainCooldownRegistry] = None,
) -> None:
    if runtime.browser_tab_concurrency <= 1 or not runtime.delivery_location_enabled:
        return
    first_url_by_domain: Dict[str, str] = {}
    for node in state.data.get("queue") or []:
        if not isinstance(node, dict):
            continue
        url = str(node.get("url") or "")
        domain = (urlparse(url).hostname or "").lower()
        if domain and domain not in first_url_by_domain:
            first_url_by_domain[domain] = url
    if not first_url_by_domain:
        return

    print(f"并发预检：串行确认 {len(first_url_by_domain)} 个 Amazon 域名的配送地址。")
    driver = start_driver(runtime)
    try:
        for domain, url in first_url_by_domain.items():
            if hasattr(runtime, "amazon_page_retry_schedule"):
                work_key = f"category-preflight:{domain}|{clean_url(url)}"
                retry_key = _retry_entry_key(work_key, "category_preflight")
                try:
                    load_category_page_with_recovery(
                        driver,
                        url,
                        runtime,
                        work_key,
                        retry_callbacks=state_retry_callbacks(state, retry_key),
                        stage="category_preflight",
                        on_manual_pause=lambda reason, page_url: state.mark_manual_pause(
                            reason,
                            page_url,
                        ),
                        on_manual_resume=state.clear_manual_pause,
                        domain_cooldowns=domain_cooldowns,
                    )
                except AmazonPageRetryExhausted as exc:
                    if failures_path is not None:
                        preflight_node = {
                            "url": url,
                            "name": infer_category_name_from_url(url),
                            "path": [infer_category_name_from_url(url)],
                            "node_id": extract_node_id(url),
                        }
                        log_amazon_retry_exhausted_once(
                            failures_path,
                            state,
                            runtime,
                            preflight_node,
                            1,
                            url,
                            exc,
                        )
                    raise UserFacingError(str(exc)) from exc
            else:
                # Compatibility for lightweight callers/tests that provide a
                # pre-recovery runtime object rather than RuntimeConfig.
                open_amazon_page(
                    driver,
                    url,
                    runtime,
                    on_manual_pause=lambda reason, page_url: state.mark_manual_pause(
                        reason,
                        page_url,
                    ),
                    on_manual_resume=state.clear_manual_pause,
                )
            print(f"配送地址已确认：{domain}")
    finally:
        try:
            driver.quit()
        except WebDriverException:
            pass


def run_crawl_concurrent(
    runtime: RuntimeConfig,
    state: StateStore,
    records_path: Path,
    failures_path: Path,
    debug_dir: Path,
) -> None:
    """Run independent category sources concurrently with main-thread persistence."""

    state.prepare_concurrent_resume()
    events: "queue.Queue[Dict[str, Any]]" = queue.Queue()
    stop_event = threading.Event()
    domain_cooldowns = DomainCooldownRegistry(
        clock=_runtime_recovery_clock(runtime),
        waiter=_runtime_recovery_waiter(runtime, stop_event),
    )
    setattr(runtime, "_amazon_domain_cooldowns", domain_cooldowns)
    preflight_category_delivery(
        runtime,
        state,
        failures_path,
        domain_cooldowns,
    )
    batch_pause = BatchPauseScheduler(runtime)
    navigation_throttle = NavigationThrottle(
        runtime.delay_seconds_min,
        runtime.delay_seconds_max,
    )
    delivery_locks = DeliveryDomainLocks()
    executor = concurrent.futures.ThreadPoolExecutor(
        max_workers=runtime.browser_tab_concurrency,
        thread_name_prefix="amazon-category",
    )
    futures: Dict[
        concurrent.futures.Future[CategoryCrawlBatch],
        Tuple[str, Dict[str, Any]],
    ] = {}
    claimed_in_this_run = 0

    try:
        while True:
            _drain_worker_events(events, state)
            limit_reached = (
                runtime.max_categories is not None
                and claimed_in_this_run >= runtime.max_categories
            )
            manual_paused = bool(
                state.data.get("manual_pause") or state.data.get("manual_pauses")
            )
            while (
                len(futures) < runtime.browser_tab_concurrency
                and not limit_reached
                and not manual_paused
                and not stop_event.is_set()
            ):
                claim = state.claim_next_category()
                if claim is None:
                    break
                claim_key, node = claim
                completed_page_keys = set(state.data.get("completed_pages") or [])
                future = executor.submit(
                    crawl_category_source,
                    runtime,
                    claim_key,
                    node,
                    completed_page_keys,
                    debug_dir,
                    events,
                    stop_event,
                    navigation_throttle,
                    delivery_locks,
                )
                futures[future] = (claim_key, node)
                claimed_in_this_run += 1
                limit_reached = (
                    runtime.max_categories is not None
                    and claimed_in_this_run >= runtime.max_categories
                )

            _drain_worker_events(events, state)
            if not futures:
                if limit_reached:
                    print("达到 max_categories，本次停止；可继续 resume。")
                else:
                    print("队列已完成。")
                break

            done, _pending = concurrent.futures.wait(
                futures,
                timeout=0.5,
                return_when=concurrent.futures.FIRST_COMPLETED,
            )
            _drain_worker_events(events, state)
            if not done:
                continue
            for future in done:
                claim_key, _node = futures.pop(future)
                try:
                    batch = future.result()
                except ConcurrentWorkerCancelled:
                    state.requeue_claimed_category(claim_key)
                    continue
                committed_pages = _commit_category_batch(
                    batch,
                    claim_key,
                    runtime,
                    state,
                    records_path,
                    failures_path,
                )
                for _index in range(committed_pages):
                    batch_pause.after_completed_page()
                _raise_concurrent_terminal(batch)
    except BaseException:
        stop_event.set()
        for future in futures:
            future.cancel()
        executor.shutdown(wait=True, cancel_futures=True)
        _drain_worker_events(events, state)
        claim_keys = list((state.data.get("in_flight_categories") or {}).keys())
        for claim_key in claim_keys:
            state.clear_manual_pause(claim_key)
        state.requeue_claimed_categories(claim_keys)
        raise
    else:
        executor.shutdown(wait=True)
        _drain_worker_events(events, state)


def _run_crawl_unlocked(runtime: RuntimeConfig, dry_run: bool) -> int:
    job_dir = ensure_dir(runtime.outputs_root / runtime.job_id)
    records_path = job_dir / "records.jsonl"
    failures_path = job_dir / "failures.jsonl"
    state_path = job_dir / "state.json"
    output_xlsx = job_dir / f"total_{runtime.job_id}_merged.xlsx"
    debug_dir = job_dir / "debug_snapshots"

    print(f"任务目录：{job_dir}")
    print(f"起始类目：{runtime.start_url}")
    print(f"浏览器后端：{runtime.browser_backend}/{runtime.browser_mode}")
    print(f"浏览器标签并发：{runtime.browser_tab_concurrency}")
    print(f"卖家精灵门禁：{'required' if runtime.sellersprite_required else 'not_required'}")
    print(
        "产品过滤："
        + json.dumps(runtime.product_filters.as_dict(), ensure_ascii=False, sort_keys=True)
    )
    print(f"输出表格：{output_xlsx}")
    if dry_run:
        print("dry-run：配置检查完成，未打开浏览器。")
        return 0
    if not runtime.resume:
        for old_file in (records_path, failures_path, state_path, output_xlsx):
            if old_file.exists():
                old_file.unlink()
        page_results_dir = job_dir / "page_results"
        if page_results_dir.exists():
            shutil.rmtree(page_results_dir)

    state = StateStore(state_path, runtime)
    state.load_or_create()
    materialize_category_records(state, records_path)
    if runtime.browser_tab_concurrency > 1:
        run_crawl_concurrent(
            runtime,
            state,
            records_path,
            failures_path,
            debug_dir,
        )
        write_workbook(records_path, failures_path, output_xlsx)
        print(f"已生成 Excel：{output_xlsx}")
        return 0
    state.recover_stale_in_flight()
    driver = start_driver(runtime)
    batch_pause = BatchPauseScheduler(runtime)
    domain_cooldowns = DomainCooldownRegistry(
        clock=_runtime_recovery_clock(runtime),
        waiter=_runtime_recovery_waiter(runtime, None),
    )
    setattr(runtime, "_amazon_domain_cooldowns", domain_cooldowns)

    def restart_plugin_driver(current_driver: WebDriver, page_url: str, wait_seconds: float) -> WebDriver:
        nonlocal driver
        try:
            current_driver.quit()
        except Exception:
            pass
        if wait_seconds > 0:
            time.sleep(wait_seconds)
        driver = start_driver(runtime)
        load_category_page_attempt(
            driver,
            page_url,
            runtime,
            on_manual_pause=lambda reason, url: state.mark_manual_pause(reason, url),
            on_manual_resume=state.clear_manual_pause,
            domain_cooldowns=domain_cooldowns,
        )
        return driver

    def recover_plugin_page(
        current_driver: WebDriver,
        target_url: str,
    ) -> WebDriver:
        assessment = load_category_page_attempt(
            current_driver,
            target_url,
            runtime,
            on_manual_pause=lambda reason, url: state.mark_manual_pause(reason, url),
            on_manual_resume=state.clear_manual_pause,
            domain_cooldowns=domain_cooldowns,
        )
        if assessment.status is PageHealthStatus.VERIFIED_EMPTY:
            raise TransientAmazonPageUnavailable(
                "页面在插件刷新期间变为明确空结果；重新执行整页事务。",
                reason="verified_empty_during_plugin_refresh",
                url=safe_driver_current_url(current_driver, target_url),
            )
        return current_driver

    try:
        processed_in_this_run = 0
        while True:
            if runtime.max_categories is not None and processed_in_this_run >= runtime.max_categories:
                print("达到 max_categories，本次停止；可继续 resume。")
                break
            current = state.next_work()
            if not current:
                print("队列已完成。")
                break

            node = current["node"]
            page_number = int(current.get("page_number") or 1)
            page_url = str(current.get("page_url") or node["url"])
            print(f"处理类目：{' > '.join(node.get('path') or [])} / 第 {page_number} 页")

            try:
                work_key = page_key(node, page_number, page_url)
                retry_key = _retry_entry_key(work_key, "category_page_work")

                def current_driver() -> WebDriver:
                    if driver is None:
                        raise WebDriverException("类目浏览器尚未启动。")
                    return driver

                def process_page_attempt(attempt: Any) -> CategoryPageWorkResult:
                    nonlocal driver
                    if attempt.attempt_number > 1:
                        try:
                            current_driver().quit()
                        except Exception:
                            pass
                        driver = start_driver(runtime)

                    active_driver = current_driver()
                    page_assessment = load_category_page_attempt(
                        active_driver,
                        page_url,
                        runtime,
                        on_manual_pause=lambda reason, url: state.mark_manual_pause(
                            reason,
                            url,
                        ),
                        on_manual_resume=state.clear_manual_pause,
                        domain_cooldowns=domain_cooldowns,
                    )
                    attempt_node = copy.deepcopy(node)
                    page_path = extract_current_category_path(active_driver)
                    if page_path:
                        attempt_node["path"] = page_path
                        attempt_node["name"] = page_path[-1]
                        attempt_node["node_id"] = (
                            attempt_node.get("node_id")
                            or extract_node_id(
                                safe_driver_current_url(active_driver, page_url)
                            )
                        )

                    children: List[Dict[str, Any]] = []
                    if page_number == 1:
                        if current.get("children_enqueued"):
                            children = copy.deepcopy(current.get("children") or [])
                        else:
                            children = discover_child_categories(
                                active_driver,
                                attempt_node,
                                strict=True,
                            )
                        if (
                            should_descend(runtime, attempt_node, children)
                            and not runtime.include_root
                        ):
                            return CategoryPageWorkResult(
                                node=attempt_node,
                                children=children,
                                skipped_intermediate=True,
                            )

                    key = page_key(attempt_node, page_number, page_url)
                    page_batch: Optional[CategoryPageBatch] = None
                    if state.is_page_completed(key):
                        print("当前页已在断点中标记完成，跳过写入。")
                    elif page_assessment.status is PageHealthStatus.VERIFIED_EMPTY:
                        page_batch = CategoryPageBatch(
                            key=key,
                            page_number=page_number,
                            page_url=page_url,
                            plugin_status="verified_empty",
                            extracted_count=0,
                            records=[],
                            rejection_counts={},
                        )
                    else:
                        plugin_status = wait_for_sellersprite_data_or_prompt(
                            active_driver,
                            runtime,
                            on_manual_pause=lambda reason, url: state.mark_manual_pause(
                                reason,
                                url,
                            ),
                            on_manual_resume=state.clear_manual_pause,
                            restart_driver=restart_plugin_driver,
                            on_readiness=state.mark_sellersprite_readiness,
                            before_navigation=lambda: domain_cooldowns.wait(
                                (urlparse(page_url).hostname or "unknown").lower()
                            ),
                            recover_amazon_page=recover_plugin_page,
                        )
                        if plugin_status == "blocked":
                            raise VerificationUnconfirmedError(
                                "sellersprite_verification_unconfirmed: "
                                "人工处理超时，任务已停止且未提取当前页数据。"
                            )
                        active_driver = current_driver()
                        post_plugin_health = wait_for_category_page_health(
                            active_driver,
                            runtime,
                            lambda reason, url: state.mark_manual_pause(reason, url),
                            state.clear_manual_pause,
                        )
                        if post_plugin_health.status is PageHealthStatus.VERIFIED_EMPTY:
                            raise TransientAmazonPageUnavailable(
                                "页面在插件等待期间变为明确空结果；重新执行整页事务。",
                                reason="verified_empty_after_plugin_wait",
                                url=safe_driver_current_url(active_driver, page_url),
                            )
                        extracted_records = merge_product_data(
                            active_driver,
                            runtime,
                            attempt_node,
                            page_number,
                            plugin_status,
                        )
                        if not extracted_records:
                            raise TransientAmazonPageUnavailable(
                                "页面已检测到商品卡片，但提取结果为空；"
                                "为避免静默漏页，保留断点并重试。",
                                reason="empty_extraction_after_expected_content",
                                url=safe_driver_current_url(active_driver, page_url),
                            )
                        records, rejection_counts = filter_product_records(
                            extracted_records,
                            runtime.product_filters,
                        )
                        page_batch = CategoryPageBatch(
                            key=key,
                            page_number=page_number,
                            page_url=page_url,
                            plugin_status=plugin_status,
                            extracted_count=len(extracted_records),
                            records=records,
                            rejection_counts=rejection_counts,
                        )
                        if plugin_status != "ok" and runtime.save_debug_snapshots:
                            save_debug_snapshot(
                                active_driver,
                                debug_dir,
                                f"plugin_{plugin_status}_{extract_node_id(page_url)}_{page_number}",
                            )

                    active_driver = current_driver()
                    return CategoryPageWorkResult(
                        node=attempt_node,
                        next_url=find_next_page_url(active_driver, strict=True),
                        page=page_batch,
                        children=children,
                    )

                page_result = run_category_page_work_with_recovery(
                    runtime,
                    page_url,
                    work_key,
                    retry_callbacks=state_retry_callbacks(state, retry_key),
                    driver_provider=current_driver,
                    operation=process_page_attempt,
                    domain_cooldowns=domain_cooldowns,
                )

                node = page_result.node
                current["node"] = node
                children = page_result.children
                if page_number == 1 and not current.get("children_enqueued"):
                    current["children"] = children
                    current["children_enqueued"] = True
                    state.set_current(current)

                if page_result.skipped_intermediate:
                    added = state.enqueue_children(children)
                    print(f"发现下级类目 {len(children)} 个，新增 {added} 个；跳过当前中间节点。")
                    state.finish_current_category()
                    processed_in_this_run += 1
                    sleep_between_pages(runtime)
                    continue

                if page_result.page is not None:
                    state.commit_page_batch(page_result.page)
                    materialize_category_records(state, records_path)
                    if page_result.page.plugin_status == "verified_empty":
                        print("页面明确显示无商品，已合法提交空页断点。")
                    else:
                        written_count = len(page_result.page.records)
                        filtered_out_count = (
                            page_result.page.extracted_count - written_count
                        )
                        print(
                            f"提取商品 {page_result.page.extracted_count} 条，"
                            f"写入 {written_count} 条，过滤 {filtered_out_count} 条，"
                            f"插件状态：{page_result.page.plugin_status}"
                        )
                    batch_pause.after_completed_page()

                if page_number == 1 and children and runtime.include_root:
                    added = state.enqueue_children(children)
                    print(f"当前节点已抓取，同时新增下级类目 {added} 个。")

                next_url = page_result.next_url
                page_limit = runtime.max_pages_per_category
                if next_url and (page_limit is None or page_number < page_limit):
                    current["page_number"] = page_number + 1
                    current["page_url"] = next_url
                    state.set_current(current)
                else:
                    state.finish_current_category()
                    processed_in_this_run += 1
                sleep_between_pages(runtime)
            except AmazonPageRetryExhausted as exc:
                log_amazon_retry_exhausted_once(
                    failures_path,
                    state,
                    runtime,
                    node,
                    page_number,
                    page_url,
                    exc,
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(
                        driver,
                        debug_dir,
                        f"amazon_page_retry_exhausted_{extract_node_id(page_url)}_{page_number}",
                    )
                raise UserFacingError(str(exc)) from exc
            except DeliveryLocationUnconfirmedError as exc:
                log_failure(
                    failures_path,
                    state,
                    runtime,
                    node,
                    page_number,
                    page_url,
                    "delivery_location_unconfirmed",
                    str(exc),
                )
                raise
            except VerificationUnconfirmedError as exc:
                log_failure(
                    failures_path,
                    state,
                    runtime,
                    node,
                    page_number,
                    page_url,
                    "verification_timeout",
                    str(exc),
                )
                if runtime.save_debug_snapshots:
                    save_debug_snapshot(driver, debug_dir, "verification_timeout")
                raise
    finally:
        try:
            driver.quit()
        except WebDriverException:
            pass

    write_workbook(records_path, failures_path, output_xlsx)
    print(f"已生成 Excel：{output_xlsx}")
    return 0


def run_crawl(runtime: RuntimeConfig, dry_run: bool) -> int:
    if dry_run:
        return _run_crawl_unlocked(runtime, dry_run=True)
    job_dir = runtime.outputs_root / runtime.job_id
    with JobRunLock(job_dir / ".run.lock"):
        return _run_crawl_unlocked(runtime, dry_run=False)


def wait_for_page_or_manual(
    driver: WebDriver,
    runtime: RuntimeConfig,
    state: StateStore,
    failures_path: Path,
    debug_dir: Path,
    node: Dict[str, Any],
    page_number: int,
    page_url: str,
) -> None:
    block_reason = detect_block(driver)
    if block_reason:
        state.mark_manual_pause(block_reason, driver.current_url)
        cleared = wait_for_manual_clear(driver, block_reason, runtime.manual_pause_timeout)
        if cleared:
            state.clear_manual_pause()
        if not cleared:
            log_failure(failures_path, state, runtime, node, page_number, page_url, block_reason, "人工处理超时")
            if runtime.save_debug_snapshots:
                save_debug_snapshot(driver, debug_dir, block_reason)
            raise VerificationUnconfirmedError(
                verification_unconfirmed_message(block_reason)
            )
    wait_for_amazon_products(driver, runtime)


def main() -> int:
    parser = argparse.ArgumentParser(description="Amazon category ranking crawler with SellerSprite extension data.")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="配置文件路径")
    parser.add_argument("--dry-run", action="store_true", help="只检查配置，不打开浏览器")
    parser.add_argument("--no-resume", action="store_true", help="忽略已有断点，重新开始任务")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser()
    if not config_path.is_absolute():
        config_path = ROOT_DIR / config_path
    runtime = build_runtime_config(load_json(config_path), config_path, args.no_resume)
    return run_crawl(runtime, args.dry_run)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except UserFacingError as exc:
        print(f"运行失败：{exc}", file=sys.stderr)
        raise SystemExit(2)
